"""Palay trade in/out report downloads (Excel + PDF)."""

import io
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone
from django.utils.html import escape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from . import models

ZERO = Decimal("0.00")


def _money(value):
    return float(value or ZERO)


def _store_name():
    try:
        from admin_panel.models import StoreProfile

        profile = StoreProfile.get()
        if profile and profile.store_name:
            return profile.store_name
    except Exception:
        pass
    try:
        from admin_panel.models import KioskConfig

        cfg = KioskConfig.get()
        if cfg and cfg.system_name:
            return cfg.system_name
    except Exception:
        pass
    return "Cooperative"


def _date_slug(date_from, date_to):
    if date_from and date_to:
        return f"{date_from.isoformat()}_{date_to.isoformat()}"
    if date_from:
        return f"from_{date_from.isoformat()}"
    if date_to:
        return f"until_{date_to.isoformat()}"
    return timezone.localdate().isoformat()


def _scope_label(*, date_from, date_to, type_filter, search_query):
    parts = []
    if date_from and date_to:
        parts.append(f"{date_from:%b %d, %Y} — {date_to:%b %d, %Y}")
    elif date_from:
        parts.append(f"From {date_from:%b %d, %Y}")
    elif date_to:
        parts.append(f"Until {date_to:%b %d, %Y}")
    else:
        parts.append("All dates")
    if type_filter == "buy":
        parts.append("Buys (IN) only")
    elif type_filter == "sell":
        parts.append("Sells (OUT) only")
    else:
        parts.append("All in/out")
    if search_query:
        parts.append(f'Search: "{search_query}"')
    return " · ".join(parts)


def _trade_rows(trades):
    rows = []
    for trade in trades:
        direction = "IN" if trade.trade_type == models.PalayTrade.TradeType.BUY else "OUT"
        rows.append(
            {
                "reference": trade.reference,
                "traded_at": timezone.localtime(trade.traded_at),
                "direction": direction,
                "trade_type": trade.get_trade_type_display(),
                "party_name": trade.party_name,
                "member": (
                    trade.member.username or trade.member.full_name if trade.member_id else ""
                ),
                "product": trade.product.name,
                "product_code": trade.product.code,
                "grade": trade.product.get_grade_display(),
                "kg": trade.net_kg,
                "unit_price": trade.unit_price,
                "amount": trade.net_amount,
                "status": trade.get_status_display(),
                "posted_by": (
                    trade.performed_by.get_full_name() or trade.performed_by.username
                    if trade.performed_by_id
                    else ""
                ),
                "notes": trade.notes or "",
            }
        )
    return rows


def _summarize(rows):
    buy_kg = sell_kg = ZERO
    buy_amt = sell_amt = ZERO
    buy_n = sell_n = 0
    for row in rows:
        if row["direction"] == "IN":
            buy_n += 1
            buy_kg += row["kg"] or ZERO
            buy_amt += row["amount"] or ZERO
        else:
            sell_n += 1
            sell_kg += row["kg"] or ZERO
            sell_amt += row["amount"] or ZERO
    return {
        "buy_n": buy_n,
        "sell_n": sell_n,
        "buy_kg": buy_kg,
        "sell_kg": sell_kg,
        "buy_amt": buy_amt,
        "sell_amt": sell_amt,
        "net_kg": buy_kg - sell_kg,
        "net_amt": buy_amt + sell_amt,
        "total_n": buy_n + sell_n,
    }


def build_excel_response(*, trades, products, date_from, date_to, type_filter, search_query, user_label):
    rows = _trade_rows(trades)
    summary = _summarize(rows)
    scope = _scope_label(
        date_from=date_from,
        date_to=date_to,
        type_filter=type_filter,
        search_query=search_query,
    )
    store = _store_name()
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    header_fill = PatternFill("solid", fgColor="00A651")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14, color="0F172A")
    muted_font = Font(color="64748B", size=10)

    wb = Workbook()

    def write_trade_sheet(ws, title, sheet_rows):
        ws["A1"] = f"Palay Trade In/Out Report — {store}"
        ws["A1"].font = title_font
        ws["A2"] = title
        ws["A2"].font = Font(bold=True, size=12)
        ws["A3"] = f"Period / filter: {scope}"
        ws["A3"].font = muted_font
        ws["A4"] = (
            f"Generated: {timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')} — {user_label}"
        )
        ws["A4"].font = muted_font
        ws["A5"] = (
            f"Tickets: {len(sheet_rows)} | "
            f"IN kg: {_money(sum((r['kg'] for r in sheet_rows if r['direction'] == 'IN'), ZERO)):,.2f} | "
            f"OUT kg: {_money(sum((r['kg'] for r in sheet_rows if r['direction'] == 'OUT'), ZERO)):,.2f} | "
            f"Amount PHP: {_money(sum((r['amount'] for r in sheet_rows), ZERO)):,.2f}"
        )
        ws["A5"].font = muted_font

        headers = [
            "#",
            "Reference",
            "Date/Time",
            "In/Out",
            "Trade type",
            "Party",
            "Member",
            "Product",
            "Code",
            "Grade",
            "Kg",
            "Unit price",
            "Amount (PHP)",
            "Status",
            "Posted by",
            "Notes",
        ]
        start = 7
        for col, val in enumerate(headers, start=1):
            cell = ws.cell(row=start, column=col, value=val)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for idx, row in enumerate(sheet_rows, start=1):
            values = [
                idx,
                row["reference"],
                row["traded_at"].strftime("%Y-%m-%d %H:%M"),
                row["direction"],
                row["trade_type"],
                row["party_name"],
                row["member"] or "—",
                row["product"],
                row["product_code"],
                row["grade"],
                _money(row["kg"]),
                _money(row["unit_price"]),
                _money(row["amount"]),
                row["status"],
                row["posted_by"] or "—",
                row["notes"],
            ]
            r = start + idx
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = thin
                if col in (11, 12, 13):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                elif col in (1, 4):
                    cell.alignment = Alignment(horizontal="center")

        widths = [5, 18, 16, 8, 16, 22, 14, 22, 14, 12, 10, 12, 14, 10, 16, 28]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    ws_all = wb.active
    ws_all.title = "All Trades"
    write_trade_sheet(ws_all, "All palay trade tickets (IN buys + OUT sells)", rows)

    ws_in = wb.create_sheet("Buys IN")
    write_trade_sheet(
        ws_in,
        "Palay IN — buys from farmers (stock increases)",
        [r for r in rows if r["direction"] == "IN"],
    )

    ws_out = wb.create_sheet("Sells OUT")
    write_trade_sheet(
        ws_out,
        "Palay OUT — sells from stock (stock decreases)",
        [r for r in rows if r["direction"] == "OUT"],
    )

    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = f"Palay In/Out Summary — {store}"
    ws_sum["A1"].font = title_font
    ws_sum["A2"] = scope
    ws_sum["A2"].font = muted_font
    summary_rows = [
        ("Buy tickets (IN)", summary["buy_n"]),
        ("Buy kg (IN)", _money(summary["buy_kg"])),
        ("Buy amount PHP (IN)", _money(summary["buy_amt"])),
        ("Sell tickets (OUT)", summary["sell_n"]),
        ("Sell kg (OUT)", _money(summary["sell_kg"])),
        ("Sell amount PHP (OUT)", _money(summary["sell_amt"])),
        ("Net kg (IN − OUT)", _money(summary["net_kg"])),
        ("Total tickets", summary["total_n"]),
        ("Total amount PHP", _money(summary["net_amt"])),
    ]
    ws_sum["A4"] = "Metric"
    ws_sum["B4"] = "Value"
    ws_sum["A4"].fill = header_fill
    ws_sum["B4"].fill = header_fill
    ws_sum["A4"].font = header_font
    ws_sum["B4"].font = header_font
    for i, (label, value) in enumerate(summary_rows, start=5):
        ws_sum.cell(row=i, column=1, value=label).border = thin
        cell = ws_sum.cell(row=i, column=2, value=value)
        cell.border = thin
        if isinstance(value, float):
            cell.number_format = "#,##0.00"
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 16

    ws_stock = wb.create_sheet("Current Stock")
    ws_stock["A1"] = "Rice product stock snapshot"
    ws_stock["A1"].font = title_font
    stock_headers = [
        "Product",
        "Code",
        "Grade",
        "Stock kg",
        "Low alert kg",
        "Buy / kg",
        "Sell / kg",
        "Active",
    ]
    for col, val in enumerate(stock_headers, start=1):
        cell = ws_stock.cell(row=3, column=col, value=val)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for idx, product in enumerate(products, start=1):
        values = [
            product.name,
            product.code,
            product.get_grade_display(),
            _money(product.stock_kg),
            _money(product.low_stock_kg),
            _money(product.buy_price_per_kg),
            _money(product.sell_price_per_kg),
            "Yes" if product.is_active else "No",
        ]
        for col, val in enumerate(values, start=1):
            cell = ws_stock.cell(row=3 + idx, column=col, value=val)
            cell.border = thin
            if col in (4, 5, 6, 7):
                cell.number_format = "#,##0.00"
    for i, width in enumerate([28, 14, 12, 12, 12, 12, 12, 10], start=1):
        ws_stock.column_dimensions[get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"palay_inout_report_{_date_slug(date_from, date_to)}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_pdf_response(*, trades, date_from, date_to, type_filter, search_query, user_label):
    rows = _trade_rows(trades)
    summary = _summarize(rows)
    scope = _scope_label(
        date_from=date_from,
        date_to=date_to,
        type_filter=type_filter,
        search_query=search_query,
    )
    store = _store_name()

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PalayTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#ED1C24"),
        alignment=TA_CENTER,
        spaceAfter=8,
        fontName="Helvetica-Bold",
    )
    normal = ParagraphStyle("PalayNormal", parent=styles["Normal"], fontSize=9, leading=12)
    small = ParagraphStyle("PalaySmall", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#64748B"))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=28,
        leftMargin=28,
        topMargin=28,
        bottomMargin=28,
    )
    elements = [
        Paragraph("Palay Trade In/Out Report", title_style),
        Paragraph(escape(store), ParagraphStyle("Store", parent=normal, alignment=TA_CENTER)),
        Spacer(1, 0.08 * inch),
        Paragraph(f"Period / filter: {escape(scope)}", normal),
        Paragraph(
            f"Generated: {escape(timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M'))} — {escape(user_label)}",
            small,
        ),
        Spacer(1, 0.1 * inch),
        Paragraph(
            f"<b>IN:</b> {summary['buy_n']} tickets · {float(summary['buy_kg']):,.2f} kg · ₱{float(summary['buy_amt']):,.2f}"
            f"&nbsp;&nbsp;|&nbsp;&nbsp;<b>OUT:</b> {summary['sell_n']} tickets · {float(summary['sell_kg']):,.2f} kg · ₱{float(summary['sell_amt']):,.2f}"
            f"&nbsp;&nbsp;|&nbsp;&nbsp;<b>Net kg:</b> {float(summary['net_kg']):,.2f}",
            normal,
        ),
        Spacer(1, 0.14 * inch),
    ]

    table_data = [["#", "Ref", "Date", "I/O", "Party", "Product", "Kg", "₱/kg", "Amount", "Posted by"]]
    for idx, row in enumerate(rows, start=1):
        table_data.append(
            [
                str(idx),
                Paragraph(escape(row["reference"]), small),
                row["traded_at"].strftime("%m/%d %H:%M"),
                row["direction"],
                Paragraph(escape(row["party_name"][:28]), small),
                Paragraph(escape(row["product"][:24]), small),
                f"{float(row['kg']):,.2f}",
                f"{float(row['unit_price']):,.2f}",
                f"{float(row['amount']):,.2f}",
                Paragraph(escape((row["posted_by"] or "—")[:18]), small),
            ]
        )

    if len(table_data) == 1:
        table_data.append(["—", "No trades in this filter", "", "", "", "", "", "", "", ""])

    table = Table(
        table_data,
        colWidths=[0.35 * inch, 1.15 * inch, 0.85 * inch, 0.4 * inch, 1.4 * inch, 1.35 * inch, 0.7 * inch, 0.7 * inch, 0.85 * inch, 1.1 * inch],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ED1C24")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (3, 0), (3, -1), "CENTER"),
                ("ALIGN", (6, 1), (8, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 0.15 * inch))
    elements.append(
        Paragraph(
            "This report lists palay trade import (IN / buy) and export (OUT / sell) details for transparency and stock monitoring.",
            small,
        )
    )

    doc.build(elements)
    buffer.seek(0)
    filename = f"palay_inout_report_{_date_slug(date_from, date_to)}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
