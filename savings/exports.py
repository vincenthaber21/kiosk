"""Excel / PDF exports for savings interest reports."""

from __future__ import annotations

import io
from datetime import date, datetime, time, timedelta
from decimal import Decimal

from django.db.models import Count, Prefetch, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from . import models, services
from .policy import ANNUAL_INTEREST_RATE, format_rate, interest_amount


ZERO = Decimal("0.00")


def parse_export_dates(request):
    """Return (date_from, date_to, range_start_aware, range_end_aware)."""
    today = timezone.localdate()
    raw_from = (request.GET.get("date_from") or "").strip()
    raw_to = (request.GET.get("date_to") or "").strip()
    try:
        date_from = date.fromisoformat(raw_from)
    except ValueError:
        date_from = date(today.year, today.month, 1)
    try:
        date_to = date.fromisoformat(raw_to)
    except ValueError:
        date_to = today
    if date_to < date_from:
        date_from, date_to = date_to, date_from

    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.combine(date_from, time.min), tz)
    end = timezone.make_aware(datetime.combine(date_to + timedelta(days=1), time.min), tz)
    return date_from, date_to, start, end


def _store_name():
    try:
        from admin_panel.models import KioskConfig

        cfg = KioskConfig.get()
        if cfg and cfg.system_name:
            return cfg.system_name
    except Exception:
        pass
    return "Cooperative"


def build_interest_report_rows(*, date_from, date_to, range_start, range_end):
    """One row per savings account with interest totals for the selected period."""
    interest_qs = models.SavingsTransaction.objects.filter(
        transaction_type=models.SavingsTransaction.TxnType.INTEREST,
        created_at__gte=range_start,
        created_at__lt=range_end,
    ).order_by("created_at")

    accounts = (
        models.MemberSavingsAccount.objects.select_related("member", "product")
        .prefetch_related(Prefetch("transactions", queryset=interest_qs, to_attr="period_interest"))
        .annotate(
            period_interest_total=Coalesce(
                Sum(
                    "transactions__amount",
                    filter=Q(
                        transactions__transaction_type=models.SavingsTransaction.TxnType.INTEREST,
                        transactions__created_at__gte=range_start,
                        transactions__created_at__lt=range_end,
                    ),
                ),
                ZERO,
            ),
            period_interest_count=Count(
                "transactions",
                filter=Q(
                    transactions__transaction_type=models.SavingsTransaction.TxnType.INTEREST,
                    transactions__created_at__gte=range_start,
                    transactions__created_at__lt=range_end,
                ),
            ),
        )
        .order_by("member__last_name", "member__first_name", "account_number")
    )

    rate = ANNUAL_INTEREST_RATE
    rate_display = format_rate(rate)
    rows = []
    for account in accounts:
        snap = services.interest_snapshot(account)
        period_total = account.period_interest_total or ZERO
        estimated = interest_amount(account.balance, rate) if account.can_transact else ZERO
        last_in_period = None
        if getattr(account, "period_interest", None):
            last_in_period = account.period_interest[-1].created_at
        rows.append(
            {
                "member": account.member.full_name if account.member_id else "—",
                "username": (
                    (account.member.username or account.member.email or "—")
                    if account.member_id
                    else "—"
                ),
                "account_number": account.account_number,
                "product": account.product.name if account.product_id else "—",
                "status": account.get_status_display(),
                "opened_at": account.opened_at,
                "balance": account.balance or ZERO,
                "annual_rate": rate,
                "annual_rate_display": rate_display,
                "period_interest": period_total,
                "period_interest_count": account.period_interest_count or 0,
                "estimated_monthly": estimated,
                "last_interest_at": snap.get("last_interest_at") or last_in_period,
                "next_credit_on": snap.get("next_credit_on"),
            }
        )
    return rows


def _excel_styles():
    header_fill = PatternFill("solid", fgColor="166534")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    return header_fill, header_font, thin


def render_interest_excel(*, rows, date_from, date_to, user_label):
    header_fill, header_font, thin = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Interest report"

    store = _store_name()
    total_interest = sum((r["period_interest"] for r in rows), ZERO)
    total_balance = sum((r["balance"] for r in rows), ZERO)
    gen_at = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")

    ws["A1"] = f"Savings Interest Report — {store}"
    ws["A1"].font = Font(size=14, bold=True, color="166534")
    ws["A2"] = f"Period: {date_from.isoformat()} to {date_to.isoformat()}"
    ws["A3"] = f"Generated: {gen_at} — {user_label}"
    ws["A4"] = f"Annual rate: {format_rate(ANNUAL_INTEREST_RATE)} (monthly = balance × rate ÷ 12)"
    ws["A5"] = (
        f"Accounts: {len(rows)} · "
        f"Total balances (PHP): {float(total_balance):,.2f} · "
        f"Interest credited in period (PHP): {float(total_interest):,.2f}"
    )

    hdr_row = 7
    headers = [
        "#",
        "Member",
        "Username / Email",
        "Account #",
        "Product",
        "Status",
        "Opened",
        "Balance (PHP)",
        "Annual rate",
        "Est. monthly interest (PHP)",
        "Interest credited (period)",
        "Credits (#)",
        "Last interest",
        "Next interest",
    ]
    for col, val in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=col, value=val)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    money_cols = {8, 10, 11}
    for idx, row in enumerate(rows, start=1):
        values = [
            idx,
            row["member"],
            row["username"],
            row["account_number"],
            row["product"],
            row["status"],
            timezone.localtime(row["opened_at"]).strftime("%Y-%m-%d") if row["opened_at"] else "",
            float(row["balance"]),
            row["annual_rate_display"],
            float(row["estimated_monthly"]),
            float(row["period_interest"]),
            row["period_interest_count"],
            (
                timezone.localtime(row["last_interest_at"]).strftime("%Y-%m-%d")
                if row["last_interest_at"]
                else ""
            ),
            (
                timezone.localtime(row["next_credit_on"]).strftime("%Y-%m-%d")
                if row["next_credit_on"]
                else ""
            ),
        ]
        excel_row = hdr_row + idx
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.border = thin
            if col in money_cols:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col in (1, 12):
                cell.alignment = Alignment(horizontal="center")

    total_row = hdr_row + len(rows) + 1
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    bal_cell = ws.cell(row=total_row, column=8, value=float(total_balance))
    bal_cell.number_format = "#,##0.00"
    bal_cell.font = Font(bold=True)
    int_cell = ws.cell(row=total_row, column=11, value=float(total_interest))
    int_cell.number_format = "#,##0.00"
    int_cell.font = Font(bold=True)
    for col in range(1, 15):
        ws.cell(row=total_row, column=col).border = thin

    widths = [5, 26, 22, 16, 18, 12, 12, 14, 12, 16, 16, 10, 12, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[hdr_row].height = 36

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"savings_interest_{date_from.isoformat()}_to_{date_to.isoformat()}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def render_interest_pdf(*, rows, date_from, date_to, user_label):
    store = _store_name()
    total_interest = sum((r["period_interest"] for r in rows), ZERO)
    total_balance = sum((r["balance"] for r in rows), ZERO)
    gen_at = timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "SavingsInterestTitle",
        parent=styles["Heading1"],
        fontSize=14,
        textColor=colors.HexColor("#166534"),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
    )
    meta_style = ParagraphStyle(
        "SavingsInterestMeta",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#475569"),
    )
    cell_style = ParagraphStyle(
        "SavingsInterestCell",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
    )
    cell_right = ParagraphStyle(
        "SavingsInterestCellRight",
        parent=cell_style,
        alignment=TA_RIGHT,
    )
    cell_center = ParagraphStyle(
        "SavingsInterestCellCenter",
        parent=cell_style,
        alignment=TA_CENTER,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=28,
        rightMargin=28,
        topMargin=28,
        bottomMargin=28,
    )

    story = [
        Paragraph(f"Savings Interest Report — {store}", title_style),
        Paragraph(
            f"Period: {date_from.isoformat()} to {date_to.isoformat()} · "
            f"Rate: {format_rate(ANNUAL_INTEREST_RATE)} / year · "
            f"Generated: {gen_at} — {user_label}",
            meta_style,
        ),
        Paragraph(
            f"Accounts: {len(rows)} · "
            f"Total balances: ₱{total_balance:,.2f} · "
            f"Interest credited in period: ₱{total_interest:,.2f}",
            meta_style,
        ),
        Spacer(1, 10),
    ]

    header = [
        Paragraph("<b>#</b>", cell_center),
        Paragraph("<b>Member</b>", cell_style),
        Paragraph("<b>Account</b>", cell_style),
        Paragraph("<b>Status</b>", cell_center),
        Paragraph("<b>Opened</b>", cell_center),
        Paragraph("<b>Balance</b>", cell_right),
        Paragraph("<b>Rate</b>", cell_center),
        Paragraph("<b>Est. monthly</b>", cell_right),
        Paragraph("<b>Interest (period)</b>", cell_right),
        Paragraph("<b>Next interest</b>", cell_center),
    ]
    data = [header]
    for idx, row in enumerate(rows, start=1):
        opened = (
            timezone.localtime(row["opened_at"]).strftime("%Y-%m-%d")
            if row["opened_at"]
            else "—"
        )
        next_on = (
            timezone.localtime(row["next_credit_on"]).strftime("%Y-%m-%d")
            if row["next_credit_on"]
            else "—"
        )
        data.append(
            [
                Paragraph(str(idx), cell_center),
                Paragraph(row["member"], cell_style),
                Paragraph(row["account_number"], cell_style),
                Paragraph(row["status"], cell_center),
                Paragraph(opened, cell_center),
                Paragraph(f"{row['balance']:,.2f}", cell_right),
                Paragraph(row["annual_rate_display"], cell_center),
                Paragraph(f"{row['estimated_monthly']:,.2f}", cell_right),
                Paragraph(f"{row['period_interest']:,.2f}", cell_right),
                Paragraph(next_on, cell_center),
            ]
        )

    data.append(
        [
            "",
            Paragraph("<b>TOTAL</b>", cell_style),
            "",
            "",
            "",
            Paragraph(f"<b>{total_balance:,.2f}</b>", cell_right),
            "",
            "",
            Paragraph(f"<b>{total_interest:,.2f}</b>", cell_right),
            "",
        ]
    )

    table = Table(
        data,
        colWidths=[28, 120, 85, 55, 62, 70, 42, 70, 78, 70],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#166534")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    filename = f"savings_interest_{date_from.isoformat()}_to_{date_to.isoformat()}.pdf"
    resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
