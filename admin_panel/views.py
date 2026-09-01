import json
import io
import csv
import os
import re
import secrets
from collections import defaultdict
from datetime import timedelta, datetime, date as date_type
from decimal import Decimal, InvalidOperation

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from functools import wraps
from django.db.models import Sum, Count, Avg, Q, F, Max, Value, DecimalField, Case, When, ExpressionWrapper, OuterRef, Subquery, Exists
from django.db.models.functions import Coalesce
from django.db import transaction as db_transaction
from django.db.models.functions import TruncDate, TruncWeek, TruncMonth
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.http import urlencode
from django.utils.html import escape
from django.views.decorators.http import require_http_methods
from django.conf import settings
from django.template.loader import render_to_string

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo

import barcode as barcode_lib
from barcode.writer import ImageWriter

from inventory.models import (
    GiveawayProduct,
    Product,
    Category,
    ProductDiscountGroup,
    ProductSaleUnit,
    ProductStockBatch,
    ProductStockHistory,
    StockTransaction,
)
from inventory.stock_history import (
    build_product_history_ledger_rows,
    capture_stock_snapshot,
    format_period_dt,
    record_stock_history,
)
from inventory.utils import (
    annotate_giveaway_units_given,
    get_giveaway_summary_for_period,
    giveaway_stock_note,
    giveaway_stock_transactions_qs,
    is_giveaway_stock_transaction,
    parse_giveaway_stock_notes,
    request_can_manage_giveaways,
    send_inventory_stock_alerts,
    serialize_giveaway_stock_transaction,
    set_product_giveaway,
)
from inventory.pricing import discounts_by_product_ids, unit_price_after_discounts, promote_new_stock_to_old_if_needed
from inventory.units import (
    UNIT_KILO,
    UNIT_PIECE,
    parse_sale_qty,
    parse_stock_qty,
    qty_json,
    retail_unit_label,
)
from django.core.exceptions import ValidationError

from members.models import (
    Member,
    MemberType,
    Role,
    BalanceTransaction,
    ShareCapitalTransaction,
    DeletedMember,
    CardBalanceRefill,
    MemberEditHistory,
    SeniorCitizenProfile,
    PWDProfile,
    SegmentProductGroupDiscount,
)
from members.utils import mask_rfid
from transactions.models import Transaction, TransactionItem, RefundReason, WalkInCustomer, CreditPayment, CreditPaymentLine
from admin_panel.models import StoreProfile, KioskConfig, CreditSettings, WebsiteAuditLog
from admin_panel.report_customers import (
    append_top_customers_pdf,
    get_top_customers_for_period,
    top_customers_excel_rows,
)
from admin_panel.report_product_summary import (
    append_product_sales_summary_pdf,
    get_product_sales_summary_for_period,
    product_sales_summary_excel_rows,
)
from admin_panel.report_utils import get_report_sale_items, get_report_transactions
from admin_panel.report_walk_in import (
    append_walk_in_summary_pdf,
    walk_in_excel_customer_rows,
    walk_in_excel_metric_rows,
)
from admin_panel.report_wholesale import (
    append_wholesale_sales_pdf,
    get_wholesale_sales_for_period,
    wholesale_excel_metric_rows,
    wholesale_excel_product_rows,
)
from admin_panel.utils import get_admin_email
from transactions.walk_in_customers import get_walk_in_summary_for_period, normalize_customer_name
from django.core.mail import EmailMessage
from mobile_api.email_utils import (
    send_refund_approval_email,
    send_welcome_member_email,
    send_credit_payment_receipt_email,
)
from helper.cookie_helper import set_secure_cookie
from inventory_helper import StockManager
from login_helper import (
    is_admin_user,
    is_cashier_or_admin,
    is_committee_only_user,
    is_loan_desk_user,
    is_loan_officer_only_user,
    is_loans_only_user,
    is_staff_role,
    restricts_member_role_to_member_only,
    can_access_django_admin,
    get_user_role,
    member_or_login_required,
    admin_required,
    get_session_member,
    is_member_session_valid,
    store_member_session,
    clear_member_session,
    login_with_credentials,
    login_with_rfid,
    validate_rfid,
    rfid_login_json_response,
    rfid_validate_json_response,
    resolve_redirect_url,
    logout_user,
)


DASHBOARD_TOP_PRODUCTS_FETCH_MAX = 100


def _store_local_today():
    """Calendar date in the active store timezone (dashboard period filters)."""
    return timezone.localdate()


def _store_local_today():
    """Calendar date in the active store timezone (dashboard period filters)."""
    return timezone.localdate()


# Completed sales plus partially refunded sales (header total_amount is net of refunded lines).
DASHBOARD_SALE_STATUSES = ('completed', 'partially_refunded')

# Manage Transactions → Void Item (creates a refund request for one active line).
VOID_ITEM_ALLOWED_STATUSES = ('completed', 'partially_refunded')


def _transaction_can_void_item(transaction):
    """True when at least one non-refunded line can still be voided."""
    return (
        transaction.status in VOID_ITEM_ALLOWED_STATUSES
        and transaction.items.filter(refunded_at__isnull=True).exists()
    )


def _exclude_test_transactions(qs):
    """Drop auto-generated Top Members dummy sales from business-facing totals."""
    return qs.exclude(transaction_number__startswith='DUMMY-')


def _transaction_net_revenue(qs):
    """Net sales from non-refunded lines; header fallback when a sale has no lines."""
    sale_qs = qs.filter(status__in=DASHBOARD_SALE_STATUSES)
    from_line_items = TransactionItem.objects.filter(
        transaction__in=sale_qs,
        refunded_at__isnull=True,
    ).aggregate(
        total=Coalesce(Sum('total_price'), Decimal('0.00')),
    )['total']
    itemless_header = sale_qs.annotate(_line_count=Count('items')).filter(_line_count=0).aggregate(
        total=Coalesce(Sum('total_amount'), Decimal('0.00')),
    )['total']
    return (from_line_items or Decimal('0.00')) + (itemless_header or Decimal('0.00'))


def _list_price_qty_sold_subtotal(
    *,
    product_ids=None,
    range_start_aware=None,
    range_end_aware=None,
):
    """
    Subtotal = current product selling price × qty sold.

    Same basis as inventory Price Summary Subtotal: completed sales only,
    non-refunded line items, priced at the product's current list selling price.
    """
    sold_qs = TransactionItem.objects.filter(
        transaction__status='completed',
        refunded_at__isnull=True,
    )
    if product_ids is not None:
        sold_qs = sold_qs.filter(product_id__in=product_ids)
    if range_start_aware is not None:
        sold_qs = sold_qs.filter(transaction__created_at__gte=range_start_aware)
    if range_end_aware is not None:
        sold_qs = sold_qs.filter(transaction__created_at__lt=range_end_aware)

    sold_rows = list(
        sold_qs.values('product_id').annotate(
            qty_sold=Coalesce(
                Sum('quantity'),
                _zero_qty(),
                output_field=DecimalField(max_digits=14, decimal_places=3),
            )
        )
    )
    product_id_list = [row['product_id'] for row in sold_rows if row['product_id']]
    if not product_id_list:
        return Decimal('0.00')

    price_map = dict(
        Product.objects.filter(id__in=product_id_list).values_list('id', 'price')
    )
    total = Decimal('0.00')
    for sold in sold_rows:
        qty = max(_as_qty(sold['qty_sold']), Decimal('0'))
        if qty <= 0:
            continue
        unit_price = Decimal(price_map.get(sold['product_id']) or 0)
        total += qty * unit_price
    return total.quantize(Decimal('0.01'))


def _list_price_qty_sold_series(range_start_aware, range_end_aware, current_tz, granularity):
    """
    Bucket selling-price × qty-sold totals by day or month for dashboard charts.
    Returns a dict keyed by local date (day) or month-start date.
    """
    items = list(
        TransactionItem.objects.filter(
            transaction__status='completed',
            refunded_at__isnull=True,
            transaction__created_at__gte=range_start_aware,
            transaction__created_at__lt=range_end_aware,
            product_id__isnull=False,
        ).values('transaction__created_at', 'product_id', 'quantity')
    )
    product_ids = {row['product_id'] for row in items if row['product_id']}
    price_map = dict(
        Product.objects.filter(id__in=product_ids).values_list('id', 'price')
    ) if product_ids else {}

    series = defaultdict(float)
    for row in items:
        try:
            qty = float(row['quantity'] or 0)
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue
        unit_price = float(price_map.get(row['product_id']) or 0)
        local_dt = row['transaction__created_at'].astimezone(current_tz)
        if granularity == 'monthly':
            key = local_dt.date().replace(day=1)
        else:
            key = local_dt.date()
        series[key] += qty * unit_price
    return series


def _dashboard_net_revenue_series(range_start_aware, range_end_aware, current_tz, granularity):
    """Bucket actual net sales (non-refunded line totals) by day or month."""
    items = list(
        TransactionItem.objects.filter(
            transaction__status__in=DASHBOARD_SALE_STATUSES,
            refunded_at__isnull=True,
            transaction__created_at__gte=range_start_aware,
            transaction__created_at__lt=range_end_aware,
        )
        .exclude(transaction__transaction_number__startswith='DUMMY-')
        .values('transaction__created_at', 'total_price')
    )

    series = defaultdict(float)
    for row in items:
        amount = float(row['total_price'] or 0)
        if amount <= 0:
            continue
        local_dt = row['transaction__created_at'].astimezone(current_tz)
        if granularity == 'monthly':
            key = local_dt.date().replace(day=1)
        else:
            key = local_dt.date()
        series[key] += amount

    # Header fallback for itemless sales in range
    itemless_txns = (
        _dashboard_sales_qs()
        .filter(
            created_at__gte=range_start_aware,
            created_at__lt=range_end_aware,
        )
        .annotate(_active_lines=Count('items', filter=Q(items__refunded_at__isnull=True)))
        .filter(_active_lines=0)
        .values('created_at', 'total_amount')
    )
    for row in itemless_txns:
        amount = float(row['total_amount'] or 0)
        if amount <= 0:
            continue
        local_dt = row['created_at'].astimezone(current_tz)
        if granularity == 'monthly':
            key = local_dt.date().replace(day=1)
        else:
            key = local_dt.date()
        series[key] += amount
    return series


def _dashboard_sales_qs():
    return _exclude_test_transactions(
        Transaction.objects.filter(status__in=DASHBOARD_SALE_STATUSES)
    )


def _dashboard_refund_ledger_qs():
    """Balance ledger rows that credit members for sale refunds."""
    return BalanceTransaction.objects.filter(
        transaction_type='deposit',
        notes__icontains='Refund',
    )


def _dashboard_refund_item_qs():
    """Refunded sale line items — source of truth for all payment types."""
    return TransactionItem.objects.filter(
        refunded_at__isnull=False,
    ).exclude(transaction__transaction_number__startswith='DUMMY-')


def _dashboard_refund_events_qs():
    """One row per refund batch (transaction + refunded_at) with line-total amount."""
    return (
        _dashboard_refund_item_qs()
        .values('transaction_id', 'refunded_at')
        .annotate(refund_amount=Coalesce(Sum('total_price'), Decimal('0.00')))
        .order_by('-refunded_at')
    )


def _dashboard_recent_refunds(range_start_aware, range_end_aware, current_tz, limit=10):
    """Latest refund batches in the selected chart period (members, walk-ins, guests)."""
    rows = list(
        _dashboard_refund_events_qs().filter(
            refunded_at__gte=range_start_aware,
            refunded_at__lt=range_end_aware,
        )[:limit]
    )
    if not rows:
        return []

    txn_map = {
        t.id: t
        for t in Transaction.objects.filter(
            id__in={row['transaction_id'] for row in rows}
        ).select_related('member', 'walk_in_customer')
    }

    recent_refunds = []
    for row in rows:
        txn = txn_map.get(row['transaction_id'])
        if not txn:
            continue
        is_partially_refunded = txn.status == 'partially_refunded'
        recent_refunds.append({
            'transaction_number': txn.transaction_number,
            'member_name': txn.customer_display_name,
            'refunded_at': timezone.localtime(row['refunded_at'], current_tz).strftime('%b %d, %Y %H:%M'),
            'refund_amount': float(row['refund_amount'] or 0),
            'is_partially_refunded': is_partially_refunded,
            'remaining_balance': float(txn.total_amount or 0) if is_partially_refunded else None,
        })
    return recent_refunds


def _dashboard_refund_stats(
    range_start_aware,
    range_end_aware,
    current_tz,
    range_granularity,
    range_start,
    range_end,
    range_days,
):
    """Refund counts/amounts from refunded line items; trend buckets use refunded_at."""
    all_events = _dashboard_refund_events_qs()
    period_events = all_events.filter(
        refunded_at__gte=range_start_aware,
        refunded_at__lt=range_end_aware,
    )

    total_refunds = all_events.count()
    total_refund_amount = float(
        _dashboard_refund_item_qs().aggregate(
            total=Coalesce(Sum('total_price'), Decimal('0.00')),
        )['total'] or 0
    )
    period_refunds = period_events.count()
    period_refund_amount = float(
        _dashboard_refund_item_qs()
        .filter(
            refunded_at__gte=range_start_aware,
            refunded_at__lt=range_end_aware,
        )
        .aggregate(total=Coalesce(Sum('total_price'), Decimal('0.00')))['total'] or 0
    )

    recent_refunds = _dashboard_recent_refunds(
        range_start_aware, range_end_aware, current_tz, limit=10
    )

    if range_granularity == 'monthly':
        refunds_in_range = period_events.values('refunded_at', 'refund_amount')
        monthly_refunds_map = defaultdict(lambda: {'amount': 0.0, 'count': 0})
        for row in refunds_in_range:
            local_dt = row['refunded_at'].astimezone(current_tz)
            month_key = local_dt.date().replace(day=1)
            monthly_refunds_map[month_key]['amount'] += float(row['refund_amount'] or 0)
            monthly_refunds_map[month_key]['count'] += 1
        daily_refund_labels = []
        daily_refund_amounts = []
        daily_refund_counts = []
        cur_month = range_start.replace(day=1)
        while cur_month <= range_end:
            daily_refund_labels.append(cur_month.strftime('%b %Y'))
            refund_data = monthly_refunds_map.get(cur_month, {'amount': 0, 'count': 0})
            daily_refund_amounts.append(round(refund_data['amount'], 2))
            daily_refund_counts.append(refund_data['count'])
            if cur_month.month == 12:
                cur_month = cur_month.replace(year=cur_month.year + 1, month=1)
            else:
                cur_month = cur_month.replace(month=cur_month.month + 1)
    else:
        refunds_in_range = period_events.values('refunded_at', 'refund_amount')
        daily_refunds_map = defaultdict(lambda: {'amount': 0.0, 'count': 0})
        for row in refunds_in_range:
            local_dt = row['refunded_at'].astimezone(current_tz)
            day_key = local_dt.date()
            daily_refunds_map[day_key]['amount'] += float(row['refund_amount'] or 0)
            daily_refunds_map[day_key]['count'] += 1
        daily_refund_labels = []
        daily_refund_amounts = []
        daily_refund_counts = []
        for offset in range(range_days):
            day = range_start + timedelta(days=offset)
            daily_refund_labels.append(day.strftime('%b %d'))
            refund_data = daily_refunds_map.get(day, {'amount': 0, 'count': 0})
            daily_refund_amounts.append(round(refund_data['amount'], 2))
            daily_refund_counts.append(refund_data['count'])

    return {
        'total_refunds': total_refunds,
        'total_refund_amount': total_refund_amount,
        'period_refunds': period_refunds,
        'period_refund_amount': period_refund_amount,
        'recent_refunds': recent_refunds,
        'daily_refund_labels': daily_refund_labels,
        'daily_refund_amounts': daily_refund_amounts,
        'daily_refund_counts': daily_refund_counts,
    }


def _parse_refund_txn_number_from_notes(notes):
    match = re.search(r'transaction\s+([A-Z0-9-]+)', notes or '', re.IGNORECASE)
    return match.group(1) if match else None


def _dashboard_recent_transactions(range_start_aware, range_end_aware, current_tz, limit=12):
    """Sales and refund events in the selected period, merged newest-first."""
    payment_label_map = dict(Transaction.PAYMENT_METHODS)
    entries = []

    sales_qs = (
        _dashboard_sales_qs()
        .filter(created_at__gte=range_start_aware, created_at__lt=range_end_aware)
        .select_related('member', 'walk_in_customer')
    )
    for txn in sales_qs:
        entries.append({
            'kind': 'sale',
            'transaction_number': txn.transaction_number,
            'customer_name': txn.customer_display_name,
            'is_walk_in': bool(txn.walk_in_customer_id or (txn.guest_customer_name or '').strip()),
            'payment_method': payment_label_map.get(txn.payment_method, (txn.payment_method or '').title()),
            'event_at': txn.created_at,
            'amount': float(txn.total_amount or 0),
        })

    refund_qs = (
        _dashboard_refund_events_qs()
        .filter(
            refunded_at__gte=range_start_aware,
            refunded_at__lt=range_end_aware,
        )
    )
    refund_txn_ids = {row['transaction_id'] for row in refund_qs}
    refund_txns = {
        t.id: t
        for t in Transaction.objects.filter(id__in=refund_txn_ids).select_related(
            'member', 'walk_in_customer'
        )
    }
    for row in refund_qs:
        txn = refund_txns.get(row['transaction_id'])
        if not txn:
            continue
        entries.append({
            'kind': 'refund',
            'transaction_number': txn.transaction_number,
            'customer_name': txn.customer_display_name,
            'is_walk_in': bool(txn.walk_in_customer_id or (txn.guest_customer_name or '').strip()),
            'payment_method': 'Refund',
            'event_at': row['refunded_at'],
            'amount': float(row['refund_amount'] or 0),
        })

    entries.sort(key=lambda e: e['event_at'], reverse=True)
    rows = []
    for entry in entries[:limit]:
        rows.append({
            'kind': entry['kind'],
            'transaction_number': entry['transaction_number'],
            'customer_name': entry['customer_name'],
            'is_walk_in': entry['is_walk_in'],
            'payment_method': entry['payment_method'],
            'created_at': timezone.localtime(entry['event_at'], current_tz).strftime('%b %d, %Y %H:%M'),
            'amount': entry['amount'],
        })
    return rows


def _dashboard_top_products(range_start_aware, range_end_aware, limit=DASHBOARD_TOP_PRODUCTS_FETCH_MAX):
    """Top products by units sold; frontend trims to visible grid capacity."""
    rows = (
        TransactionItem.objects.filter(
            transaction__status__in=DASHBOARD_SALE_STATUSES,
            transaction__created_at__gte=range_start_aware,
            transaction__created_at__lt=range_end_aware,
            refunded_at__isnull=True,
        )
        .values('product_name')
        .annotate(
            total_sold=Sum('quantity'),
            total_revenue=Sum('total_price'),
        )
        .order_by('-total_sold')
    )
    if limit:
        rows = rows[:limit]
    return [
        {
            'product_name': (row['product_name'] or '').strip() or 'Unknown',
            'total_sold': int(row['total_sold'] or 0),
            'total_revenue': float(row['total_revenue'] or 0),
        }
        for row in rows
    ]


def _dashboard_walk_in_insights(period_completed_txns, range_start_aware, range_end_aware, current_tz):
    """Walk-in customer registry metrics for the dashboard (from kiosk / admin registry)."""
    # Prefer linked WalkInCustomer rows; also count guest-name-only sales (legacy / admin).
    walk_in_filter = Q(member__isnull=True) & (
        Q(walk_in_customer__isnull=False) | ~Q(guest_customer_name='')
    )
    period_walk_in_txns = period_completed_txns.filter(walk_in_filter)
    period_stats = period_walk_in_txns.aggregate(
        txn_count=Count('id'),
        revenue=Sum('total_amount'),
        # Fallback when tax was disabled and header total_amount was left at 0.
        revenue_subtotal=Sum('subtotal'),
    )
    period_revenue = float(period_stats['revenue'] or 0)
    if period_revenue == 0:
        period_revenue = float(period_stats['revenue_subtotal'] or 0)

    txn_period_filter = Q(
        transactions__status__in=DASHBOARD_SALE_STATUSES,
        transactions__created_at__gte=range_start_aware,
        transactions__created_at__lt=range_end_aware,
        transactions__member__isnull=True,
    )
    top_customers_qs = (
        WalkInCustomer.objects.annotate(
            period_txn_count=Count('transactions', filter=txn_period_filter),
            period_revenue=Sum('transactions__total_amount', filter=txn_period_filter),
            period_revenue_subtotal=Sum('transactions__subtotal', filter=txn_period_filter),
            all_time_txn_count=Count(
                'transactions',
                filter=Q(
                    transactions__status__in=DASHBOARD_SALE_STATUSES,
                    transactions__member__isnull=True,
                ),
            ),
        )
        .filter(period_txn_count__gt=0)
        .order_by('-period_revenue', '-period_revenue_subtotal', 'display_name')[:10]
    )
    top_customers = []
    for c in top_customers_qs:
        rev = float(c.period_revenue or 0)
        if rev == 0:
            rev = float(c.period_revenue_subtotal or 0)
        top_customers.append(
            {
                'display_name': c.display_name,
                'txn_count': c.period_txn_count,
                'all_time_txn_count': c.all_time_txn_count,
                'revenue': round(rev, 2),
                'last_seen_at': timezone.localtime(c.last_seen_at, current_tz).strftime('%b %d, %Y'),
            }
        )

    # Also surface guest-name-only sales (no WalkInCustomer FK) in the top table.
    linked_names = {normalize_customer_name(c['display_name']) for c in top_customers}
    orphan_rows = (
        period_walk_in_txns.filter(walk_in_customer__isnull=True)
        .exclude(guest_customer_name='')
        .values('guest_customer_name')
        .annotate(
            txn_count=Count('id'),
            revenue=Sum('total_amount'),
            revenue_subtotal=Sum('subtotal'),
            last_seen=Max('created_at'),
        )
    )
    for row in orphan_rows:
        name = (row['guest_customer_name'] or '').strip()
        if not name:
            continue
        key = normalize_customer_name(name)
        if key in linked_names:
            continue
        rev = float(row['revenue'] or 0)
        if rev == 0:
            rev = float(row['revenue_subtotal'] or 0)
        top_customers.append(
            {
                'display_name': name,
                'txn_count': row['txn_count'],
                'all_time_txn_count': row['txn_count'],
                'revenue': round(rev, 2),
                'last_seen_at': timezone.localtime(row['last_seen'], current_tz).strftime('%b %d, %Y')
                if row['last_seen']
                else '—',
            }
        )
    top_customers.sort(key=lambda c: (-c['revenue'], c['display_name'].casefold()))
    top_customers = top_customers[:10]

    recent_period_filter = Q(
        last_seen_at__gte=range_start_aware,
        last_seen_at__lt=range_end_aware,
    )
    recent_qs = (
        WalkInCustomer.objects.filter(recent_period_filter)
        .annotate(
            all_time_txn_count=Count(
                'transactions',
                filter=Q(
                    transactions__status__in=DASHBOARD_SALE_STATUSES,
                    transactions__member__isnull=True,
                ),
            ),
            period_txn_count=Count('transactions', filter=txn_period_filter),
        )
        .order_by('-last_seen_at')[:8]
    )
    # If nobody was active in-range, fall back to global recent so the panel is not empty.
    if not recent_qs.exists():
        recent_qs = (
            WalkInCustomer.objects.annotate(
                all_time_txn_count=Count(
                    'transactions',
                    filter=Q(
                        transactions__status__in=DASHBOARD_SALE_STATUSES,
                        transactions__member__isnull=True,
                    ),
                ),
                period_txn_count=Count('transactions', filter=txn_period_filter),
            ).order_by('-last_seen_at')[:8]
        )

    recent_customers = [
        {
            'display_name': c.display_name,
            'txn_count': c.period_txn_count or c.all_time_txn_count,
            'last_seen_at': timezone.localtime(c.last_seen_at, current_tz).strftime('%b %d, %Y %H:%M'),
            'created_at': timezone.localtime(c.created_at, current_tz).strftime('%b %d, %Y'),
        }
        for c in recent_qs
    ]

    return {
        'total_registered': WalkInCustomer.objects.count(),
        'new_in_period': WalkInCustomer.objects.filter(
            created_at__gte=range_start_aware,
            created_at__lt=range_end_aware,
        ).count(),
        'period_txn_count': period_stats['txn_count'] or 0,
        'period_revenue': round(period_revenue, 2),
        'top_customers': top_customers,
        'recent_customers': recent_customers,
        'admin_url': reverse('admin:transactions_walkincustomer_changelist'),
    }


def _dashboard_payment_mix(sales_qs):
    """Payment Mix from actual DB net sales, grouped by payment method."""
    totals_by = {}

    for entry in (
        TransactionItem.objects.filter(
            transaction__in=sales_qs,
            refunded_at__isnull=True,
        )
        .values('transaction__payment_method')
        .annotate(total=Coalesce(Sum('total_price'), Decimal('0.00')))
    ):
        pm = entry['transaction__payment_method']
        totals_by[pm] = totals_by.get(pm, 0.0) + float(entry['total'] or 0)

    # Header fallback when a sale has no active line items
    for entry in (
        sales_qs.annotate(_active_lines=Count('items', filter=Q(items__refunded_at__isnull=True)))
        .filter(_active_lines=0)
        .values('payment_method')
        .annotate(total=Coalesce(Sum('total_amount'), Decimal('0.00')))
    ):
        pm = entry['payment_method']
        amt = float(entry['total'] or 0)
        if amt:
            totals_by[pm] = totals_by.get(pm, 0.0) + amt

    payment_labels = []
    payment_totals = []
    for key, label in Transaction.PAYMENT_METHODS:
        payment_labels.append(label)
        payment_totals.append(round(totals_by.get(key, 0.0), 2))
    return payment_labels, payment_totals


def _dashboard_operational_insights(period_completed_txns, period_refund_count, range_start_aware, range_end_aware):
    """
    Metrics for the dashboard Operational Insights panel.
    Period-scoped: completed sales in range, new members joined in range, refund rate vs period sales.
    Live snapshot: open transaction queues and active product count.
    """
    total_txn = period_completed_txns.count()
    totals = period_completed_txns.aggregate(
        rev=Sum('total_amount'),
        rev_sub=Sum('subtotal'),
        guest_rev=Sum('total_amount', filter=Q(member__isnull=True)),
        guest_rev_sub=Sum('subtotal', filter=Q(member__isnull=True)),
        member_rev=Sum('total_amount', filter=Q(member__isnull=False)),
        member_rev_sub=Sum('subtotal', filter=Q(member__isnull=False)),
    )
    total_rev = float(totals['rev'] or 0) or float(totals['rev_sub'] or 0)
    guest_rev = float(totals['guest_rev'] or 0) or float(totals['guest_rev_sub'] or 0)
    member_rev = float(totals['member_rev'] or 0) or float(totals['member_rev_sub'] or 0)
    guest_count = period_completed_txns.filter(member__isnull=True).count()
    member_count = period_completed_txns.filter(member__isnull=False).count()
    avg_order = (total_rev / total_txn) if total_txn else 0.0
    guest_share_pct = (100.0 * guest_count / total_txn) if total_txn else 0.0
    member_share_pct = (100.0 * member_count / total_txn) if total_txn else 0.0
    refund_rate_pct = (100.0 * period_refund_count / total_txn) if total_txn else 0.0

    new_members_in_period = Member.objects.filter(
        date_joined__gte=range_start_aware,
        date_joined__lt=range_end_aware,
    ).count()

    pending_checkouts = Transaction.objects.filter(status='pending').count()
    refund_requested_open = Transaction.objects.filter(status='refund_requested').count()
    return_window_open = Transaction.objects.filter(status='return_window').count()
    active_products = Product.objects.filter(is_active=True).count()

    return {
        'avg_order_value': round(avg_order, 2),
        'guest_txn_count': guest_count,
        'member_txn_count': member_count,
        'guest_revenue': round(guest_rev, 2),
        'member_revenue': round(member_rev, 2),
        'guest_share_pct': round(guest_share_pct, 1),
        'member_share_pct': round(member_share_pct, 1),
        'new_members_in_period': new_members_in_period,
        'pending_checkouts': pending_checkouts,
        'refund_requested_open': refund_requested_open,
        'return_window_open': return_window_open,
        'active_products': active_products,
        'refund_rate_pct': round(refund_rate_pct, 1),
    }


def handle_login(request, redirect_to_dashboard=False):
    """Shared login logic — delegates to login_helper for all auth logic."""
    if request.user.is_authenticated:
        if is_loans_only_user(request.user):
            return redirect('loans_overview')
        return redirect('dashboard' if is_cashier_or_admin(request.user) else 'user_choice')

    # Honour an existing member-only session — always show user-choice hub
    if is_member_session_valid(request):
        return redirect('user_choice')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        next_url = request.POST.get('next') or request.GET.get('next', '')

        if not username or not password:
            messages.error(request, 'Please enter both username and password.')
            store_profile = StoreProfile.get()
            return render(request, 'admin_panel/login.html', {'store_profile': store_profile})

        result = login_with_credentials(request, username, password, next_url)
        if result['success']:
            return redirect(result['redirect_url'])

        messages.error(request, result['error'])

    store_profile = StoreProfile.get()
    return render(request, 'admin_panel/login.html', {'store_profile': store_profile})


@require_http_methods(["GET", "POST"])
def root_login(request):
    """Root login page - first page users see."""
    return handle_login(request)


@require_http_methods(["GET", "POST"])
def redirect_to_root_login(request):
    """Redirect /admin/login/ to root login page, preserving query parameters."""
    root_login_url = reverse('root_login')
    query_string = request.META.get('QUERY_STRING', '')
    if query_string:
        root_login_url = f"{root_login_url}?{query_string}"
    return redirect(root_login_url)



# is_staff_role is now imported from login_helper above.


def admin_role_badge_context(request):
    """Keys for `partials/admin_role_badge.html` — role pill on admin UI.

    Priority: superuser / member admin → Admin Verified; member cashier → Cashier;
    member staff or Django staff (non-superuser) → Staff.
    """
    user = request.user
    member = None
    try:
        member = Member.objects.get(user=user)
    except Member.DoesNotExist:
        pass

    role_badge = ''
    if user.is_superuser or (member and member.is_active and member.role == 'admin'):
        role_badge = 'admin_verified'
    elif member and member.is_active and member.role == 'cashier':
        role_badge = 'cashier'
    elif member and member.is_active and member.role == 'committee':
        role_badge = 'committee'
    elif member and member.is_active and member.role == 'loan_officer':
        role_badge = 'loan_officer'
    elif (member and member.is_active and member.role == 'staff') or (
        user.is_staff and not user.is_superuser
    ):
        role_badge = 'staff'

    return {
        'is_admin': is_cashier_or_admin(request.user),
        'is_admin_user': is_admin_user(request.user),
        'role_badge': role_badge,
    }


@login_required
def loans_overview(request):
    """Loan desk landing page: view all loans members have acquired, plus pipeline shortcuts.

    The primary feature is a searchable/filterable table of every LoanApplication
    (member, product, amount, status, outstanding balance). Pipeline step cards
    remain as secondary shortcuts into the 13-step workflow.
    """
    if not is_loan_desk_user(request.user):
        messages.warning(request, 'You do not have permission to access loan management.')
        return redirect('kiosk_home')

    from django.urls import reverse
    from loans.models import LoanApplication, Disbursement

    inquiry_url = reverse('loans:inquiry-create')
    apply_url = reverse('loans:application-create')
    list_url = reverse('loans:application-list')

    Status = LoanApplication.Status
    # "Acquired" = funds released / repayment in progress / settled.
    ACQUIRED_STATUSES = (
        Status.DISBURSED,
        Status.ACTIVE,
        Status.FULLY_PAID,
        Status.CLOSED,
    )

    search_query = (request.GET.get('search') or '').strip()
    status_filter = (request.GET.get('status') or 'acquired').strip().lower()
    valid_status_keys = {s.value.lower() for s in Status} | {'all', 'acquired'}
    if status_filter not in valid_status_keys:
        status_filter = 'acquired'

    loans_qs = (
        LoanApplication.objects
        .select_related('member', 'loan_product', 'disbursement', 'payment_option')
        .prefetch_related('amortization_schedules')
        .order_by('-created_at')
    )

    if status_filter == 'acquired':
        loans_qs = loans_qs.filter(status__in=ACQUIRED_STATUSES)
    elif status_filter != 'all':
        loans_qs = loans_qs.filter(status=status_filter.upper())

    if search_query:
        loans_qs = loans_qs.filter(
            Q(member__username__icontains=search_query)
            | Q(member__first_name__icontains=search_query)
            | Q(member__last_name__icontains=search_query)
            | Q(member__email__icontains=search_query)
            | Q(loan_product__name__icontains=search_query)
            | Q(purpose__icontains=search_query)
        )

    loans_page = Paginator(loans_qs, 25).get_page(request.GET.get('page') or 1)

    # Attach outstanding balance without N+1 inside the template.
    loans_with_balance = []
    for loan in loans_page.object_list:
        loan.outstanding_balance = loan.total_outstanding_balance()
        loans_with_balance.append(loan)

    all_loans = LoanApplication.objects.all()
    total_loans = all_loans.count()
    acquired_count = all_loans.filter(status__in=ACQUIRED_STATUSES).count()
    active_count = all_loans.filter(status=Status.ACTIVE).count()
    pending_count = all_loans.exclude(
        status__in=ACQUIRED_STATUSES + (Status.REJECTED, Status.VERIFICATION_FAILED, Status.DRAFT)
    ).count()
    total_disbursed = (
        Disbursement.objects.aggregate(total=Sum('amount_released')).get('total')
        or Decimal('0')
    )
    # Outstanding across ACTIVE acquired loans only.
    outstanding_total = Decimal('0')
    for loan in all_loans.filter(status=Status.ACTIVE).prefetch_related('amortization_schedules'):
        outstanding_total += loan.total_outstanding_balance()

    steps = [
        {'n': 1, 'title': 'Loan Inquiry', 'icon': 'bi-question-circle',
         'desc': 'Log a member inquiry and match them to a suitable loan product.',
         'url': inquiry_url, 'action': 'Start inquiry'},
        {'n': 2, 'title': 'Loan Application & Documents', 'icon': 'bi-file-earmark-text',
         'desc': 'Capture the application details and upload required documents.',
         'url': apply_url, 'action': 'New application'},
        {'n': 3, 'title': 'Eligibility & Document Verification', 'icon': 'bi-patch-check',
         'desc': 'Staff confirm membership standing and document completeness.',
         'url': list_url, 'action': 'Verify'},
        {'n': 4, 'title': 'Credit Investigation & Evaluation', 'icon': 'bi-search',
         'desc': 'Credit officer scores repayment capacity and recommends a decision.',
         'url': list_url, 'action': 'Investigate'},
        {'n': 5, 'title': 'Credit Committee Approval', 'icon': 'bi-people',
         'desc': 'The committee reviews the evaluation and records a decision.',
         'url': list_url, 'action': 'Review'},
        {'n': 6, 'title': 'Approval Decision', 'icon': 'bi-check2-circle',
         'desc': 'Approve or reject — the member is emailed the decision automatically.',
         'url': list_url, 'action': 'Decide'},
        {'n': 7, 'title': 'Insurance Enrollment', 'icon': 'bi-shield-check',
         'desc': 'Enroll credit-life insurance when the product requires it.',
         'url': list_url, 'action': 'Enroll'},
        {'n': 8, 'title': 'Documentation & Signing', 'icon': 'bi-pen',
         'desc': 'Generate and sign the loan agreement with the borrower.',
         'url': list_url, 'action': 'Sign documents'},
        {'n': 9, 'title': 'Loan Disbursement', 'icon': 'bi-cash-stack',
         'desc': 'Cashier releases funds and records the disbursement reference.',
         'url': list_url, 'action': 'Disburse'},
        {'n': 10, 'title': 'Payment Collection & Monitoring', 'icon': 'bi-wallet2',
         'desc': 'Record repayments and monitor the loan while it is active.',
         'url': list_url, 'action': 'Collect payment'},
        {'n': 11, 'title': 'Loan Fully Paid?', 'icon': 'bi-cash-coin',
         'desc': 'Once every installment is settled the loan is marked fully paid.',
         'url': list_url, 'action': 'View status'},
        {'n': 12, 'title': 'Settlement & Account Closure', 'icon': 'bi-file-earmark-check',
         'desc': 'Issue the clearance certificate, release collateral and close the account.',
         'url': list_url, 'action': 'Close account'},
    ]

    status_choices = [
        ('acquired', 'Acquired loans'),
        ('all', 'All applications'),
        ('active', 'Active'),
        ('disbursed', 'Disbursed'),
        ('fully_paid', 'Fully paid'),
        ('closed', 'Closed'),
        ('pending_committee_approval', 'Pending committee'),
        ('submitted', 'Submitted'),
        ('under_verification', 'Under verification'),
        ('under_investigation', 'Under investigation'),
        ('approved', 'Approved'),
        ('rejected', 'Rejected'),
        ('draft', 'Draft'),
    ]

    context = {
        'steps': steps,
        'loans_list_url': list_url,
        'apply_url': apply_url,
        'loans': loans_with_balance,
        'page_obj': loans_page,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': status_choices,
        'total_loans': total_loans,
        'acquired_count': acquired_count,
        'active_count': active_count,
        'pending_count': pending_count,
        'total_disbursed': total_disbursed,
        'outstanding_total': outstanding_total,
        **admin_role_badge_context(request),
    }
    return render(request, 'admin_panel/loans_overview.html', context)


@login_required
def dashboard(request):
    # Loan-only roles land on loans; other staff use the full console.
    if is_loans_only_user(request.user):
        return redirect('loans_overview')
    # Ensure only admin/cashier users can access dashboard
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to access the admin dashboard.')
        return redirect('kiosk_home')
    kiosk_config = KioskConfig.get()
    today = _store_local_today()

    # --- Period type selection: week / month / year ---
    import calendar
    selected_range = request.GET.get('range', 'custom')
    if selected_range not in ('week', 'month', 'year', 'custom', 'all'):
        selected_range = 'custom'

    # Build list of available years from first transaction up to current year
    first_txn = Transaction.objects.order_by('created_at').first()
    first_year = first_txn.created_at.year if first_txn else today.year
    available_years = list(range(today.year, first_year - 1, -1))
    min_dashboard_date = date_type(first_year, 1, 1)
    if first_txn:
        min_dashboard_date = min(min_dashboard_date, first_txn.created_at.astimezone(timezone.get_current_timezone()).date())

    # Max allowed values (never beyond today)
    max_week = today.strftime('%G-W%V')
    max_month = today.strftime('%Y-%m')
    today_iso = today.strftime('%Y-%m-%d')
    min_date_iso = min_dashboard_date.strftime('%Y-%m-%d')

    if selected_range == 'all':
        range_start = min_dashboard_date
        range_end = today
        span_days = (range_end - range_start).days + 1
        range_granularity = 'monthly' if span_days > 60 else 'daily'
        range_days = span_days
        selected_week = today.strftime('%G-W%V')
        selected_month = today.strftime('%Y-%m')
        selected_year = today.year
        selected_date_from = range_start.strftime('%Y-%m-%d')
        selected_date_to = range_end.strftime('%Y-%m-%d')
        range_label = 'All Data'

    elif selected_range == 'week':
        week_param = request.GET.get('week', '')
        try:
            year_str, week_str = week_param.split('-W')
            range_start = datetime.strptime(
                f'{year_str}-W{int(week_str):02d}-1', '%G-W%V-%u'
            ).date()
            # Clamp: week must not start after today's Monday
            today_monday = today - timedelta(days=today.weekday())
            if range_start > today_monday:
                range_start = today_monday
        except Exception:
            range_start = today - timedelta(days=today.weekday())
        range_end = min(range_start + timedelta(days=6), today)
        selected_week = range_start.strftime('%G-W%V')
        selected_month = today.strftime('%Y-%m')
        selected_year = today.year
        range_label = f"{range_start.strftime('%b %d')} \u2013 {range_end.strftime('%b %d, %Y')}"
        range_granularity = 'daily'
        range_days = (range_end - range_start).days + 1
        selected_date_from = range_start.strftime('%Y-%m-%d')
        selected_date_to = range_end.strftime('%Y-%m-%d')

    elif selected_range == 'month':
        month_param = request.GET.get('month', '')
        try:
            year_str, mon_str = month_param.split('-')
            range_start = date_type(int(year_str), int(mon_str), 1)
            # Clamp: month must not be in the future
            this_month_start = today.replace(day=1)
            if range_start > this_month_start:
                range_start = this_month_start
        except Exception:
            range_start = today.replace(day=1)
        last_day = calendar.monthrange(range_start.year, range_start.month)[1]
        range_end = min(range_start.replace(day=last_day), today)
        selected_month = range_start.strftime('%Y-%m')
        selected_week = today.strftime('%G-W%V')
        selected_year = today.year
        range_label = range_start.strftime('%B %Y')
        range_granularity = 'daily'
        range_days = (range_end - range_start).days + 1
        selected_date_from = range_start.strftime('%Y-%m-%d')
        selected_date_to = range_end.strftime('%Y-%m-%d')

    elif selected_range == 'custom':
        df_raw = request.GET.get('date_from', '').strip()
        dt_raw = request.GET.get('date_to', '').strip()
        try:
            d_from = datetime.strptime(df_raw, '%Y-%m-%d').date() if df_raw else min_dashboard_date
        except ValueError:
            d_from = min_dashboard_date
        try:
            d_to = datetime.strptime(dt_raw, '%Y-%m-%d').date() if dt_raw else today
        except ValueError:
            d_to = today
        if d_from > d_to:
            d_from, d_to = d_to, d_from
        if d_to > today:
            d_to = today
        if d_from > today:
            d_from = today
        if d_from < min_dashboard_date:
            d_from = min_dashboard_date
        if d_to < min_dashboard_date:
            d_to = min_dashboard_date
        if d_from > d_to:
            d_from = d_to
        range_start = d_from
        range_end = d_to
        span_days = (range_end - range_start).days + 1
        if span_days > 60:
            range_granularity = 'monthly'
            range_days = span_days
        else:
            range_granularity = 'daily'
            range_days = span_days
        selected_week = today.strftime('%G-W%V')
        selected_month = today.strftime('%Y-%m')
        selected_year = today.year
        selected_date_from = range_start.strftime('%Y-%m-%d')
        selected_date_to = range_end.strftime('%Y-%m-%d')
        if range_start == range_end:
            range_label = range_start.strftime('%b %d, %Y')
        else:
            range_label = f"{range_start.strftime('%b %d, %Y')} \u2013 {range_end.strftime('%b %d, %Y')}"

    else:  # year
        try:
            year_val = int(request.GET.get('year', str(today.year)))
        except (ValueError, TypeError):
            year_val = today.year
        if year_val not in available_years:
            year_val = today.year
        # Clamp: future year not allowed
        if year_val > today.year:
            year_val = today.year
        range_start = date_type(year_val, 1, 1)
        range_end = min(date_type(year_val, 12, 31), today)
        selected_year = year_val
        selected_week = today.strftime('%G-W%V')
        selected_month = today.strftime('%Y-%m')
        range_label = str(year_val)
        range_granularity = 'monthly'
        range_days = (range_end - range_start).days + 1
        selected_date_from = range_start.strftime('%Y-%m-%d')
        selected_date_to = range_end.strftime('%Y-%m-%d')

    base_qs = _dashboard_sales_qs()

    all_time_transactions = base_qs.count()
    all_time_revenue = float(_transaction_net_revenue(base_qs))

    # Build timezone-aware today range (Django 5.2 + Python 3.14 + SQLite: __date lookups are broken)
    current_tz = timezone.get_current_timezone()
    today_start = timezone.make_aware(datetime(today.year, today.month, today.day, 0, 0, 0), current_tz)
    tomorrow = today + timedelta(days=1)
    tomorrow_start = timezone.make_aware(datetime(tomorrow.year, tomorrow.month, tomorrow.day, 0, 0, 0), current_tz)

    today_qs = base_qs.filter(created_at__gte=today_start, created_at__lt=tomorrow_start)
    today_transactions = today_qs.count()
    today_revenue = float(_transaction_net_revenue(today_qs))

    total_members = Member.objects.filter(is_active=True).count()

    # Current inventory snapshot (not tied to sales period — stock levels are "now")
    _active_products = Product.objects.filter(is_active=True)
    low_stock_products = _active_products.filter(
        stock_quantity__gt=0,
        stock_quantity__lte=F('low_stock_threshold'),
    ).count()
    out_of_stock_products = _active_products.filter(stock_quantity=0).count()
    inventory_alert_total = low_stock_products + out_of_stock_products

    # --- Chart data calculations ---
    # Build timezone-aware range boundaries for filtering
    range_start_aware = timezone.make_aware(datetime(range_start.year, range_start.month, range_start.day, 0, 0, 0), current_tz)
    range_end_next = range_end + timedelta(days=1)
    range_end_aware = timezone.make_aware(datetime(range_end_next.year, range_end_next.month, range_end_next.day, 0, 0, 0), current_tz)

    period_txns = base_qs.filter(
        created_at__gte=range_start_aware,
        created_at__lt=range_end_aware,
    )

    top_products = _dashboard_top_products(range_start_aware, range_end_aware)

    recent_transactions = _dashboard_recent_transactions(
        range_start_aware, range_end_aware, current_tz, limit=12
    )
    walk_in_insights = _dashboard_walk_in_insights(
        period_txns, range_start_aware, range_end_aware, current_tz
    )
    total_transactions = period_txns.count()
    total_revenue = float(_transaction_net_revenue(period_txns))

    sales_series = _dashboard_net_revenue_series(
        range_start_aware, range_end_aware, current_tz, range_granularity
    )
    if range_granularity == 'monthly':
        # Monthly aggregation for year view — Python-side grouping (avoids TruncMonth SQLite bug)
        daily_labels = []
        daily_totals = []
        cur_month = range_start.replace(day=1)
        while cur_month <= range_end:
            daily_labels.append(cur_month.strftime('%b %Y'))
            daily_totals.append(round(sales_series.get(cur_month, 0), 2))
            if cur_month.month == 12:
                cur_month = cur_month.replace(year=cur_month.year + 1, month=1)
            else:
                cur_month = cur_month.replace(month=cur_month.month + 1)
    else:
        # Daily aggregation — Python-side grouping (avoids TruncDate SQLite bug)
        daily_labels = []
        daily_totals = []
        for offset in range(range_days):
            day = range_start + timedelta(days=offset)
            daily_labels.append(day.strftime('%b %d'))
            daily_totals.append(round(sales_series.get(day, 0), 2))

    payment_labels, payment_totals = _dashboard_payment_mix(
        base_qs.filter(
            created_at__gte=range_start_aware,
            created_at__lt=range_end_aware,
        )
    )

    category_sales = TransactionItem.objects.filter(
        transaction__status__in=DASHBOARD_SALE_STATUSES,
        transaction__created_at__gte=range_start_aware,
        transaction__created_at__lt=range_end_aware,
        refunded_at__isnull=True,
        product__category__isnull=False
    ).values('product__category__name').annotate(
        total=Sum('total_price')
    ).order_by('-total')[:6]
    category_labels = [entry['product__category__name'] or 'Uncategorized' for entry in category_sales]
    category_totals = [float(entry['total'] or 0) for entry in category_sales]

    sale_txn_filter = Q(transactions__status__in=DASHBOARD_SALE_STATUSES)
    top_members_qs = (
        Member.objects.filter(sale_txn_filter)
        .annotate(total_spent=Sum('transactions__total_amount', filter=sale_txn_filter))
        .filter(total_spent__gt=0)
        .order_by('-total_spent')
    )
    top_members_per_page = 7
    top_members_paginator = Paginator(top_members_qs, top_members_per_page)
    top_members_page_number = request.GET.get('top_members_page', 1)
    top_members = top_members_paginator.get_page(top_members_page_number)

    top_members_query_params = request.GET.copy()
    if 'top_members_page' in top_members_query_params:
        del top_members_query_params['top_members_page']

    refund_stats = _dashboard_refund_stats(
        range_start_aware,
        range_end_aware,
        current_tz,
        range_granularity,
        range_start,
        range_end,
        range_days,
    )
    total_refunds = refund_stats['total_refunds']
    total_refund_amount = refund_stats['total_refund_amount']
    period_refunds = refund_stats['period_refunds']
    period_refund_amount = refund_stats['period_refund_amount']
    recent_refunds = refund_stats['recent_refunds']
    daily_refund_labels = refund_stats['daily_refund_labels']
    daily_refund_amounts = refund_stats['daily_refund_amounts']
    daily_refund_counts = refund_stats['daily_refund_counts']

    operational_insights = _dashboard_operational_insights(
        period_txns, period_refunds, range_start_aware, range_end_aware
    )

    context = {
        'total_transactions': total_transactions,
        'total_revenue': total_revenue,
        'all_time_transactions': all_time_transactions,
        'all_time_revenue': all_time_revenue,
        'today_transactions': today_transactions,
        'today_revenue': today_revenue,
        'total_members': total_members,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'inventory_alert_total': inventory_alert_total,
        'recent_transactions': recent_transactions,
        'top_products': top_products,
        'top_products_json': json.dumps(top_products),
        'top_members': top_members,
        'top_members_query_params': top_members_query_params.urlencode(),
        'daily_sales_labels': json.dumps(daily_labels),
        'daily_sales_totals': json.dumps(daily_totals),
        'payment_labels': json.dumps(payment_labels),
        'payment_totals': json.dumps(payment_totals),
        'selected_range': selected_range,
        'range_label': range_label,
        'selected_week': selected_week,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'available_years': available_years,
        'max_week': max_week,
        'max_month': max_month,
        'today_iso': today_iso,
        'min_date_iso': min_date_iso,
        'selected_date_from': selected_date_from,
        'selected_date_to': selected_date_to,
        'category_labels': json.dumps(category_labels),
        'category_totals': json.dumps(category_totals),
        'user_display_name': request.user.get_full_name() or request.user.username,
        'dashboard_hero_description': kiosk_config.admin_dashboard_description,
        **admin_role_badge_context(request),
        # Refund statistics
        'total_refunds': total_refunds,
        'total_refund_amount': total_refund_amount,
        'period_refunds': period_refunds,
        'period_refund_amount': period_refund_amount,
        'recent_refunds': recent_refunds,
        'daily_refund_labels': json.dumps(daily_refund_labels),
        'daily_refund_amounts': json.dumps(daily_refund_amounts),
        'daily_refund_counts': json.dumps(daily_refund_counts),
        'operational_insights': operational_insights,
        'walk_in_insights': walk_in_insights,
    }

    return render(request, 'admin_panel/dashboard.html', context)


def _dashboard_period_bounds_from_get(request, default_range='custom'):
    """Parse dashboard-style GET params into local dates and aware [start, end) bounds.

    Same rules as ``api_dashboard_period_data`` / main dashboard period filters
    (transaction ``created_at`` in the store timezone, half-open interval).
    """
    import calendar

    today = _store_local_today()
    raw = (request.GET.get('range') or '').strip()
    selected_range = raw if raw in ('week', 'month', 'year', 'custom', 'all') else default_range
    if selected_range not in ('week', 'month', 'year', 'custom', 'all'):
        selected_range = 'custom'

    first_txn = Transaction.objects.order_by('created_at').first()
    first_year = first_txn.created_at.year if first_txn else today.year
    available_years = list(range(today.year, first_year - 1, -1))
    min_dashboard_date = date_type(first_year, 1, 1)
    if first_txn:
        min_dashboard_date = min(
            min_dashboard_date,
            first_txn.created_at.astimezone(timezone.get_current_timezone()).date(),
        )

    if selected_range == 'all':
        range_start = min_dashboard_date
        range_end = today
        span_days = (range_end - range_start).days + 1
        range_granularity = 'monthly' if span_days > 60 else 'daily'
        range_days = span_days
        range_label = 'All Data'
    elif selected_range == 'week':
        week_param = request.GET.get('week', '')
        try:
            year_str, week_str = week_param.split('-W')
            range_start = datetime.strptime(
                f'{year_str}-W{int(week_str):02d}-1', '%G-W%V-%u'
            ).date()
            today_monday = today - timedelta(days=today.weekday())
            if range_start > today_monday:
                range_start = today_monday
        except Exception:
            range_start = today - timedelta(days=today.weekday())
        range_end = min(range_start + timedelta(days=6), today)
        range_label = f"{range_start.strftime('%b %d')} \u2013 {range_end.strftime('%b %d, %Y')}"
        range_granularity = 'daily'
        range_days = (range_end - range_start).days + 1
    elif selected_range == 'month':
        month_param = request.GET.get('month', '')
        try:
            year_str, mon_str = month_param.split('-')
            range_start = date_type(int(year_str), int(mon_str), 1)
            this_month_start = today.replace(day=1)
            if range_start > this_month_start:
                range_start = this_month_start
        except Exception:
            range_start = today.replace(day=1)
        last_day = calendar.monthrange(range_start.year, range_start.month)[1]
        range_end = min(range_start.replace(day=last_day), today)
        range_label = range_start.strftime('%B %Y')
        range_granularity = 'daily'
        range_days = (range_end - range_start).days + 1
    elif selected_range == 'custom':
        df_raw = request.GET.get('date_from', '').strip()
        dt_raw = request.GET.get('date_to', '').strip()
        try:
            d_from = datetime.strptime(df_raw, '%Y-%m-%d').date() if df_raw else min_dashboard_date
        except ValueError:
            d_from = min_dashboard_date
        try:
            d_to = datetime.strptime(dt_raw, '%Y-%m-%d').date() if dt_raw else today
        except ValueError:
            d_to = today
        if d_from > d_to:
            d_from, d_to = d_to, d_from
        if d_to > today:
            d_to = today
        if d_from > today:
            d_from = today
        if d_from < min_dashboard_date:
            d_from = min_dashboard_date
        if d_to < min_dashboard_date:
            d_to = min_dashboard_date
        if d_from > d_to:
            d_from = d_to
        range_start, range_end = d_from, d_to
        span_days = (range_end - range_start).days + 1
        range_granularity = 'monthly' if span_days > 60 else 'daily'
        range_days = span_days
        if range_start == range_end:
            range_label = range_start.strftime('%b %d, %Y')
        else:
            range_label = f"{range_start.strftime('%b %d, %Y')} \u2013 {range_end.strftime('%b %d, %Y')}"
    else:
        try:
            year_val = int(request.GET.get('year', str(today.year)))
        except (ValueError, TypeError):
            year_val = today.year
        if year_val not in available_years or year_val > today.year:
            year_val = today.year
        range_start = date_type(year_val, 1, 1)
        range_end = min(date_type(year_val, 12, 31), today)
        range_label = str(year_val)
        range_granularity = 'monthly'
        range_days = (range_end - range_start).days + 1

    current_tz = timezone.get_current_timezone()
    range_start_aware = timezone.make_aware(
        datetime(range_start.year, range_start.month, range_start.day, 0, 0, 0),
        current_tz,
    )
    range_end_next = range_end + timedelta(days=1)
    range_end_aware = timezone.make_aware(
        datetime(range_end_next.year, range_end_next.month, range_end_next.day, 0, 0, 0),
        current_tz,
    )

    return {
        'today': today,
        'selected_range': selected_range,
        'range_label': range_label,
        'range_start': range_start,
        'range_end': range_end,
        'range_start_aware': range_start_aware,
        'range_end_aware': range_end_aware,
        'range_granularity': range_granularity,
        'range_days': range_days,
        'current_tz': current_tz,
        'min_dashboard_date': min_dashboard_date,
        'available_years': available_years,
    }


@login_required
def api_dashboard_period_data(request):
    """Return dashboard chart payload based on selected chart period."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

    b = _dashboard_period_bounds_from_get(request)
    selected_range = b['selected_range']
    range_label = b['range_label']
    range_start = b['range_start']
    range_end = b['range_end']
    range_start_aware = b['range_start_aware']
    range_end_aware = b['range_end_aware']
    range_granularity = b['range_granularity']
    range_days = b['range_days']
    current_tz = b['current_tz']

    base_qs = _dashboard_sales_qs()
    txns_filtered = base_qs.filter(created_at__gte=range_start_aware, created_at__lt=range_end_aware)
    total_transactions = txns_filtered.count()
    total_revenue = float(_transaction_net_revenue(txns_filtered))
    _active_products = Product.objects.filter(is_active=True)
    low_stock_products = _active_products.filter(
        stock_quantity__gt=0,
        stock_quantity__lte=F('low_stock_threshold'),
    ).count()
    out_of_stock_products = _active_products.filter(stock_quantity=0).count()
    inventory_alert_total = low_stock_products + out_of_stock_products
    recent_transactions = _dashboard_recent_transactions(
        range_start_aware, range_end_aware, current_tz, limit=12
    )

    walk_in_insights = _dashboard_walk_in_insights(
        txns_filtered, range_start_aware, range_end_aware, current_tz
    )

    sales_series = _dashboard_net_revenue_series(
        range_start_aware, range_end_aware, current_tz, range_granularity
    )
    if range_granularity == 'monthly':
        daily_labels, daily_totals = [], []
        cur_month = range_start.replace(day=1)
        while cur_month <= range_end:
            daily_labels.append(cur_month.strftime('%b %Y'))
            daily_totals.append(round(sales_series.get(cur_month, 0), 2))
            if cur_month.month == 12:
                cur_month = cur_month.replace(year=cur_month.year + 1, month=1)
            else:
                cur_month = cur_month.replace(month=cur_month.month + 1)
    else:
        daily_labels, daily_totals = [], []
        for offset in range(range_days):
            day = range_start + timedelta(days=offset)
            daily_labels.append(day.strftime('%b %d'))
            daily_totals.append(round(sales_series.get(day, 0), 2))

    payment_labels, payment_totals = _dashboard_payment_mix(txns_filtered)

    category_sales = TransactionItem.objects.filter(
        transaction__status__in=DASHBOARD_SALE_STATUSES,
        transaction__created_at__gte=range_start_aware,
        transaction__created_at__lt=range_end_aware,
        refunded_at__isnull=True,
        product__category__isnull=False,
    ).values('product__category__name').annotate(total=Sum('total_price')).order_by('-total')[:6]
    category_labels = [entry['product__category__name'] or 'Uncategorized' for entry in category_sales]
    category_totals = [float(entry['total'] or 0) for entry in category_sales]

    refund_stats = _dashboard_refund_stats(
        range_start_aware,
        range_end_aware,
        current_tz,
        range_granularity,
        range_start,
        range_end,
        range_days,
    )
    all_time_refunds = refund_stats['total_refunds']
    all_time_refund_amount = refund_stats['total_refund_amount']
    period_refunds = refund_stats['period_refunds']
    period_refund_amount = refund_stats['period_refund_amount']
    daily_refund_labels = refund_stats['daily_refund_labels']
    daily_refund_amounts = refund_stats['daily_refund_amounts']
    daily_refund_counts = refund_stats['daily_refund_counts']

    operational_insights = _dashboard_operational_insights(
        txns_filtered, period_refunds, range_start_aware, range_end_aware
    )

    top_products_api = _dashboard_top_products(range_start_aware, range_end_aware)

    return JsonResponse({
        'success': True,
        'selected_range': selected_range,
        'range_label': range_label,
        'total_transactions': total_transactions,
        'total_revenue': total_revenue,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'inventory_alert_total': inventory_alert_total,
        'all_time_refunds': all_time_refunds,
        'all_time_refund_amount': all_time_refund_amount,
        'period_refunds': period_refunds,
        'period_refund_amount': period_refund_amount,
        'recent_refunds': refund_stats['recent_refunds'],
        'recent_transactions': recent_transactions,
        'daily_sales_labels': daily_labels,
        'daily_sales_totals': daily_totals,
        'payment_labels': payment_labels,
        'payment_totals': payment_totals,
        'category_labels': category_labels,
        'category_totals': category_totals,
        'daily_refund_labels': daily_refund_labels,
        'daily_refund_amounts': daily_refund_amounts,
        'daily_refund_counts': daily_refund_counts,
        'operational_insights': operational_insights,
        'walk_in_insights': walk_in_insights,
        'top_products': top_products_api,
    })


@login_required
def inventory_management(request):
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to access this page.')
        return redirect('kiosk_home')
    
    # Get search query and filter from request
    search_query = request.GET.get('search', '').strip()
    filter_type = request.GET.get('filter', 'all')  # 'all', 'low_stock', 'out_of_stock'
    # Keep inventory search behavior on All Products.
    if search_query:
        filter_type = 'all'
    
    # Start with all products
    products = Product.objects.select_related('category', 'giveaway').prefetch_related(
        'stock_batches', 'sale_units',
    ).all()
    if request_can_manage_giveaways(request) or filter_type == 'giveaway':
        products = annotate_giveaway_units_given(products)
    
    # Apply filter
    if filter_type == 'low_stock':
        # Low stock: stock <= threshold but > 0
        products = products.filter(is_active=True, stock_quantity__lte=F('low_stock_threshold'), stock_quantity__gt=0)
    elif filter_type == 'out_of_stock':
        # Out of stock: stock = 0
        products = products.filter(is_active=True, stock_quantity=0)
    elif filter_type == 'giveaway':
        products = products.filter(giveaway_units_given__gt=0)
    
    # Apply search filter if query exists
    if search_query:
        search_filters = (
            Q(name__icontains=search_query) |
            Q(barcode__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )
        products = products.filter(search_filters)

    products = products.order_by('name')
    inventory_price_summary = _compute_inventory_price_summary(products)
    
    # Pagination: 10 products per page
    paginator = Paginator(products, 10)
    page_number = request.GET.get('page', 1)
    try:
        products_page = paginator.get_page(page_number)
    except:
        products_page = paginator.get_page(1)

    for product in products_page.object_list:
        if promote_new_stock_to_old_if_needed(product):
            product.refresh_from_db()
    
    all_categories = Category.objects.all().order_by('name')
    category_paginator = Paginator(all_categories, 10)
    category_page_number = request.GET.get('category_page', 1)
    categories_page = category_paginator.get_page(category_page_number)
    
    # Calculate statistics (from all products, not filtered)
    all_products = Product.objects.all()
    total_products = all_products.count()
    low_stock_products = all_products.filter(is_active=True, stock_quantity__lte=F('low_stock_threshold'), stock_quantity__gt=0).count()
    out_of_stock_products = all_products.filter(is_active=True, stock_quantity=0).count()
    total_categories = all_categories.count()
    
    all_active_products = Product.objects.filter(is_active=True).select_related('category').order_by('name')

    can_manage_giveaways = request_can_manage_giveaways(request)
    giveaway_products = 0
    if can_manage_giveaways:
        products_with_giveaways = annotate_giveaway_units_given(Product.objects.all())
        giveaway_products = products_with_giveaways.filter(giveaway_units_given__gt=0).count()

    context = {
        'products': products_page,
        'page_obj': products_page,
        'categories': all_categories,
        'categories_page_obj': categories_page,
        'all_active_products': all_active_products,
        'total_products': total_products,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'total_categories': total_categories,
        'search_query': search_query,
        'filter_type': filter_type,
        'can_manage_giveaways': can_manage_giveaways,
        'giveaway_products': giveaway_products,
        'inventory_price_summary': inventory_price_summary,
        **admin_role_badge_context(request),
    }
    
    return render(request, 'admin_panel/inventory.html', context)


@login_required
@require_http_methods(["POST"])
def api_send_inventory_stock_alerts(request):
    """Email current low-stock and out-of-stock products to staff and admin roles."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    result = send_inventory_stock_alerts()
    status = 200 if result.get('success') else 400
    payload = {k: v for k, v in result.items() if k != 'recipients'}
    return JsonResponse(payload, status=status)


def _money_expr(qty_field, price_field):
    return ExpressionWrapper(
        F(qty_field) * F(price_field),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )


def _zero_money():
    return Value(Decimal('0.00'), output_field=DecimalField(max_digits=18, decimal_places=2))


def _zero_qty():
    """Coalesce default for Decimal stock / sale quantities (pieces or kg)."""
    return Value(Decimal('0'), output_field=DecimalField(max_digits=14, decimal_places=3))


def _as_qty(value):
    try:
        return Decimal(str(value or 0))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def _compute_inventory_price_summary(product_qs):
    """
    Buying Value = buying price × qty/stock (batch cost×qty when tiers exist).
    Selling Value = selling price × qty/stock.
    REVENUE PER ITEM (total_subtotal) = selling value − buying value (per product, summed).
    """
    product_ids = list(product_qs.values_list('id', flat=True))
    product_count = len(product_ids)
    empty = {
        'product_count': 0,
        'total_units': 0,
        'total_qty_sold': 0,
        'total_buying': Decimal('0.00'),
        'total_selling': Decimal('0.00'),
        'total_subtotal': Decimal('0.00'),
    }
    if not product_ids:
        return empty

    money_field = DecimalField(max_digits=18, decimal_places=2)
    qty_field = DecimalField(max_digits=14, decimal_places=3)
    batch_totals = ProductStockBatch.objects.filter(
        product_id__in=product_ids,
        quantity__gt=0,
    ).aggregate(
        total_units=Coalesce(Sum('quantity'), _zero_qty(), output_field=qty_field),
        total_buying=Coalesce(
            Sum(_money_expr('quantity', 'cost')),
            _zero_money(),
            output_field=money_field,
        ),
    )

    batched_ids = set(
        ProductStockBatch.objects.filter(
            product_id__in=product_ids,
            quantity__gt=0,
        ).values_list('product_id', flat=True)
    )
    leftover_qs = product_qs.exclude(id__in=batched_ids).filter(stock_quantity__gt=0)
    leftover_totals = leftover_qs.aggregate(
        total_units=Coalesce(Sum('stock_quantity'), _zero_qty(), output_field=qty_field),
        total_buying=Coalesce(
            Sum(_money_expr('stock_quantity', 'cost')),
            _zero_money(),
            output_field=money_field,
        ),
    )

    total_units = _as_qty(batch_totals['total_units']) + _as_qty(leftover_totals['total_units'])
    total_buying = (batch_totals['total_buying'] or Decimal('0.00')) + (
        leftover_totals['total_buying'] or Decimal('0.00')
    )
    total_buying = Decimal(total_buying).quantize(Decimal('0.01'))

    # Selling value of current stock (same unit counts as buying)
    units_by_product = defaultdict(lambda: Decimal('0'))
    for row in ProductStockBatch.objects.filter(
        product_id__in=product_ids,
        quantity__gt=0,
    ).values('product_id', 'quantity'):
        units_by_product[row['product_id']] += _as_qty(row['quantity'])
    for row in leftover_qs.values('id', 'stock_quantity'):
        units_by_product[row['id']] = _as_qty(row['stock_quantity'])

    price_map = dict(product_qs.values_list('id', 'price'))
    total_selling = Decimal('0.00')
    for pid, units in units_by_product.items():
        total_selling += units * Decimal(price_map.get(pid) or 0)
    total_selling = total_selling.quantize(Decimal('0.01'))

    sold_by_price = _qty_sold_by_price(product_ids)
    sold_map = {
        pid: sum((_as_qty(t['qty_sold']) for t in tiers), Decimal('0'))
        for pid, tiers in sold_by_price.items()
    }
    total_qty_sold = sum(sold_map.values(), Decimal('0'))
    history_original_map = _original_qty_from_history(product_ids)
    total_subtotal = Decimal('0.00')
    for product in product_qs.prefetch_related('stock_batches'):
        row = _product_inventory_value_row(product)
        _apply_qty_sold_values(
            row,
            sold_map.get(product.id, 0),
            history_original=history_original_map.get(product.id),
            price_tiers=sold_by_price.get(product.id),
        )
        total_subtotal += row['subtotal']
    total_subtotal = total_subtotal.quantize(Decimal('0.01'))

    return {
        'product_count': product_count,
        'total_units': total_units,
        'total_qty_sold': total_qty_sold,
        'total_buying': total_buying,
        'total_selling': total_selling,
        'total_subtotal': total_subtotal,
    }


def _inventory_products_queryset_from_request(request):
    """Shared product list filters for inventory page and PDF/Excel export."""
    search_query = (request.GET.get('search') or '').strip()
    filter_type = (request.GET.get('filter') or 'all').strip() or 'all'
    if search_query:
        filter_type = 'all'

    products = Product.objects.select_related('category', 'giveaway').prefetch_related(
        'stock_batches', 'sale_units',
    ).all()
    if request_can_manage_giveaways(request) or filter_type == 'giveaway':
        products = annotate_giveaway_units_given(products)

    if filter_type == 'low_stock':
        products = products.filter(
            is_active=True,
            stock_quantity__lte=F('low_stock_threshold'),
            stock_quantity__gt=0,
        )
    elif filter_type == 'out_of_stock':
        products = products.filter(is_active=True, stock_quantity=0)
    elif filter_type == 'giveaway':
        products = products.filter(giveaway_units_given__gt=0)

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query)
            | Q(barcode__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(category__name__icontains=search_query)
        )

    return products.order_by('name'), search_query, filter_type


def _product_inventory_value_row(product):
    """
    Per-product stock and list buy/sell prices.
    Buying Value is based on original qty after qty sold is applied.
    Selling value and REVENUE PER ITEM (subtotal) are filled later from qty sold.
    ``stock`` is current units on hand (stock left).
    """
    units = Decimal('0')
    buy_value = Decimal('0.00')
    old_batch = product.old_stock_batch
    new_batch = product.new_stock_batch
    if old_batch and old_batch.quantity > 0:
        qty = Decimal(str(old_batch.quantity))
        units += qty
        buy_value += qty * Decimal(old_batch.cost or 0)
    if new_batch and new_batch.quantity > 0:
        qty = Decimal(str(new_batch.quantity))
        units += qty
        buy_value += qty * Decimal(new_batch.cost or 0)

    buying_price = Decimal(product.cost or 0).quantize(Decimal('0.01'))
    selling_price = Decimal(product.price or 0).quantize(Decimal('0.01'))

    if units <= 0 and (product.stock_quantity or 0) > 0:
        units = Decimal(str(product.stock_quantity))
        buy_value = (units * buying_price).quantize(Decimal('0.01'))
    else:
        buy_value = buy_value.quantize(Decimal('0.01'))

    if units <= 0:
        status = 'Out of Stock'
    elif product.is_low_stock:
        status = 'Low Stock'
    else:
        status = 'In Stock'
    if not product.is_active:
        status = f'{status} / Inactive'

    sell_value = (units * selling_price).quantize(Decimal('0.01'))

    return {
        'name': product.name or '',
        'barcode': product.barcode or '',
        'category': product.category.name if product.category else 'Uncategorized',
        'stock': units,  # stock left (current on hand)
        'qty': units,  # original qty — updated in _apply_qty_sold_values
        'qty_sold': 0,
        'sale_price_tiers': [],
        'buying_price': buying_price,
        'selling_price': selling_price,
        'buy_value': buy_value,
        'sell_value': sell_value,
        'subtotal': Decimal('0.00'),
        'profit': Decimal('0.00'),
        'status': status,
        'is_active': bool(product.is_active),
    }


def _original_qty_from_history(product_ids):
    """
    Map product_id → earliest known stock total from ProductStockHistory.
    Used so the report ``qty`` column stays at the original stock (e.g. 50)
    even after sales reduce stock left.
    """
    if not product_ids:
        return {}
    result = {}
    for entry in (
        ProductStockHistory.objects
        .filter(product_id__in=product_ids)
        .order_by('created_at', 'id')
        .only('product_id', 'change_type', 'total_before', 'total_after')
    ):
        if entry.product_id in result:
            continue
        if entry.change_type == ProductStockHistory.CHANGE_CREATED:
            result[entry.product_id] = entry.total_after or 0
        else:
            result[entry.product_id] = entry.total_before or 0
    return result


def _qty_sold_by_price(product_ids):
    """
    Map product_id → list of {unit_price, qty_sold, sell_value} from completed
    (non-refunded) sales, one entry per distinct transaction unit_price.
    """
    if not product_ids:
        return {}
    tiers_by_product = defaultdict(list)
    for row in (
        TransactionItem.objects.filter(
            product_id__in=product_ids,
            transaction__status='completed',
            refunded_at__isnull=True,
        )
        .values('product_id', 'unit_price')
        .annotate(
            qty_sold=Coalesce(
                Sum('quantity'),
                _zero_qty(),
                output_field=DecimalField(max_digits=14, decimal_places=3),
            )
        )
        .order_by('product_id', 'unit_price')
    ):
        qty = max(_as_qty(row['qty_sold']), Decimal('0'))
        if qty <= 0:
            continue
        price = Decimal(row['unit_price'] or 0).quantize(Decimal('0.01'))
        tiers_by_product[row['product_id']].append({
            'unit_price': price,
            'qty_sold': qty,
            'sell_value': (price * qty).quantize(Decimal('0.01')),
        })
    return dict(tiers_by_product)


def _format_sale_price_lines(row, *, currency_symbol='', multiline=True):
    """
    Human-readable selling price / qty sold / selling value lines
    (one line per distinct sale price) for PDF or Excel cells.
    Returns (price_text, qty_text, sell_value_text).
    """
    tiers = row.get('sale_price_tiers') or []
    sep = '\n' if multiline else '; '
    if not tiers:
        price = Decimal(row.get('selling_price') or 0).quantize(Decimal('0.01'))
        sell_value = Decimal(row.get('sell_value') or 0).quantize(Decimal('0.01'))
        return (
            f'{currency_symbol}{float(price):,.2f}',
            '0',
            f'{currency_symbol}{float(sell_value):,.2f}',
        )
    price_lines = [
        f'{currency_symbol}{float(t["unit_price"]):,.2f}' for t in tiers
    ]
    qty_lines = [str(int(t['qty_sold'])) for t in tiers]
    value_lines = [
        f'{currency_symbol}{float(t["sell_value"]):,.2f}' for t in tiers
    ]
    return sep.join(price_lines), sep.join(qty_lines), sep.join(value_lines)


def _apply_qty_sold_values(row, qty_sold, *, history_original=None, price_tiers=None):
    """
    Set qty sold, original qty, and revalue buying/selling amounts.

    ``qty`` (original) = max(history original, stock left + qty sold)
    so the report keeps the starting inventory (e.g. 50) while
    ``stock`` remains the current stock left (e.g. 49).

    When ``price_tiers`` is provided (sales at one or more unit prices),
    selling value uses actual sale prices × qty at each price.
    """
    tiers = [
        {
            'unit_price': Decimal(t['unit_price']).quantize(Decimal('0.01')),
            'qty_sold': max(int(t.get('qty_sold') or 0), 0),
            'sell_value': Decimal(t.get('sell_value') or 0).quantize(Decimal('0.01')),
        }
        for t in (price_tiers or [])
        if max(int(t.get('qty_sold') or 0), 0) > 0
    ]
    if tiers:
        sold = sum(t['qty_sold'] for t in tiers)
        sell_value = sum((t['sell_value'] for t in tiers), Decimal('0.00')).quantize(Decimal('0.01'))
        # Prefer a single display price when only one tier; otherwise keep list price
        # as the primary and rely on sale_price_tiers for the breakdown.
        if len(tiers) == 1:
            row['selling_price'] = tiers[0]['unit_price']
    else:
        sold = max(int(qty_sold or 0), 0)
        sell_value = (Decimal(sold) * Decimal(row['selling_price'])).quantize(Decimal('0.01'))

    stock_left = int(row.get('stock') or 0)
    reconstructed = stock_left + sold
    if history_original is not None:
        original = max(int(history_original), reconstructed)
    else:
        original = reconstructed

    row['qty_sold'] = sold
    row['qty'] = original
    row['sale_price_tiers'] = tiers
    # Buying value = list price × original qty (history qty)
    row['buy_value'] = (Decimal(original) * Decimal(row['buying_price'])).quantize(Decimal('0.01'))
    row['sell_value'] = sell_value
    # Profit on sold units only: selling value − (buying price × qty sold)
    sold_buy_value = (Decimal(sold) * Decimal(row['buying_price'])).quantize(Decimal('0.01'))
    row['profit'] = (sell_value - sold_buy_value).quantize(Decimal('0.01'))
    row['subtotal'] = (row['sell_value'] - row['buy_value']).quantize(Decimal('0.01'))
    return row


def _register_pdf_unicode_fonts():
    """
    Register a Unicode TTF so ReportLab can render ₱ (Helvetica cannot).
    Returns (regular_font_name, bold_font_name).
    """
    regular_name = 'GenGlowPdf'
    bold_name = 'GenGlowPdf-Bold'
    if regular_name in pdfmetrics.getRegisteredFontNames():
        registered = set(pdfmetrics.getRegisteredFontNames())
        return regular_name, bold_name if bold_name in registered else regular_name

    candidates = [
        (Path(r'C:\Windows\Fonts\arial.ttf'), Path(r'C:\Windows\Fonts\arialbd.ttf')),
        (Path(r'C:\Windows\Fonts\calibri.ttf'), Path(r'C:\Windows\Fonts\calibrib.ttf')),
        (Path(r'C:\Windows\Fonts\segoeui.ttf'), Path(r'C:\Windows\Fonts\segoeuib.ttf')),
        (Path(r'C:\Windows\Fonts\tahoma.ttf'), Path(r'C:\Windows\Fonts\tahomabd.ttf')),
    ]
    for regular_path, bold_path in candidates:
        if not regular_path.exists():
            continue
        try:
            pdfmetrics.registerFont(TTFont(regular_name, str(regular_path)))
            if bold_path.exists():
                pdfmetrics.registerFont(TTFont(bold_name, str(bold_path)))
                pdfmetrics.registerFontFamily(
                    regular_name,
                    normal=regular_name,
                    bold=bold_name,
                )
                return regular_name, bold_name
            pdfmetrics.registerFontFamily(
                regular_name,
                normal=regular_name,
                bold=regular_name,
            )
            return regular_name, regular_name
        except Exception:
            continue
    return 'Helvetica', 'Helvetica-Bold'


@login_required
@require_http_methods(['GET'])
def export_inventory_price_report(request):
    """
    Download product list price report (PDF or Excel).

    Columns: Product | Buying price | qty | buying value | selling price |
    qty sold | selling value | PROFIT PER ITEM | stock left.

    qty = original stock from history (or stock left + qty sold).
    stock left = units currently on hand.
    Buying Value = buying price × qty (original).
    Selling price / qty sold = each distinct sale unit_price with units sold
    at that price (from completed, non-refunded transaction lines).
    Selling Value = sum of (sale price × qty at that price).
    PROFIT PER ITEM = selling value − (buying price × qty sold).

    Rows with qty sold > 0 are listed first (highest qty sold first) so
    multi-price sale breakdowns appear at the top of the download.
    """
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to export this report.')
        return redirect('kiosk_home')

    requested_format = (request.GET.get('format') or 'pdf').strip().lower()
    if requested_format not in ('pdf', 'excel'):
        requested_format = 'pdf'

    products_qs, search_query, filter_type = _inventory_products_queryset_from_request(request)
    products_list = list(products_qs)
    product_ids = [p.id for p in products_list]
    sold_by_price = _qty_sold_by_price(product_ids)
    sold_map = {
        pid: sum(t['qty_sold'] for t in tiers)
        for pid, tiers in sold_by_price.items()
    }
    history_original_map = _original_qty_from_history(product_ids)
    rows = []
    for product in products_list:
        row = _product_inventory_value_row(product)
        _apply_qty_sold_values(
            row,
            sold_map.get(product.id, 0),
            history_original=history_original_map.get(product.id),
            price_tiers=sold_by_price.get(product.id),
        )
        rows.append(row)
    # Products with sales first (multi-price breakdown visible at top of PDF/Excel),
    # then unsold products alphabetically.
    rows.sort(
        key=lambda r: (
            0 if int(r.get('qty_sold') or 0) > 0 else 1,
            -int(r.get('qty_sold') or 0),
            -float(r.get('sell_value') or 0),
            (r.get('name') or '').lower(),
        )
    )
    total_qty = sum(int(r['qty'] or 0) for r in rows)
    total_stock_left = sum(int(r['stock'] or 0) for r in rows)
    total_qty_sold = sum(int(r['qty_sold'] or 0) for r in rows)
    total_buying_value = sum((r['buy_value'] for r in rows), Decimal('0.00')).quantize(Decimal('0.01'))
    total_selling_value = sum((r['sell_value'] for r in rows), Decimal('0.00')).quantize(Decimal('0.01'))
    total_profit = sum((r['profit'] for r in rows), Decimal('0.00')).quantize(Decimal('0.01'))
    summary = _compute_inventory_price_summary(products_qs)
    # Prefer row-derived totals so export matches the table exactly.
    summary['total_units'] = total_stock_left
    summary['total_qty'] = total_qty
    summary['total_qty_sold'] = total_qty_sold
    summary['total_buying'] = total_buying_value
    summary['total_selling'] = total_selling_value
    summary['total_profit'] = total_profit

    filter_labels = {
        'all': 'All products',
        'low_stock': 'Low stock',
        'out_of_stock': 'Out of stock',
        'giveaway': 'With giveaways recorded',
    }
    scope_parts = [filter_labels.get(filter_type, filter_type)]
    if search_query:
        scope_parts.append(f'Search: "{search_query}"')
    scope_label = ' · '.join(scope_parts)

    gen_at = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
    user_label = request.user.get_full_name() or request.user.username
    date_slug = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    currency_symbol = '₱'
    pdf_font, pdf_font_bold = _register_pdf_unicode_fonts()

    report_headers = [
        'Product',
        'Buying price',
        'qty',
        'buying value',
        'selling price',
        'qty sold',
        'selling value',
        'PROFIT PER ITEM',
        'stock left',
    ]

    if requested_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Price report'
        header_fill = PatternFill(start_color='1F7A3A', end_color='1F7A3A', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin = Side(style='thin', color='D1D5DB')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws['A1'] = 'INVENTORY PRICE REPORT'
        ws['A1'].font = Font(size=14, bold=True, color='166534')
        ws['A2'] = f'Scope: {scope_label}'
        ws['A3'] = f'Generated: {gen_at} — {user_label}'
        ws['A4'] = (
            f'Products: {summary["product_count"]} | Units: {summary["total_units"]} | '
            f'Qty sold: {summary["total_qty_sold"]} | '
            f'Total buying value: {float(summary["total_buying"]):,.2f} | '
            f'Total selling value: {float(summary["total_selling"]):,.2f} | '
            f'Total profit: {float(summary["total_profit"]):,.2f}'
        )

        hdr_row = 6
        for col, val in enumerate(report_headers, start=1):
            c = ws.cell(row=hdr_row, column=col, value=val)
            c.fill = header_fill
            c.font = header_font
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in rows:
            price_text, qty_text, value_text = _format_sale_price_lines(
                row, currency_symbol='', multiline=True,
            )
            ws.append([
                row['name'],
                float(row['buying_price']),
                row['qty'],
                float(row['buy_value']),
                price_text,
                qty_text,
                value_text,
                float(row['profit']),
                row['stock'],
            ])

        money_cols = {2, 4, 8}
        qty_cols = {3, 9}
        for r in range(hdr_row + 1, ws.max_row + 1):
            for col in range(1, 10):
                cell = ws.cell(row=r, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='right' if col in money_cols or col in (5, 7) else (
                        'center' if col in qty_cols or col == 6 else 'left'
                    ),
                    vertical='center',
                    wrap_text=True,
                )
                if col in money_cols:
                    cell.number_format = '#,##0.00'

        totals_row = ws.max_row + 1
        ws.cell(row=totals_row, column=1, value='TOTAL')
        ws.cell(row=totals_row, column=3, value=summary['total_qty'])
        ws.cell(row=totals_row, column=4, value=float(summary['total_buying']))
        ws.cell(row=totals_row, column=6, value=summary['total_qty_sold'])
        ws.cell(row=totals_row, column=7, value=float(summary['total_selling']))
        ws.cell(row=totals_row, column=8, value=float(summary['total_profit']))
        ws.cell(row=totals_row, column=9, value=summary['total_units'])
        for col in range(1, 10):
            cell = ws.cell(row=totals_row, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0B5F3A', end_color='0B5F3A', fill_type='solid')
            cell.border = thin_border
            if col in (4, 7, 8):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col in (3, 6, 9):
                cell.alignment = Alignment(horizontal='center')

        widths = {
            'A': 30, 'B': 12, 'C': 8, 'D': 13, 'E': 12, 'F': 10,
            'G': 13, 'H': 14, 'I': 11,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="inventory_price_report_{date_slug}.xlsx"'
        return resp

    # PDF (landscape)
    buffer = io.BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page,
        rightMargin=22,
        leftMargin=22,
        topMargin=28,
        bottomMargin=28,
    )
    styles = getSampleStyleSheet()
    pdf_primary = colors.HexColor('#ED1C24')
    pdf_primary_dark = colors.HexColor('#C4121A')
    pdf_row_alt = colors.HexColor('#FEF7D5')
    title_style = ParagraphStyle(
        'InvPriceTitle',
        parent=styles['Heading1'],
        fontName=pdf_font_bold,
        fontSize=16,
        textColor=pdf_primary_dark,
        alignment=TA_LEFT,
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        'InvPriceMeta',
        parent=styles['Normal'],
        fontName=pdf_font,
        fontSize=9,
        textColor=colors.HexColor('#475569'),
        leading=12,
        spaceAfter=3,
    )
    cell_style = ParagraphStyle(
        'InvPriceCell',
        parent=styles['Normal'],
        fontName=pdf_font,
        fontSize=7,
        leading=8.5,
    )
    money_style = ParagraphStyle(
        'InvPriceMoney',
        parent=styles['Normal'],
        fontName=pdf_font,
        fontSize=7,
        leading=8.5,
        alignment=TA_RIGHT,
    )
    header_cell_style = ParagraphStyle(
        'InvPriceHeaderCell',
        parent=styles['Normal'],
        fontName=pdf_font_bold,
        fontSize=7,
        leading=8.5,
        textColor=colors.whitesmoke,
        alignment=TA_CENTER,
    )

    elements = [
        Paragraph('INVENTORY PRICE REPORT', title_style),
        Paragraph(f'Scope: {scope_label}', meta_style),
        Paragraph(f'Generated: {gen_at} &mdash; {escape(user_label)}', meta_style),
        Paragraph(
            f'Products: {summary["product_count"]} &nbsp;|&nbsp; Units: {summary["total_units"]:,} '
            f'&nbsp;|&nbsp; Qty sold: {summary["total_qty_sold"]:,} '
            f'&nbsp;|&nbsp; Total buying value: {currency_symbol}{float(summary["total_buying"]):,.2f} '
            f'&nbsp;|&nbsp; Total selling value: {currency_symbol}{float(summary["total_selling"]):,.2f} '
            f'&nbsp;|&nbsp; Total profit: {currency_symbol}{float(summary["total_profit"]):,.2f}',
            meta_style,
        ),
        Spacer(1, 0.15 * inch),
    ]

    table_data = [[Paragraph(escape(h), header_cell_style) for h in report_headers]]
    for row in rows:
        price_text, qty_text, value_text = _format_sale_price_lines(
            row, currency_symbol=currency_symbol, multiline=True,
        )
        price_html = '<br/>'.join(escape(line) for line in price_text.split('\n'))
        qty_html = '<br/>'.join(escape(line) for line in qty_text.split('\n'))
        value_html = '<br/>'.join(escape(line) for line in value_text.split('\n'))
        table_data.append([
            Paragraph(escape((row['name'] or '—')[:42]), cell_style),
            Paragraph(f"{currency_symbol}{float(row['buying_price']):,.2f}", money_style),
            f"{int(row['qty'] or 0):,}",
            Paragraph(f"{currency_symbol}{float(row['buy_value']):,.2f}", money_style),
            Paragraph(price_html, money_style),
            Paragraph(qty_html, ParagraphStyle(
                'InvPriceQtyCell',
                parent=cell_style,
                alignment=TA_CENTER,
            )),
            Paragraph(value_html, money_style),
            Paragraph(f"{currency_symbol}{float(row['profit']):,.2f}", money_style),
            f"{int(row['stock'] or 0):,}",
        ])

    table_data.append([
        Paragraph('<b>TOTAL</b>', cell_style),
        '',
        f"{summary['total_qty']:,}",
        Paragraph(f"<b>{currency_symbol}{float(summary['total_buying']):,.2f}</b>", money_style),
        '',
        f"{summary['total_qty_sold']:,}",
        Paragraph(f"<b>{currency_symbol}{float(summary['total_selling']):,.2f}</b>", money_style),
        Paragraph(f"<b>{currency_symbol}{float(summary['total_profit']):,.2f}</b>", money_style),
        f"{summary['total_units']:,}",
    ])
    totals_idx = len(table_data) - 1

    # Landscape usable width ~11.3" with 22pt margins
    col_widths = [
        1.75 * inch,  # Product
        0.95 * inch,  # Buying price
        0.55 * inch,  # qty
        1.05 * inch,  # buying value
        1.00 * inch,  # selling price
        0.65 * inch,  # qty sold
        1.05 * inch,  # selling value
        1.15 * inch,  # PROFIT PER ITEM
        0.75 * inch,  # stock left
    ]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), pdf_primary),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), pdf_font_bold),
        ('FONTNAME', (0, 1), (-1, -1), pdf_font),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, totals_idx - 1), 7),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
        ('ALIGN', (5, 0), (5, -1), 'CENTER'),
        ('ALIGN', (6, 0), (7, -1), 'RIGHT'),
        ('ALIGN', (8, 0), (8, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('ROWBACKGROUNDS', (0, 1), (-1, totals_idx - 1), [colors.white, pdf_row_alt]),
        ('BACKGROUND', (0, totals_idx), (-1, totals_idx), pdf_primary_dark),
        ('TEXTCOLOR', (0, totals_idx), (-1, totals_idx), colors.whitesmoke),
        ('FONTNAME', (0, totals_idx), (-1, totals_idx), pdf_font_bold),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(
        f'Notes: Products with sales are listed first (highest qty sold first), '
        f'then products with no sales alphabetically. '
        f'qty = original stock from history (stays at starting inventory; '
        f'falls back to stock left + qty sold when no history exists). '
        f'stock left = units currently on hand. '
        f'selling price / qty sold / selling value = each distinct sale price with '
        f'units sold and value at that price (completed, non-refunded sales). '
        f'buying value = buying price x qty. '
        f'PROFIT PER ITEM = selling value - (buying price x qty sold). '
        f'Amounts in {currency_symbol}.',
        meta_style,
    ))

    doc.build(elements)
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="inventory_price_report_{date_slug}.pdf"'
    return resp


_PURCHASE_HISTORY_HEADERS = [
    'Date',
    'Product',
    'Barcode',
    'Type',
    'Description',
    'Qty',
    'Buying price',
    'Buying value',
    'Selling price',
    'Sale amount',
    'Stock before',
    'Stock after',
    'Note',
]

_PURCHASE_SHEET_HEADERS = [
    'Date',
    'Product',
    'Barcode',
    'Description',
    'Qty purchased',
    'Buying price',
    'Buying value',
    'List selling price',
    'Stock after',
    'Note',
]

_SALE_SHEET_HEADERS = [
    'Date',
    'Product',
    'Barcode',
    'Description',
    'Qty sold',
    'Selling price',
    'Sale amount',
    'Buying price (cost)',
    'Stock before',
    'Stock after',
    'Note',
]


def _money_cell(value):
    if value is None or value == '':
        return None
    return float(value)


def _collect_product_history_ledger_rows(products):
    """Flatten event-level purchase/sale history for one or many products."""
    product_ids = [p.id for p in products]
    history_by_product = defaultdict(list)
    if product_ids:
        for entry in (
            ProductStockHistory.objects
            .filter(product_id__in=product_ids)
            .order_by('created_at', 'id')
        ):
            history_by_product[entry.product_id].append(entry)

    rows = []
    for product in products:
        rows.extend(
            build_product_history_ledger_rows(
                product,
                history_entries=history_by_product.get(product.id, []),
            )
        )
    return rows


def _purchase_history_excel_response(rows, *, title, scope_label, user_label, filename_stem):
    gen_at = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
    date_slug = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')

    purchase_rows = [r for r in rows if r.get('event_type') == 'PURCHASE']
    sale_rows = [r for r in rows if r.get('event_type') == 'SALE']
    total_qty_purchased = sum(int(r.get('qty') or 0) for r in purchase_rows)
    total_buy = sum(
        (r.get('buying_value') or Decimal('0.00') for r in purchase_rows),
        Decimal('0.00'),
    ).quantize(Decimal('0.01'))
    total_qty_sold = sum(int(r.get('qty') or 0) for r in sale_rows)
    total_sale = sum(
        (r.get('sale_value') or Decimal('0.00') for r in sale_rows),
        Decimal('0.00'),
    ).quantize(Decimal('0.01'))

    wb = Workbook()
    header_fill = PatternFill(start_color='1F7A3A', end_color='1F7A3A', fill_type='solid')
    sale_header_fill = PatternFill(start_color='B45309', end_color='B45309', fill_type='solid')
    purchase_fill = PatternFill(start_color='ECFDF3', end_color='ECFDF3', fill_type='solid')
    sale_fill = PatternFill(start_color='FFF7ED', end_color='FFF7ED', fill_type='solid')
    price_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D1D5DB')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    type_fills = {
        'PURCHASE': purchase_fill,
        'SALE': sale_fill,
        'PRICE': price_fill,
    }

    def _write_header(ws, headers, fill=header_fill, start_row=7):
        for col, val in enumerate(headers, start=1):
            c = ws.cell(row=start_row, column=col, value=val)
            c.fill = fill
            c.font = header_font
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        return start_row

    def _style_body(ws, hdr_row, money_cols, qty_cols, end_col):
        for r in range(hdr_row + 1, ws.max_row + 1):
            for col in range(1, end_col + 1):
                cell = ws.cell(row=r, column=col)
                cell.border = thin_border
                if col in money_cols and cell.value is not None:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col in qty_cols:
                    cell.alignment = Alignment(horizontal='center')

    # --- Sheet 1: Purchases only ---
    ws_buy = wb.active
    ws_buy.title = 'Purchases'
    ws_buy['A1'] = title
    ws_buy['A1'].font = Font(size=14, bold=True, color='166534')
    ws_buy['A2'] = f'Scope: {scope_label}'
    ws_buy['A3'] = f'Generated: {gen_at} — {user_label}'
    ws_buy['A4'] = (
        f'Purchase rows: {len(purchase_rows)} | Qty purchased: {total_qty_purchased:,} | '
        f'Total buying value: {float(total_buy):,.2f}'
    )
    ws_buy['A5'] = 'PURCHASE only — stock in / buying. Sales are on the Sales sheet.'
    hdr_buy = _write_header(ws_buy, _PURCHASE_SHEET_HEADERS)
    for row in purchase_rows:
        ws_buy.append([
            format_period_dt(row.get('created_at')),
            row.get('name') or '',
            row.get('barcode') or '',
            row.get('description') or '',
            int(row.get('qty') or 0),
            _money_cell(row.get('buying_price')),
            _money_cell(row.get('buying_value')),
            _money_cell(row.get('selling_price')),
            row.get('stock_after') if row.get('stock_after') is not None else '',
            row.get('note') or '',
        ])
        for col in range(1, 11):
            ws_buy.cell(row=ws_buy.max_row, column=col).fill = purchase_fill
    _style_body(ws_buy, hdr_buy, money_cols={6, 7, 8}, qty_cols={5, 9}, end_col=10)
    totals_buy = ws_buy.max_row + 1
    ws_buy.cell(row=totals_buy, column=1, value='TOTAL PURCHASES')
    ws_buy.cell(row=totals_buy, column=5, value=total_qty_purchased)
    ws_buy.cell(row=totals_buy, column=7, value=float(total_buy))
    for col in range(1, 11):
        cell = ws_buy.cell(row=totals_buy, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0B5F3A', end_color='0B5F3A', fill_type='solid')
        cell.border = thin_border
        if col == 7:
            cell.number_format = '#,##0.00'
    for col, width in {
        'A': 18, 'B': 28, 'C': 14, 'D': 32, 'E': 12, 'F': 12, 'G': 13, 'H': 14, 'I': 11, 'J': 28,
    }.items():
        ws_buy.column_dimensions[col].width = width

    # --- Sheet 2: Sales only ---
    ws_sell = wb.create_sheet('Sales')
    ws_sell['A1'] = title.replace('Purchase', 'Sales') if 'Purchase' in title else f'{title} — Sales'
    ws_sell['A1'].font = Font(size=14, bold=True, color='9A3412')
    ws_sell['A2'] = f'Scope: {scope_label}'
    ws_sell['A3'] = f'Generated: {gen_at} — {user_label}'
    ws_sell['A4'] = (
        f'Sale rows: {len(sale_rows)} | Qty sold: {total_qty_sold:,} | '
        f'Total sale amount: {float(total_sale):,.2f}'
    )
    ws_sell['A5'] = 'SALE only — e.g. Sold 1 unit at selling price. Purchases are on the Purchases sheet.'
    hdr_sell = _write_header(ws_sell, _SALE_SHEET_HEADERS, fill=sale_header_fill)
    for row in sale_rows:
        ws_sell.append([
            format_period_dt(row.get('created_at')),
            row.get('name') or '',
            row.get('barcode') or '',
            row.get('description') or '',
            int(row.get('qty') or 0),
            _money_cell(row.get('selling_price')),
            _money_cell(row.get('sale_value')),
            _money_cell(row.get('buying_price')),
            row.get('stock_before') if row.get('stock_before') is not None else '',
            row.get('stock_after') if row.get('stock_after') is not None else '',
            row.get('note') or '',
        ])
        for col in range(1, 12):
            ws_sell.cell(row=ws_sell.max_row, column=col).fill = sale_fill
    _style_body(ws_sell, hdr_sell, money_cols={6, 7, 8}, qty_cols={5, 9, 10}, end_col=11)
    totals_sell = ws_sell.max_row + 1
    ws_sell.cell(row=totals_sell, column=1, value='TOTAL SALES')
    ws_sell.cell(row=totals_sell, column=5, value=total_qty_sold)
    ws_sell.cell(row=totals_sell, column=7, value=float(total_sale))
    for col in range(1, 12):
        cell = ws_sell.cell(row=totals_sell, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='9A3412', end_color='9A3412', fill_type='solid')
        cell.border = thin_border
        if col == 7:
            cell.number_format = '#,##0.00'
    for col, width in {
        'A': 18, 'B': 28, 'C': 14, 'D': 28, 'E': 10, 'F': 12, 'G': 12, 'H': 14, 'I': 11, 'J': 11, 'K': 24,
    }.items():
        ws_sell.column_dimensions[col].width = width

    # --- Sheet 3: All events timeline ---
    ws_all = wb.create_sheet('All events')
    ws_all['A1'] = f'{title} — Full timeline'
    ws_all['A1'].font = Font(size=14, bold=True, color='166534')
    ws_all['A2'] = f'Scope: {scope_label}'
    ws_all['A3'] = f'Generated: {gen_at} — {user_label}'
    ws_all['A4'] = (
        f'Events: {len(rows)} | Purchases: {len(purchase_rows)} | Sales: {len(sale_rows)} | '
        f'Buy value: {float(total_buy):,.2f} | Sale amount: {float(total_sale):,.2f}'
    )
    ws_all['A5'] = (
        'PURCHASE = stock bought/restocked (buying price). '
        'SALE = units sold (selling price / sale amount). '
        'PRICE = price change only.'
    )
    hdr_all = _write_header(ws_all, _PURCHASE_HISTORY_HEADERS)
    for row in rows:
        ws_all.append([
            format_period_dt(row.get('created_at')),
            row.get('name') or '',
            row.get('barcode') or '',
            row.get('event_type') or '',
            row.get('description') or '',
            int(row.get('qty') or 0),
            _money_cell(row.get('buying_price')),
            _money_cell(row.get('buying_value')),
            _money_cell(row.get('selling_price')),
            _money_cell(row.get('sale_value')),
            row.get('stock_before') if row.get('stock_before') is not None else '',
            row.get('stock_after') if row.get('stock_after') is not None else '',
            row.get('note') or '',
        ])
        fill = type_fills.get(row.get('event_type'))
        if fill:
            for col in range(1, 14):
                ws_all.cell(row=ws_all.max_row, column=col).fill = fill
    _style_body(ws_all, hdr_all, money_cols={7, 8, 9, 10}, qty_cols={6, 11, 12}, end_col=13)
    for col, width in {
        'A': 18, 'B': 26, 'C': 14, 'D': 10, 'E': 34, 'F': 8,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 11, 'L': 11, 'M': 24,
    }.items():
        ws_all.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename_stem}_{date_slug}.xlsx"'
    return resp


def _purchase_history_pdf_response(rows, *, title, scope_label, user_label, filename_stem):
    gen_at = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
    date_slug = timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')
    currency_symbol = '₱'
    pdf_font, pdf_font_bold = _register_pdf_unicode_fonts()

    purchase_rows = [r for r in rows if r.get('event_type') == 'PURCHASE']
    sale_rows = [r for r in rows if r.get('event_type') == 'SALE']
    total_qty_purchased = sum(int(r.get('qty') or 0) for r in purchase_rows)
    total_buy = sum(
        (r.get('buying_value') or Decimal('0.00') for r in purchase_rows),
        Decimal('0.00'),
    ).quantize(Decimal('0.01'))
    total_qty_sold = sum(int(r.get('qty') or 0) for r in sale_rows)
    total_sale = sum(
        (r.get('sale_value') or Decimal('0.00') for r in sale_rows),
        Decimal('0.00'),
    ).quantize(Decimal('0.01'))

    buffer = io.BytesIO()
    page = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page,
        rightMargin=18,
        leftMargin=18,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    pdf_primary = colors.HexColor('#ED1C24')
    pdf_primary_dark = colors.HexColor('#C4121A')
    pdf_sale = colors.HexColor('#b45309')
    pdf_row_alt = colors.HexColor('#FEF7D5')
    pdf_sale_alt = colors.HexColor('#FEF7D5')
    title_style = ParagraphStyle(
        'PurchaseHistTitle',
        parent=styles['Heading1'],
        fontName=pdf_font_bold,
        fontSize=14,
        textColor=pdf_primary_dark,
        alignment=TA_LEFT,
        spaceAfter=4,
    )
    section_style = ParagraphStyle(
        'PurchaseHistSection',
        parent=styles['Heading2'],
        fontName=pdf_font_bold,
        fontSize=11,
        textColor=pdf_primary_dark,
        spaceBefore=10,
        spaceAfter=4,
    )
    sale_section_style = ParagraphStyle(
        'SaleHistSection',
        parent=section_style,
        textColor=pdf_sale,
    )
    meta_style = ParagraphStyle(
        'PurchaseHistMeta',
        parent=styles['Normal'],
        fontName=pdf_font,
        fontSize=8,
        textColor=colors.HexColor('#475569'),
        leading=11,
        spaceAfter=2,
    )
    cell_style = ParagraphStyle(
        'PurchaseHistCell',
        parent=styles['Normal'],
        fontName=pdf_font,
        fontSize=6.5,
        leading=8,
    )
    money_style = ParagraphStyle(
        'PurchaseHistMoney',
        parent=styles['Normal'],
        fontName=pdf_font,
        fontSize=6.5,
        leading=8,
        alignment=TA_RIGHT,
    )
    header_cell_style = ParagraphStyle(
        'PurchaseHistHeader',
        parent=styles['Normal'],
        fontName=pdf_font_bold,
        fontSize=6.5,
        leading=8,
        textColor=colors.whitesmoke,
        alignment=TA_CENTER,
    )

    def _money_pdf(value):
        if value is None:
            return '—'
        return f'{currency_symbol}{float(value):,.2f}'

    elements = [
        Paragraph(escape(title), title_style),
        Paragraph(f'Scope: {escape(scope_label)}', meta_style),
        Paragraph(f'Generated: {gen_at} &mdash; {escape(user_label)}', meta_style),
        Paragraph(
            f'Purchases: {len(purchase_rows)} (qty {total_qty_purchased:,}, '
            f'buy {currency_symbol}{float(total_buy):,.2f}) &nbsp;|&nbsp; '
            f'Sales: {len(sale_rows)} (qty {total_qty_sold:,}, '
            f'sale {currency_symbol}{float(total_sale):,.2f})',
            meta_style,
        ),
    ]

    # Purchases table
    elements.append(Paragraph('Purchases (buying)', section_style))
    buy_headers = [
        'Date', 'Product', 'Type', 'Description', 'Qty', 'Buying price', 'Buying value', 'List sell', 'Stock after',
    ]
    buy_data = [[Paragraph(escape(h), header_cell_style) for h in buy_headers]]
    for row in purchase_rows:
        buy_data.append([
            Paragraph(escape(format_period_dt(row.get('created_at')) or '—'), cell_style),
            Paragraph(escape((row.get('name') or '—')[:28]), cell_style),
            Paragraph('PURCHASE', cell_style),
            Paragraph(escape((row.get('description') or '—')[:36]), cell_style),
            f"{int(row.get('qty') or 0):,}",
            Paragraph(_money_pdf(row.get('buying_price')), money_style),
            Paragraph(_money_pdf(row.get('buying_value')), money_style),
            Paragraph(_money_pdf(row.get('selling_price')), money_style),
            '' if row.get('stock_after') is None else f"{int(row.get('stock_after')):,}",
        ])
    if len(buy_data) == 1:
        buy_data.append([Paragraph('No purchase events', cell_style), '', '', '', '', '', '', '', ''])
    buy_table = Table(
        buy_data,
        colWidths=[1.15 * inch, 1.4 * inch, 0.7 * inch, 1.7 * inch, 0.45 * inch, 0.8 * inch, 0.85 * inch, 0.75 * inch, 0.7 * inch],
        repeatRows=1,
    )
    buy_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), pdf_primary),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), pdf_font),
        ('FONTSIZE', (0, 0), (-1, -1), 6.5),
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),
        ('ALIGN', (5, 1), (7, -1), 'RIGHT'),
        ('ALIGN', (8, 1), (8, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, pdf_row_alt]),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.grey),
    ]))
    elements.append(buy_table)

    # Sales table
    elements.append(Paragraph('Sales (selling)', sale_section_style))
    sell_headers = [
        'Date', 'Product', 'Type', 'Description', 'Qty', 'Selling price', 'Sale amount', 'Cost', 'Stock after',
    ]
    sell_data = [[Paragraph(escape(h), header_cell_style) for h in sell_headers]]
    for row in sale_rows:
        sell_data.append([
            Paragraph(escape(format_period_dt(row.get('created_at')) or '—'), cell_style),
            Paragraph(escape((row.get('name') or '—')[:28]), cell_style),
            Paragraph('SALE', cell_style),
            Paragraph(escape((row.get('description') or '—')[:36]), cell_style),
            f"{int(row.get('qty') or 0):,}",
            Paragraph(_money_pdf(row.get('selling_price')), money_style),
            Paragraph(_money_pdf(row.get('sale_value')), money_style),
            Paragraph(_money_pdf(row.get('buying_price')), money_style),
            '' if row.get('stock_after') is None else f"{int(row.get('stock_after')):,}",
        ])
    if len(sell_data) == 1:
        sell_data.append([Paragraph('No sale events', cell_style), '', '', '', '', '', '', '', ''])
    sell_table = Table(
        sell_data,
        colWidths=[1.15 * inch, 1.4 * inch, 0.55 * inch, 1.7 * inch, 0.45 * inch, 0.85 * inch, 0.85 * inch, 0.75 * inch, 0.7 * inch],
        repeatRows=1,
    )
    sell_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), pdf_sale),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), pdf_font),
        ('FONTSIZE', (0, 0), (-1, -1), 6.5),
        ('ALIGN', (4, 1), (4, -1), 'CENTER'),
        ('ALIGN', (5, 1), (7, -1), 'RIGHT'),
        ('ALIGN', (8, 1), (8, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, pdf_sale_alt]),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.grey),
    ]))
    elements.append(sell_table)
    elements.append(Spacer(1, 0.12 * inch))
    elements.append(Paragraph(
        'Purchases and sales are listed separately. Example: a purchase at list sell ₱2,000.00 '
        'and a sale of 1 unit at ₱1,200.00 appear as different rows. '
        f'Amounts in {currency_symbol}.',
        meta_style,
    ))

    doc.build(elements)
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename_stem}_{date_slug}.pdf"'
    return resp


@login_required
@require_http_methods(['GET'])
def export_inventory_purchase_history(request):
    """
    Download product history with purchases and sales separated.
    Excel: Purchases sheet + Sales sheet + All events timeline.
    """
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to export this report.')
        return redirect('kiosk_home')

    requested_format = (request.GET.get('format') or 'pdf').strip().lower()
    if requested_format not in ('pdf', 'excel'):
        requested_format = 'pdf'

    products_qs, search_query, filter_type = _inventory_products_queryset_from_request(request)
    products_list = list(products_qs)
    rows = _collect_product_history_ledger_rows(products_list)

    filter_labels = {
        'all': 'All products',
        'low_stock': 'Low stock',
        'out_of_stock': 'Out of stock',
        'giveaway': 'With giveaways recorded',
    }
    scope_parts = [filter_labels.get(filter_type, filter_type)]
    if search_query:
        scope_parts.append(f'Search: "{search_query}"')
    scope_label = ' · '.join(scope_parts)
    user_label = request.user.get_full_name() or request.user.username
    title = 'Inventory Purchase & Sales History'

    if requested_format == 'excel':
        return _purchase_history_excel_response(
            rows,
            title=title,
            scope_label=scope_label,
            user_label=user_label,
            filename_stem='inventory_purchase_sales_history',
        )
    return _purchase_history_pdf_response(
        rows,
        title=title,
        scope_label=scope_label,
        user_label=user_label,
        filename_stem='inventory_purchase_sales_history',
    )


@login_required
@require_http_methods(['GET'])
def export_product_purchase_history(request, product_id):
    """
    Download one product's history from the History modal.
    Purchases (buying) and sales (e.g. Sold 1 unit @ ₱1,200) are separate rows/sheets.
    """
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to export this report.')
        return redirect('kiosk_home')

    try:
        product = Product.objects.select_related('category').get(id=product_id)
    except Product.DoesNotExist:
        messages.warning(request, 'Product not found.')
        return redirect('inventory_management')

    requested_format = (request.GET.get('format') or 'excel').strip().lower()
    if requested_format not in ('pdf', 'excel'):
        requested_format = 'excel'

    rows = _collect_product_history_ledger_rows([product])
    scope_label = f'Product: {product.name}'
    if product.barcode:
        scope_label += f' ({product.barcode})'
    user_label = request.user.get_full_name() or request.user.username
    title = 'Product Purchase & Sales History'
    safe_name = re.sub(r'[^a-zA-Z0-9_-]+', '_', (product.name or 'product'))[:40].strip('_') or 'product'
    filename_stem = f'product_history_{product.id}_{safe_name}'

    if requested_format == 'pdf':
        return _purchase_history_pdf_response(
            rows,
            title=title,
            scope_label=scope_label,
            user_label=user_label,
            filename_stem=filename_stem,
        )
    return _purchase_history_excel_response(
        rows,
        title=title,
        scope_label=scope_label,
        user_label=user_label,
        filename_stem=filename_stem,
    )


def _member_discount_usage_for_period(member_ids, range_start_aware, range_end_aware):
    """Per-member segment + manual discount totals on completed sales in the period."""
    from inventory.pricing import (
        _best_product_discount_price,
        _member_price_segment,
        _money,
        _segment_rules_index,
        discounts_by_product_ids,
    )

    usage = {
        mid: {
            'segment_discount_php': Decimal('0.00'),
            'manual_discount_php': Decimal('0.00'),
            'transaction_count': 0,
        }
        for mid in member_ids
    }
    if not member_ids:
        return usage

    segment_rules = list(
        SegmentProductGroupDiscount.objects.filter(is_active=True).select_related('discount_group')
    )
    amount_map, _ = _segment_rules_index(segment_rules)

    items_qs = TransactionItem.objects.filter(
        transaction__member_id__in=member_ids,
        transaction__status__in=DASHBOARD_SALE_STATUSES,
        transaction__created_at__gte=range_start_aware,
        transaction__created_at__lt=range_end_aware,
        refunded_at__isnull=True,
    ).exclude(
        transaction__transaction_number__startswith='DUMMY-',
    ).select_related('product__discount_group', 'transaction__member')

    items = list(items_qs)
    product_ids = list({i.product_id for i in items if i.product_id})
    discounts_map = discounts_by_product_ids(product_ids)
    txn_seen = {mid: set() for mid in member_ids}

    for item in items:
        mid = item.transaction.member_id
        if mid not in usage:
            continue

        if item.transaction_id not in txn_seen[mid]:
            txn_seen[mid].add(item.transaction_id)
            usage[mid]['transaction_count'] += 1

        manual = Decimal(str(item.manual_discount_php or 0)).quantize(Decimal('0.01'))
        usage[mid]['manual_discount_php'] += manual

        if not item.product_id or not item.product:
            continue
        member = item.transaction.member
        if not member or _member_price_segment(member) != SegmentProductGroupDiscount.SEG_SENIOR_PWD:
            continue
        code = (item.product.discount_group_code or '').strip()
        if not code:
            continue
        off = amount_map.get((SegmentProductGroupDiscount.SEG_SENIOR_PWD, code))
        if not off or off <= 0:
            continue

        disc_list = discounts_map.get(item.product_id) or []
        regular = _money(item.product.price)
        promo_price, _ = _best_product_discount_price(regular, disc_list)
        actual = _money(item.unit_price)
        per_unit = max(Decimal('0'), promo_price - actual)
        if per_unit > off:
            per_unit = off
        usage[mid]['segment_discount_php'] += _money(per_unit * item.quantity)

    for mid in usage:
        seg = usage[mid]['segment_discount_php']
        man = usage[mid]['manual_discount_php']
        usage[mid]['segment_discount_php'] = _money(seg)
        usage[mid]['manual_discount_php'] = _money(man)
        usage[mid]['total_discount_php'] = _money(seg + man)

    return usage


def _registered_discount_members_context(range_start_aware, range_end_aware):
    """Members with active Senior or PWD concession registration (kiosk segment pricing)."""
    rows = []
    senior_qs = (
        SeniorCitizenProfile.objects.filter(is_active=True)
        .select_related('member', 'member__member_role')
        .order_by('member__last_name', 'member__first_name')
    )
    for sp in senior_qs:
        m = sp.member
        rows.append(
            {
                'member_id': m.id,
                'full_name': m.full_name,
                'rfid_masked': mask_rfid(m.rfid_card_number),
                'kind': 'senior',
                'kind_label': 'Senior',
                'id_reference': (sp.osca_id_number or '').strip(),
                'registered_at': sp.created_at,
                'member_is_active': m.is_active,
                'role_name': m.member_role.name if getattr(m, 'member_role_id', None) else '',
            }
        )
    pwd_qs = (
        PWDProfile.objects.filter(is_active=True)
        .select_related('member', 'member__member_role')
        .order_by('member__last_name', 'member__first_name')
    )
    for pp in pwd_qs:
        m = pp.member
        rows.append(
            {
                'member_id': m.id,
                'full_name': m.full_name,
                'rfid_masked': mask_rfid(m.rfid_card_number),
                'kind': 'pwd',
                'kind_label': 'PWD',
                'id_reference': (pp.pwd_id_number or '').strip(),
                'registered_at': pp.created_at,
                'member_is_active': m.is_active,
                'role_name': m.member_role.name if getattr(m, 'member_role_id', None) else '',
            }
        )
    rows.sort(key=lambda r: (r['full_name'] or '').lower())

    usage_by_id = _member_discount_usage_for_period(
        [r['member_id'] for r in rows],
        range_start_aware,
        range_end_aware,
    )
    grand_total = Decimal('0.00')
    for row in rows:
        usage = usage_by_id.get(row['member_id'], {})
        row['segment_discount_php'] = usage.get('segment_discount_php', Decimal('0.00'))
        row['manual_discount_php'] = usage.get('manual_discount_php', Decimal('0.00'))
        row['total_discount_php'] = usage.get('total_discount_php', Decimal('0.00'))
        row['transaction_count'] = usage.get('transaction_count', 0)
        grand_total += row['total_discount_php']
    grand_total = grand_total.quantize(Decimal('0.01'))

    senior_pwd_rules = list(
        SegmentProductGroupDiscount.objects.filter(
            is_active=True,
            segment=SegmentProductGroupDiscount.SEG_SENIOR_PWD,
        )
        .select_related('discount_group')
        .order_by('discount_group__sort_order', 'discount_group__code')
    )

    senior_count = sum(1 for r in rows if r['kind'] == 'senior')
    pwd_count = sum(1 for r in rows if r['kind'] == 'pwd')
    active_rows = [
        r for r in rows
        if (r.get('total_discount_php') or Decimal('0.00')) > 0
    ]
    return {
        'registered_discount_members': active_rows,
        'registered_discount_senior_count': senior_count,
        'registered_discount_pwd_count': pwd_count,
        'registered_discount_total_count': len(rows),
        'registered_discount_usage_total': grand_total,
        'senior_pwd_segment_rules': senior_pwd_rules,
    }


def _serialize_registered_discount_members(rows):
    return [
        {
            'member_id': r['member_id'],
            'full_name': r['full_name'],
            'rfid_masked': r['rfid_masked'],
            'kind': r['kind'],
            'kind_label': r['kind_label'],
            'id_reference': r['id_reference'],
            'role_name': r['role_name'],
            'registered_at': r['registered_at'].strftime('%Y-%m-%d') if r.get('registered_at') else '',
            'member_is_active': r['member_is_active'],
            'segment_discount_php': str(r.get('segment_discount_php', Decimal('0.00'))),
            'manual_discount_php': str(r.get('manual_discount_php', Decimal('0.00'))),
            'total_discount_php': str(r.get('total_discount_php', Decimal('0.00'))),
            'transaction_count': r.get('transaction_count', 0),
        }
        for r in rows
    ]


def _manual_discount_line_items_qs(range_start_aware, range_end_aware):
    """Sale line items in the dashboard period (matches main dashboard sale rules)."""
    return TransactionItem.objects.filter(
        transaction__status__in=DASHBOARD_SALE_STATUSES,
        transaction__created_at__gte=range_start_aware,
        transaction__created_at__lt=range_end_aware,
        refunded_at__isnull=True,
    ).exclude(transaction__transaction_number__startswith='DUMMY-')


def _walk_in_sale_filter(range_start_aware, range_end_aware):
    """Non-member walk-in / guest sales in the period (active lines only)."""
    return Q(
        transaction__status__in=DASHBOARD_SALE_STATUSES,
        transaction__member__isnull=True,
        transaction__created_at__gte=range_start_aware,
        transaction__created_at__lt=range_end_aware,
        refunded_at__isnull=True,
    ) & (
        Q(transaction__walk_in_customer__isnull=False)
        | ~Q(transaction__guest_customer_name='')
    ) & ~Q(transaction__transaction_number__startswith='DUMMY-')


def _walk_in_discount_dashboard_context(range_start_aware, range_end_aware):
    """
    Walk-in / guest buyers with manual Discount ₱ in the period only (no zero-discount rows).
    """
    from transactions.walk_in_customers import normalize_customer_name

    item_period = _walk_in_sale_filter(range_start_aware, range_end_aware)
    txn_period = Q(
        status__in=DASHBOARD_SALE_STATUSES,
        member__isnull=True,
        created_at__gte=range_start_aware,
        created_at__lt=range_end_aware,
    ) & (
        Q(walk_in_customer__isnull=False) | ~Q(guest_customer_name='')
    ) & ~Q(transaction_number__startswith='DUMMY-')

    discounted_items = TransactionItem.objects.filter(
        item_period,
        manual_discount_php__gt=0,
    )
    period_totals = discounted_items.aggregate(total_manual=Sum('manual_discount_php'))
    walk_in_manual_total = (period_totals['total_manual'] or Decimal('0.00')).quantize(
        Decimal('0.01')
    )
    discounted_txn_count = (
        Transaction.objects.filter(
            txn_period,
            items__manual_discount_php__gt=0,
            items__refunded_at__isnull=True,
        )
        .distinct()
        .count()
    )

    customer_rows = {}
    product_buckets = defaultdict(
        lambda: defaultdict(lambda: Decimal('0.00'))
    )

    def _customer_key(txn):
        if txn.walk_in_customer_id:
            return ('wi', txn.walk_in_customer_id)
        name = (txn.guest_customer_name or '').strip()
        if len(name) >= 2:
            return ('guest', normalize_customer_name(name))
        return None

    for item in discounted_items.select_related('transaction', 'transaction__walk_in_customer'):
        txn = item.transaction
        key = _customer_key(txn)
        if not key:
            continue
        disc = Decimal(str(item.manual_discount_php or 0)).quantize(Decimal('0.01'))
        product_name = (item.product_name or '').strip() or '—'
        product_buckets[key][product_name] += disc
        if key not in customer_rows:
            display = txn.walk_in_customer.display_name if txn.walk_in_customer_id else (
                (txn.guest_customer_name or '').strip()
            )
            customer_rows[key] = {
                'walk_in_customer_id': txn.walk_in_customer_id,
                'display_name': display,
                'manual_discount_php': Decimal('0.00'),
                'transaction_ids': set(),
            }
        customer_rows[key]['manual_discount_php'] += disc
        customer_rows[key]['transaction_ids'].add(txn.id)

    rows = []
    for key, data in customer_rows.items():
        manual = data['manual_discount_php'].quantize(Decimal('0.01'))
        if manual <= 0:
            continue
        products = [
            {
                'product_name': pname,
                'manual_discount_php': ptotal.quantize(Decimal('0.01')),
            }
            for pname, ptotal in product_buckets[key].items()
        ]
        products.sort(key=lambda p: (-p['manual_discount_php'], p['product_name'].lower()))
        rows.append(
            {
                'walk_in_customer_id': data['walk_in_customer_id'],
                'display_name': data['display_name'],
                'manual_discount_php': manual,
                'transaction_count': len(data['transaction_ids']),
                'products': products,
            }
        )

    rows.sort(
        key=lambda r: (
            -r['manual_discount_php'],
            -(r['transaction_count'] or 0),
            (r['display_name'] or '').lower(),
        )
    )

    return {
        'walk_in_discount_customers': rows,
        'walk_in_with_discount_count': len(rows),
        'walk_in_manual_discount_total': walk_in_manual_total,
        'walk_in_discounted_txn_count': discounted_txn_count,
        'walk_in_admin_url': reverse('admin:transactions_walkincustomer_changelist'),
    }


def _serialize_walk_in_discount_customers(rows):
    return [
        {
            'walk_in_customer_id': r.get('walk_in_customer_id'),
            'display_name': r['display_name'],
            'manual_discount_php': str(r.get('manual_discount_php', Decimal('0.00'))),
            'transaction_count': r.get('transaction_count', 0),
            'products': [
                {
                    'product_name': p['product_name'],
                    'manual_discount_php': str(p['manual_discount_php']),
                }
                for p in r.get('products', [])
            ],
        }
        for r in rows
    ]


@login_required
def inventory_discount_dashboard(request):
    """
    Chart of kiosk cart manual discounts (manual_discount_php) on completed sales,
    grouped by product name. Period matches dashboard (completed txn ``created_at``).
    """
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to access this page.')
        return redirect('kiosk_home')

    b = _dashboard_period_bounds_from_get(request, default_range='all')
    today = b['today']
    min_dashboard_date = b['min_dashboard_date']
    available_years = b['available_years']
    max_week = today.strftime('%G-W%V')
    max_month = today.strftime('%Y-%m')
    today_iso = today.strftime('%Y-%m-%d')
    min_date_iso = min_dashboard_date.strftime('%Y-%m-%d')
    selected_range = b['selected_range']
    range_label = b['range_label']
    range_start = b['range_start']
    range_end = b['range_end']
    range_start_aware = b['range_start_aware']
    range_end_aware = b['range_end_aware']
    selected_date_from = range_start.strftime('%Y-%m-%d')
    selected_date_to = range_end.strftime('%Y-%m-%d')
    if selected_range == 'week':
        selected_week = range_start.strftime('%G-W%V')
    else:
        selected_week = today.strftime('%G-W%V')
    if selected_range == 'month':
        selected_month = range_start.strftime('%Y-%m')
    else:
        selected_month = today.strftime('%Y-%m')
    if selected_range == 'year':
        selected_year = range_start.year
    else:
        selected_year = today.year

    now = timezone.now()
    manual_base = _manual_discount_line_items_qs(range_start_aware, range_end_aware)
    manual_agg = manual_base.aggregate(
        total_manual=Sum('manual_discount_php'),
    )
    total_manual_discount_php = manual_agg['total_manual'] or Decimal('0.00')

    top_manual_by_product = list(
        manual_base.values('product_id', 'product_name')
        .annotate(
            total_manual=Sum('manual_discount_php'),
        )
        .filter(total_manual__gt=0)
        .order_by('-total_manual')[:25]
    )

    manual_chart_points = [
        {
            'label': ((row['product_name'] or '—').strip() or '—')[:80],
            'total': float(row['total_manual'] or 0),
        }
        for row in top_manual_by_product
    ]

    context = {
        'now': now,
        'total_manual_discount_php': total_manual_discount_php,
        'top_manual_by_product': top_manual_by_product,
        'manual_chart_points': manual_chart_points,
        'selected_range': selected_range,
        'range_label': range_label,
        'selected_week': selected_week,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'available_years': available_years,
        'max_week': max_week,
        'max_month': max_month,
        'today_iso': today_iso,
        'min_date_iso': min_date_iso,
        'selected_date_from': selected_date_from,
        'selected_date_to': selected_date_to,
        **_registered_discount_members_context(range_start_aware, range_end_aware),
        **_walk_in_discount_dashboard_context(range_start_aware, range_end_aware),
        **admin_role_badge_context(request),
    }
    return render(request, 'admin_panel/inventory_discount_dashboard.html', context)


@login_required
@require_http_methods(['GET'])
def api_inventory_manual_discount_period(request):
    """JSON: manual line discount totals for a period (same GET params as dashboard period API)."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

    b = _dashboard_period_bounds_from_get(request, default_range='all')
    range_start_aware = b['range_start_aware']
    range_end_aware = b['range_end_aware']
    manual_base = _manual_discount_line_items_qs(range_start_aware, range_end_aware)
    manual_agg = manual_base.aggregate(total_manual=Sum('manual_discount_php'))
    total_manual = manual_agg['total_manual'] or Decimal('0.00')

    top_rows = list(
        manual_base.values('product_id', 'product_name')
        .annotate(total_manual=Sum('manual_discount_php'))
        .filter(total_manual__gt=0)
        .order_by('-total_manual')[:25]
    )
    chart_points = [
        {
            'label': ((row['product_name'] or '—').strip() or '—')[:80],
            'total': float(row['total_manual'] or 0),
        }
        for row in top_rows
    ]
    table_rows = [
        {
            'product_id': row['product_id'],
            'product_name': row['product_name'],
            'total_manual': str(row['total_manual'] or Decimal('0.00')),
        }
        for row in top_rows
    ]

    reg_ctx = _registered_discount_members_context(range_start_aware, range_end_aware)
    reg_rows = _serialize_registered_discount_members(reg_ctx['registered_discount_members'])
    wi_ctx = _walk_in_discount_dashboard_context(range_start_aware, range_end_aware)
    wi_rows = _serialize_walk_in_discount_customers(wi_ctx['walk_in_discount_customers'])

    return JsonResponse(
        {
            'success': True,
            'selected_range': b['selected_range'],
            'range_label': b['range_label'],
            'date_from': b['range_start'].strftime('%Y-%m-%d'),
            'date_to': b['range_end'].strftime('%Y-%m-%d'),
            'total_manual_discount_php': str(total_manual),
            'manual_chart_points': chart_points,
            'top_manual_by_product': table_rows,
            'registered_discount_members': reg_rows,
            'registered_discount_usage_total': str(reg_ctx['registered_discount_usage_total']),
            'walk_in_discount_customers': wi_rows,
            'walk_in_with_discount_count': wi_ctx['walk_in_with_discount_count'],
            'walk_in_manual_discount_total': str(wi_ctx['walk_in_manual_discount_total']),
            'walk_in_discounted_txn_count': wi_ctx['walk_in_discounted_txn_count'],
        }
    )


@login_required
@require_http_methods(['GET'])
def export_inventory_manual_discount_report(request):
    """Download manual kiosk line discounts (completed sales) as PDF or Excel for the selected period."""
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to export this report.')
        return redirect('kiosk_home')

    requested_format = (request.GET.get('format') or 'pdf').strip().lower()
    if requested_format not in ('pdf', 'excel'):
        requested_format = 'pdf'

    b = _dashboard_period_bounds_from_get(request, default_range='all')
    range_start = b['range_start']
    range_end = b['range_end']
    range_label = b['range_label']
    range_start_aware = b['range_start_aware']
    range_end_aware = b['range_end_aware']

    manual_base = _manual_discount_line_items_qs(range_start_aware, range_end_aware)
    manual_agg = manual_base.aggregate(total_manual=Sum('manual_discount_php'))
    total_manual = manual_agg['total_manual'] or Decimal('0.00')

    rows_qs = (
        manual_base.values('product_id', 'product_name')
        .annotate(total_manual=Sum('manual_discount_php'))
        .filter(total_manual__gt=0)
        .order_by('-total_manual')
    )
    detail_rows = list(rows_qs)

    date_slug = f'{range_start.strftime("%Y%m%d")}_to_{range_end.strftime("%Y%m%d")}'
    gen_at = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
    user_label = request.user.get_full_name() or request.user.username

    if requested_format == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Manual discounts'
        header_fill = PatternFill(start_color='1F7A3A', end_color='1F7A3A', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin = Side(style='thin', color='D1D5DB')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws['A1'] = 'Manual discount report (kiosk Discount ₱)'
        ws['A1'].font = Font(size=14, bold=True, color='166534')
        ws['A2'] = f'Period: {range_label}'
        ws['A3'] = f'Sale dates (completed / partially refunded, net of voided lines): {range_start.isoformat()} to {range_end.isoformat()}'
        ws['A4'] = f'Generated: {gen_at} — {user_label}'
        ws['A5'] = f'Grand total manual discount (PHP): {float(total_manual):,.2f}'
        hdr_row = 6
        headers = ['#', 'Product ID', 'Product name', 'Total discount (PHP)']
        for col, val in enumerate(headers, start=1):
            c = ws.cell(row=hdr_row, column=col, value=val)
            c.fill = header_fill
            c.font = header_font
            c.border = thin_border
            c.alignment = Alignment(
                horizontal='center' if col in (1, 2) else 'left',
                vertical='center',
            )

        for idx, row in enumerate(detail_rows, start=1):
            pid = row['product_id']
            ws.append(
                [
                    idx,
                    pid if pid is not None else '',
                    (row['product_name'] or '').strip() or '—',
                    float(row['total_manual'] or 0),
                ]
            )
        for r in range(hdr_row + 1, ws.max_row + 1):
            for col in range(1, 5):
                cell = ws.cell(row=r, column=col)
                cell.border = thin_border
                if col == 4:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 1:
                    cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 22

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="manual_discounts_{date_slug}.xlsx"'
        return resp

    from xml.sax.saxutils import escape as xml_escape

    pdf_primary = colors.HexColor('#ED1C24')
    pdf_primary_dark = colors.HexColor('#C4121A')
    pdf_heading = colors.HexColor('#166534')
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'MDTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=pdf_primary_dark,
        spaceAfter=14,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
    )
    heading_style = ParagraphStyle(
        'MDHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=pdf_heading,
        spaceAfter=8,
        spaceBefore=10,
        fontName='Helvetica-Bold',
    )
    cell_norm = ParagraphStyle('MDCell', parent=styles['Normal'], fontSize=8, leading=10)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    elements = []
    elements.append(Paragraph('Manual discount report', title_style))
    elements.append(Paragraph('<i>Kiosk Discount ₱ per line · completed &amp; partially refunded sales (voided lines excluded)</i>', styles['Normal']))
    elements.append(Spacer(1, 0.15 * inch))
    elements.append(Paragraph(f'Period: {escape(range_label)}', styles['Normal']))
    elements.append(
        Paragraph(
            f'Sale dates: {range_start.strftime("%Y-%m-%d")} to {range_end.strftime("%Y-%m-%d")}',
            styles['Normal'],
        )
    )
    elements.append(Paragraph(f'Generated: {escape(gen_at)} — {escape(user_label)}', styles['Normal']))
    elements.append(Spacer(1, 0.12 * inch))
    elements.append(
        Paragraph(
            f'<b>Grand total (PHP):</b> {float(total_manual):,.2f}',
            styles['Normal'],
        )
    )
    elements.append(Spacer(1, 0.2 * inch))

    table_data = [['#', 'Product ID', 'Product name', 'Total discount (PHP)']]
    for idx, row in enumerate(detail_rows, start=1):
        pid = row['product_id']
        pid_s = str(pid) if pid is not None else '—'
        name = ((row['product_name'] or '').strip() or '—')[:120]
        amt = row['total_manual'] or Decimal('0.00')
        table_data.append(
            [
                str(idx),
                pid_s,
                Paragraph(xml_escape(name), cell_norm),
                f'{float(amt):,.2f}',
            ]
        )

    if len(table_data) == 1:
        table_data.append(['—', '—', Paragraph('No manual discounts in this period', cell_norm), '0.00'])

    tbl = Table(table_data, colWidths=[0.45 * inch, 0.85 * inch, 3.5 * inch, 1.35 * inch])
    tbl.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), pdf_primary),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FEF7D5')]),
            ]
        )
    )
    elements.append(Paragraph('All products with manual discount (ranked)', heading_style))
    elements.append(tbl)
    doc.build(elements)
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="manual_discounts_{date_slug}.pdf"'
    return resp


@login_required
@require_http_methods(["GET"])
def api_generate_barcode(request):
    """Return a unique EAN-13 barcode (prefix 200) not already in the database."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    exclude_barcodes = _parse_exclude_barcodes_from_request(request.GET)
    exclude_product_id = _parse_optional_product_id(request.GET.get('product_id'))
    if request.GET.get('product_id') and exclude_product_id is None:
        return JsonResponse({'success': False, 'error': 'Invalid product ID'}, status=400)

    barcode = _generate_unique_barcode(
        exclude_barcodes=exclude_barcodes,
        exclude_product_id=exclude_product_id,
    )
    if not barcode:
        return JsonResponse(
            {'success': False, 'error': 'Could not generate a unique barcode. Please try again.'},
            status=500,
        )
    return JsonResponse({'success': True, 'barcode': barcode})


@login_required
@require_http_methods(["GET"])
def api_generate_wholesale_barcode(request):
    """
    Return a unique wholesale barcode that does not match the piece barcode
    or any barcode already used by another product / sale unit.
    """
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    exclude_barcodes = _parse_exclude_barcodes_from_request(request.GET)
    piece_barcode = (request.GET.get('piece_barcode') or '').strip()
    if piece_barcode:
        exclude_barcodes.append(piece_barcode)
    current_barcode = (request.GET.get('current') or '').strip()
    if current_barcode:
        exclude_barcodes.append(current_barcode)

    exclude_product_id = _parse_optional_product_id(request.GET.get('product_id'))
    if request.GET.get('product_id') and exclude_product_id is None:
        return JsonResponse({'success': False, 'error': 'Invalid product ID'}, status=400)

    barcode = _generate_unique_barcode(
        exclude_barcodes=exclude_barcodes,
        exclude_product_id=exclude_product_id,
    )
    if not barcode:
        return JsonResponse(
            {'success': False, 'error': 'Could not generate a unique wholesale barcode. Please try again.'},
            status=500,
        )
    return JsonResponse({'success': True, 'barcode': barcode})


def _parse_exclude_barcodes_param(raw):
    if not raw:
        return []
    return [part.strip() for part in str(raw).split(',') if part.strip()]


def _parse_exclude_barcodes_from_request(get):
    codes = []
    if hasattr(get, 'getlist'):
        codes.extend(get.getlist('exclude'))
    raw = get.get('exclude')
    if raw:
        codes.append(raw)
    cleaned = []
    for code in codes:
        cleaned.extend(_parse_exclude_barcodes_param(code))
    return [code for code in cleaned if code]


def _parse_optional_product_id(raw):
    if not raw:
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def _generate_ean13_barcode():
    import random

    digits = '200' + ''.join([str(random.randint(0, 9)) for _ in range(9)])
    total = sum(int(d) * (1 if i % 2 == 0 else 3) for i, d in enumerate(digits))
    check = (10 - (total % 10)) % 10
    return digits + str(check)


def _generate_unique_barcode(*, exclude_barcodes=None, exclude_product_id=None, max_attempts=40):
    blocked = {code.strip() for code in (exclude_barcodes or []) if code and str(code).strip()}
    for _ in range(max_attempts):
        barcode = _generate_ean13_barcode()
        if barcode in blocked:
            continue
        if not _barcode_is_taken(barcode, exclude_product_id=exclude_product_id):
            return barcode
    return None


def _barcode_is_taken(barcode, *, exclude_product_id=None, exclude_sale_unit_id=None):
    if not barcode:
        return False
    product_qs = Product.objects.filter(barcode=barcode)
    if exclude_product_id:
        product_qs = product_qs.exclude(id=exclude_product_id)
    if product_qs.exists():
        return True
    unit_qs = ProductSaleUnit.objects.filter(barcode=barcode)
    if exclude_product_id:
        unit_qs = unit_qs.exclude(product_id=exclude_product_id)
    if exclude_sale_unit_id:
        unit_qs = unit_qs.exclude(id=exclude_sale_unit_id)
    return unit_qs.exists()


def _parse_sale_units_config(get):
    raw = get('sale_units_json')
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return None


def _sale_units_payload(product):
    return {'sale_units': product.dashboard_sale_units_payload()}


def _save_product_sale_unit(unit, *, product, sale_mode, unit_label, barcode, price, units_per_package, is_active):
    """Create or update one ProductSaleUnit with model validation."""
    from django.core.exceptions import ValidationError

    if unit is None:
        unit = ProductSaleUnit(product=product)
    unit.sale_mode = sale_mode
    unit.unit_label = unit_label
    unit.barcode = barcode
    unit.price = price
    unit.units_per_package = units_per_package
    unit.is_active = is_active
    try:
        unit.full_clean()
        unit.save()
    except ValidationError as exc:
        msg = '; '.join(
            f'{field}: {", ".join(errors)}'
            for field, errors in exc.message_dict.items()
        )
        return False, msg or 'Invalid sale unit data'
    except Exception as exc:
        return False, str(exc) or 'Could not save sale unit'
    return True, None


def _sync_product_sale_units(product, *, piece_barcode, piece_price, is_active, sale_units_config, unit_type=UNIT_PIECE):
    """
    Keep retail (piece/kg) and optional wholesale sale units in sync with dashboard forms.
    Returns (success, error_message).
    """
    if sale_units_config is None:
        return True, None

    if unit_type == UNIT_KILO:
        sale_units_config = {**(sale_units_config or {}), 'wholesale_enabled': False}

    wholesale_enabled = bool(sale_units_config.get('wholesale_enabled'))
    wholesale = sale_units_config.get('wholesale') or {}
    retail_label = retail_unit_label(unit_type)

    retail_unit = (
        product.sale_units.filter(sale_mode=ProductSaleUnit.SALE_MODE_RETAIL)
        .order_by('id')
        .first()
    )
    retail_exclude_id = retail_unit.id if retail_unit else None
    if _barcode_is_taken(
        piece_barcode,
        exclude_product_id=product.id,
        exclude_sale_unit_id=retail_exclude_id,
    ):
        return False, 'Barcode is already used by another product or sale unit'

    if retail_unit:
        ok, err = _save_product_sale_unit(
            retail_unit,
            product=product,
            sale_mode=ProductSaleUnit.SALE_MODE_RETAIL,
            unit_label=retail_label,
            barcode=piece_barcode,
            price=piece_price,
            units_per_package=1,
            is_active=is_active,
        )
        if not ok:
            return False, err
    else:
        ok, err = _save_product_sale_unit(
            None,
            product=product,
            sale_mode=ProductSaleUnit.SALE_MODE_RETAIL,
            unit_label=retail_label,
            barcode=piece_barcode,
            price=piece_price,
            units_per_package=1,
            is_active=is_active,
        )
        if not ok:
            return False, err

    if not wholesale_enabled:
        product.sale_units.filter(sale_mode=ProductSaleUnit.SALE_MODE_WHOLESALE).delete()
        return True, None

    label = (wholesale.get('unit_label') or '').strip()
    wholesale_barcode = (wholesale.get('barcode') or '').strip()
    if not label:
        return False, 'Wholesale unit label is required'
    if not wholesale_barcode:
        return False, 'Wholesale barcode is required'
    if wholesale_barcode == piece_barcode:
        return False, 'Wholesale barcode must differ from the piece barcode'

    try:
        wholesale_price = Decimal(str(wholesale.get('price', '0')))
        units_per_package = int(wholesale.get('units_per_package', 0))
    except (InvalidOperation, TypeError, ValueError):
        return False, 'Invalid wholesale price or units per package'

    if wholesale_price < 0:
        return False, 'Wholesale price cannot be negative'
    if units_per_package < 2:
        return False, 'Wholesale units per package must be at least 2'

    wholesale_id = wholesale.get('id')
    existing_wholesale = None
    if wholesale_id:
        existing_wholesale = product.sale_units.filter(
            id=wholesale_id,
            sale_mode=ProductSaleUnit.SALE_MODE_WHOLESALE,
        ).first()
    if not existing_wholesale:
        existing_wholesale = (
            product.sale_units.filter(sale_mode=ProductSaleUnit.SALE_MODE_WHOLESALE)
            .order_by('id')
            .first()
        )

    exclude_sale_unit_id = existing_wholesale.id if existing_wholesale else None
    if _barcode_is_taken(
        wholesale_barcode,
        exclude_product_id=product.id,
        exclude_sale_unit_id=exclude_sale_unit_id,
    ):
        return False, 'Wholesale barcode is already used by another product or sale unit'

    if existing_wholesale:
        ok, err = _save_product_sale_unit(
            existing_wholesale,
            product=product,
            sale_mode=ProductSaleUnit.SALE_MODE_WHOLESALE,
            unit_label=label,
            barcode=wholesale_barcode,
            price=wholesale_price,
            units_per_package=units_per_package,
            is_active=is_active,
        )
        if not ok:
            return False, err
        product.sale_units.filter(
            sale_mode=ProductSaleUnit.SALE_MODE_WHOLESALE,
        ).exclude(id=existing_wholesale.id).delete()
    else:
        product.sale_units.filter(sale_mode=ProductSaleUnit.SALE_MODE_WHOLESALE).delete()
        ok, err = _save_product_sale_unit(
            None,
            product=product,
            sale_mode=ProductSaleUnit.SALE_MODE_WHOLESALE,
            unit_label=label,
            barcode=wholesale_barcode,
            price=wholesale_price,
            units_per_package=units_per_package,
            is_active=is_active,
        )
        if not ok:
            return False, err

    return True, None


def _parse_stock_batch_qty(value, unit_type, default=0):
    return parse_stock_qty(value, unit_type, default=default)


def _parse_stock_batch_decimal(value):
    if value is None or value == '':
        return None
    return Decimal(str(value))


def _stock_batch_payload(product):
    old_batch = product.old_stock_batch
    new_batch = product.new_stock_batch
    return {
        'old_stock_quantity': qty_json(old_batch.quantity) if old_batch else 0,
        'old_stock_price': str(old_batch.unit_price) if old_batch else '',
        'old_stock_cost': str(old_batch.cost) if old_batch else '',
        'new_stock_quantity': qty_json(new_batch.quantity) if new_batch else 0,
        'new_stock_price': str(new_batch.unit_price) if new_batch else '',
        'new_stock_cost': str(new_batch.cost) if new_batch else '',
    }


def _apply_product_stock_batches(product, get, *, default_price=None, default_cost=None):
    """
    Persist old/new stock tiers from dashboard product forms.
    Each tier records quantity, selling price (unit_price), and buying price (cost).
    Returns combined batch quantity when batch fields are present.

    Product.price / Product.cost are synced from the active shelf tier:
    - New stock qty > 0 → use new-stock prices
    - Only old stock on hand → use old-stock prices
    - Otherwise keep the form list price/cost defaults
    Stale new-stock price fields must NOT overwrite the product when new qty is 0.
    """
    has_batch_input = any(
        get(key) not in (None, '')
        for key in (
            'old_stock_quantity',
            'old_stock_price',
            'old_stock_cost',
            'new_stock_quantity',
            'new_stock_price',
            'new_stock_cost',
        )
    )
    if not has_batch_input:
        return None

    old_qty = _parse_stock_batch_qty(get('old_stock_quantity'), product.unit_type, 0)
    new_qty = _parse_stock_batch_qty(get('new_stock_quantity'), product.unit_type, 0)
    old_selling = _parse_stock_batch_decimal(get('old_stock_price'))
    old_buying = _parse_stock_batch_decimal(get('old_stock_cost'))
    new_selling = _parse_stock_batch_decimal(get('new_stock_price'))
    new_buying = _parse_stock_batch_decimal(get('new_stock_cost'))

    fallback_selling = default_price if default_price is not None else product.price
    fallback_buying = (
        default_cost if default_cost is not None
        else (product.cost if product.cost is not None else Decimal('0.00'))
    )

    product_update_fields = []
    # Active shelf price: new stock first, else old stock, else form list price.
    # Never let a leftover new-stock price field change the product when new_qty is 0.
    if new_qty > 0 and new_selling is not None:
        if product.price != new_selling:
            product.price = new_selling
            product_update_fields.append('price')
        fallback_selling = new_selling
    elif old_qty > 0 and new_qty == 0 and old_selling is not None:
        if product.price != old_selling:
            product.price = old_selling
            product_update_fields.append('price')
        fallback_selling = old_selling
    elif default_price is not None and product.price != default_price:
        product.price = default_price
        product_update_fields.append('price')
        fallback_selling = default_price

    if new_qty > 0 and new_buying is not None:
        if product.cost != new_buying:
            product.cost = new_buying
            product_update_fields.append('cost')
        fallback_buying = new_buying
    elif old_qty > 0 and new_qty == 0 and old_buying is not None:
        if product.cost != old_buying:
            product.cost = old_buying
            product_update_fields.append('cost')
        fallback_buying = old_buying
    elif default_cost is not None and product.cost != default_cost:
        product.cost = default_cost
        product_update_fields.append('cost')
        fallback_buying = default_cost

    if product_update_fields:
        product_update_fields.append('updated_at')
        product.save(update_fields=product_update_fields)

    if old_qty > 0:
        ProductStockBatch.objects.update_or_create(
            product=product,
            tier=ProductStockBatch.TIER_OLD,
            defaults={
                'quantity': old_qty,
                'unit_price': old_selling if old_selling is not None else fallback_selling,
                'cost': old_buying if old_buying is not None else fallback_buying,
            },
        )
    else:
        ProductStockBatch.objects.filter(product=product, tier=ProductStockBatch.TIER_OLD).delete()

    if new_qty > 0:
        ProductStockBatch.objects.update_or_create(
            product=product,
            tier=ProductStockBatch.TIER_NEW,
            defaults={
                'quantity': new_qty,
                'unit_price': new_selling if new_selling is not None else fallback_selling,
                'cost': new_buying if new_buying is not None else fallback_buying,
            },
        )
    else:
        ProductStockBatch.objects.filter(product=product, tier=ProductStockBatch.TIER_NEW).delete()

    promote_new_stock_to_old_if_needed(product)
    product.refresh_from_db()
    old_batch = product.old_stock_batch
    new_batch = product.new_stock_batch
    return (old_batch.quantity if old_batch else 0) + (new_batch.quantity if new_batch else 0)


@login_required
@require_http_methods(["POST"])
def api_create_product(request):
    """Create a product without using the Django admin UI"""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    content_type = request.content_type or ''
    if 'multipart/form-data' in content_type:
        data = request.POST
        image_file = request.FILES.get('image')
        def get(key, default=None):
            return data.get(key, default)
    else:
        try:
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)
        image_file = None
        def get(key, default=None):
            return data.get(key, default)

    name = (get('name') or '').strip()
    barcode = (get('barcode') or '').strip()
    description = (get('description') or '').strip()
    category_id = get('category_id')
    is_active = get('is_active', 'true')
    if isinstance(is_active, str):
        is_active = is_active.lower() not in ('false', '0', '')
    else:
        is_active = bool(is_active)
    if not name:
        return JsonResponse({'success': False, 'error': 'Product name is required'}, status=400)
    if not barcode:
        return JsonResponse({'success': False, 'error': 'Barcode is required'}, status=400)
    if _barcode_is_taken(barcode):
        return JsonResponse({'success': False, 'error': 'A product with this barcode already exists'}, status=400)

    sale_units_config = _parse_sale_units_config(get)
    if sale_units_config is None:
        return JsonResponse({'success': False, 'error': 'Invalid sale units data'}, status=400)

    unit_type = (get('unit_type') or UNIT_PIECE).strip().lower()
    if unit_type not in (UNIT_PIECE, UNIT_KILO):
        return JsonResponse({'success': False, 'error': 'Invalid unit type. Choose piece or kilogram.'}, status=400)

    try:
        price = Decimal(str(get('price', '0')))
        cost = Decimal(str(get('cost', '0'))) if get('cost') else Decimal('0.00')
    except (InvalidOperation, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid price or cost value'}, status=400)

    try:
        stock_quantity = parse_stock_qty(get('stock_quantity', 0), unit_type, default=0)
        low_stock_threshold = parse_stock_qty(get('low_stock_threshold', 10), unit_type, default=10)
    except ValueError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    discount_group_code = (get('discount_group') or '').strip()
    valid_codes = frozenset(ProductDiscountGroup.objects.values_list('code', flat=True))
    if discount_group_code and discount_group_code not in valid_codes:
        return JsonResponse({'success': False, 'error': 'Invalid discount_group'}, status=400)
    discount_group_obj = None
    if discount_group_code:
        discount_group_obj = ProductDiscountGroup.objects.get(code=discount_group_code)

    category = None
    if category_id:
        try:
            category = Category.objects.get(id=category_id)
        except Category.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Selected category does not exist'}, status=400)

    product = Product(
        name=name,
        barcode=barcode,
        description=description,
        category=category,
        price=price,
        cost=cost,
        unit_type=unit_type,
        stock_quantity=stock_quantity,
        low_stock_threshold=low_stock_threshold,
        is_active=is_active,
        discount_group=discount_group_obj,
    )
    if image_file:
        product.image = image_file
    product.save()
    set_product_giveaway(product)

    try:
        batch_total = _apply_product_stock_batches(product, get, default_price=price, default_cost=cost)
    except ValueError as exc:
        product.delete()
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)
    if batch_total is not None and batch_total != product.stock_quantity:
        product.stock_quantity = batch_total
        product.save(update_fields=['stock_quantity', 'updated_at'])

    record_stock_history(
        product,
        ProductStockHistory.CHANGE_CREATED,
        {'old': 0, 'new': 0, 'total': 0},
        note='Product created via admin panel',
        user=request.user,
        skip_if_unchanged=False,
    )

    units_ok, units_err = _sync_product_sale_units(
        product,
        piece_barcode=barcode,
        piece_price=product.price,
        is_active=is_active,
        sale_units_config=sale_units_config,
        unit_type=unit_type,
    )
    if not units_ok:
        product.delete()
        return JsonResponse({'success': False, 'error': units_err}, status=400)

    return JsonResponse({
        'success': True,
        'message': 'Product created successfully',
        'product': {
            'id': product.id,
            'name': product.name,
            'barcode': product.barcode,
            'price': str(product.price),
            'cost': str(product.cost),
            'unit_type': product.unit_type,
            'stock_quantity': qty_json(product.stock_quantity),
            'category': product.category.name if product.category else None,
            'is_active': product.is_active,
            'is_giveaway': product.is_giveaway,
            'discount_group': product.discount_group_code,
            **_stock_batch_payload(product),
            **_sale_units_payload(product),
        }
    })


@login_required
@require_http_methods(["POST"])
def api_create_category(request):
    """Create a category without using the Django admin UI"""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    is_active = bool(data.get('is_active', True))

    if not name:
        return JsonResponse({'success': False, 'error': 'Category name is required'}, status=400)

    category = Category.objects.create(
        name=name,
        description=description,
        is_active=is_active,
    )

    return JsonResponse({
        'success': True,
        'message': 'Category created successfully',
        'category': {
            'id': category.id,
            'name': category.name,
            'description': category.description,
            'is_active': category.is_active,
        }
    })


@login_required
@require_http_methods(["POST"])
def api_update_product(request):
    """Update a product without using the Django admin UI"""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    content_type = request.content_type or ''
    if 'multipart/form-data' in content_type:
        data = request.POST
        image_file = request.FILES.get('image')
        def get(key, default=None):
            return data.get(key, default)
    else:
        try:
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)
        image_file = None
        def get(key, default=None):
            return data.get(key, default)

    product_id = get('id')
    if not product_id:
        return JsonResponse({'success': False, 'error': 'Product ID is required'}, status=400)

    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'}, status=404)

    name = (get('name') or '').strip()
    barcode = (get('barcode') or '').strip()
    description = (get('description') or '').strip()
    category_id = get('category_id')
    is_active = get('is_active', 'true')
    if isinstance(is_active, str):
        is_active = is_active.lower() not in ('false', '0', '')
    else:
        is_active = bool(is_active)
    if not name:
        return JsonResponse({'success': False, 'error': 'Product name is required'}, status=400)
    if not barcode:
        return JsonResponse({'success': False, 'error': 'Barcode is required'}, status=400)
    
    # Check if barcode is already used by another product or sale unit
    if _barcode_is_taken(barcode, exclude_product_id=product_id):
        return JsonResponse({'success': False, 'error': 'A product with this barcode already exists'}, status=400)

    sale_units_raw = get('sale_units_json')
    sale_units_config = None
    if sale_units_raw not in (None, ''):
        sale_units_config = _parse_sale_units_config(get)
        if sale_units_config is None:
            return JsonResponse({'success': False, 'error': 'Invalid sale units data'}, status=400)

    unit_type = (get('unit_type') or product.unit_type or UNIT_PIECE)
    unit_type = str(unit_type).strip().lower()
    if unit_type not in (UNIT_PIECE, UNIT_KILO):
        return JsonResponse({'success': False, 'error': 'Invalid unit type. Choose piece or kilogram.'}, status=400)

    try:
        price = Decimal(str(get('price', '0')))
        cost = Decimal(str(get('cost', '0'))) if get('cost') else Decimal('0.00')
    except (InvalidOperation, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid price or cost value'}, status=400)

    try:
        stock_quantity = parse_stock_qty(get('stock_quantity', 0), unit_type, default=0)
        low_stock_threshold = parse_stock_qty(get('low_stock_threshold', 10), unit_type, default=10)
    except ValueError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    discount_group_obj = None
    if get('discount_group') is not None:
        discount_group_code = (get('discount_group') or '').strip()
        valid_codes = frozenset(ProductDiscountGroup.objects.values_list('code', flat=True))
        if discount_group_code and discount_group_code not in valid_codes:
            return JsonResponse({'success': False, 'error': 'Invalid discount_group'}, status=400)
        if discount_group_code:
            discount_group_obj = ProductDiscountGroup.objects.get(code=discount_group_code)

    category = None
    if category_id:
        try:
            category = Category.objects.get(id=category_id)
        except Category.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Selected category does not exist'}, status=400)

    before_stock_snapshot = capture_stock_snapshot(product)

    try:
        with db_transaction.atomic():
            product.name = name
            product.barcode = barcode
            product.description = description
            product.category = category
            product.price = price
            product.cost = cost
            product.unit_type = unit_type
            product.low_stock_threshold = low_stock_threshold
            product.is_active = is_active
            if get('discount_group') is not None:
                product.discount_group = discount_group_obj
            if image_file:
                product.image = image_file
            product.save()
            set_product_giveaway(product)

            batch_total = _apply_product_stock_batches(
                product, get, default_price=price, default_cost=cost,
            )
            if batch_total is not None:
                stock_quantity = batch_total

            if product.stock_quantity != stock_quantity:
                StockManager.adjust_stock(product, stock_quantity, notes='Stock updated via admin panel')

            if sale_units_config is not None:
                units_ok, units_err = _sync_product_sale_units(
                    product,
                    piece_barcode=barcode,
                    piece_price=product.price,
                    is_active=is_active,
                    sale_units_config=sale_units_config,
                    unit_type=unit_type,
                )
                if not units_ok:
                    raise ValueError(units_err or 'Could not save sale units')
    except ValueError as exc:
        return JsonResponse({'success': False, 'error': str(exc)}, status=400)

    product.refresh_from_db()

    record_stock_history(
        product,
        ProductStockHistory.CHANGE_EDIT,
        before_stock_snapshot,
        note='Product updated via admin panel',
        user=request.user,
    )

    return JsonResponse({
        'success': True,
        'message': 'Product updated successfully',
        'product': {
            'id': product.id,
            'name': product.name,
            'barcode': product.barcode,
            'price': str(product.price),
            'cost': str(product.cost),
            'unit_type': product.unit_type,
            'stock_quantity': qty_json(product.stock_quantity),
            'category': product.category.name if product.category else None,
            'is_active': product.is_active,
            'is_giveaway': product.is_giveaway,
            'discount_group': product.discount_group_code,
            **_stock_batch_payload(product),
            **_sale_units_payload(product),
        }
    })


@login_required
@require_http_methods(["POST"])
def api_delete_product(request):
    """Permanently delete a product. Admin role only; written to the audit trail."""
    if not is_admin_user(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    product_id = data.get('id')
    if not product_id:
        return JsonResponse({'success': False, 'error': 'Product ID is required'}, status=400)

    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'}, status=404)

    product_name = product.name
    product_barcode = product.barcode or ''
    product_pk = product.pk
    stock_qty = str(product.stock_quantity)

    try:
        with db_transaction.atomic():
            product.delete()
    except Exception:
        return JsonResponse({
            'success': False,
            'error': 'Could not delete this product. It may still be linked to other records.',
        }, status=400)

    try:
        from admin_panel.audit import mark_audit_recorded, record_audit
        record_audit(
            WebsiteAuditLog.Action.INVENTORY,
            actor=request.user,
            description=f'Deleted product "{product_name}" (barcode {product_barcode})',
            request=request,
            object_type='Product',
            object_id=product_pk,
            metadata={
                'name': product_name,
                'barcode': product_barcode,
                'stock_quantity': stock_qty,
            },
        )
        mark_audit_recorded(request)
    except Exception:
        pass

    return JsonResponse({
        'success': True,
        'message': f'Product "{product_name}" deleted successfully.',
    })


@login_required
@require_http_methods(["GET"])
def api_product_stock_history(request, product_id):
    """
    Return the stock-change history for a single product plus a summary of the
    original quantity, current stock, total sold/restocked, and price history.
    Powers the "History" modal on the inventory dashboard.
    """
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'}, status=404)

    history_qs = (
        ProductStockHistory.objects
        .filter(product=product)
        .select_related('changed_by')
        .order_by('-created_at')
    )

    # Cumulative sales / restocks are derived from StockTransaction so they also
    # reflect activity from before the history feature was enabled.
    sold_total = (
        StockTransaction.objects
        .filter(product=product, transaction_type='out')
        .exclude(notes__icontains='Giveaway')
        .aggregate(total=Sum('quantity'))['total']
    ) or 0
    given_away_total = (
        StockTransaction.objects
        .filter(product=product, transaction_type='out', notes__icontains='Giveaway')
        .aggregate(total=Sum('quantity'))['total']
    ) or 0
    restocked_total = (
        StockTransaction.objects
        .filter(product=product, transaction_type='in')
        .aggregate(total=Sum('quantity'))['total']
    ) or 0

    first_record = history_qs.order_by('created_at').first()
    if first_record is None:
        original_qty = None
        original_selling = None
        original_buying = None
    elif first_record.change_type == ProductStockHistory.CHANGE_CREATED:
        original_qty = first_record.total_after
        original_selling = first_record.unit_price
        original_buying = first_record.cost
    else:
        original_qty = first_record.total_before
        original_selling = first_record.unit_price_before or first_record.unit_price
        original_buying = first_record.cost_before or first_record.cost

    price_change_count = history_qs.filter(
        change_type=ProductStockHistory.CHANGE_PRICE,
    ).count()
    # Also count edits that moved prices (legacy rows before CHANGE_PRICE existed)
    for entry in history_qs.exclude(change_type=ProductStockHistory.CHANGE_PRICE)[:200]:
        if entry.price_changed:
            price_change_count += 1

    def _money_str(value):
        return str(value) if value is not None else ''

    rows = []
    for entry in history_qs[:200]:
        rows.append({
            'id': entry.id,
            'change_type': entry.change_type,
            'change_type_display': entry.get_change_type_display(),
            'old_stock_before': qty_json(entry.old_stock_before),
            'old_stock_after': qty_json(entry.old_stock_after),
            'new_stock_before': qty_json(entry.new_stock_before),
            'new_stock_after': qty_json(entry.new_stock_after),
            'total_before': qty_json(entry.total_before),
            'total_after': qty_json(entry.total_after),
            'total_change': qty_json(entry.total_change),
            'quantity_sold': qty_json(entry.quantity_sold),
            'unit_price_before': _money_str(entry.unit_price_before),
            'unit_price': _money_str(entry.unit_price),
            'cost_before': _money_str(entry.cost_before),
            'cost': _money_str(entry.cost),
            'old_stock_price_before': _money_str(entry.old_stock_price_before),
            'old_stock_price_after': _money_str(entry.old_stock_price_after),
            'new_stock_price_before': _money_str(entry.new_stock_price_before),
            'new_stock_price_after': _money_str(entry.new_stock_price_after),
            'price_changed': bool(entry.price_changed),
            'note': entry.note,
            'changed_by': entry.changed_by.get_username() if entry.changed_by else '',
            'created_at': timezone.localtime(entry.created_at).strftime('%Y-%m-%d %I:%M %p'),
        })

    return JsonResponse({
        'success': True,
        'product': {
            'id': product.id,
            'name': product.name,
            'barcode': product.barcode or '',
            'unit_type': product.unit_type,
        },
        'summary': {
            'original_qty': qty_json(original_qty),
            'current_qty': qty_json(product.stock_quantity),
            'total_sold': qty_json(sold_total),
            'total_given_away': qty_json(given_away_total),
            'total_restocked': qty_json(restocked_total),
            'original_selling_price': _money_str(original_selling),
            'current_selling_price': _money_str(product.price),
            'original_buying_price': _money_str(original_buying),
            'current_buying_price': _money_str(product.cost),
            'price_change_count': price_change_count,
        },
        'history': rows,
    })


def _giveaway_product_search_payload(product):
    category_name = ''
    try:
        category_name = product.category.name if product.category_id else ''
    except Exception:
        category_name = ''
    return {
        'id': str(product.id),
        'name': product.name,
        'barcode': product.barcode or '',
        'stock_quantity': product.stock_quantity,
        'category': category_name,
        'text': f'{product.name} · {product.barcode or "—"} — stock: {product.stock_quantity}',
    }


@login_required
@require_http_methods(["GET"])
def api_search_giveaway_products(request):
    """Search active in-stock products for the Record Giveaway modal."""
    if not request_can_manage_giveaways(request):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    query = request.GET.get('q', '').strip()
    initial = request.GET.get('initial', '').lower() in ('1', 'true', 'yes')
    try:
        limit = int(request.GET.get('limit', 150))
    except (TypeError, ValueError):
        limit = 150
    limit = max(20, min(limit, 400))

    base_qs = (
        Product.objects.filter(is_active=True, stock_quantity__gt=0)
        .select_related('category')
        .order_by('name')
    )

    if initial and not query:
        total_available = base_qs.count()
        products_qs = base_qs[:limit]
    elif not query:
        return JsonResponse({'success': True, 'products': [], 'total': 0, 'limited': False})
    else:
        search_filters = (
            Q(name__icontains=query)
            | Q(barcode__icontains=query)
            | Q(description__icontains=query)
            | Q(category__name__icontains=query)
        )
        filtered = base_qs.filter(search_filters)
        total_available = filtered.count()
        products_qs = filtered[:limit]

    products = [_giveaway_product_search_payload(product) for product in products_qs]
    return JsonResponse({
        'success': True,
        'products': products,
        'total': total_available,
        'limited': total_available > len(products),
    })


@login_required
@require_http_methods(["POST"])
def api_record_giveaway(request):
    """Deduct stock for one or more giveaway products (free distribution)."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    items = data.get('items') or []
    if not items:
        return JsonResponse({'success': False, 'error': 'Add at least one product and quantity.'}, status=400)

    extra_notes = (data.get('notes') or '').strip()
    stock_notes = giveaway_stock_note(request.user.get_username(), extra_notes)

    parsed = []
    seen_ids = set()
    for raw in items:
        try:
            product_id = int(raw.get('product_id'))
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid product or quantity.'}, status=400)
        if product_id in seen_ids:
            return JsonResponse(
                {'success': False, 'error': 'Each product can only appear once. Combine quantities on one line.'},
                status=400,
            )
        seen_ids.add(product_id)
        parsed.append((product_id, raw.get('quantity')))

    results = []
    try:
        with db_transaction.atomic():
            for product_id, quantity_raw in parsed:
                try:
                    product = Product.objects.select_for_update().get(
                        pk=product_id,
                        is_active=True,
                    )
                except Product.DoesNotExist:
                    return JsonResponse(
                        {'success': False, 'error': 'Product not found or is inactive.'},
                        status=400,
                    )
                try:
                    quantity = parse_sale_qty(quantity_raw, product)
                except ValueError as exc:
                    return JsonResponse({'success': False, 'error': str(exc)}, status=400)
                if product.stock_quantity < quantity:
                    return JsonResponse(
                        {
                            'success': False,
                'error': (
                    f'Not enough stock for "{product.name}". '
                    f'Requested {format_product_qty(product, quantity)}, '
                    f'available {format_product_qty(product, product.stock_quantity)}.'
                ),
                        },
                        status=400,
                    )
                stock_result = StockManager.reduce_stock(product, quantity, notes=stock_notes)
                if not stock_result.success:
                    return JsonResponse({'success': False, 'error': stock_result.error or 'Stock update failed.'}, status=400)
                results.append({
                    'product_id': product.id,
                    'name': product.name,
                    'quantity': qty_json(quantity),
                    'stock_after': stock_result.stock_after,
                })
    except Exception:
        return JsonResponse({'success': False, 'error': 'Could not record giveaway. Please try again.'}, status=500)

    total_units = sum(r['quantity'] for r in results)
    return JsonResponse({
        'success': True,
        'message': f'Giveaway recorded: {total_units} unit{"s" if total_units != 1 else ""} from {len(results)} product{"s" if len(results) != 1 else ""}.',
        'results': results,
    })


@login_required
@require_http_methods(["GET"])
def api_list_giveaways(request):
    """List recorded giveaway stock transactions for management."""
    if not request_can_manage_giveaways(request):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    product_id = request.GET.get('product_id', '').strip()
    search = request.GET.get('search', '').strip()
    page = request.GET.get('page', '1')

    qs = giveaway_stock_transactions_qs().order_by('-created_at')

    if product_id.isdigit():
        qs = qs.filter(product_id=int(product_id))
    if search:
        qs = qs.filter(
            Q(product__name__icontains=search)
            | Q(product__barcode__icontains=search)
            | Q(notes__icontains=search)
        )

    paginator = Paginator(qs, 15)
    try:
        page_obj = paginator.get_page(page)
    except Exception:
        page_obj = paginator.get_page(1)

    return JsonResponse({
        'success': True,
        'giveaways': [serialize_giveaway_stock_transaction(row) for row in page_obj],
        'pagination': {
            'page': page_obj.number,
            'num_pages': paginator.num_pages,
            'total': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
        },
    })


@login_required
@require_http_methods(["POST"])
def api_update_giveaway(request):
    """Edit quantity and/or notes on a recorded giveaway."""
    if not request_can_manage_giveaways(request):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    giveaway_id = data.get('id')
    if not giveaway_id:
        return JsonResponse({'success': False, 'error': 'Giveaway record id is required.'}, status=400)

    has_quantity = 'quantity' in data and data.get('quantity') is not None
    has_notes = 'notes' in data
    if not has_quantity and not has_notes:
        return JsonResponse({'success': False, 'error': 'Provide quantity and/or notes to update.'}, status=400)

    new_quantity = None
    if has_quantity:
        try:
            new_quantity = int(data.get('quantity'))
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid quantity.'}, status=400)
        if new_quantity < 1:
            return JsonResponse({'success': False, 'error': 'Quantity must be at least 1.'}, status=400)

    new_extra_notes = (data.get('notes') or '').strip() if has_notes else None

    try:
        with db_transaction.atomic():
            stock_txn = (
                StockTransaction.objects.select_for_update()
                .select_related('product')
                .get(pk=giveaway_id)
            )
            if not is_giveaway_stock_transaction(stock_txn):
                return JsonResponse({'success': False, 'error': 'Giveaway record not found.'}, status=404)

            product = Product.objects.select_for_update().get(pk=stock_txn.product_id)
            parsed = parse_giveaway_stock_notes(stock_txn.notes)
            old_quantity = stock_txn.quantity

            if new_quantity is not None and new_quantity != old_quantity:
                delta = new_quantity - old_quantity
                if delta > 0:
                    if product.stock_quantity < delta:
                        return JsonResponse(
                            {
                                'success': False,
                                'error': (
                                    f'Not enough stock for "{product.name}". '
                                    f'Need {delta} more, available {product.stock_quantity}.'
                                ),
                            },
                            status=400,
                        )
                    product.stock_quantity -= delta
                else:
                    product.stock_quantity += abs(delta)
                product.save(update_fields=['stock_quantity', 'updated_at'])
                stock_txn.quantity = new_quantity
                stock_txn.stock_after = product.stock_quantity

            if new_extra_notes is not None:
                recorder = parsed['recorded_by'] or request.user.get_username()
                stock_txn.notes = giveaway_stock_note(recorder, new_extra_notes)

            stock_txn.save()
    except StockTransaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Giveaway record not found.'}, status=404)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Could not update giveaway. Please try again.'}, status=500)

    stock_txn.refresh_from_db()
    stock_txn.product.refresh_from_db()
    return JsonResponse({
        'success': True,
        'message': 'Giveaway updated.',
        'giveaway': serialize_giveaway_stock_transaction(stock_txn),
    })


@login_required
@require_http_methods(["POST"])
def api_delete_giveaway(request):
    """Delete a giveaway record and restore deducted stock."""
    if not request_can_manage_giveaways(request):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    giveaway_id = data.get('id')
    if not giveaway_id:
        return JsonResponse({'success': False, 'error': 'Giveaway record id is required.'}, status=400)

    try:
        with db_transaction.atomic():
            stock_txn = (
                StockTransaction.objects.select_for_update()
                .select_related('product')
                .get(pk=giveaway_id)
            )
            if not is_giveaway_stock_transaction(stock_txn):
                return JsonResponse({'success': False, 'error': 'Giveaway record not found.'}, status=404)

            product = Product.objects.select_for_update().get(pk=stock_txn.product_id)
            quantity = stock_txn.quantity
            product.stock_quantity += quantity
            product.save(update_fields=['stock_quantity', 'updated_at'])
            stock_txn.delete()
    except StockTransaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Giveaway record not found.'}, status=404)
    except Exception:
        return JsonResponse({'success': False, 'error': 'Could not delete giveaway. Please try again.'}, status=500)

    return JsonResponse({
        'success': True,
        'message': f'Giveaway deleted. Restored {quantity} unit{"s" if quantity != 1 else ""} to stock.',
    })


@login_required
@require_http_methods(["POST"])
def api_update_category(request):
    """Update a category without using the Django admin UI"""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    category_id = data.get('id')
    if not category_id:
        return JsonResponse({'success': False, 'error': 'Category ID is required'}, status=400)

    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Category not found'}, status=404)

    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    is_active = bool(data.get('is_active', True))

    if not name:
        return JsonResponse({'success': False, 'error': 'Category name is required'}, status=400)

    # Update category
    category.name = name
    category.description = description
    category.is_active = is_active
    category.save()

    return JsonResponse({
        'success': True,
        'message': 'Category updated successfully',
        'category': {
            'id': category.id,
            'name': category.name,
            'description': category.description,
            'is_active': category.is_active,
        }
    })


@login_required
@require_http_methods(["POST"])
def api_create_member_type(request):
    """Create a member type without the Django admin UI."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    is_active = bool(data.get('is_active', True))

    if not name:
        return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)

    member_type = MemberType.objects.create(
        name=name,
        description=description,
        is_active=is_active,
    )

    return JsonResponse({
        'success': True,
        'message': 'Member type created successfully',
        'member_type': {
            'id': member_type.id,
            'name': member_type.name,
            'description': member_type.description,
            'is_active': member_type.is_active,
        }
    })


@login_required
@require_http_methods(["POST"])
def api_update_member_type(request):
    """Update a member type without the Django admin UI."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    member_type_id = data.get('id')
    if not member_type_id:
        return JsonResponse({'success': False, 'error': 'Member type ID is required'}, status=400)

    try:
        member_type = MemberType.objects.get(id=member_type_id)
    except MemberType.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Member type not found'}, status=404)

    name = (data.get('name') or '').strip()
    description = (data.get('description') or '').strip()
    is_active = bool(data.get('is_active', member_type.is_active))

    if not name:
        return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)

    # Update the member type
    member_type.name = name
    member_type.description = description
    member_type.is_active = is_active
    member_type.save()

    return JsonResponse({
        'success': True,
        'message': 'Member type updated successfully',
        'member_type': {
            'id': member_type.id,
            'name': member_type.name,
            'description': member_type.description,
            'is_active': member_type.is_active,
        }
    })


@login_required
@require_http_methods(["GET"])
def api_generate_username(request):
    """Return a unique username derived from first_name + last_name."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    from helper.members_helper import generate_unique_member_username

    username = generate_unique_member_username(
        request.GET.get('first_name', ''),
        request.GET.get('last_name', ''),
    )
    if not username:
        return JsonResponse({'success': False, 'username': ''})

    return JsonResponse({'success': True, 'username': username})


def _format_model_validation_error(exc):
    if isinstance(exc, ValidationError) and hasattr(exc, 'message_dict') and exc.message_dict:
        parts = []
        for msgs in exc.message_dict.values():
            if isinstance(msgs, (list, tuple)):
                parts.extend(str(m) for m in msgs)
            else:
                parts.append(str(msgs))
        return '; '.join(parts) if parts else str(exc)
    return str(exc)


def _apply_member_concession_from_request(member, concession_raw):
    """Apply ``SeniorCitizenProfile`` / ``PWDProfile`` from API JSON. Returns None or an error string."""
    if concession_raw is None:
        return None
    if not isinstance(concession_raw, dict):
        return 'Invalid concession payload'

    ctype = (concession_raw.get('type') or 'none').strip().lower()
    if ctype not in ('none', 'senior', 'pwd'):
        return 'Concession type must be none, senior, or pwd'

    senior_in = concession_raw.get('senior') or {}
    pwd_in = concession_raw.get('pwd') or {}
    osca = (senior_in.get('osca_id_number') or '').strip()
    senior_notes = (senior_in.get('notes') or '').strip()
    pwd_id = (pwd_in.get('pwd_id_number') or '').strip()
    pwd_notes = (pwd_in.get('notes') or '').strip()

    def _senior_profile():
        return member.get_senior_profile_safe()

    def _pwd_profile():
        return member.get_pwd_profile_safe()

    try:
        with db_transaction.atomic():
            if ctype == 'none':
                sp = _senior_profile()
                if sp and sp.is_active:
                    sp.is_active = False
                    sp.save(update_fields=['is_active', 'updated_at'])
                pp = _pwd_profile()
                if pp and pp.is_active:
                    pp.is_active = False
                    pp.save(update_fields=['is_active', 'updated_at'])
                return None

            if ctype == 'senior':
                pp = _pwd_profile()
                if pp and pp.is_active:
                    pp.is_active = False
                    pp.save(update_fields=['is_active', 'updated_at'])
                sp, _created = SeniorCitizenProfile.objects.get_or_create(
                    member=member,
                    defaults={
                        'is_active': True,
                        'osca_id_number': osca,
                        'notes': senior_notes,
                    },
                )
                if not _created:
                    sp.osca_id_number = osca
                    sp.notes = senior_notes
                    sp.is_active = True
                sp.full_clean()
                sp.save()
                return None

            # pwd
            sp = _senior_profile()
            if sp and sp.is_active:
                sp.is_active = False
                sp.save(update_fields=['is_active', 'updated_at'])
            pp, _created = PWDProfile.objects.get_or_create(
                member=member,
                defaults={
                    'is_active': True,
                    'pwd_id_number': pwd_id,
                    'notes': pwd_notes,
                },
            )
            if not _created:
                pp.pwd_id_number = pwd_id
                pp.notes = pwd_notes
                pp.is_active = True
            pp.full_clean()
            pp.save()
            return None
    except ValidationError as exc:
        return _format_model_validation_error(exc)


@login_required
@require_http_methods(["POST"])
def api_create_member(request):
    """Create a member without redirecting to the admin site."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    first_name = (data.get('first_name') or '').strip()
    last_name = (data.get('last_name') or '').strip()
    username = (data.get('username') or '').strip() or None
    from helper.members_helper import (
        apply_member_complete_details,
        extract_member_complete_details,
        generate_unique_member_username,
        normalize_rfid,
        parse_member_date_joined,
        resolve_inactive_remark,
        rfid_is_taken_by_other,
    )

    rfid = normalize_rfid(data.get('rfid'))
    email = (data.get('email') or '').strip() or None
    phone = (data.get('phone') or '').strip()
    member_type_id = data.get('member_type_id')
    role = (data.get('role') or 'member').strip() or 'member'
    is_active = bool(data.get('is_active', True))
    balance_raw = data.get('balance', '0.00')
    pin = (data.get('pin') or '').strip()
    pin_attempts_raw = data.get('pin_attempts', 0)
    is_pin_locked = bool(data.get('is_pin_locked', False))

    # Staff and cashier (Member role) may only create plain 'member' accounts
    if restricts_member_role_to_member_only(request.user):
        if role not in ['member']:
            return JsonResponse({'success': False, 'error': 'You can only create members with the "member" role'}, status=403)
        role = 'member'  # Force to member role

    if (role or '').strip().lower() == 'committee':
        balance_raw = '0.00'

    if not first_name or not last_name:
        return JsonResponse({'success': False, 'error': 'First and last name are required'}, status=400)
    if not username:
        username = generate_unique_member_username(first_name, last_name) or None
    if username and Member.objects.filter(username=username).exists():
        return JsonResponse({'success': False, 'error': f'Username "{username}" already exists'}, status=400)
    if rfid and rfid_is_taken_by_other(rfid):
        return JsonResponse({'success': False, 'error': 'RFID card number already exists'}, status=400)
    if email and Member.objects.filter(email=email).exists():
        return JsonResponse({'success': False, 'error': 'Email already exists'}, status=400)
    if pin and (len(pin) != 4 or not pin.isdigit()):
        return JsonResponse({'success': False, 'error': 'PIN must be exactly 4 digits'}, status=400)

    try:
        balance = Decimal(str(balance_raw or '0.00'))
    except (InvalidOperation, TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid balance value'}, status=400)
    if balance < 0:
        return JsonResponse({'success': False, 'error': 'Balance cannot be negative'}, status=400)

    share_capital_raw = data.get('share_capital', '0.00')
    try:
        share_capital = Decimal(str(share_capital_raw or '0.00'))
    except (InvalidOperation, TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid share capital value'}, status=400)
    if share_capital < 0:
        return JsonResponse({'success': False, 'error': 'Share capital cannot be negative'}, status=400)

    try:
        pin_attempts = int(pin_attempts_raw or 0)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid PIN attempts value'}, status=400)
    if pin_attempts < 0:
        pin_attempts = 0

    try:
        date_joined = parse_member_date_joined(data.get('date_joined'))
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid registration date format'}, status=400)

    detail_fields, detail_error = extract_member_complete_details(data)
    if detail_error:
        return JsonResponse({'success': False, 'error': detail_error}, status=400)

    member_type = None
    if member_type_id:
        try:
            member_type = MemberType.objects.get(id=member_type_id)
        except MemberType.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Selected member type does not exist'}, status=400)

    inactive_remark, remark_error = resolve_inactive_remark(is_active, data.get('inactive_remark'))
    if remark_error:
        return JsonResponse({'success': False, 'error': remark_error}, status=400)

    member = Member(
        first_name=first_name,
        last_name=last_name,
        username=username,
        rfid_card_number=rfid or None,
        email=email,
        phone=phone,
        member_type=member_type,
        member_role=Role.resolve_slug(role),
        balance=balance,
        share_capital=share_capital,
        pin_attempts=pin_attempts,
        is_pin_locked=is_pin_locked,
        is_active=is_active,
        inactive_remark=inactive_remark,
        date_joined=date_joined,
    )
    apply_member_complete_details(member, detail_fields or {})
    member.save()
    if pin:
        member.set_pin(pin)

    if share_capital > 0:
        ShareCapitalTransaction.objects.create(
            member=member,
            transaction_type='opening',
            amount=share_capital,
            balance_before=Decimal('0.00'),
            balance_after=share_capital,
            notes='Opening share capital on member registration',
            performed_by=request.user if request.user.is_authenticated else None,
        )

    concession_err = _apply_member_concession_from_request(member, data.get('concession'))
    if concession_err:
        member.delete()
        return JsonResponse({'success': False, 'error': concession_err}, status=400)

    # Handle user account creation if requested
    create_user_account = data.get('create_user_account', False)
    if create_user_account:
        username = (data.get('username') or '').strip()
        password = data.get('password', '').strip()
        
        if not username:
            # Delete member if user creation fails
            member.delete()
            return JsonResponse({'success': False, 'error': 'Username is required when creating a user account'}, status=400)
        
        if not password:
            member.delete()
            return JsonResponse({'success': False, 'error': 'Password is required when creating a user account'}, status=400)
        
        # Check if username already exists
        if User.objects.filter(username=username).exists():
            member.delete()
            return JsonResponse({'success': False, 'error': f'Username "{username}" already exists'}, status=400)
        
        # Create User account
        try:
            user = User.objects.create_user(
                username=username,
                email=email or '',
                password=password,
                first_name=first_name,
                last_name=last_name,
            )
            # Link user to member
            member.user = user
            member.save()
            from members.role_permissions import sync_member_loan_permissions
            sync_member_loan_permissions(member)
        except Exception as e:
            member.delete()
            return JsonResponse({'success': False, 'error': f'Failed to create user account: {str(e)}'}, status=400)

    # Send welcome notification email to the new member (fire-and-forget, non-blocking)
    notification_sent = False
    if member.email:
        try:
            store_profile = StoreProfile.get()
            store_name = store_profile.store_name if store_profile else 'Cooperative Store'
        except Exception:
            store_name = 'Cooperative Store'
        added_by = request.user.get_full_name() or request.user.username
        send_welcome_member_email(member, store_name=store_name, added_by=added_by)
        notification_sent = True

    member.refresh_from_db()

    return JsonResponse({
        'success': True,
        'message': 'Member created successfully',
        'notification_sent': notification_sent,
        'member': {
            'id': member.id,
            'name': member.full_name,
            'rfid': member.rfid_card_number,
            'email': member.email or '',
            'phone': member.phone,
            'member_type': member.member_type.name if member.member_type else '',
            'role': member.role,
            'is_active': member.is_active,
            'balance': str(member.balance),
            'share_capital': str(member.share_capital),
            'username': member.username or '',
            'pin_set': bool(member.pin_hash),
        }
    })


def _member_edit_pin_error():
    """User-facing error when an edit-authorisation PIN check fails."""
    if not _member_edit_pin_available():
        return (
            'No member PIN is saved yet. Open a member, set a 4-digit PIN, '
            'save it, then use that PIN to authorize edits.'
        )
    return 'Invalid PIN. Enter a 4-digit PIN that is saved on a member account.'


@login_required
@require_http_methods(["POST"])
def api_verify_member_edit_pin(request):
    """Verify an edit-authorisation PIN before the edit form is opened."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    pin = (data.get('pin') or '').strip()
    if not re.fullmatch(r'\d{4}', pin):
        return JsonResponse({
            'success': False,
            'error': 'A valid 4-digit PIN is required.',
        }, status=400)

    session_member = _get_active_member_for_refill_user(request.user)
    authorising_member = _resolve_member_edit_pin_authorizer(pin, session_member)
    if not authorising_member:
        return JsonResponse({
            'success': False,
            'error': _member_edit_pin_error(),
        }, status=403)

    return JsonResponse({'success': True})


@login_required
@require_http_methods(["POST"])
def api_update_member(request):
    """Update a member without redirecting to the admin site."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    member_id = data.get('member_id')
    if not member_id:
        return JsonResponse({'success': False, 'error': 'Member ID is required'}, status=400)

    try:
        member_id = int(member_id)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid member ID'}, status=400)

    try:
        member = Member.objects.get(pk=member_id)
    except Member.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Member not found'}, status=404)

    authorizer_pin = (data.get('authorizer_pin') or '').strip()
    session_member = _get_active_member_for_refill_user(request.user)
    if not authorizer_pin:
        return JsonResponse({
            'success': False,
            'error': 'A 4-digit PIN is required to edit a member.',
        }, status=403)
    authorising_member = _resolve_member_edit_pin_authorizer(authorizer_pin, session_member)
    if not authorising_member:
        return JsonResponse({
            'success': False,
            'error': _member_edit_pin_error(),
        }, status=403)

    from helper.members_helper import (
        apply_member_complete_details,
        extract_member_complete_details,
        normalize_rfid,
        parse_member_date_joined,
        resolve_inactive_remark,
        rfid_is_taken_by_other,
        rfids_equivalent,
    )

    first_name = (data.get('first_name') or member.first_name).strip()
    last_name = (data.get('last_name') or member.last_name).strip()
    rfid_raw = data.get('rfid')
    if isinstance(rfid_raw, str):
        rfid = normalize_rfid(rfid_raw)
    else:
        rfid = normalize_rfid(member.rfid_card_number)
    current_rfid = normalize_rfid(member.rfid_card_number)
    email = (data.get('email') or '').strip() or None
    phone = (data.get('phone') or member.phone).strip()
    member_type_id = data.get('member_type_id')
    role = (data.get('role') or member.role).strip()
    is_active = bool(data.get('is_active', member.is_active))

    # Staff and cashier (Member role) have the same role edit rules
    if restricts_member_role_to_member_only(request.user):
        # Can only set role to 'member' for non-privileged records
        # If member already has admin/cashier/staff role, keep it (cannot promote or change elevated roles)
        if member.role in ['admin', 'cashier', 'staff']:
            role = member.role  # Keep existing role, don't allow change
        elif role not in ['member']:
            return JsonResponse({'success': False, 'error': 'You can only set role to "member"'}, status=403)
        else:
            role = 'member'  # Force to member role

    if not first_name or not last_name:
        return JsonResponse({'success': False, 'error': 'First and last name are required'}, status=400)

    if rfid and not rfids_equivalent(rfid, current_rfid) and rfid_is_taken_by_other(rfid, exclude_member_pk=member.pk):
        return JsonResponse({'success': False, 'error': 'RFID card number already exists'}, status=400)
    if email and Member.objects.filter(email=email).exclude(pk=member.pk).exists():
        return JsonResponse({'success': False, 'error': 'Email already exists'}, status=400)

    inactive_remark, remark_error = resolve_inactive_remark(is_active, data.get('inactive_remark'))
    if remark_error:
        return JsonResponse({'success': False, 'error': remark_error}, status=400)

    detail_fields, detail_error = extract_member_complete_details(data)
    if detail_error:
        return JsonResponse({'success': False, 'error': detail_error}, status=400)

    member_type_new = None
    update_member_type = 'member_type_id' in data
    if update_member_type:
        if member_type_id in (None, '', 0, '0'):
            member_type_new = None
        else:
            try:
                member_type_new = MemberType.objects.get(id=member_type_id)
            except (MemberType.DoesNotExist, ValueError, TypeError):
                return JsonResponse({'success': False, 'error': 'Selected member type does not exist'}, status=400)

    old_values = {
        'username': member.username or '',
        'rfid': member.rfid_card_number or '',
        'first_name': member.first_name or '',
        'last_name': member.last_name or '',
        'email': member.email or '',
        'phone': member.phone or '',
        'role': member.role or '',
        'is_active': bool(member.is_active),
        'inactive_remark': member.inactive_remark or '',
    }

    # Save a snapshot of current values so the edit can be undone later.
    MemberEditHistory.objects.create(
        member=member,
        username=member.username,
        first_name=member.first_name,
        last_name=member.last_name,
        email=member.email,
        phone=member.phone or '',
        rfid_card_number=member.rfid_card_number,
        role=member.role or '',
        edited_by=request.user.username,
    )

    member.first_name = first_name
    member.last_name = last_name
    member.rfid_card_number = rfid
    member.email = email
    member.phone = phone
    apply_member_complete_details(member, detail_fields or {})
    if update_member_type:
        member.member_type = member_type_new
    requested = (role or "").strip().lower()
    if requested and Role.objects.filter(slug__iexact=requested, is_active=True).exists():
        slug_to_apply = requested
    else:
        slug_to_apply = (member.role or "member")
    member.member_role = Role.resolve_slug(slug_to_apply)
    member.is_active = is_active
    member.inactive_remark = inactive_remark

    if 'date_joined' in data:
        try:
            member.date_joined = parse_member_date_joined(data.get('date_joined'), default_now=False)
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid registration date format'}, status=400)
        if member.date_joined is None:
            return JsonResponse({'success': False, 'error': 'Registration date is required'}, status=400)

    # Update username (Member.username CharField) directly
    username_raw = data.get('username')
    if username_raw is not None:
        new_username = username_raw.strip() or None
        if new_username and Member.objects.filter(username=new_username).exclude(id=member.id).exists():
            return JsonResponse({'success': False, 'error': f'Username "{new_username}" already exists'}, status=400)
        member.username = new_username

    # Update security fields
    pin_attempts_raw = data.get('pin_attempts')
    if pin_attempts_raw is not None:
        try:
            member.pin_attempts = int(pin_attempts_raw)
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid pin_attempts value'}, status=400)

    is_pin_locked_raw = data.get('is_pin_locked')
    if is_pin_locked_raw is not None:
        member.is_pin_locked = bool(is_pin_locked_raw)

    # PIN change flow: require old PIN when setting a new PIN.
    old_pin = (data.get('old_pin') or '').strip()
    new_pin = (data.get('new_pin') or '').strip()
    if new_pin:
        if not new_pin.isdigit() or len(new_pin) != 4:
            return JsonResponse({'success': False, 'error': 'New PIN must be exactly 4 digits'}, status=400)
        if member.pin_hash:
            if not old_pin:
                return JsonResponse({'success': False, 'error': 'Old PIN is required to set a new PIN'}, status=400)
            if not old_pin.isdigit() or len(old_pin) != 4:
                return JsonResponse({'success': False, 'error': 'Old PIN must be exactly 4 digits'}, status=400)
            if not member.check_pin(old_pin):
                return JsonResponse({'success': False, 'error': 'Old PIN is incorrect'}, status=403)
        member.set_pin(new_pin)

    share_capital_changed = False
    share_capital_before = member.share_capital or Decimal('0.00')
    share_capital_after = share_capital_before
    if 'share_capital' in data:
        try:
            share_capital_after = Decimal(str(data.get('share_capital') or '0.00'))
        except (InvalidOperation, TypeError, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid share capital value'}, status=400)
        if share_capital_after < 0:
            return JsonResponse({'success': False, 'error': 'Share capital cannot be negative'}, status=400)
        if share_capital_after != share_capital_before:
            member.share_capital = share_capital_after
            share_capital_changed = True

    changed_labels = []
    if old_values['username'] != (member.username or ''):
        changed_labels.append('Username')
    if old_values['rfid'] != (member.rfid_card_number or ''):
        changed_labels.append('RFID Card Number')
    if old_values['first_name'] != (member.first_name or '') or old_values['last_name'] != (member.last_name or ''):
        changed_labels.append('Full Name')
    if old_values['email'] != (member.email or ''):
        changed_labels.append('Email Address')
    if old_values['phone'] != (member.phone or ''):
        changed_labels.append('Phone Number')
    if old_values['role'] != (member.role or ''):
        changed_labels.append('Role')
    if old_values['is_active'] != bool(member.is_active):
        changed_labels.append('Account Status')
    if old_values['inactive_remark'] != (member.inactive_remark or ''):
        changed_labels.append('Inactive Remark')
    if new_pin:
        changed_labels.append('PIN')
    if share_capital_changed:
        changed_labels.append('Share Capital')

    # Save all member changes (except PIN, already persisted via member.set_pin)
    member.save()

    if share_capital_changed:
        delta = share_capital_after - share_capital_before
        ShareCapitalTransaction.objects.create(
            member=member,
            transaction_type='adjustment',
            amount=abs(delta),
            balance_before=share_capital_before,
            balance_after=share_capital_after,
            notes=(
                f'Share capital adjusted from ₱{share_capital_before} to ₱{share_capital_after} '
                'via Edit Member'
            ),
            performed_by=request.user if request.user.is_authenticated else None,
        )

    concession_err = _apply_member_concession_from_request(member, data.get('concession'))
    if concession_err:
        return JsonResponse({'success': False, 'error': concession_err}, status=400)

    member_email_sent = False
    if member.email:
        import logging
        import threading
        logger = logging.getLogger(__name__)
        changed_text = '\n'.join([f'  - {label}' for label in changed_labels]) or '  - Profile details'
        actor_name = request.user.get_full_name().strip() or request.user.username
        actor_role = getattr(authorising_member, 'role', 'unknown').capitalize()
        timestamp_str = timezone.localtime(timezone.now()).strftime('%B %d, %Y %I:%M %p')
        inactive_reason_block = ''
        if not member.is_active and (member.inactive_remark or '').strip():
            inactive_reason_block = (
                f"\nReason for inactive status:\n  {member.inactive_remark.strip()}\n"
            )

        def send_member_update_email():
            try:
                subject = 'Your Account Details Were Updated'
                body = f"""Hi {member.first_name},

Your account details were updated in BAGNOS MPC.

Updated fields:
{changed_text}
{inactive_reason_block}
Updated by:
  - User: {actor_name}
  - Role: {actor_role}

Date and time: {timestamp_str}

If you did not request or authorize this change, please contact admin immediately.
"""
                email_msg = EmailMessage(
                    subject=subject,
                    body=body.strip(),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[member.email],
                )
                email_msg.send(fail_silently=False)
            except Exception as exc:
                logger.error(f'Failed to send member update notification email: {exc}', exc_info=True)

        threading.Thread(target=send_member_update_email, daemon=True).start()
        member_email_sent = True

    return JsonResponse({
        'success': True,
        'message': 'Member updated successfully',
        'member_notified': member_email_sent,
        'member': {
            'id': member.id,
            'name': member.full_name,
            'rfid': member.rfid_card_number,
            'email': member.email or '',
            'phone': member.phone,
            'member_type': member.member_type.name if member.member_type else '',
            'role': member.role,
            'is_active': member.is_active,
            'inactive_remark': member.inactive_remark or '',
            'balance': str(member.balance),
            'share_capital': str(member.share_capital),
        }
    })


@login_required
@require_http_methods(["GET"])
def api_get_member_last_edit(request):
    """Return the most recent pre-edit snapshot for a member so the UI can preview what will be restored."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    member_id = request.GET.get('member_id')
    if not member_id:
        return JsonResponse({'success': False, 'error': 'Member ID is required'}, status=400)

    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Member not found'}, status=404)

    snapshot = MemberEditHistory.objects.filter(member=member).first()
    if not snapshot:
        return JsonResponse({'success': False, 'error': 'No edit history found for this member'}, status=404)

    def _role_label(slug):
        s = (slug or "").strip()
        if not s:
            return ""
        r = Role.objects.filter(slug__iexact=s).first()
        return r.name if r else s

    return JsonResponse({
        'success': True,
        'snapshot': {
            'username': snapshot.username or '',
            'first_name': snapshot.first_name,
            'last_name': snapshot.last_name,
            'email': snapshot.email or '',
            'phone': snapshot.phone or '',
            'rfid_card_number': snapshot.rfid_card_number or '',
            'role': snapshot.role,
            'role_display': _role_label(snapshot.role),
            'edited_at': snapshot.edited_at.strftime('%Y-%m-%d %H:%M:%S'),
            'edited_by': snapshot.edited_by,
        },
        'current': {
            'username': member.username or '',
            'first_name': member.first_name,
            'last_name': member.last_name,
            'email': member.email or '',
            'phone': member.phone or '',
            'rfid_card_number': member.rfid_card_number or '',
            'role': member.role,
            'role_display': member.get_role_display(),
        },
    })


@login_required
@require_http_methods(["POST"])
def api_restore_member_last_edit(request):
    """Restore a member's profile fields to the most recent pre-edit snapshot."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    member_id = data.get('member_id')
    if not member_id:
        return JsonResponse({'success': False, 'error': 'Member ID is required'}, status=400)

    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Member not found'}, status=404)

    snapshot = MemberEditHistory.objects.filter(member=member).first()
    if not snapshot:
        return JsonResponse({'success': False, 'error': 'No edit history found for this member'}, status=404)

    # Check for conflicts before applying
    if snapshot.username and Member.objects.filter(username=snapshot.username).exclude(id=member.id).exists():
        return JsonResponse({
            'success': False,
            'error': f'Username "{snapshot.username}" is now used by another member and cannot be restored.',
        }, status=400)
    if snapshot.email and Member.objects.filter(email=snapshot.email).exclude(id=member.id).exists():
        return JsonResponse({
            'success': False,
            'error': f'Email "{snapshot.email}" is now used by another member and cannot be restored.',
        }, status=400)
    if snapshot.rfid_card_number and Member.objects.filter(rfid_card_number=snapshot.rfid_card_number).exclude(id=member.id).exists():
        return JsonResponse({
            'success': False,
            'error': f'RFID card "{snapshot.rfid_card_number}" is now used by another member and cannot be restored.',
        }, status=400)

    # Save a new snapshot of current state before overwriting (so this restore can itself be undone)
    MemberEditHistory.objects.create(
        member=member,
        username=member.username,
        first_name=member.first_name,
        last_name=member.last_name,
        email=member.email,
        phone=member.phone or '',
        rfid_card_number=member.rfid_card_number,
        role=member.role or '',
        edited_by=request.user.username,
    )

    # Apply the snapshot values
    member.username = snapshot.username or None
    member.first_name = snapshot.first_name
    member.last_name = snapshot.last_name
    member.email = snapshot.email or None
    member.phone = snapshot.phone or ''
    from helper.members_helper import normalize_rfid

    member.rfid_card_number = normalize_rfid(snapshot.rfid_card_number)
    if (snapshot.role or "").strip():
        member.member_role = Role.resolve_slug(snapshot.role)
    member.save()

    # Delete the consumed snapshot so the next "restore" uses the one before it
    snapshot.delete()

    return JsonResponse({
        'success': True,
        'message': f'Member "{member.full_name}" has been restored to the previous values.',
        'member': {
            'id': member.id,
            'name': member.full_name,
            'username': member.username or '',
            'first_name': member.first_name,
            'last_name': member.last_name,
            'email': member.email or '',
            'phone': member.phone,
            'rfid_card_number': member.rfid_card_number or '',
            'role': member.role,
        },
    })


@login_required
@require_http_methods(["POST"])
def api_restore_all_last_edit(request):
    """Restore every member that has edit history to their most-recent pre-edit snapshot.

    Returns a summary of how many were restored vs skipped (due to unique-field conflicts).
    """
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    from helper.members_helper import normalize_rfid

    # Collect the most recent snapshot per member (MemberEditHistory is ordered by -edited_at)
    seen_members = set()
    snapshots = []
    for snap in MemberEditHistory.objects.select_related('member').order_by('member_id', '-edited_at'):
        if snap.member_id not in seen_members:
            seen_members.add(snap.member_id)
            snapshots.append(snap)

    restored = []
    skipped = []

    for snapshot in snapshots:
        member = snapshot.member

        # Conflict checks (same as single-member restore)
        if snapshot.username and Member.objects.filter(username=snapshot.username).exclude(id=member.id).exists():
            skipped.append({'name': member.full_name, 'reason': f'Username "{snapshot.username}" already taken'})
            continue
        if snapshot.email and Member.objects.filter(email=snapshot.email).exclude(id=member.id).exists():
            skipped.append({'name': member.full_name, 'reason': f'Email "{snapshot.email}" already taken'})
            continue
        if snapshot.rfid_card_number and Member.objects.filter(rfid_card_number=snapshot.rfid_card_number).exclude(id=member.id).exists():
            skipped.append({'name': member.full_name, 'reason': f'RFID "{snapshot.rfid_card_number}" already taken'})
            continue

        # Snapshot current state so this restore can itself be undone
        MemberEditHistory.objects.create(
            member=member,
            username=member.username,
            first_name=member.first_name,
            last_name=member.last_name,
            email=member.email,
            phone=member.phone or '',
            rfid_card_number=member.rfid_card_number,
            role=member.role or '',
            edited_by=request.user.username,
        )

        member.username = snapshot.username or None
        member.first_name = snapshot.first_name
        member.last_name = snapshot.last_name
        member.email = snapshot.email or None
        member.phone = snapshot.phone or ''
        member.rfid_card_number = normalize_rfid(snapshot.rfid_card_number)
        if (snapshot.role or "").strip():
            member.member_role = Role.resolve_slug(snapshot.role)
        member.save()

        snapshot.delete()
        restored.append(member.full_name)

    return JsonResponse({
        'success': True,
        'restored_count': len(restored),
        'skipped_count': len(skipped),
        'restored': restored,
        'skipped': skipped,
        'message': f'{len(restored)} member(s) restored, {len(skipped)} skipped.',
    })


@login_required
@require_http_methods(["POST"])
def api_reset_pin_attempts(request):
    """Reset a member's PIN lockout counter from the dashboard."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    member_id = data.get('member_id')
    if not member_id:
        return JsonResponse({'success': False, 'error': 'Member ID is required'}, status=400)

    try:
        member = Member.objects.get(id=member_id)
    except Member.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Member not found'}, status=404)

    member.pin_attempts = 0
    member.is_pin_locked = False
    member.save(update_fields=['pin_attempts', 'is_pin_locked'])

    return JsonResponse({
        'success': True,
        'message': f'PIN lockout reset for {member.full_name}. They can now log in again.',
    })


@login_required
def member_management(request):
    # Allow admin, cashier and staff role users to access member management
    is_admin = is_cashier_or_admin(request.user)
    is_staff_role_user = is_staff_role(request.user)
    
    if not (is_admin or is_staff_role_user):
        messages.warning(request, 'You do not have permission to access this page.')
        return redirect('kiosk_home')
    
    # Cashier/staff users may manage only plain member accounts.
    restrict_member_role = restricts_member_role_to_member_only(request.user)

    # Get active filters from request
    search_query = request.GET.get('search', '').strip()
    pin_lock_filter = request.GET.get('pin_lock', '').strip().lower()
    status_filter = request.GET.get('status', 'all').strip().lower()
    if status_filter not in ('all', 'active', 'inactive'):
        status_filter = 'all'
    credit_filter = request.GET.get('credit', '').strip().lower()
    sort_filter = request.GET.get('sort', 'az').strip().lower()
    if sort_filter not in ('az', 'za'):
        sort_filter = 'az'

    kiosk_config = KioskConfig.get()
    from kiosk_helper import get_member_max_credit_limit
    from helper.credit_interest_helper import ensure_credit_interest_up_to_date

    member_max_credit = get_member_max_credit_limit()
    credit_limit_enabled = member_max_credit > 0
    # Accrue any due monthly interest so dashboard totals stay current
    ensure_credit_interest_up_to_date()
    credit_settings = CreditSettings.get()

    # List members like Django admin: default shows everyone; optional active/inactive filters.
    members = Member.objects.select_related(
        'member_role', 'member_type', 'user', 'senior_profile', 'pwd_profile',
    )
    if restrict_member_role:
        members = members.filter(member_role__slug='member')

    from helper.credit_settlement_helper import (
        annotate_members_credit_outstanding,
        members_with_unsettled_credit_filter,
        unsettled_credit_items_queryset,
    )

    members = annotate_members_credit_outstanding(members)
    if credit_limit_enabled:
        members = members.annotate(
            credit_available=Case(
                When(
                    credit_outstanding__gte=member_max_credit,
                    then=Value(Decimal('0')),
                ),
                default=ExpressionWrapper(
                    Value(member_max_credit) - F('credit_outstanding'),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
            ),
        )

    if credit_filter == 'owed':
        # Item-level outstanding; include active and inactive members with utang.
        members = members.filter(members_with_unsettled_credit_filter())

    if pin_lock_filter == 'locked':
        members = members.filter(is_pin_locked=True)
    elif credit_filter != 'owed':
        if status_filter == 'active':
            members = members.filter(is_active=True)
        elif status_filter == 'inactive':
            members = members.filter(is_active=False)
    
    # Apply search filter if query exists
    if search_query:
        # Build base query for all non-name fields
        search_filters = Q(
            Q(rfid_card_number__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(middle_name__icontains=search_query) |
            Q(rsbsa_number__icontains=search_query) |
            Q(tin__icontains=search_query) |
            Q(member_type__name__icontains=search_query) |
            Q(member_role__slug__icontains=search_query) |
            Q(member_role__name__icontains=search_query)
        )
        
        # Handle name search - check if query contains spaces (full name search)
        if ' ' in search_query:
            # Split the query into parts (handle multiple spaces)
            name_parts = [part.strip() for part in search_query.split() if part.strip()]
            
            if len(name_parts) >= 2:
                # Full name search: "John Doe" or "John Michael Doe"
                first_part = name_parts[0]
                remaining_parts = ' '.join(name_parts[1:])  # Join remaining parts for last name
                
                # Match combinations:
                # 1. First part in first_name AND remaining in last_name
                # 2. First part in last_name AND remaining in first_name (reverse)
                # 3. Full query in first_name or last_name (for partial matches)
                name_filter = (
                    (Q(first_name__icontains=first_part) & Q(last_name__icontains=remaining_parts)) |
                    (Q(first_name__icontains=remaining_parts) & Q(last_name__icontains=first_part)) |
                    Q(first_name__icontains=search_query) |
                    Q(middle_name__icontains=search_query) |
                    Q(last_name__icontains=search_query)
                )
                search_filters |= name_filter
            else:
                # Single word, search in both name fields
                search_filters |= Q(first_name__icontains=name_parts[0]) | Q(middle_name__icontains=name_parts[0]) | Q(last_name__icontains=name_parts[0])
        else:
            # No spaces - single word search in first_name, last_name, or full name
            # Search individual fields
            search_filters |= Q(first_name__icontains=search_query) | Q(middle_name__icontains=search_query) | Q(last_name__icontains=search_query)
            
            # Also try to match full name by checking if query matches start of first_name + last_name
            # This handles cases where user types "JohnDoe" (no space)
            # We'll search for members where first_name starts with query or last_name starts with query
            # This is already covered by the icontains above, but we can be more specific
        
        members = members.filter(search_filters)
    
    if sort_filter == 'za':
        members = members.order_by('-last_name', '-first_name', '-id')
    else:
        members = members.order_by('last_name', 'first_name', 'id')
    members_per_page = 20
    members_paginator = Paginator(members, members_per_page)
    members_page = members_paginator.get_page(request.GET.get('page') or 1)
    members_elided_pages = list(
        members_paginator.get_elided_page_range(
            members_page.number, on_each_side=2, on_ends=1
        )
    )
    members_list_params = request.GET.copy()
    members_list_params.pop('page', None)
    members_list_query = members_list_params.urlencode()
    members_base_params = request.GET.copy()
    members_base_params.pop('page', None)
    members_base_params.pop('sort', None)
    members_base_query = members_base_params.urlencode()
    members_filter_params = request.GET.copy()
    members_filter_params.pop('page', None)
    members_filter_params.pop('search', None)
    members_filter_query = members_filter_params.urlencode()

    member_types = MemberType.objects.filter(is_active=True).order_by('name')
    assignable_roles = Role.objects.filter(is_active=True).order_by('sort_order', 'name')
    
    # Calculate statistics (scope depends on caller role restrictions).
    all_members = Member.objects.all()
    if restrict_member_role:
        all_members = all_members.filter(member_role__slug='member')
    total_members = all_members.count()
    active_members = all_members.filter(is_active=True).count()
    inactive_members = all_members.filter(is_active=False).count()
    total_balances = all_members.aggregate(Sum('balance'))['balance__sum'] or 0
    total_locked_members = all_members.filter(is_pin_locked=True).count()
    active_locked_members = all_members.filter(is_active=True, is_pin_locked=True).count()

    credit_item_qs = unsettled_credit_items_queryset().filter(
        transaction__member__isnull=False,
    )
    if restrict_member_role:
        credit_item_qs = credit_item_qs.filter(
            transaction__member__member_role__slug='member',
        )
    total_principal = credit_item_qs.aggregate(t=Sum('outstanding'))['t'] or Decimal('0')
    from helper.credit_settlement_helper import _interest_outstanding_expression
    interest_qs = Transaction.objects.filter(
        payment_method='credit',
        status='completed',
        credit_settled_at__isnull=True,
        member__isnull=False,
    ).annotate(unpaid=_interest_outstanding_expression())
    if restrict_member_role:
        interest_qs = interest_qs.filter(member__member_role__slug='member')
    total_interest = interest_qs.aggregate(t=Sum('unpaid'))['t'] or Decimal('0')
    total_credit_outstanding = (
        Decimal(total_principal) + Decimal(total_interest)
    ).quantize(Decimal('0.01'))
    members_with_credit = (
        Member.objects.filter(members_with_unsettled_credit_filter())
    )
    if restrict_member_role:
        members_with_credit = members_with_credit.filter(member_role__slug='member')
    members_with_credit = members_with_credit.count()

    segment_discount_rules = list(
        SegmentProductGroupDiscount.objects.filter(is_active=True)
        .select_related('discount_group')
        .order_by('segment', 'discount_group__sort_order', 'discount_group__code')
    )
    segment_discount_rules_all = list(
        SegmentProductGroupDiscount.objects.select_related('discount_group').order_by(
            'segment', 'discount_group__sort_order', 'discount_group__code'
        )
    )

    from helper.members_helper import member_complete_details_dict

    context = {
        'members': members_page,
        'page_obj': members_page,
        'members_filtered_total': members_paginator.count,
        'members_elided_pages': members_elided_pages,
        'members_list_query': members_list_query,
        'members_base_query': members_base_query,
        'members_filter_query': members_filter_query,
        'member_types': member_types,
        'member_complete_details_map': {
            str(m.pk): member_complete_details_dict(m)
            for m in members_page.object_list
        },
        'total_members': total_members,
        'active_members': active_members,
        'inactive_members': inactive_members,
        'total_balances': total_balances,
        'total_locked_members': total_locked_members,
        'active_locked_members': active_locked_members,
        'restrict_member_role': restrict_member_role,
        'is_staff': restrict_member_role,
        'is_admin': is_admin,
        'search_query': search_query,
        'pin_lock_filter': pin_lock_filter,
        'status_filter': status_filter,
        'assignable_roles': assignable_roles,
        'segment_discount_rules': segment_discount_rules,
        'segment_discount_rules_all': segment_discount_rules_all,
        'can_edit_segment_discount_rules': is_admin_user(request.user),
        'member_max_credit': member_max_credit,
        'credit_limit_enabled': credit_limit_enabled,
        'total_credit_outstanding': total_credit_outstanding,
        'total_credit_interest': Decimal(total_interest).quantize(Decimal('0.01')),
        'members_with_credit': members_with_credit,
        'credit_filter': credit_filter,
        'sort_filter': sort_filter,
        'credit_settings': credit_settings,
        'credit_interest_enabled': bool(credit_settings.is_enabled and credit_settings.interest_rate > 0),
        **admin_role_badge_context(request),
    }
    
    return render(request, 'admin_panel/members.html', context)


@login_required
def credit_unpaid_history(request):
    """
    Transparency view: unpaid credit (utang) sales with dates since unpaid.
    """
    is_admin = is_cashier_or_admin(request.user)
    is_staff_role_user = is_staff_role(request.user)
    if not (is_admin or is_staff_role_user):
        messages.warning(request, 'You do not have permission to access this page.')
        return redirect('kiosk_home')

    restrict_member_role = restricts_member_role_to_member_only(request.user)
    search_query = request.GET.get('search', '').strip()

    from helper.credit_interest_helper import (
        ensure_credit_interest_up_to_date,
        interest_start_date,
        first_interest_due_date,
        monthly_periods_due,
        interest_rate_as_multiplier,
        _principal_outstanding_for_sale,
        _add_months,
    )

    ensure_credit_interest_up_to_date()
    credit_settings = CreditSettings.get()
    grace_days = int(credit_settings.grace_period_days or 0)
    rate_mult = interest_rate_as_multiplier(credit_settings.interest_rate or 0)
    interest_on = bool(credit_settings.is_enabled and rate_mult > 0)
    today = timezone.localdate()

    sales_qs = (
        Transaction.objects.filter(
            payment_method='credit',
            status='completed',
            credit_settled_at__isnull=True,
            member__isnull=False,
        )
        .select_related('member', 'member__member_role')
        .prefetch_related('items')
        .order_by('created_at', 'id')
    )
    if restrict_member_role:
        sales_qs = sales_qs.filter(member__member_role__slug='member')
    if search_query:
        sales_qs = sales_qs.filter(
            Q(member__first_name__icontains=search_query)
            | Q(member__last_name__icontains=search_query)
            | Q(transaction_number__icontains=search_query)
            | Q(member__rfid_card_number__icontains=search_query)
        )

    unpaid_rows = []
    total_principal = Decimal('0.00')
    total_interest = Decimal('0.00')
    total_next_month = Decimal('0.00')

    for sale in sales_qs:
        principal = _principal_outstanding_for_sale(sale)
        interest = sale.credit_interest_outstanding
        if principal <= 0 and interest <= 0:
            continue
        sale_date = timezone.localtime(sale.created_at).date()
        days_unpaid = (today - sale_date).days
        grace_ends = interest_start_date(sale, grace_days)
        first_due = first_interest_due_date(sale, grace_days)
        months_late = monthly_periods_due(first_due, today) if interest_on else 0
        within_grace = today < grace_ends
        awaiting_first_month = (not within_grace) and today < first_due
        days_past_grace = max((today - grace_ends).days, 0) if not within_grace else 0
        total = (principal + interest).quantize(Decimal('0.01'))

        # Projected total if still unpaid when the next monthly charge applies
        next_interest_add = Decimal('0.00')
        next_month_due = total
        next_charge_on = None
        if interest_on and principal > 0:
            last_applied = sale.credit_interest_last_applied_on
            already = monthly_periods_due(first_due, last_applied) if last_applied else 0
            next_charge_on = _add_months(first_due, already)
            if next_charge_on <= today:
                next_charge_on = _add_months(next_charge_on, 1)
            next_interest_add = (principal * rate_mult).quantize(Decimal('0.01'))
            next_month_due = (total + next_interest_add).quantize(Decimal('0.01'))

        total_principal += principal
        total_interest += interest
        total_next_month += next_month_due
        unpaid_rows.append({
            'sale': sale,
            'member': sale.member,
            'sale_date': sale_date,
            'sale_datetime': timezone.localtime(sale.created_at),
            'days_unpaid': days_unpaid,
            'grace_ends': grace_ends,
            'first_interest_due': first_due,
            'within_grace': within_grace,
            'awaiting_first_month': awaiting_first_month,
            'days_past_grace': days_past_grace,
            'months_late': months_late,
            'principal': principal,
            'interest': interest,
            'total': total,
            'next_interest_add': next_interest_add,
            'next_month_due': next_month_due,
            'next_charge_on': next_charge_on,
            'receipt_url': reverse('view_credit_receipt', kwargs={'transaction_id': sale.id}),
        })

    # Newest unpaid first for scanning overdue
    unpaid_rows.sort(key=lambda r: (r['days_unpaid'], r['sale_date']), reverse=True)

    recent_payments = (
        CreditPayment.objects.select_related('member', 'performed_by')
        .order_by('-created_at')[:40]
    )
    if restrict_member_role:
        recent_payments = recent_payments.filter(member__member_role__slug='member')

    context = {
        'unpaid_rows': unpaid_rows,
        'unpaid_count': len(unpaid_rows),
        'total_principal': total_principal.quantize(Decimal('0.01')),
        'total_interest': total_interest.quantize(Decimal('0.01')),
        'total_outstanding': (total_principal + total_interest).quantize(Decimal('0.01')),
        'total_next_month': total_next_month.quantize(Decimal('0.01')),
        'credit_settings': credit_settings,
        'credit_interest_enabled': interest_on,
        'search_query': search_query,
        'recent_payments': recent_payments,
        'is_admin': is_admin,
        **admin_role_badge_context(request),
    }
    return render(request, 'admin_panel/credit_unpaid_history.html', context)


def _serialize_credit_sale_for_pay_modal(sale: Transaction) -> dict:
    unsettled_items = [
        i for i in sale.items.all()
        if i.credit_settled_at is None and i.credit_line_outstanding > 0
    ]
    unsettled_principal = sum((i.credit_line_outstanding for i in unsettled_items), Decimal('0'))
    interest_out = sale.credit_interest_outstanding
    unsettled_total = (unsettled_principal + interest_out).quantize(Decimal('0.01'))
    sale_local = timezone.localtime(sale.created_at)
    days_unpaid = (timezone.localdate() - sale_local.date()).days
    return {
        'id': sale.id,
        'transaction_number': sale.transaction_number,
        'created_at': sale_local.strftime('%Y-%m-%d %I:%M %p'),
        'sale_date': sale_local.strftime('%Y-%m-%d'),
        'days_unpaid': days_unpaid,
        'total_amount': str(unsettled_total),
        'principal_amount': str(unsettled_principal.quantize(Decimal('0.01'))),
        'interest_amount': str(interest_out.quantize(Decimal('0.01'))),
        'receipt_url': reverse('view_credit_receipt', kwargs={'transaction_id': sale.id}),
        'items': [
            {
                'id': item.id,
                'product_name': item.product_name,
                'quantity': item.quantity,
                'unit_price': str(item.unit_price),
                'line_amount': str(item.credit_line_outstanding),
                'line_total': str(item.credit_line_amount),
                'line_paid': str(Decimal(item.credit_amount_paid or 0).quantize(Decimal('0.01'))),
            }
            for item in unsettled_items
        ],
    }


@require_http_methods(['GET', 'POST'])
def api_credit_settings(request):
    """GET/POST CreditSettings (interest rate + grace period) with Admin PIN on save."""
    if not request.user.is_authenticated:
        return JsonResponse(
            {'success': False, 'error': 'Authentication required. Please log in again.'},
            status=401,
        )

    is_admin = is_cashier_or_admin(request.user)
    is_staff_role_user = is_staff_role(request.user)
    if not (is_admin or is_staff_role_user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    config = CreditSettings.get()

    if request.method == 'GET':
        return JsonResponse({
            'success': True,
            'interest_rate': str(Decimal(config.interest_rate or 0).quantize(Decimal('0.001'))),
            'grace_period_days': int(config.grace_period_days or 0),
            'is_enabled': bool(config.is_enabled),
        })

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    admin_pin = (data.get('admin_pin') or '').strip()
    if not admin_pin or len(admin_pin) != 4 or not admin_pin.isdigit():
        return JsonResponse({
            'success': False,
            'error': 'A valid 4-digit Admin PIN is required to save credit interest settings.',
        })

    session_member = _get_active_member_for_refill_user(request.user)
    if not _resolve_refill_pin_authorizer(admin_pin, session_member):
        return JsonResponse({
            'success': False,
            'error': 'Incorrect PIN. Enter the 4-digit PIN of an active Admin member with a PIN set.',
        })

    raw_rate = data.get('interest_rate')
    try:
        new_rate = Decimal(str(raw_rate if raw_rate is not None else '0'))
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid interest rate'})

    if new_rate < 0 or new_rate > Decimal('100'):
        return JsonResponse({
            'success': False,
            'error': 'Interest rate must be between 0 and 100.',
        })

    try:
        new_grace = int(data.get('grace_period_days', 0))
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid grace period days'})

    if new_grace < 0 or new_grace > 3650:
        return JsonResponse({
            'success': False,
            'error': 'Grace period must be between 0 and 3650 days.',
        })

    is_enabled = bool(data.get('is_enabled', True))
    # Enabling with 0% rate is allowed but effectively no charge
    if new_rate == 0:
        is_enabled = False

    new_rate = new_rate.quantize(Decimal('0.001'))
    config.interest_rate = new_rate
    config.grace_period_days = new_grace
    config.is_enabled = is_enabled
    config.save(update_fields=['interest_rate', 'grace_period_days', 'is_enabled', 'updated_at'])

    # Accrue immediately so existing overdue utang picks up the new rate schedule
    from helper.credit_interest_helper import ensure_credit_interest_up_to_date
    ensure_credit_interest_up_to_date()

    return JsonResponse({
        'success': True,
        'interest_rate': str(new_rate),
        'grace_period_days': new_grace,
        'is_enabled': is_enabled,
    })


@require_http_methods(['GET', 'POST'])
def api_kiosk_credit_limit(request):
    """GET current member_max_credit; POST update with Admin PIN."""
    from kiosk_helper import CREDIT_LIMIT_FEATURE_ENABLED

    if not CREDIT_LIMIT_FEATURE_ENABLED:
        return JsonResponse(
            {'success': False, 'error': 'Credit limit settings are not available right now.'},
            status=404,
        )

    if not request.user.is_authenticated:
        return JsonResponse(
            {'success': False, 'error': 'Authentication required. Please log in again.'},
            status=401,
        )

    is_admin = is_cashier_or_admin(request.user)
    is_staff_role_user = is_staff_role(request.user)
    if not (is_admin or is_staff_role_user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    kiosk_config = KioskConfig.get()

    if request.method == 'GET':
        limit = Decimal(str(kiosk_config.member_max_credit or 0))
        return JsonResponse({
            'success': True,
            'member_max_credit': str(limit.quantize(Decimal('0.01'))),
            'credit_limit_enabled': limit > 0,
        })

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    admin_pin = (data.get('admin_pin') or '').strip()
    if not admin_pin or len(admin_pin) != 4 or not admin_pin.isdigit():
        return JsonResponse({
            'success': False,
            'error': 'A valid 4-digit Admin PIN is required to save credit limit settings.',
        })

    session_member = _get_active_member_for_refill_user(request.user)
    if not _resolve_refill_pin_authorizer(admin_pin, session_member):
        return JsonResponse({
            'success': False,
            'error': 'Incorrect PIN. Enter the 4-digit PIN of an active Admin member with a PIN set.',
        })

    raw_limit = data.get('member_max_credit')
    try:
        new_limit = Decimal(str(raw_limit if raw_limit is not None else '0'))
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid credit limit amount'})

    if new_limit < 0:
        return JsonResponse({'success': False, 'error': 'Credit limit cannot be negative'})

    new_limit = new_limit.quantize(Decimal('0.01'))
    kiosk_config.member_max_credit = new_limit
    kiosk_config.save(update_fields=['member_max_credit', 'updated_at'])

    return JsonResponse({
        'success': True,
        'member_max_credit': str(new_limit),
        'credit_limit_enabled': new_limit > 0,
    })


@login_required
@require_http_methods(["GET"])
def api_member_credit_details(request):
    """Unsettled credit sales for Pay Credit modal."""
    is_admin = is_cashier_or_admin(request.user)
    is_staff_role_user = is_staff_role(request.user)
    if not (is_admin or is_staff_role_user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    member_id = request.GET.get('member_id', '').strip()
    if not member_id:
        return JsonResponse({'success': False, 'error': 'member_id is required'})
    try:
        member_id = int(member_id)
    except (ValueError, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid member_id'})

    try:
        member = Member.objects.get(pk=member_id, is_active=True)
    except Member.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Member not found'})

    from helper.credit_settlement_helper import member_credit_outstanding_amount, unsettled_credit_sales

    sales = list(unsettled_credit_sales(member))
    outstanding = member_credit_outstanding_amount(member)
    from kiosk_helper import get_member_max_credit_limit
    from helper.credit_interest_helper import member_credit_interest_outstanding
    from helper.credit_settlement_helper import member_credit_principal_outstanding
    member_max_credit = get_member_max_credit_limit()
    principal = member_credit_principal_outstanding(member)
    interest = member_credit_interest_outstanding(member.pk)
    credit_settings = CreditSettings.get()

    return JsonResponse({
        'success': True,
        'member': {
            'id': member.id,
            'name': member.full_name,
            'email': member.email or '',
            'balance': str(member.balance),
        },
        'credit_outstanding': str(outstanding),
        'credit_principal': str(principal),
        'credit_interest': str(interest),
        'member_max_credit': str(member_max_credit),
        'credit_interest_enabled': bool(credit_settings.is_enabled and credit_settings.interest_rate > 0),
        'interest_rate': str(Decimal(credit_settings.interest_rate or 0).quantize(Decimal('0.001'))),
        'grace_period_days': int(credit_settings.grace_period_days or 0),
        'sales': [_serialize_credit_sale_for_pay_modal(s) for s in sales],
    })


@login_required
@require_http_methods(["POST"])
def api_pay_member_credit(request):
    """Settle all outstanding credit for a member; email receipt to member and admin."""
    import logging
    import threading

    logger = logging.getLogger(__name__)

    is_admin = is_cashier_or_admin(request.user)
    is_staff_role_user = is_staff_role(request.user)
    if not (is_admin or is_staff_role_user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body)
        member_id = data.get('member_id')
        payment_method = (data.get('payment_method') or 'cash').strip().lower()
        notes = (data.get('notes') or '').strip()
        admin_pin = (data.get('admin_pin') or '').strip()

        if not admin_pin or len(admin_pin) != 4 or not admin_pin.isdigit():
            return JsonResponse({
                'success': False,
                'error': 'A valid 4-digit Admin PIN is required to confirm this payment.',
            })

        session_member = _get_active_member_for_refill_user(request.user)
        authorising_member = _resolve_refill_pin_authorizer(admin_pin, session_member)
        if not authorising_member:
            return JsonResponse({
                'success': False,
                'error': 'Incorrect PIN. Enter the 4-digit PIN of an active Admin member with a PIN set.',
            })

        if payment_method not in ('cash', 'debit'):
            return JsonResponse({'success': False, 'error': 'Payment method must be cash or debit'})

        try:
            member_pk = int(member_id)
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': 'Member not found'})

        raw_item_ids = data.get('item_ids')
        item_ids = None
        if raw_item_ids is not None:
            if not isinstance(raw_item_ids, list) or not raw_item_ids:
                return JsonResponse({
                    'success': False,
                    'error': 'Select at least one product to pay.',
                })
            try:
                item_ids = [int(x) for x in raw_item_ids]
            except (TypeError, ValueError):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid product selection.',
                })

        raw_transaction_ids = data.get('transaction_ids')
        transaction_ids = None
        if item_ids is None and raw_transaction_ids is not None:
            if not isinstance(raw_transaction_ids, list) or not raw_transaction_ids:
                return JsonResponse({
                    'success': False,
                    'error': 'Select at least one credit sale to pay.',
                })
            try:
                transaction_ids = [int(x) for x in raw_transaction_ids]
            except (TypeError, ValueError):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid transaction selection.',
                })

        amount_paid = None
        raw_amount = data.get('amount_paid')
        if raw_amount is not None and str(raw_amount).strip() != '':
            try:
                amount_paid = Decimal(str(raw_amount).strip()).quantize(Decimal('0.01'))
            except (InvalidOperation, TypeError):
                return JsonResponse({'success': False, 'error': 'Invalid amount to pay.'})

        from helper.credit_settlement_helper import settle_member_credit

        try:
            member = Member.objects.get(pk=member_pk, is_active=True)
        except Member.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Member not found'})

        try:
            payment, settled_sales, settled_items = settle_member_credit(
                member=member,
                payment_method=payment_method,
                performed_by=request.user,
                authorizing_member=authorising_member,
                notes=notes,
                transaction_ids=transaction_ids,
                item_ids=item_ids,
                amount_paid=amount_paid,
            )
        except ValueError as exc:
            return JsonResponse({'success': False, 'error': str(exc)})

        member.refresh_from_db(fields=['balance'])
        receipt_url = request.build_absolute_uri(
            reverse('view_credit_payment_receipt', kwargs={'payment_id': payment.pk})
        )

        def _send_emails():
            try:
                send_credit_payment_receipt_email(
                    payment=payment,
                    settled_sales=settled_sales,
                    settled_items=settled_items,
                    member=member,
                    performed_by_user=request.user,
                    authorizing_member=authorising_member,
                    receipt_url=receipt_url,
                )
            except Exception as exc:
                logger.error('Failed to send credit payment emails: %s', exc, exc_info=True)

        threading.Thread(target=_send_emails, daemon=True).start()

        from helper.credit_settlement_helper import member_credit_outstanding_amount

        remaining_credit = member_credit_outstanding_amount(member)

        try:
            from admin_panel.audit import mark_audit_recorded, record_audit
            record_audit(
                'CREDIT_PAYMENT',
                actor=request.user,
                description=(
                    f"Credit payment ₱{payment.amount_paid} for {member.full_name} "
                    f"via {payment.payment_method}"
                ),
                request=request,
                object_type='CreditPayment',
                object_id=payment.pk,
                metadata={
                    'settlement_number': payment.settlement_number,
                    'amount_paid': str(payment.amount_paid),
                    'payment_method': payment.payment_method,
                    'member': member.full_name,
                    'remaining_credit': str(remaining_credit),
                },
            )
            mark_audit_recorded(request)
        except Exception:
            pass

        return JsonResponse({
            'success': True,
            'message': (
                f'Successfully recorded credit payment of ₱{payment.amount_paid} '
                f'for {member.full_name}.'
            ),
            'settlement_number': payment.settlement_number,
            'amount_paid': str(payment.amount_paid),
            'payment_method': payment.payment_method,
            'receipt_url': receipt_url,
            'member_notified': bool(member.email),
            'member': {
                'id': member.id,
                'name': member.full_name,
                'new_balance': str(member.balance),
                'credit_outstanding': str(remaining_credit),
            },
            'settled_count': len(settled_items),
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as exc:
        return JsonResponse({'success': False, 'error': f'Server error: {exc}'})


@login_required
def backup_members_data(request):
    """Export member data as Excel (.xlsx) or CSV backup file."""
    is_admin = is_cashier_or_admin(request.user)
    is_staff_role_user = is_staff_role(request.user)

    if not (is_admin or is_staff_role_user):
        messages.warning(request, 'You do not have permission to access this page.')
        return redirect('kiosk_home')

    backup_date_str = request.GET.get('date', '')
    if backup_date_str:
        try:
            backup_date = datetime.strptime(backup_date_str, '%Y-%m-%d').date()
        except ValueError:
            backup_date = timezone.now().date()
    else:
        backup_date = timezone.now().date()

    export_format = (request.GET.get('format') or 'xlsx').strip().lower()
    if export_format not in ('xlsx', 'csv'):
        export_format = 'xlsx'

    backup_datetime_end = timezone.make_aware(datetime.combine(backup_date, datetime.max.time()))
    members = Member.objects.select_related('member_type', 'member_role', 'user').filter(
        created_at__lte=backup_datetime_end
    ).order_by('id')

    headers = [
        'ID',
        'RFID Card Number',
        'First Name',
        'Last Name',
        'Full Name',
        'Email',
        'Phone',
        'Member Type',
        'Role',
        'Balance',
        'Share Capital',
        'Is Active',
        'Inactive Remark',
        'Username',
        'Has PIN Set',
        'Date Joined',
        'Last Transaction',
        'Created At',
        'Updated At',
    ]

    def _fmt_dt(value):
        if not value:
            return ''
        return timezone.localtime(value).strftime('%Y-%m-%d %H:%M:%S') if timezone.is_aware(value) else value.strftime('%Y-%m-%d %H:%M:%S')

    rows = []
    for member in members:
        rows.append([
            member.id,
            member.rfid_card_number or '',
            member.first_name or '',
            member.last_name or '',
            member.full_name,
            member.email or '',
            member.phone or '',
            member.member_type.name if member.member_type else '',
            member.get_role_display(),
            str(member.balance),
            str(member.share_capital),
            'Yes' if member.is_active else 'No',
            member.inactive_remark or '',
            member.username or (member.user.username if member.user_id else ''),
            'Yes' if member.pin_hash else 'No',
            _fmt_dt(member.date_joined),
            _fmt_dt(member.last_transaction),
            _fmt_dt(member.created_at),
            _fmt_dt(member.updated_at),
        ])

    date_str = backup_date.strftime('%Y%m%d')

    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="members_backup_{date_str}.csv"'
        response.write('\ufeff')  # Excel-friendly UTF-8 BOM
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response

    from io import BytesIO
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Members'
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    # Light header styling + reasonable column widths
    from openpyxl.styles import Font
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col[:50]:
            max_len = max(max_len, len(str(cell.value or '')))
        sheet.column_dimensions[col_letter].width = min(max(12, max_len + 2), 40)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="members_backup_{date_str}.xlsx"'
    return response


@login_required
@require_http_methods(["POST"])
def restore_members_data(request):
    """Restore deleted/inactive members from a backup date by reactivating them."""
    # Check if user has permission (admin or cashier for restore)
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied. Only admins can restore members.'}, status=403)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)
    
    restore_date_str = data.get('date', '').strip()
    restore_all = data.get('restore_all', False)  # Optional flag to restore all inactive members
    
    # Initialize variables
    inactive_members = Member.objects.none()
    deleted_members_log = DeletedMember.objects.none()
    
    # If restore_all is True, restore all inactive members regardless of date
    if restore_all:
        # Restore ALL inactive members - no date filtering
        inactive_members = Member.objects.filter(is_active=False).order_by('id')
        deleted_members_log = DeletedMember.objects.filter(restored=False).order_by('deleted_at')
        restore_date_str = 'ALL'  # For display purposes
    elif restore_date_str:
        try:
            restore_date = datetime.strptime(restore_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)
        
        # Convert restore_date to datetime for comparison (end of day)
        restore_datetime_end = timezone.make_aware(datetime.combine(restore_date, datetime.max.time()))
        
        # Find inactive members that were created on or before the restore date
        inactive_members = Member.objects.filter(
            is_active=False,
            created_at__lte=restore_datetime_end
        ).order_by('id')
        
        # Also find deleted members from log that were deleted on or before restore date
        deleted_members_log = DeletedMember.objects.filter(
            restored=False,
            deleted_at__lte=restore_datetime_end
        ).order_by('deleted_at')
    else:
        return JsonResponse({'success': False, 'error': 'Backup date is required (or set restore_all=true)'}, status=400)
    
    # Get total counts for debugging
    all_inactive_count = Member.objects.filter(is_active=False).count()
    all_deleted_log_count = DeletedMember.objects.filter(restored=False).count()
    matching_inactive_count = inactive_members.count()
    matching_deleted_log_count = deleted_members_log.count()
    
    # Print restore operation header to terminal
    print("\n" + "="*80)
    if restore_all:
        print(f"RESTORE MEMBERS OPERATION - Restoring ALL inactive members and from deletion log")
    else:
        print(f"RESTORE MEMBERS OPERATION - Backup Date: {restore_date_str}")
    print(f"Requested by: {request.user.username} ({request.user.get_full_name() or 'N/A'})")
    print(f"Timestamp: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("-"*80)
    print(f"DEBUG: Total inactive members in database: {all_inactive_count}")
    print(f"DEBUG: Inactive members matching restore criteria: {matching_inactive_count}")
    print(f"DEBUG: Total deleted members in log: {all_deleted_log_count}")
    print(f"DEBUG: Deleted members from log matching restore criteria: {matching_deleted_log_count}")
    print("="*80)
    
    restored_count = 0
    restored_members = []
    
    # Step 1: Restore inactive members (soft-deleted)
    inactive_members_list = list(inactive_members)
    print(f"\n[Step 1] Processing {len(inactive_members_list)} inactive member(s)...")
    
    for member in inactive_members_list:
        try:
            if not member.is_active:
                member.is_active = True
                member.save(update_fields=['is_active'])
                restored_count += 1
                
                member_info = {
                    'id': member.id,
                    'name': member.full_name,
                    'rfid': member.rfid_card_number,
                    'email': member.email or '',
                    'source': 'inactive'
                }
                restored_members.append(member_info)
                
                print(f"  [{restored_count}] ID: {member.id:4d} | Name: {member.full_name:30s} | "
                      f"RFID: {member.rfid_card_number:15s} | Email: {member.email or 'N/A'} | Source: Inactive")
        except Exception as e:
            print(f"  ERROR: Failed to restore member ID {member.id}: {str(e)}")
            continue
    
    # Step 2: Restore from deletion log
    deleted_members_list = list(deleted_members_log)
    print(f"\n[Step 2] Processing {len(deleted_members_list)} deleted member(s) from log...")
    
    for deleted_member in deleted_members_list:
        try:
            # Check if member already exists
            if Member.objects.filter(rfid_card_number=deleted_member.rfid_card_number).exists():
                print(f"  SKIP: Member with RFID {deleted_member.rfid_card_number} already exists, skipping...")
                continue
            
            if deleted_member.email and Member.objects.filter(email=deleted_member.email).exists():
                print(f"  SKIP: Member with email {deleted_member.email} already exists, skipping...")
                continue
            
            # Get member type if it exists
            member_type = None
            if deleted_member.member_type_name:
                try:
                    member_type = MemberType.objects.get(name=deleted_member.member_type_name)
                except MemberType.DoesNotExist:
                    pass
            
            # Get user if username was provided
            user = None
            if deleted_member.username:
                try:
                    user = User.objects.get(username=deleted_member.username)
                except User.DoesNotExist:
                    pass
            
            # Restore member
            restored_member = Member.objects.create(
                rfid_card_number=deleted_member.rfid_card_number,
                first_name=deleted_member.first_name,
                last_name=deleted_member.last_name,
                email=deleted_member.email,
                phone=deleted_member.phone,
                member_type=member_type,
                member_role=Role.resolve_slug(deleted_member.role),
                balance=deleted_member.balance,
                user=user,
                pin_hash=deleted_member.pin_hash,
                is_active=True,
                date_joined=deleted_member.original_date_joined or timezone.now(),
                last_transaction=deleted_member.original_last_transaction,
                created_at=deleted_member.original_created_at or timezone.now(),
                updated_at=timezone.now(),
            )
            
            # Mark as restored in log
            deleted_member.restored = True
            deleted_member.restored_at = timezone.now()
            deleted_member.restored_by = request.user.username
            deleted_member.save()
            
            restored_count += 1
            member_info = {
                'id': restored_member.id,
                'name': restored_member.full_name,
                'rfid': restored_member.rfid_card_number,
                'email': restored_member.email or '',
                'source': 'deletion_log'
            }
            restored_members.append(member_info)
            
            print(f"  [{restored_count}] ID: {restored_member.id:4d} | Name: {restored_member.full_name:30s} | "
                  f"RFID: {restored_member.rfid_card_number:15s} | Email: {restored_member.email or 'N/A'} | Source: Deletion Log")
        except Exception as e:
            print(f"  ERROR: Failed to restore deleted member {deleted_member.first_name} {deleted_member.last_name}: {str(e)}")
            continue
    
    if restored_count == 0:
        print("\n  No members found to restore for the selected criteria.")
        if all_inactive_count > 0 or all_deleted_log_count > 0:
            print(f"  NOTE: There are {all_inactive_count} inactive member(s) and {all_deleted_log_count} deleted member(s) in log,")
            print(f"        but they don't match the restore date criteria ({restore_date_str}).")
            print(f"        Try using 'Restore all' option or a more recent date.")
        print("="*80 + "\n")
        return JsonResponse({
            'success': True,
            'message': f'No inactive members found to restore for the date {restore_date_str}',
            'restored_count': 0,
            'restored_members': []
        })
    
    # Print summary to terminal
    print("-"*80)
    print(f"SUMMARY: Successfully restored {restored_count} member(s)")
    print("="*80 + "\n")
    
    return JsonResponse({
        'success': True,
        'message': f'Successfully restored {restored_count} member(s) from backup date {restore_date_str}',
        'restored_count': restored_count,
        'restored_members': restored_members
    })


@login_required
def transaction_history(request):
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to access this page.')
        return redirect('kiosk_home')

    # First visit: treat existing refills as already "seen" so the badge only tracks new activity.
    if 'balance_refills_last_seen_id' not in request.session:
        request.session['balance_refills_last_seen_id'] = (
            CardBalanceRefill.objects.aggregate(m=Max('id'))['m'] or 0
        )
    if request.GET.get('tab') == 'balance-refills':
        request.session['balance_refills_last_seen_id'] = (
            CardBalanceRefill.objects.aggregate(m=Max('id'))['m'] or 0
        )
        request.session.modified = True

    last_seen_refill_id = int(request.session.get('balance_refills_last_seen_id', 0) or 0)
    unseen_refills_count = CardBalanceRefill.objects.filter(id__gt=last_seen_refill_id).count()

    payment_method_filter = request.GET.get('payment_method', '').strip().lower()
    if payment_method_filter and payment_method_filter not in dict(Transaction.PAYMENT_METHODS):
        payment_method_filter = ''

    valid_statuses = {c[0] for c in Transaction.STATUS_CHOICES}
    status_filter = request.GET.get('status', '').strip().lower()
    if status_filter not in valid_statuses:
        status_filter = ''
    status_filter_label = dict(Transaction.STATUS_CHOICES).get(status_filter, '') if status_filter else ''

    # Base queryset (payment filter only) — stats pills use this; list may also filter by status.
    transactions_base_qs = (
        _exclude_test_transactions(Transaction.objects.all())
        .select_related('member')
        .prefetch_related('items', 'refund_reason')
        .order_by('-created_at', '-id')
    )
    if payment_method_filter:
        transactions_base_qs = transactions_base_qs.filter(payment_method=payment_method_filter)

    transactions_qs = transactions_base_qs
    if status_filter:
        transactions_qs = transactions_qs.filter(status=status_filter)
    paginator = Paginator(transactions_qs, 10)
    page_number = request.GET.get('page', 1)
    transactions_page = paginator.get_page(page_number)

    # Calculate statistics (always from base qs so filter pills stay accurate)
    txn_stats = transactions_base_qs.aggregate(
        total_transactions=Count('id'),
        completed_transactions=Count('id', filter=Q(status='completed')),
        pending_transactions=Count('id', filter=Q(status='pending')),
        cancelled_transactions=Count('id', filter=Q(status='cancelled')),
        refunded_transactions=Count('id', filter=Q(status='refunded')),
        partially_refunded_transactions=Count('id', filter=Q(status='partially_refunded')),
        return_expired_transactions=Count('id', filter=Q(status='return_expired')),
        refund_requested_count=Count('id', filter=Q(status='refund_requested')),
        return_window_count=Count('id', filter=Q(status='return_window')),
    )
    total_transactions = txn_stats['total_transactions']
    completed_transactions = txn_stats['completed_transactions']
    pending_transactions = txn_stats['pending_transactions']
    cancelled_transactions = txn_stats['cancelled_transactions']
    refunded_transactions = txn_stats['refunded_transactions']
    partially_refunded_transactions = txn_stats['partially_refunded_transactions']
    return_expired_transactions = txn_stats['return_expired_transactions']
    refund_requested_count = txn_stats['refund_requested_count']
    return_window_count = txn_stats['return_window_count']
    total_revenue = _transaction_net_revenue(transactions_base_qs)
    partial_refund_net_amount = transactions_base_qs.filter(
        status='partially_refunded',
    ).aggregate(
        total=Coalesce(Sum('total_amount'), Decimal('0.00')),
    )['total'] or Decimal('0.00')

    refund_requested_transactions = (
        transactions_base_qs.filter(status='refund_requested')
        .order_by('-created_at', '-id')
    )
    return_window_transactions = (
        transactions_base_qs.filter(status='return_window')
        .select_related('return_window')
        .order_by('return_window__return_deadline', 'id')
    )

    refills_qs = CardBalanceRefill.objects.select_related(
        'member', 'performed_by', 'balance_transaction', 'reversal_balance_transaction'
    ).order_by('-created_at')
    total_refills = refills_qs.count()
    refill_paginator = Paginator(refills_qs, 10)
    refill_page_number = request.GET.get('refill_page') or 1
    refills_page = refill_paginator.get_page(refill_page_number)
    refill_elided_pages = list(
        refill_paginator.get_elided_page_range(
            refills_page.number, on_each_side=2, on_ends=1
        )
    )

    credit_payments_qs = CreditPayment.objects.select_related(
        'member', 'performed_by'
    ).prefetch_related('payment_lines__item__transaction').order_by('-created_at')
    total_credit_payments = credit_payments_qs.count()
    credit_payment_paginator = Paginator(credit_payments_qs, 10)
    credit_payment_page_number = request.GET.get('credit_payment_page') or 1
    credit_payments_page = credit_payment_paginator.get_page(credit_payment_page_number)
    credit_payment_elided_pages = list(
        credit_payment_paginator.get_elided_page_range(
            credit_payments_page.number, on_each_side=2, on_ends=1
        )
    )

    from admin_panel.models import ReportScheduleConfig as _RSC
    _config = _RSC.get()

    context = {
        'transactions': transactions_page,
        'page_obj': transactions_page,
        'total_transactions': total_transactions,
        'completed_transactions': completed_transactions,
        'pending_transactions': pending_transactions,
        'cancelled_transactions': cancelled_transactions,
        'refunded_transactions': refunded_transactions,
        'refund_requested_transactions': refund_requested_transactions,
        'refund_requested_count': refund_requested_count,
        'return_window_transactions': return_window_transactions,
        'return_window_count': return_window_count,
        'return_expired_transactions': return_expired_transactions,
        'return_window_days': _config.return_window_days,
        'total_revenue': total_revenue,
        'refills_page': refills_page,
        'total_refills': total_refills,
        'unseen_refills_count': unseen_refills_count,
        'refill_elided_pages': refill_elided_pages,
        'credit_payments_page': credit_payments_page,
        'total_credit_payments': total_credit_payments,
        'credit_payment_elided_pages': credit_payment_elided_pages,
        'payment_method_filter': payment_method_filter,
        'status_filter': status_filter,
        'status_filter_label': status_filter_label,
        'partially_refunded_transactions': partially_refunded_transactions,
        'partial_refund_net_amount': partial_refund_net_amount,
        'txn_export_date_from': _store_local_today().isoformat(),
        'txn_export_date_to': _store_local_today().isoformat(),
        **admin_role_badge_context(request),
    }
    
    return render(request, 'admin_panel/transactions.html', context)


def _transaction_export_filters_from_request(request):
    """Parse payment/status GET filters the same way as the transactions page."""
    payment_method_filter = request.GET.get('payment_method', '').strip().lower()
    if payment_method_filter and payment_method_filter not in dict(Transaction.PAYMENT_METHODS):
        payment_method_filter = ''

    valid_statuses = {c[0] for c in Transaction.STATUS_CHOICES}
    status_filter = request.GET.get('status', '').strip().lower()
    if status_filter not in valid_statuses:
        status_filter = ''
    return payment_method_filter, status_filter


def _transaction_export_queryset(date_from, date_to, payment_method_filter='', status_filter=''):
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    current_tz = timezone.get_current_timezone()
    range_start_aware = timezone.make_aware(
        datetime.combine(date_from, datetime.min.time()), current_tz
    )
    range_end_aware = timezone.make_aware(
        datetime.combine(date_to + timedelta(days=1), datetime.min.time()), current_tz
    )
    qs = (
        _exclude_test_transactions(Transaction.objects.all())
        .filter(created_at__gte=range_start_aware, created_at__lt=range_end_aware)
        .select_related('member', 'processed_by')
        .order_by('-created_at', '-id')
    )
    if payment_method_filter:
        qs = qs.filter(payment_method=payment_method_filter)
    if status_filter:
        qs = qs.filter(status=status_filter)
    return date_from, date_to, qs


def _transaction_export_row_values(txn):
    created = timezone.localtime(txn.created_at) if txn.created_at else None
    return {
        'number': txn.transaction_number or '',
        'customer': txn.customer_display_name,
        'created': created,
        'amount': Decimal(str(txn.total_amount or 0)),
        'payment': txn.get_payment_method_display(),
        'status': txn.get_status_display(),
        'processed_by': txn.processed_by_display or '—',
    }


@login_required
@require_http_methods(['GET'])
def export_transaction_history(request):
    """Download transaction history as PDF or Excel for a selected date range."""
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to download transactions.')
        return redirect('kiosk_home')

    requested_format = (request.GET.get('format') or 'excel').strip().lower()
    if requested_format not in ('pdf', 'excel'):
        requested_format = 'excel'

    date_from, date_to, _, _ = _staff_sales_date_range_from_request(request)
    payment_method_filter, status_filter = _transaction_export_filters_from_request(request)
    date_from, date_to, qs = _transaction_export_queryset(
        date_from, date_to, payment_method_filter, status_filter
    )

    kiosk_config = KioskConfig.get()
    store_name = kiosk_config.system_name if kiosk_config else 'Admin'
    user_label = request.user.get_full_name() or request.user.username
    date_slug = _staff_sales_date_slug(date_from, date_to)
    filter_bits = []
    if status_filter:
        filter_bits.append(dict(Transaction.STATUS_CHOICES).get(status_filter, status_filter))
    if payment_method_filter:
        filter_bits.append(dict(Transaction.PAYMENT_METHODS).get(payment_method_filter, payment_method_filter))
    filter_label = ' · '.join(filter_bits) if filter_bits else 'All statuses and payment methods'
    txn_count = qs.count()
    total_amount = qs.aggregate(total=Coalesce(Sum('total_amount'), Decimal('0.00')))['total'] or Decimal('0.00')

    if requested_format == 'excel':
        header_fill, header_font, thin_border = _staff_sales_excel_styles()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Transactions'
        hdr_row = _staff_sales_write_excel_header(
            ws,
            f'Transaction History — {store_name}',
            date_from,
            date_to,
            user_label,
            extra_lines=[
                f'Filters: {filter_label}',
                f'Transactions: {txn_count}',
                f'Total amount (PHP): {float(total_amount):,.2f}',
            ],
        )
        hdr_row += 1
        headers = [
            '#', 'Transaction #', 'Customer', 'Date & Time',
            'Amount (PHP)', 'Payment', 'Status', 'Processed by',
        ]
        for col, val in enumerate(headers, start=1):
            c = ws.cell(row=hdr_row, column=col, value=val)
            c.fill = header_fill
            c.font = header_font
            c.border = thin_border
            c.alignment = Alignment(
                horizontal='center' if col in (1, 5) else 'left',
                vertical='center',
            )

        for idx, txn in enumerate(qs.iterator(chunk_size=500), start=1):
            row = _transaction_export_row_values(txn)
            created_str = row['created'].strftime('%Y-%m-%d %H:%M:%S') if row['created'] else ''
            ws.append([
                idx,
                row['number'],
                row['customer'],
                created_str,
                float(row['amount']),
                row['payment'],
                row['status'],
                row['processed_by'],
            ])

        last_row = ws.max_row
        for r in range(hdr_row + 1, last_row + 1):
            for col in range(1, 9):
                cell = ws.cell(row=r, column=col)
                cell.border = thin_border
                if col == 5:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 1:
                    cell.alignment = Alignment(horizontal='center')
        total_row = last_row + 1
        ws.cell(row=total_row, column=2, value='Total')
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=5, value=float(total_amount))
        total_cell.number_format = '#,##0.00'
        total_cell.font = Font(bold=True)
        for col in range(1, 9):
            ws.cell(row=total_row, column=col).border = thin_border
        for col, width in zip('ABCDEFGH', [6, 20, 28, 20, 16, 22, 22, 22]):
            ws.column_dimensions[col].width = width

        return _staff_sales_excel_response(wb, f'transactions_{date_slug}.xlsx')

    pdf_primary_dark = colors.HexColor('#C4121A')
    pdf_heading = colors.HexColor('#166534')
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TxnExportTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=pdf_primary_dark,
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
    )
    cell_style = ParagraphStyle(
        'TxnExportCell',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
    )
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=28,
        leftMargin=28,
        topMargin=28,
        bottomMargin=28,
    )
    elements = [
        Paragraph('Transaction History', title_style),
        Paragraph(f'<i>{escape(store_name)}</i>', styles['Normal']),
        Spacer(1, 0.1 * inch),
        Paragraph(
            f'Period: {date_from.isoformat()} to {date_to.isoformat()}',
            styles['Normal'],
        ),
        Paragraph(
            f'Generated: {escape(timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M"))}'
            f' — {escape(user_label)}',
            styles['Normal'],
        ),
        Paragraph(f'Filters: {escape(filter_label)}', styles['Normal']),
        Spacer(1, 0.08 * inch),
        Paragraph(
            f'<b>Transactions:</b> {txn_count} &nbsp;|&nbsp; '
            f'<b>Total amount:</b> PHP {float(total_amount):,.2f}',
            styles['Normal'],
        ),
        Spacer(1, 0.16 * inch),
    ]

    table_data = [[
        '#', 'Transaction #', 'Customer', 'Date & Time',
        'Amount (PHP)', 'Payment', 'Status', 'Processed by',
    ]]
    for idx, txn in enumerate(qs.iterator(chunk_size=500), start=1):
        row = _transaction_export_row_values(txn)
        created_str = row['created'].strftime('%Y-%m-%d %H:%M') if row['created'] else ''
        table_data.append([
            str(idx),
            Paragraph(escape(row['number']), cell_style),
            Paragraph(escape(row['customer']), cell_style),
            created_str,
            f'{float(row["amount"]):,.2f}',
            Paragraph(escape(row['payment']), cell_style),
            Paragraph(escape(row['status']), cell_style),
            Paragraph(escape(row['processed_by']), cell_style),
        ])
    if txn_count == 0:
        table_data.append(['', 'No transactions in this date range.', '', '', '', '', '', ''])
    else:
        table_data.append([
            '', 'Total', '', '', f'{float(total_amount):,.2f}', '', '', '',
        ])

    tbl = Table(
        table_data,
        colWidths=[28, 95, 110, 78, 70, 90, 90, 90],
        repeatRows=1,
    )
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), pdf_heading),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d1d5db')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(tbl)
    doc.build(elements)
    pdf = buffer.getvalue()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="transactions_{date_slug}.pdf"'
    return resp


@login_required
@require_http_methods(['POST'])
def api_mark_balance_refills_seen(request):
    """Clear the Balance refills notification badge after the user opens that tab."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Forbidden'}, status=403)
    max_id = CardBalanceRefill.objects.aggregate(m=Max('id'))['m'] or 0
    request.session['balance_refills_last_seen_id'] = max_id
    request.session.modified = True
    return JsonResponse({'success': True})


@login_required
@require_http_methods(['POST'])
def api_reverse_balance_refill(request):
    """Undo a card balance refill: deduct the same amount after Admin PIN verification."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)

    admin_pin = (data.get('admin_pin') or '').strip()
    if not admin_pin or len(admin_pin) != 4 or not admin_pin.isdigit():
        return JsonResponse({
            'success': False,
            'error': 'A valid 4-digit Admin PIN is required to confirm this action.',
        })

    session_member = _get_active_member_for_refill_user(request.user)
    authorising_member = _resolve_refill_pin_authorizer(admin_pin, session_member)
    if not authorising_member:
        return JsonResponse({
            'success': False,
            'error': 'Incorrect PIN. Enter the 4-digit PIN of an active Admin member with a PIN set.',
        })

    try:
        refill_pk = int(data.get('refill_id'))
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid refill.'}, status=400)

    reversal_txn_number = None
    try:
        with db_transaction.atomic():
            refill = (
                CardBalanceRefill.objects.select_related('member', 'balance_transaction')
                .select_for_update()
                .get(pk=refill_pk)
            )
            if refill.reversed_at is not None:
                return JsonResponse({'success': False, 'error': 'This refill has already been reversed.'})

            if not refill.balance_transaction_id:
                return JsonResponse({
                    'success': False,
                    'error': 'This refill has no ledger row and cannot be reversed.',
                })

            member = Member.objects.select_for_update().get(pk=refill.member_id, is_active=True)
            amount = refill.amount
            balance_before = member.balance
            if balance_before < amount:
                return JsonResponse({
                    'success': False,
                    'error': (
                        f'Insufficient balance to reverse. Member has ₱{balance_before:.2f}; '
                        f'refill was ₱{amount:.2f}.'
                    ),
                })

            if not member.deduct_balance(amount):
                return JsonResponse({
                    'success': False,
                    'error': 'Could not deduct balance. Please try again.',
                })

            member.refresh_from_db(fields=['balance'])
            balance_after = member.balance
            orig_num = refill.balance_transaction.transaction_number
            auth_note = ''
            if not session_member or authorising_member.pk != session_member.pk:
                auth_note = (
                    f' PIN authorised by {authorising_member.full_name} '
                    f'({authorising_member.get_role_display()}).'
                )

            rev_bt = BalanceTransaction.objects.create(
                member=member,
                transaction_type='deduction',
                amount=amount,
                balance_before=balance_before,
                balance_after=balance_after,
                notes=(
                    f'Reversal of balance refill {orig_num} (₱{amount:.2f}).'
                    f' Reversed by {request.user.username}.{auth_note}'
                ),
            )
            reversal_txn_number = rev_bt.transaction_number
            refill.reversal_balance_transaction = rev_bt
            refill.reversed_at = timezone.now()
            refill.reversed_by = request.user
            refill.save(
                update_fields=['reversal_balance_transaction', 'reversed_at', 'reversed_by']
            )
    except CardBalanceRefill.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Refill not found.'}, status=404)
    except Member.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Member not found or inactive.'}, status=400)

    return JsonResponse({
        'success': True,
        'message': 'Balance refill reversed successfully.',
        'reversal_transaction_number': reversal_txn_number,
    })


@require_http_methods(["GET", "POST"])
def admin_logout(request):
    """Log out every role (admin, cashier, staff, committee, member) to the login page."""
    return _logout_to_login(request)


@require_http_methods(["GET", "POST"])
def kiosk_logout(request):
    """Logout endpoint that clears all session state and shows the login page."""
    return _logout_to_login(request)


def _logout_to_login(request):
    logout_user(request)
    messages.success(request, 'You have been successfully logged out.')
    store_profile = StoreProfile.get()
    response = render(
        request,
        'admin_panel/login.html',
        {
            'store_profile': store_profile,
            'clear_kiosk_browser_state': True,
        },
    )
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@member_or_login_required
def user_choice(request):
    """Choice page for regular users after login - view transactions or go to kiosk"""
    # Check if user is authenticated (Django user)
    if request.user.is_authenticated:
        if is_loans_only_user(request.user):
            return redirect('loans_overview')
        if is_cashier_or_admin(request.user):
            return redirect('dashboard')
        # Get member associated with user for template
        member = None
        try:
            member = Member.objects.get(user=request.user, is_active=True)
        except (Member.DoesNotExist, Member.MultipleObjectsReturned):
            pass
        
        context = {
            'user': request.user,
            'member': member,
        }
        return render(request, 'admin_panel/user_choice.html', context)
    
    # Member without Django user account (session-based)
    member_id = request.session.get('member_id')
    if member_id:
        try:
            member = Member.objects.get(id=member_id, is_active=True)
            # Create a mock user object for template compatibility
            class MockUser:
                def __init__(self, member):
                    self.username = member.rfid_card_number
                    self.first_name = member.first_name
                    self.last_name = member.last_name
                    self.is_authenticated = True
                
                def get_full_name(self):
                    return f"{self.first_name} {self.last_name}".strip()
            
            mock_user = MockUser(member)
            context = {
                'user': mock_user,
                'member': member,
            }
            return render(request, 'admin_panel/user_choice.html', context)
        except Member.DoesNotExist:
            pass
    
    # Should not reach here due to decorator, but just in case
    messages.warning(request, 'Please log in to access this page.')
    return redirect('root_login')


USER_TRANSACTIONS_LIMIT = 10
# Recent rows per table used to build merged "all activity" (enough to determine global latest N).
USER_TRANSACTIONS_MERGE_POOL = 100


@member_or_login_required
def user_transactions(request):
    """View all transactions for the logged-in user or member (purchases, refunds, transfers, deposits)."""

    def _build_context(member, user_obj):
        import itertools

        # Recent purchase-type rows only (for lists + merge)
        tx_recent = list(
            Transaction.objects.filter(member=member)
            .select_related('member')
            .prefetch_related('items')
            .order_by('-created_at')[:USER_TRANSACTIONS_MERGE_POOL]
        )
        purchases_all = [t for t in tx_recent if t.status == 'completed']
        refunds_all = [t for t in tx_recent if t.status == 'refunded']
        pending_all = [t for t in tx_recent if t.status in ('pending', 'cancelled')]

        # Recent balance rows only
        balance_txns = list(
            BalanceTransaction.objects.filter(member=member).order_by('-created_at')[
                :USER_TRANSACTIONS_MERGE_POOL
            ]
        )

        # Categorise balance transactions
        transfers_sent = []
        transfers_received = []
        top_ups = []
        other_balance = []
        for bt in balance_txns:
            notes_lower = (bt.notes or '').lower()
            if 'fund transfer to' in notes_lower:
                transfers_sent.append(bt)
            elif 'fund transfer from' in notes_lower:
                transfers_received.append(bt)
            elif 'refund' in notes_lower:
                # refund credit to balance – shown alongside refund section
                other_balance.append(bt)
            elif bt.transaction_type == 'deposit':
                top_ups.append(bt)
            else:
                other_balance.append(bt)

        # Single chronological list for Transfers tab (max N total, not N sent + N received)
        transfer_entries = sorted(
            [('sent', bt) for bt in transfers_sent] + [('received', bt) for bt in transfers_received],
            key=lambda x: x[1].created_at,
            reverse=True,
        )[:USER_TRANSACTIONS_LIMIT]

        # Merged "all activity" list sorted newest first (pool is capped; then show latest N only)
        all_activity = sorted(
            itertools.chain(
                [{'kind': 'purchase', 'obj': t} for t in purchases_all],
                [{'kind': 'refund',   'obj': t} for t in refunds_all],
                [{'kind': 'pending',  'obj': t} for t in pending_all],
                [{'kind': 'transfer_sent',     'obj': bt} for bt in transfers_sent],
                [{'kind': 'transfer_received', 'obj': bt} for bt in transfers_received],
                [{'kind': 'top_up',   'obj': bt} for bt in top_ups],
                [{'kind': 'balance',  'obj': bt} for bt in other_balance],
            ),
            key=lambda x: x['obj'].created_at,
            reverse=True,
        )[:USER_TRANSACTIONS_LIMIT]

        return {
            'member': member,
            'user': user_obj,
            'all_activity': all_activity,
            'purchases': purchases_all[:USER_TRANSACTIONS_LIMIT],
            'refunds': refunds_all[:USER_TRANSACTIONS_LIMIT],
            'pending': pending_all[:USER_TRANSACTIONS_LIMIT],
            'transfer_entries': transfer_entries,
            'top_ups': top_ups[:USER_TRANSACTIONS_LIMIT],
            'other_balance': other_balance[:USER_TRANSACTIONS_LIMIT],
        }

    # ── Django-authenticated user ──────────────────────────────────────────
    if request.user.is_authenticated:
        if is_cashier_or_admin(request.user):
            return redirect('dashboard')

        member = None
        try:
            member = Member.objects.get(user=request.user, is_active=True)
        except Member.DoesNotExist:
            pass
        except Member.MultipleObjectsReturned:
            member = Member.objects.filter(user=request.user, is_active=True).first()
        except Exception as e:
            import logging
            logging.getLogger(__name__).error(f"Error fetching member: {e}", exc_info=True)

        ctx = _build_context(member, request.user) if member else {'member': None, 'user': request.user}
        return render(request, 'admin_panel/user_transactions.html', ctx)

    # ── Session-based member (no Django user) ─────────────────────────────
    member_id = request.session.get('member_id')
    if member_id:
        try:
            member = Member.objects.get(id=member_id, is_active=True)

            class MockUser:
                def __init__(self, m):
                    self.username = m.rfid_card_number
                    self.first_name = m.first_name
                    self.last_name = m.last_name
                    self.is_authenticated = True

                def get_full_name(self):
                    return f"{self.first_name} {self.last_name}".strip()

            ctx = _build_context(member, MockUser(member))
            return render(request, 'admin_panel/user_transactions.html', ctx)
        except Member.DoesNotExist:
            pass

    messages.warning(request, 'Please log in to access this page.')
    return redirect('root_login')


@require_http_methods(["POST"])
def api_rfid_login(request):
    """Login directly using RFID card — delegates to login_helper."""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})

    rfid = (data.get('rfid') or '').strip()
    next_url = data.get('next', '')
    return rfid_login_json_response(request, rfid, next_url)


@login_required
@require_http_methods(["GET"])
def api_search_members(request):
    """Search members by RFID card number or name for balance refill - accessible to admin, cashier and staff"""
    # Allow admin, cashier and staff role users to search members
    is_admin = is_cashier_or_admin(request.user)
    is_staff_role_user = is_staff_role(request.user)
    
    if not (is_admin or is_staff_role_user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({'success': True, 'members': []})
    
    try:
        # Search by RFID (exact or partial) or by name
        members = Member.objects.filter(
            Q(rfid_card_number__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        ).filter(is_active=True)[:20]
        
        results = []
        show_full_rfid = is_admin_user(request.user)
        for member in members:
            rfid_value = member.rfid_card_number or ''
            results.append({
                'id': member.id,
                'rfid': rfid_value if show_full_rfid else mask_rfid(rfid_value),
                'name': member.full_name,
                'email': member.email or '',
                'current_balance': str(member.balance),
            })
        
        return JsonResponse({'success': True, 'members': results})
    except Exception as e:
        return JsonResponse({'success': False, 'error': 'Server error occurred'})


_REFILL_PIN_AUTHORIZER_ROLES = frozenset({'admin'})
_EDIT_SELF_PIN_ROLES = frozenset({'admin', 'cashier', 'staff'})
_REFUND_PIN_AUTHORIZER_ROLES = frozenset({'admin', 'cashier', 'staff'})


def _get_active_member_for_refill_user(user):
    """Resolve Member for a Django user (FK or username fallback; link FK if missing)."""
    if not user or not getattr(user, 'is_active', False):
        return None
    try:
        return Member.objects.get(user=user, is_active=True)
    except Member.DoesNotExist:
        pass
    try:
        uname = (getattr(user, 'username', '') or '').strip()
        if not uname:
            return None
        m = (
            Member.objects
            .filter(username__iexact=uname, is_active=True)
            .order_by('id')
            .first()
        )
        if not m:
            return None
        if m.user_id is None:
            m.user = user
            m.save(update_fields=['user'])
        return m
    except Exception:
        return None


def _resolve_refill_pin_authorizer(admin_pin, session_member):
    """
    Find an **Admin** member whose stored PIN matches *admin_pin*.
    Prefers *session_member* when they are Admin and their PIN matches.
    """
    if session_member and session_member.role in _REFILL_PIN_AUTHORIZER_ROLES:
        if session_member.check_pin(admin_pin):
            return session_member
    qs = (
        Member.objects.filter(is_active=True, member_role__slug__in=_REFILL_PIN_AUTHORIZER_ROLES)
        .exclude(pin_hash__isnull=True)
        .exclude(pin_hash='')
        .select_related('member_role')
        .order_by('id')
    )
    skip_pk = session_member.pk if session_member else None
    for m in qs:
        if skip_pk is not None and m.pk == skip_pk:
            continue
        if m.check_pin(admin_pin):
            return m
    return None


def _resolve_member_edit_pin_authorizer(pin, session_member):
    """
    Accept edit authorisation PIN by verifying against pin_hash in the database.

    Prefers the logged-in member when their role is admin/cashier/staff.
    Otherwise accepts any active member that has a PIN set whose hash matches.
    """
    pin = (pin or '').strip()
    if not pin or not pin.isdigit() or len(pin) != 4:
        return None

    if session_member and session_member.role in _EDIT_SELF_PIN_ROLES:
        if session_member.check_pin(pin):
            return session_member

    qs = (
        Member.objects.filter(is_active=True)
        .exclude(pin_hash__isnull=True)
        .exclude(pin_hash='')
        .select_related('member_role')
        .order_by('id')
    )
    skip_pk = session_member.pk if session_member else None
    for m in qs:
        if skip_pk is not None and m.pk == skip_pk:
            continue
        if m.check_pin(pin):
            return m
    return None


def _member_edit_pin_available():
    """True when at least one active member has a PIN saved in the database."""
    return (
        Member.objects.filter(is_active=True)
        .exclude(pin_hash__isnull=True)
        .exclude(pin_hash='')
        .exists()
    )


def _resolve_refund_pin_authorizer(pin, session_member):
    """
    Find an active admin/cashier/staff member whose PIN matches *pin*.
    Prefers *session_member* when their role is authorized and PIN matches.
    """
    if session_member and session_member.role in _REFUND_PIN_AUTHORIZER_ROLES:
        if session_member.pin_hash and session_member.check_pin(pin):
            return session_member
    qs = (
        Member.objects.filter(is_active=True, member_role__slug__in=_REFUND_PIN_AUTHORIZER_ROLES)
        .exclude(pin_hash__isnull=True)
        .exclude(pin_hash='')
        .select_related('member_role')
        .order_by('id')
    )
    skip_pk = session_member.pk if session_member else None
    for m in qs:
        if skip_pk is not None and m.pk == skip_pk:
            continue
        if m.check_pin(pin):
            return m
    return None


@login_required
@require_http_methods(["POST"])
def api_refill_balance(request):
    """Refill/add balance to a member's card - accessible to admin and staff role users"""
    import logging
    import threading
    logger = logging.getLogger(__name__)

    # Check if user is admin, cashier or staff (staff role from Member model)
    is_admin = is_cashier_or_admin(request.user)
    is_staff_role_user = is_staff_role(request.user)

    if not (is_admin or is_staff_role_user):
        return JsonResponse({'success': False, 'error': 'Permission denied. Only admin and staff can refill balances.'}, status=403)

    try:
        data = json.loads(request.body)
        member_id = data.get('member_id')
        amount = data.get('amount')
        notes = data.get('notes', '').strip()
        admin_pin = data.get('admin_pin', '').strip()

        # --- Security: verify Admin-role 4-digit PIN before processing ---
        if not admin_pin or len(admin_pin) != 4 or not admin_pin.isdigit():
            return JsonResponse({
                'success': False,
                'error': 'A valid 4-digit Admin PIN is required to confirm this transaction.',
                'pin_wrong': True,
            })

        session_member = _get_active_member_for_refill_user(request.user)
        authorising_member = _resolve_refill_pin_authorizer(admin_pin, session_member)
        if not authorising_member:
            return JsonResponse({
                'success': False,
                'error': 'Incorrect PIN. Enter the 4-digit PIN of an active Admin member with a PIN set.',
                'pin_wrong': True,
            })

        if not member_id:
            return JsonResponse({'success': False, 'error': 'Member ID is required'})

        if not amount:
            return JsonResponse({'success': False, 'error': 'Amount is required'})

        try:
            amount = Decimal(str(amount))
            if amount <= 0:
                return JsonResponse({'success': False, 'error': 'Amount must be greater than zero'})
        except (InvalidOperation, ValueError):
            return JsonResponse({'success': False, 'error': 'Invalid amount format'})

        try:
            member_pk = int(member_id)
        except (TypeError, ValueError):
            return JsonResponse({'success': False, 'error': 'Member not found'})

        # Get the staff member who is performing the refill (if staff role)
        staff_member = None
        performed_by_role = 'admin'
        performed_by_name = request.user.get_full_name() or request.user.username

        if is_staff_role_user and not is_admin:
            try:
                staff_member = Member.objects.get(user=request.user, is_active=True)
                if staff_member.role == 'staff':
                    performed_by_role = 'staff'
                    performed_by_name = staff_member.full_name
            except Member.DoesNotExist:
                pass

        with db_transaction.atomic():
            try:
                member = Member.objects.select_for_update().get(id=member_pk, is_active=True)
            except Member.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Member not found'})

            balance_before = member.balance
            member.add_balance(amount)
            member.refresh_from_db(fields=['balance'])
            balance_after = member.balance

            transaction_notes = f"Balance refill by {performed_by_role}"
            if notes:
                transaction_notes += f". {notes}"
            if not session_member or authorising_member.pk != session_member.pk:
                transaction_notes += (
                    f" (PIN authorised by {authorising_member.full_name} — "
                    f"{authorising_member.get_role_display()})"
                )

            bt = BalanceTransaction.objects.create(
                member=member,
                transaction_type='deposit',
                amount=amount,
                balance_before=balance_before,
                balance_after=balance_after,
                notes=transaction_notes,
            )
            CardBalanceRefill.objects.create(
                member=member,
                amount=amount,
                balance_before=balance_before,
                balance_after=balance_after,
                notes=notes,
                balance_transaction=bt,
                performed_by=request.user,
            )
            refill_transaction_number = bt.transaction_number

        timestamp_str = timezone.localtime(timezone.now()).strftime('%B %d, %Y %I:%M %p')

        pin_audit_block = ''
        if not session_member or authorising_member.pk != session_member.pk:
            pin_audit_block = (
                f"PIN Authorised By:\n"
                f"  • Name : {authorising_member.full_name}\n"
                f"  • Role : {authorising_member.get_role_display()}\n\n"
            )

        # --- Always send email notification to admin ---
        def send_admin_email():
            try:
                admin_email = get_admin_email()
                if admin_email:
                    subject = f'[Balance Refill] {member.full_name} — ₱{amount:.2f} Added'
                    body = f"""Balance Refill Confirmation
{'=' * 45}

A card balance refill has been processed successfully.

Transaction Details:
  • Transaction #  : {refill_transaction_number}
  • Member Name  : {member.full_name}
  • RFID Card    : {member.rfid_card_number}
  • Amount Added : ₱{amount:.2f}
  • Balance Before: ₱{balance_before:.2f}
  • Balance After : ₱{balance_after:.2f}

Processed By:
  • Name : {performed_by_name}
  • Role : {performed_by_role.capitalize()}
  • Login: {request.user.username}

{pin_audit_block}Additional Notes: {notes if notes else 'None'}

Timestamp: {timestamp_str}
{'=' * 45}
This notification was sent automatically by the system.
"""
                    email_msg = EmailMessage(
                        subject=subject,
                        body=body.strip(),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[admin_email],
                    )
                    email_msg.send(fail_silently=False)
            except Exception as exc:
                logger.error(f'Failed to send admin balance refill email: {exc}', exc_info=True)

        # --- Send email notification to member if they have an email ---
        def send_member_email():
            try:
                if member.email:
                    subject = f'Your Card Balance Has Been Refilled — ₱{amount:.2f}'
                    body = f"""Hi {member.first_name},

Great news! Your BAGNOS MPC card balance has been successfully refilled.

Transaction Summary:
  • Transaction # : {refill_transaction_number}
  • Amount Added  : ₱{amount:.2f}
  • Balance Before: ₱{balance_before:.2f}
  • New Balance   : ₱{balance_after:.2f}

If you did not request this refill or believe this is an error,
please contact the admin immediately.

Transaction Date: {timestamp_str}

Thank you for using BAGNOS MPC!
"""
                    email_msg = EmailMessage(
                        subject=subject,
                        body=body.strip(),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[member.email],
                    )
                    email_msg.send(fail_silently=False)
            except Exception as exc:
                logger.error(f'Failed to send member balance refill email: {exc}', exc_info=True)

        # Send both emails in background threads so they don't block the response
        threading.Thread(target=send_admin_email, daemon=True).start()
        threading.Thread(target=send_member_email, daemon=True).start()

        member_email_sent = bool(member.email)

        try:
            from admin_panel.audit import mark_audit_recorded, record_audit
            record_audit(
                'BALANCE_REFILL',
                actor=request.user,
                description=(
                    f"Refilled ₱{amount:.2f} to {member.full_name} "
                    f"(new balance ₱{member.balance})"
                ),
                request=request,
                object_type='Member',
                object_id=member.pk,
                metadata={
                    'amount': str(amount),
                    'member': member.full_name,
                    'new_balance': str(member.balance),
                    'transaction_number': refill_transaction_number,
                },
            )
            mark_audit_recorded(request)
        except Exception:
            pass

        return JsonResponse({
            'success': True,
            'message': f'Successfully added ₱{amount:.2f} to {member.full_name}\'s balance.',
            'transaction_number': refill_transaction_number,
            'member_notified': member_email_sent,
            'member': {
                'id': member.id,
                'name': member.full_name,
                'rfid': member.rfid_card_number,
                'new_balance': str(member.balance),
                'email': member.email or '',
            }
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'})


def generate_refund_receipt_data(
    transaction,
    refund_reason,
    member,
    balance_before=None,
    balance_after=None,
    request=None,
    receipt_item_ids=None,
):
    """Generate refund receipt text data"""
    lines = []
    
    def money(v):
        if v is None:
            return '₱0.00'
        return '₱' + str(Decimal(str(v)).quantize(Decimal('0.01')))
    
    # Header (uses system name from Admin → Kiosk Config)
    lines.append(KioskConfig.get().system_name.upper())
    lines.append('REFUND RECEIPT')
    lines.append('')
    
    # Transaction info
    lines.append('Original Txn:')
    lines.append(transaction.transaction_number)
    lines.append('Refund Date:')
    lines.append(timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S'))
    lines.append('')
    
    # Member info
    if member:
        lines.append('Member:')
        lines.append(member.full_name)
        if hasattr(member, 'member_id') and member.member_id:
            lines.append(f'Member ID: {member.member_id}')
        lines.append('')
    
    display = _refund_receipt_display_context(transaction, item_ids=receipt_item_ids)
    refunded_items = display['refunded_items']

    # Items refunded
    lines.append('ITEMS REFUNDED:')
    for item in refunded_items:
        name = item.product_name
        qty = item.quantity
        total = money(item.total_price)
        lines.append(f'{name} x{qty}')
        lines.append(total)
    lines.append('')
    
    # Amounts
    lines.append('Subtotal:')
    lines.append(money(display['refund_subtotal']))
    lines.append('Total Refund:')
    lines.append(money(display['refund_total_amount']))
    lines.append('')
    
    # Payment method refund info - All refunds now go to balance
    lines.append('REFUND METHOD:')
    if member and balance_before is not None:
        lines.append('Refunded to Member Balance')
        lines.append(f'Balance Before: {money(balance_before)}')
        lines.append(f'Balance After: {money(balance_after)}')
    else:
        lines.append('Refunded to Member Balance')
    lines.append('')
    
    # Reason if provided
    if refund_reason:
        lines.append('Reason:')
        lines.append(refund_reason)
        lines.append('')
    
    lines.append('Thank you!')
    
    return {
        'text': '\r\n'.join(lines),
        'html': generate_refund_receipt_html(
            transaction,
            refund_reason,
            member,
            balance_before,
            balance_after,
            request=request,
            receipt_item_ids=receipt_item_ids,
        ),
    }


def _match_items_by_product_names(all_items, names):
    matched = []
    for name in names:
        for item in all_items:
            if item.product_name == name and item not in matched:
                matched.append(item)
                break
    return matched


def _parse_refund_receipt_item_ids(raw):
    """Normalize optional receipt line filters from query params or API payloads."""
    if raw is None:
        return None
    if isinstance(raw, str):
        raw = [x.strip() for x in raw.split(',') if x.strip()]
    try:
        ids = {int(x) for x in raw}
    except (TypeError, ValueError):
        return None
    return ids or None


def _latest_refund_batch_items(refunded_lines):
    """Return only the most recent refund batch when lines were refunded at different times."""
    if not refunded_lines:
        return []
    timestamps = {item.refunded_at for item in refunded_lines if item.refunded_at}
    if len(timestamps) <= 1:
        return refunded_lines
    latest = max(timestamps)
    return [item for item in refunded_lines if item.refunded_at == latest]


def _items_refunded_for_receipt(transaction, item_ids=None):
    """Transaction line items that were actually refunded (for receipt display)."""
    all_items = list(transaction.items.order_by('id'))
    if not all_items:
        return []

    item_id_set = _parse_refund_receipt_item_ids(item_ids)
    if item_id_set is not None:
        picked = [i for i in all_items if i.id in item_id_set]
        if picked:
            return picked

    refunded_lines = list(transaction.items.filter(refunded_at__isnull=False).order_by('id'))
    if refunded_lines:
        return _latest_refund_batch_items(refunded_lines)

    try:
        selected = list(transaction.refund_reason.refund_items.order_by('id'))
        if selected:
            return selected
    except RefundReason.DoesNotExist:
        pass

    notes = transaction.notes or ''

    id_match = re.search(r'\[refund_items:([\d,]+)\]', notes)
    if id_match:
        ids = {int(x) for x in id_match.group(1).split(',') if x.strip().isdigit()}
        picked = [i for i in all_items if i.id in ids]
        if picked:
            return picked

    if 'item(s):' in notes:
        match = re.search(r'item\(s\):\s*(.+?)\)(?:\.|\s*\[)', notes)
        if match:
            names = [n.strip() for n in match.group(1).split(',') if n.strip()]
            matched = _match_items_by_product_names(all_items, names)
            if matched:
                return matched

    # Legacy partial refunds: item names recorded on the member balance transaction.
    balance_txn = BalanceTransaction.objects.filter(
        notes__icontains=transaction.transaction_number,
    ).filter(
        Q(notes__icontains='Partial refund') | Q(notes__icontains='Refund for'),
    ).order_by('-created_at').first()
    if balance_txn and 'item(s):' in (balance_txn.notes or ''):
        match = re.search(r'item\(s\):\s*(.+?)\)\s*\(Original', balance_txn.notes)
        if match:
            names = [n.strip() for n in match.group(1).split(',') if n.strip()]
            matched = _match_items_by_product_names(all_items, names)
            if matched:
                return matched

    partial_m = re.search(r'Partially refunded \((\d+) of (\d+) item', notes)
    if partial_m and int(partial_m.group(1)) < int(partial_m.group(2)):
        # Partial refund but line items were not persisted — avoid showing the full sale.
        return []

    # Legacy full refunds processed before per-line refunded_at existed.
    if transaction.status in ('refunded', 'cancelled'):
        return all_items

    return []


def _extract_refund_reason_from_transaction_notes(transaction):
    notes = (transaction.notes or '').strip()
    if not notes:
        return ''
    if notes.startswith('Partially refunded'):
        parts = notes.split('. ', 1)
        if len(parts) > 1 and 'item(s)' in parts[0]:
            return parts[1].strip()
        return ''
    if 'Refunded' in notes:
        parts = notes.split('.', 1)
        if len(parts) > 1:
            return parts[1].strip()
    return ''


def _refund_receipt_display_context(transaction, balance_txn=None, item_ids=None):
    """Amounts and line items shown on refund receipts (partial vs full)."""
    refunded_items = _items_refunded_for_receipt(transaction, item_ids=item_ids)
    if balance_txn is None:
        balance_txn = BalanceTransaction.objects.filter(
            notes__icontains=transaction.transaction_number,
        ).filter(
            Q(notes__icontains='Refund') | Q(notes__icontains='Partial refund')
        ).order_by('-created_at').first()

    refund_subtotal = sum(
        (item.total_price for item in refunded_items),
        Decimal('0.00'),
    ).quantize(Decimal('0.01'))
    refund_total = refund_subtotal
    if balance_txn and balance_txn.amount is not None:
        balance_amount = Decimal(str(balance_txn.amount)).quantize(Decimal('0.01'))
        if balance_amount == refund_subtotal:
            refund_total = balance_amount
    all_count = transaction.items.count()
    notes = transaction.notes or ''
    partial_m = re.search(r'Partially refunded \((\d+) of (\d+) item', notes)
    if partial_m:
        is_partial = int(partial_m.group(1)) < int(partial_m.group(2))
    else:
        is_partial = (
            transaction.status == 'partially_refunded'
            or (bool(refunded_items) and len(refunded_items) < all_count)
        )
    return {
        'refunded_items': refunded_items,
        'refund_subtotal': refund_subtotal,
        'refund_total_amount': refund_total,
        'is_partial_refund': is_partial,
    }


def _receipt_merchandise_discount_context(transaction, items=None):
    """
    Receipt breakdown using the same pricing stack as checkout: ProductDiscount
    promos plus SegmentProductGroupDiscount fixed amounts for the transaction member.

    Per line: list amount (regular unit × qty from unit_price_after_discounts) minus
    charged line total. Positive gaps are summed as discount. Labels from pricing
    meta are de-duplicated for the receipt summary line.
    """
    items = list(items if items is not None else transaction.items.all())
    member = transaction.member
    product_ids = [i.product_id for i in items if i.product_id]
    disc_map = discounts_by_product_ids(product_ids)
    segment_rules = None
    if member:
        segment_rules = list(SegmentProductGroupDiscount.objects.filter(is_active=True).select_related('discount_group'))

    gross = Decimal("0.00")
    discount = Decimal("0.00")
    summary_parts = []
    manual_discount_total = Decimal("0.00")

    for item in items:
        qty = Decimal(item.quantity)
        manual_line = Decimal(str(item.manual_discount_php or 0)).quantize(Decimal("0.01"))
        if manual_line > 0:
            manual_discount_total += manual_line

        prod = getattr(item, "product", None)
        charged_gross = (Decimal(str(item.unit_price)) * qty).quantize(Decimal("0.01"))
        if prod is None:
            gross += charged_gross
            continue

        _eff, regular, meta = unit_price_after_discounts(
            prod,
            discount_list=disc_map.get(prod.id, []),
            member=member,
            segment_rules=segment_rules,
        )
        list_line = (regular * qty).quantize(Decimal("0.01"))
        gross += list_line
        promo_disc = (list_line - charged_gross).quantize(Decimal("0.01"))
        if promo_disc > 0:
            discount += promo_disc
            name = (meta or {}).get("discount_name")
            if name and name not in summary_parts:
                summary_parts.append(name)

    if manual_discount_total > 0:
        discount += manual_discount_total
        if 'Discount (₱)' not in summary_parts:
            summary_parts.append('Discount (₱)')

    discount = discount.quantize(Decimal("0.01"))
    if discount < 0:
        discount = Decimal("0.00")
    has_discount = discount > 0
    return {
        "receipt_has_discount": has_discount,
        "receipt_gross_subtotal": gross,
        "receipt_discount_total": discount,
        "receipt_discount_summary": "; ".join(summary_parts) if summary_parts else "",
    }


# Statuses where cashiers/admins may reprint the original sale receipt from the dashboard.
_REPRINT_SALE_STATUSES = (
    'completed',
    'refund_requested',
    'return_window',
    'return_expired',
    'refunded',
    'cancelled',
)


def _is_receipt_reprint(request):
    """True when receipt is opened for reprint from the transaction history dashboard."""
    return request.GET.get('reprint', '').lower() in ('1', 'true', 'yes')


def _sale_receipt_lookup(request, transaction_id, payment_method):
    """Resolve a transaction for cash/debit/credit sale receipt printing."""
    filters = {'id': transaction_id}
    if isinstance(payment_method, (list, tuple)):
        filters['payment_method__in'] = payment_method
    else:
        filters['payment_method'] = payment_method
    if request.user.is_authenticated and is_cashier_or_admin(request.user):
        filters['status__in'] = _REPRINT_SALE_STATUSES
    else:
        filters['status'] = 'completed'
    return Transaction.objects.select_related(
        'member', 'member__senior_profile', 'member__pwd_profile', 'processed_by'
    ).prefetch_related('items__product').get(**filters)


def _receipt_print_context():
    """Build shared printer/paper context used by all receipt templates."""
    from admin_panel.models import PrinterSettings

    printer_settings = PrinterSettings.get()
    paper_size = (printer_settings.paper_size or "").strip()

    # Keep values explicit so templates can reuse a consistent print layout.
    # @page margin 0 — content uses full roll width; inner padding is on receipt sections.
    layout_map = {
        "57mm": {"page_size": "57mm auto", "page_margin": "0", "paper_width": "57mm"},
        "58mm": {"page_size": "58mm auto", "page_margin": "3.5mm", "paper_width": "51mm"},
        "80mm": {"page_size": "80mm auto", "page_margin": "3.5mm", "paper_width": "73mm"},
        "A4": {"page_size": "A4", "page_margin": "10mm", "paper_width": "180mm"},
    }
    layout = layout_map.get(paper_size, layout_map["57mm"])

    return {
        "printer_settings": printer_settings,
        "receipt_page_size": layout["page_size"],
        "receipt_page_margin": layout["page_margin"],
        "receipt_paper_width": layout["paper_width"],
    }


def _receipt_shop_header_context(kiosk_config):
    """Store header lines from KioskConfig for on-screen and printed receipts."""
    return {
        'shop_name': kiosk_config.receipt_header_store_name,
        'shop_description': kiosk_config.receipt_header_store_description,
        'shop_address': kiosk_config.receipt_header_address,
        'shop_phone': kiosk_config.receipt_header_phone,
    }


def _active_tax_label():
    """
    Return a human-readable tax label from the first active TaxRate record.
    Used on all receipts so the label automatically reflects the rate set in
    Admin → Inventory → Tax rates without touching any template.
    e.g.  TaxRate(name="VAT 12%", rate=12) → "VAT 12%"
          TaxRate(name="VAT",     rate=12) → "VAT 12%"
    Falls back to "VAT (12%)" when no active rate exists.
    """
    from inventory.models import TaxRate
    tax = TaxRate.objects.filter(is_active=True).order_by('name').first()
    if not tax:
        return 'VAT (12%)'
    rate_pct = float(tax.rate)
    pct_str = f'{int(rate_pct)}%' if rate_pct == int(rate_pct) else f'{rate_pct}%'
    if '%' in tax.name:
        return tax.name
    return f'{tax.name} ({pct_str})'


def _append_processed_by_lines(lines, transaction):
    name = transaction.processed_by_display
    if name:
        lines.extend(['', f'Processed By: {name}'])
    return lines


def _build_cash_receipt_print_text(transaction, change_amount):
    """Build plain text specifically for cash receipt printing."""
    lines = [
        "CASH RECEIPT",
        f"Txn No: {transaction.transaction_number}",
        f"Date: {timezone.localtime(transaction.created_at).strftime('%Y-%m-%d %I:%M %p')}",
        "",
        "ITEMS",
    ]

    for item in transaction.items.all():
        lines.append(f"- {item.product_name} x{item.quantity}  P{item.total_price}")

    disc = _receipt_merchandise_discount_context(transaction)
    lines.append("")
    if disc["receipt_has_discount"]:
        lines.append(f"Subtotal: P{disc['receipt_gross_subtotal']}")
        lines.append(f"Discount: P-{disc['receipt_discount_total']}")
        if disc.get("receipt_discount_summary"):
            lines.append(f"          {disc['receipt_discount_summary']}")
    else:
        lines.append(f"Subtotal: P{transaction.subtotal}")
    if transaction.vat_amount:
        tax_label = _active_tax_label()
        lines.append(f"Vatable Sale: P{transaction.vatable_sale}")
        lines.append(f"{tax_label}: P{transaction.vat_amount}")
    lines.append(f"Total:    P{transaction.total_amount}")
    lines.extend(
        [
            f"Cash:     P{transaction.amount_paid}",
            f"Change:   P{change_amount}",
        ]
    )

    if transaction.member:
        lines.extend(
            [
                "",
                f"Member: {transaction.member.full_name}",
            ]
        )

    _append_processed_by_lines(lines, transaction)
    lines.extend(["", "Thank you!", "", ""])
    return "\r\n".join(lines)


def _build_credit_receipt_print_text(transaction):
    """Build plain text specifically for credit receipt printing."""
    lines = [
        "CREDIT RECEIPT",
        f"Txn No: {transaction.transaction_number}",
        f"Date: {timezone.localtime(transaction.created_at).strftime('%Y-%m-%d %I:%M %p')}",
        "Status: APPROVED",
        "Payment: CREDIT",
        "",
        "ITEMS",
    ]

    for item in transaction.items.all():
        lines.append(f"- {item.product_name} x{item.quantity}  P{item.total_price}")

    disc = _receipt_merchandise_discount_context(transaction)
    lines.append("")
    if disc["receipt_has_discount"]:
        lines.append(f"List:      P{disc['receipt_gross_subtotal']}")
        lines.append(f"Discount:  P-{disc['receipt_discount_total']}")
        if disc.get("receipt_discount_summary"):
            lines.append(f"           {disc['receipt_discount_summary']}")
    lines.append(f"Subtotal:  P{transaction.subtotal}")
    if transaction.vat_amount:
        tax_label = _active_tax_label()
        lines.append(f"Vatable Sale: P{transaction.vatable_sale}")
        lines.append(f"{tax_label}: P{transaction.vat_amount}")
    lines.extend(
        [
            f"Total:     P{transaction.total_amount}",
            f"Credit:    P{transaction.total_amount}",
        ]
    )

    if transaction.member:
        lines.extend(
            [
                "",
                f"Member: {transaction.member.full_name}",
                f"Avail Bal: P{transaction.member.available_balance}",
            ]
        )

    _append_processed_by_lines(lines, transaction)
    lines.extend(["", "Thank you!", "", ""])
    return "\r\n".join(lines)


def _build_debit_receipt_print_text(transaction):
    """Build plain text specifically for debit receipt printing."""
    lines = [
        "DEBIT RECEIPT",
        f"Txn No: {transaction.transaction_number}",
        f"Date: {timezone.localtime(transaction.created_at).strftime('%Y-%m-%d %I:%M %p')}",
        "Status: APPROVED",
        "",
        "ITEMS",
    ]

    for item in transaction.items.all():
        lines.append(f"- {item.product_name} x{item.quantity}  P{item.total_price}")

    disc = _receipt_merchandise_discount_context(transaction)
    lines.append("")
    if disc["receipt_has_discount"]:
        lines.append(f"List:      P{disc['receipt_gross_subtotal']}")
        lines.append(f"Discount:  P-{disc['receipt_discount_total']}")
        if disc.get("receipt_discount_summary"):
            lines.append(f"           {disc['receipt_discount_summary']}")
    lines.append(f"Subtotal:  P{transaction.subtotal}")
    if transaction.vat_amount:
        tax_label = _active_tax_label()
        lines.append(f"Vatable Sale: P{transaction.vatable_sale}")
        lines.append(f"{tax_label}: P{transaction.vat_amount}")
    lines.append(f"Total:     P{transaction.total_amount}")

    if transaction.amount_from_balance and transaction.amount_from_balance > 0:
        lines.append(f"From Bal:  P{transaction.amount_from_balance}")

    if transaction.member:
        lines.extend(
            [
                "",
                f"Member: {transaction.member.full_name}",
                f"Avail Bal: P{transaction.member.available_balance}",
            ]
        )

    _append_processed_by_lines(lines, transaction)
    lines.extend(["", "Thank you!", "", ""])
    return "\r\n".join(lines)


def generate_refund_receipt_html(
    transaction,
    refund_reason,
    member,
    balance_before=None,
    balance_after=None,
    request=None,
    receipt_item_ids=None,
):
    """Generate HTML version of refund receipt using template"""
    # Get shop information and printer preferences
    from admin_panel.models import KioskConfig
    kiosk_config = KioskConfig.get()
    print_ctx = _receipt_print_context()
    shop_header = _receipt_shop_header_context(kiosk_config)
    
    # Determine refund method display - All refunds now go to balance
    show_balance_refund = (member and balance_before is not None)
    show_cash_refund = False  # Cash refunds also go to balance now
    
    store_logo_url = ''
    if request:
        store_profile = StoreProfile.get()
        store_logo_url = request.build_absolute_uri(store_profile.logo.url) if store_profile.logo else ''

    refund_display = _refund_receipt_display_context(transaction, item_ids=receipt_item_ids)

    context = {
        'transaction': transaction,
        'member': member,
        'refund_reason': refund_reason,
        'refund_date': timezone.localtime(timezone.now()),
        'balance_before': balance_before,
        'balance_after': balance_after,
        'show_balance_refund': show_balance_refund,
        'show_cash_refund': show_cash_refund,
        **shop_header,
        'store_logo_url': store_logo_url,
        'receipt_thank_you': kiosk_config.receipt_thank_you,
        'receipt_footer_customer_tagline': kiosk_config.receipt_footer_customer_tagline,
        'receipt_footer_merchant_note': kiosk_config.receipt_footer_merchant_note,
        'active_tax_label': _active_tax_label(),
        **_receipt_merchandise_discount_context(
            transaction,
            items=refund_display['refunded_items'],
        ),
        **refund_display,
        **print_ctx,
        'reprint_mode': True,
    }
    
    # Render the template - use request if provided for proper context
    if request:
        html = render_to_string('admin_panel/refund_receipt.html', context, request=request)
    else:
        html = render_to_string('admin_panel/refund_receipt.html', context)
    
    return html


@login_required
def process_refund(request):
    """Refund management page - accessible to all logged-in users
    
    Access control:
    - Regular members: can only search and refund their own transactions
    - Cashiers and admins: can search and refund all transactions
    """
    # Check if user is cashier or admin
    has_full_access = is_cashier_or_admin(request.user)
    
    # Get today's date range in local timezone
    today = timezone.localtime(timezone.now()).date()
    today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
    today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()))

    # Only show transactions within the admin-configured refund window
    from admin_panel.models import ReportScheduleConfig
    _refund_config = ReportScheduleConfig.get()
    refund_cutoff = timezone.now() - timedelta(days=_refund_config.refund_window_days)
    
    # Sales that still have at least one line not refunded (completed or partial).
    today_transactions = (
        Transaction.objects.filter(
            created_at__gte=refund_cutoff,
            status__in=['completed', 'partially_refunded'],
        )
        .annotate(
            has_refundable_items=Exists(
                TransactionItem.objects.filter(
                    transaction_id=OuterRef('pk'),
                    refunded_at__isnull=True,
                )
            ),
        )
        .filter(has_refundable_items=True)
        .select_related('member')
        .prefetch_related('items')
        .order_by('-created_at')[:20]
    )
    
    # If user is not cashier/admin, filter to only their own transactions
    if not has_full_access:
        try:
            member = Member.objects.get(user=request.user, is_active=True)
            today_transactions = today_transactions.filter(member=member)
        except Member.DoesNotExist:
            today_transactions = Transaction.objects.none()
        except Member.MultipleObjectsReturned:
            member = Member.objects.filter(user=request.user, is_active=True).first()
            if member:
                today_transactions = today_transactions.filter(member=member)
            else:
                today_transactions = Transaction.objects.none()
    
    # Prepare transaction data for template
    transactions_data = []
    for transaction in today_transactions:
        items = list(transaction.items.all())
        items_remaining = sum(1 for i in items if not i.refunded_at)
        transactions_data.append({
            'id': transaction.id,
            'transaction_number': transaction.transaction_number,
            'member_name': transaction.customer_display_name,
            'total_amount': transaction.total_amount,
            'payment_method': transaction.get_payment_method_display(),
            'created_at': timezone.localtime(transaction.created_at).strftime('%Y-%m-%d %H:%M:%S'),
            'items_count': len(items),
            'items_remaining': items_remaining,
            'is_partially_refunded': transaction.status == 'partially_refunded',
            'status_display': transaction.get_status_display(),
        })
    
    context = {
        'today_transactions': transactions_data,
        'refund_uses_staff_pin': has_full_access,
        **admin_role_badge_context(request),
    }
    
    return render(request, 'admin_panel/refund.html', context)


@login_required
@require_http_methods(["GET"])
def view_refund_receipt(request, transaction_id):
    """View refund receipt for a cancelled transaction
    
    Access control:
    - Regular members: can only view receipts for their own transactions
    - Cashiers and admins: can view any transaction receipt
    """
    try:
        # Get the transaction - must be refunded
        transaction = Transaction.objects.select_related('member', 'processed_by').prefetch_related(
            'items',
            'refund_reason__refund_items',
        ).get(
            id=transaction_id,
            status__in=['refunded', 'partially_refunded', 'cancelled']
        )
        
        # Check access control
        has_full_access = is_cashier_or_admin(request.user)
        if not has_full_access:
            # Get member associated with the logged-in user
            try:
                user_member = Member.objects.get(user=request.user, is_active=True)
            except Member.DoesNotExist:
                messages.error(request, 'You do not have permission to view this receipt')
                return redirect('process_refund')
            except Member.MultipleObjectsReturned:
                user_member = Member.objects.filter(user=request.user, is_active=True).first()
                if not user_member:
                    messages.error(request, 'You do not have permission to view this receipt')
                    return redirect('process_refund')
            
            # Check if the transaction belongs to the user
            if transaction.member != user_member:
                messages.error(request, 'You can only view receipts for your own transactions')
                return redirect('process_refund')
        
        member = transaction.member
        
        refund_reason = _extract_refund_reason_from_transaction_notes(transaction)
        
        # Try to get balance information from BalanceTransaction
        balance_before = None
        balance_after = None
        # Look for the most recent balance transaction related to this refund
        balance_txn = BalanceTransaction.objects.filter(
            notes__icontains=f'transaction {transaction.transaction_number}'
        ).filter(
            Q(notes__icontains='Refund') | Q(notes__icontains='Partial refund')
        ).order_by('-created_at').first()
        
        if balance_txn:
            balance_before = balance_txn.balance_before
            balance_after = balance_txn.balance_after
        elif member:
            # For cash refunds or if balance transaction not found, show current balance
            # Balance doesn't change for cash refunds, so before = after = current balance
            if transaction.payment_method == 'cash':
                balance_before = member.balance
                balance_after = member.balance
            else:
                # For other cases, try to get current balance as fallback
                balance_after = member.balance
        
        # Prepare context for template
        from admin_panel.models import KioskConfig
        kiosk_config = KioskConfig.get()
        print_ctx = _receipt_print_context()
        store_profile = StoreProfile.get()
        shop_header = _receipt_shop_header_context(kiosk_config)
        store_logo_url = request.build_absolute_uri(store_profile.logo.url) if store_profile.logo else ''
        
        # All refunds now go to balance, regardless of original payment method
        show_balance_refund = (member and balance_before is not None)
        show_cash_refund = False  # Cash refunds also go to balance now        
        if member:
            member.refresh_from_db()

        receipt_item_ids = _parse_refund_receipt_item_ids(request.GET.get('items'))
        refund_display = _refund_receipt_display_context(
            transaction,
            balance_txn,
            item_ids=receipt_item_ids,
        )

        context = {
            'transaction': transaction,
            'member': member,
            'refund_reason': refund_reason,
            'refund_date': timezone.localtime(transaction.updated_at) if transaction.updated_at else timezone.localtime(timezone.now()),  # Use when transaction was cancelled, converted to local timezone
            'balance_before': balance_before,
            'balance_after': balance_after,
            'show_balance_refund': show_balance_refund,
            'show_cash_refund': show_cash_refund,
            **shop_header,
            'store_logo_url': store_logo_url,
            'receipt_thank_you': kiosk_config.receipt_thank_you,
            'receipt_footer_customer_tagline': kiosk_config.receipt_footer_customer_tagline,
            'receipt_footer_merchant_note': kiosk_config.receipt_footer_merchant_note,
            'active_tax_label': _active_tax_label(),
            **_receipt_merchandise_discount_context(
                transaction,
                items=refund_display['refunded_items'],
            ),
            **refund_display,
            **print_ctx,
            'reprint_mode': True,
        }
        
        return render(request, 'admin_panel/refund_receipt.html', context)
        
    except Transaction.DoesNotExist:
        messages.error(request, 'Refund receipt not found')
        return redirect('process_refund')
    except Exception as e:
        messages.error(request, f'Error loading receipt: {str(e)}')
        return redirect('process_refund')


@require_http_methods(["GET"])
def view_cash_receipt(request, transaction_id):
    """View cash receipt for a completed cash transaction
    
    Access control:
    - Regular members: can only view receipts for their own transactions
    - Cashiers and admins: can view any transaction receipt
    - Supports both Django-authenticated users and session-based member logins
    """
    # Allow access if Django-authenticated OR if a member session exists
    if not request.user.is_authenticated and not request.session.get('member_id'):
        messages.warning(request, 'Please log in to view this receipt.')
        return redirect('root_login')

    try:
        transaction = _sale_receipt_lookup(request, transaction_id, 'cash')

        # Check access control
        has_full_access = request.user.is_authenticated and is_cashier_or_admin(request.user)
        if not has_full_access:
            # Try to get member via Django auth first, then fall back to session
            user_member = None
            if request.user.is_authenticated:
                try:
                    user_member = Member.objects.get(user=request.user, is_active=True)
                except Member.DoesNotExist:
                    pass
                except Member.MultipleObjectsReturned:
                    user_member = Member.objects.filter(user=request.user, is_active=True).first()

            # Fallback: look up member from session (for RFID/PIN session-only logins)
            if not user_member:
                member_id = request.session.get('member_id')
                if member_id:
                    try:
                        user_member = Member.objects.get(id=member_id, is_active=True)
                    except Member.DoesNotExist:
                        pass

            # Cash transactions may not have a member — allow access if no member on txn
            if user_member and transaction.member and transaction.member != user_member:
                messages.error(request, 'You can only view receipts for your own transactions')
                return redirect('user_choice')
        
        # Calculate change amount
        change_amount = Decimal('0.00')
        if transaction.amount_paid > transaction.total_amount:
            change_amount = transaction.amount_paid - transaction.total_amount

        # Build thermal-safe plain text specifically for cash receipts.
        raw_print_text = _build_cash_receipt_print_text(transaction, change_amount)
        
        # Get shop information and printer preferences (admin-configurable)
        from admin_panel.models import KioskConfig
        kiosk_config = KioskConfig.get()
        print_ctx = _receipt_print_context()
        store_profile = StoreProfile.get()
        shop_header = _receipt_shop_header_context(kiosk_config)
        store_logo_url = request.build_absolute_uri(store_profile.logo.url) if store_profile.logo else ''
        
        context = {
            'transaction': transaction,
            'change_amount': change_amount,
            **shop_header,
            'is_admin_or_staff': has_full_access,
            'receipt_thank_you': kiosk_config.receipt_thank_you,
            'receipt_footer_customer_tagline': kiosk_config.receipt_footer_customer_tagline,
            'receipt_footer_merchant_note': kiosk_config.receipt_footer_merchant_note,
            'store_logo_url': store_logo_url,
            'raw_print_text': raw_print_text,
            'active_tax_label': _active_tax_label(),
            **_receipt_merchandise_discount_context(transaction),
            **print_ctx,
            'reprint_mode': _is_receipt_reprint(request),
        }
        
        return render(request, 'admin_panel/cash_receipt.html', context)
        
    except Transaction.DoesNotExist:
        messages.error(request, 'Cash receipt not found')
        return redirect('transaction_history')
    except Exception as e:
        messages.error(request, f'Error loading receipt: {str(e)}')
        return redirect('transaction_history')


@require_http_methods(["GET"])
def view_debit_credit_receipt(request, transaction_id):
    """View debit receipt for a completed debit transaction
    
    Access control:
    - Regular members: can only view receipts for their own transactions
    - Cashiers and admins: can view any transaction receipt
    - Supports both Django-authenticated users and session-based member logins
    """
    # Allow access if Django-authenticated OR if a member session exists
    if not request.user.is_authenticated and not request.session.get('member_id'):
        messages.warning(request, 'Please log in to view this receipt.')
        return redirect('root_login')

    try:
        transaction = _sale_receipt_lookup(request, transaction_id, ['debit', 'credit'])
        if transaction.payment_method == 'credit':
            qs = request.GET.urlencode()
            url = reverse('view_credit_receipt', kwargs={'transaction_id': transaction_id})
            if qs:
                url = f'{url}?{qs}'
            return redirect(url)
        return _render_debit_credit_receipt(request, transaction)

    except Transaction.DoesNotExist:
        messages.error(request, 'Receipt not found')
        return redirect('transaction_history')
    except Exception as e:
        messages.error(request, f'Error loading receipt: {str(e)}')
        return redirect('transaction_history')


def _render_sale_receipt_access(request, transaction):
    """Shared access check for debit/credit sale receipts. Returns has_full_access."""
    has_full_access = request.user.is_authenticated and is_cashier_or_admin(request.user)
    if has_full_access:
        return has_full_access

    user_member = None
    if request.user.is_authenticated:
        try:
            user_member = Member.objects.get(user=request.user, is_active=True)
        except Member.DoesNotExist:
            pass
        except Member.MultipleObjectsReturned:
            user_member = Member.objects.filter(user=request.user, is_active=True).first()

    if not user_member:
        member_id = request.session.get('member_id')
        if member_id:
            try:
                user_member = Member.objects.get(id=member_id, is_active=True)
            except Member.DoesNotExist:
                pass

    if not user_member:
        messages.error(request, 'You do not have permission to view this receipt')
        return None

    if transaction.member != user_member:
        messages.error(request, 'You can only view receipts for your own transactions')
        return None

    return has_full_access


def _render_debit_credit_receipt(request, transaction):
    has_full_access = _render_sale_receipt_access(request, transaction)
    if has_full_access is None:
        return redirect('user_choice')

    from admin_panel.models import KioskConfig
    kiosk_config = KioskConfig.get()
    print_ctx = _receipt_print_context()
    store_profile = StoreProfile.get()
    shop_header = _receipt_shop_header_context(kiosk_config)
    store_logo_url = request.build_absolute_uri(store_profile.logo.url) if store_profile.logo else ''
    merchant_id = getattr(settings, 'MERCHANT_ID', None)
    terminal_id = getattr(settings, 'TERMINAL_ID', None)
    approval_code = getattr(settings, 'APPROVAL_CODE', None)

    if transaction.member:
        transaction.member.refresh_from_db()

    context = {
        'transaction': transaction,
        **shop_header,
        'merchant_id': merchant_id,
        'terminal_id': terminal_id,
        'approval_code': approval_code,
        'is_admin_or_staff': has_full_access,
        'receipt_thank_you': kiosk_config.receipt_thank_you,
        'receipt_footer_customer_tagline': kiosk_config.receipt_footer_customer_tagline,
        'receipt_footer_merchant_note': kiosk_config.receipt_footer_merchant_note,
        'store_logo_url': store_logo_url,
        'raw_print_text': _build_debit_receipt_print_text(transaction),
        'active_tax_label': _active_tax_label(),
        **_receipt_merchandise_discount_context(transaction),
        **print_ctx,
        'reprint_mode': _is_receipt_reprint(request),
    }
    return render(request, 'admin_panel/debit_credit_receipt.html', context)


def _render_credit_receipt(request, transaction):
    has_full_access = _render_sale_receipt_access(request, transaction)
    if has_full_access is None:
        return redirect('user_choice')

    from admin_panel.models import KioskConfig
    kiosk_config = KioskConfig.get()
    print_ctx = _receipt_print_context()
    store_profile = StoreProfile.get()
    shop_header = _receipt_shop_header_context(kiosk_config)
    store_logo_url = request.build_absolute_uri(store_profile.logo.url) if store_profile.logo else ''
    merchant_id = getattr(settings, 'MERCHANT_ID', None)
    terminal_id = getattr(settings, 'TERMINAL_ID', None)
    approval_code = getattr(settings, 'APPROVAL_CODE', None)

    if transaction.member:
        transaction.member.refresh_from_db()

    context = {
        'transaction': transaction,
        **shop_header,
        'merchant_id': merchant_id,
        'terminal_id': terminal_id,
        'approval_code': approval_code,
        'is_admin_or_staff': has_full_access,
        'receipt_thank_you': kiosk_config.receipt_thank_you,
        'receipt_footer_customer_tagline': kiosk_config.receipt_footer_customer_tagline,
        'receipt_footer_merchant_note': kiosk_config.receipt_footer_merchant_note,
        'store_logo_url': store_logo_url,
        'raw_print_text': _build_credit_receipt_print_text(transaction),
        'active_tax_label': _active_tax_label(),
        **_receipt_merchandise_discount_context(transaction),
        **print_ctx,
        'reprint_mode': _is_receipt_reprint(request),
    }
    return render(request, 'admin_panel/credit_receipt.html', context)


def _build_credit_payment_receipt_print_text(
    payment, settled_sales, settled_items=None, payment_lines=None
):
    from decimal import Decimal as _D
    _zero = _D('0.00')
    lines = [
        "CREDIT PAYMENT RECEIPT",
        f"Settlement: {payment.settlement_number}",
        f"Date: {timezone.localtime(payment.created_at).strftime('%Y-%m-%d %I:%M %p')}",
        f"Member: {payment.member.full_name}",
        f"Payment: {payment.get_payment_method_display()}",
        f"Amount Paid: P{payment.amount_paid}",
        "",
        "SETTLED ITEMS",
    ]
    _vat_total = _zero
    _vatable_total = _zero
    if payment_lines:
        for line in payment_lines:
            item = line.item
            lines.append(
                f"- {item.transaction.transaction_number} | {item.product_name} "
                f"x{item.quantity}  P{line.amount_applied}"
            )
            _vat_total += item.vat_amount or _zero
            _vatable_total += item.vatable_sale or _zero
    elif settled_items:
        for item in settled_items:
            lines.append(
                f"- {item.transaction.transaction_number} | {item.product_name} "
                f"x{item.quantity}  P{item.credit_line_amount}"
            )
            _vat_total += item.vat_amount or _zero
            _vatable_total += item.vatable_sale or _zero
    else:
        for sale in settled_sales:
            lines.append(
                f"- {sale.transaction_number}  P{sale.total_amount}  "
                f"({timezone.localtime(sale.created_at).strftime('%Y-%m-%d')})"
            )
            _vat_total += sale.vat_amount or _zero
            _vatable_total += sale.vatable_sale or _zero
    if _vat_total > _zero:
        tax_label = _active_tax_label()
        lines.extend([
            "",
            "TAX BREAKDOWN (included in settled amount):",
            f"Vatable Sale: P{_vatable_total:.2f}",
            f"{tax_label}: P{_vat_total:.2f}",
        ])
    if payment.balance_before is not None:
        lines.extend([
            "",
            f"Balance Before: P{payment.balance_before}",
            f"Balance After:  P{payment.balance_after}",
        ])
    if payment.performed_by_display:
        lines.extend(["", f"Processed By: {payment.performed_by_display}"])
    lines.extend(["", "Thank you!", ""])
    return "\r\n".join(lines)


@login_required
@require_http_methods(["GET"])
def view_credit_payment_receipt(request, payment_id):
    """Printable receipt for a credit settlement (Pay Credit)."""
    if not (is_cashier_or_admin(request.user) or is_staff_role(request.user)):
        messages.warning(request, 'You do not have permission to view this receipt.')
        return redirect('member_management')

    payment = get_object_or_404(
        CreditPayment.objects.select_related('member', 'performed_by'),
        pk=payment_id,
    )
    payment_lines = list(
        payment.payment_lines.select_related(
            'item', 'item__transaction'
        ).order_by(
            'item__transaction__created_at', 'item__transaction_id', 'item_id'
        )
    )
    settled_items = [line.item for line in payment_lines] or list(
        payment.settled_items.select_related('transaction').order_by(
            'transaction__created_at', 'transaction_id', 'id'
        )
    )
    settled_sales = list(
        payment.settled_sales.prefetch_related('items').order_by('created_at', 'id')
    )

    kiosk_config = KioskConfig.get()
    store_profile = StoreProfile.get()
    print_ctx = _receipt_print_context()
    store_logo_url = (
        request.build_absolute_uri(store_profile.logo.url) if store_profile.logo else ''
    )

    # Aggregate VAT from settled items for the transparent tax breakdown.
    from decimal import Decimal as _D
    _zero = _D('0.00')
    _settled_vat = _zero
    _settled_vatable = _zero
    if payment_lines:
        for _pl in payment_lines:
            _settled_vat += _pl.item.vat_amount or _zero
            _settled_vatable += _pl.item.vatable_sale or _zero
    elif settled_items:
        for _si in settled_items:
            _settled_vat += _si.vat_amount or _zero
            _settled_vatable += _si.vatable_sale or _zero
    elif settled_sales:
        for _ss in settled_sales:
            _settled_vat += _ss.vat_amount or _zero
            _settled_vatable += _ss.vatable_sale or _zero

    context = {
        'payment': payment,
        'payment_lines': payment_lines,
        'settled_sales': settled_sales,
        'settled_items': settled_items,
        'member': payment.member,
        **_receipt_shop_header_context(kiosk_config),
        'receipt_thank_you': kiosk_config.receipt_thank_you,
        'receipt_footer_customer_tagline': kiosk_config.receipt_footer_customer_tagline,
        'store_logo_url': store_logo_url,
        'is_admin_or_staff': True,
        'active_tax_label': _active_tax_label(),
        'settled_vat_amount': _settled_vat if _settled_vat > _zero else None,
        'settled_vatable_sale': _settled_vatable,
        'raw_print_text': _build_credit_payment_receipt_print_text(
            payment, settled_sales, settled_items, payment_lines
        ),
        **print_ctx,
    }
    return render(request, 'admin_panel/credit_payment_receipt.html', context)


def view_credit_receipt(request, transaction_id):
    """View credit receipt for a completed credit transaction."""
    if not request.user.is_authenticated and not request.session.get('member_id'):
        messages.warning(request, 'Please log in to view this receipt.')
        return redirect('root_login')

    try:
        transaction = _sale_receipt_lookup(request, transaction_id, 'credit')
        return _render_credit_receipt(request, transaction)
    except Transaction.DoesNotExist:
        messages.error(request, 'Credit receipt not found')
        return redirect('transaction_history')
    except Exception as e:
        messages.error(request, f'Error loading receipt: {str(e)}')
        return redirect('transaction_history')


@login_required
@require_http_methods(["GET"])
def api_search_transactions_for_refund(request):
    """Search transactions by transaction number for refund processing
    
    Access control:
    - Regular members: can only search their own transactions
    - Cashiers and admins: can search all transactions
    """
    query = request.GET.get('q', '').strip()
    
    if not query or len(query) < 2:
        return JsonResponse({'success': True, 'transactions': []})
    
    try:
        # Check if user is cashier or admin
        has_full_access = is_cashier_or_admin(request.user)
        
        # Base query for completed transactions within the admin-configured refund window
        from admin_panel.models import ReportScheduleConfig
        _refund_config = ReportScheduleConfig.get()
        cutoff_time = timezone.now() - timedelta(days=_refund_config.refund_window_days)
        transactions = Transaction.objects.filter(
            transaction_number__icontains=query,
            status__in=['completed', 'partially_refunded'],
            created_at__gte=cutoff_time
        ).select_related('member').prefetch_related('items')
        
        # If user is not cashier/admin, filter to only their own transactions
        if not has_full_access:
            # Get member associated with the logged-in user
            try:
                member = Member.objects.get(user=request.user, is_active=True)
                transactions = transactions.filter(member=member)
            except Member.DoesNotExist:
                # User doesn't have a member account, return empty results
                return JsonResponse({'success': True, 'transactions': []})
            except Member.MultipleObjectsReturned:
                # Multiple members found, use the first one
                member = Member.objects.filter(user=request.user, is_active=True).first()
                if member:
                    transactions = transactions.filter(member=member)
                else:
                    return JsonResponse({'success': True, 'transactions': []})
        
        # Order and limit results
        transactions = transactions.order_by('-created_at', '-id')[:20]
        
        results = []
        for transaction in transactions:
            if not transaction.items.filter(refunded_at__isnull=True).exists():
                continue
            # Get transaction items (include refund state for the process-refund UI)
            items = []
            for item in transaction.items.all():
                items.append({
                    'id': item.id,
                    'product_name': item.product_name,
                    'quantity': item.quantity,
                    'total_price': str(item.total_price),
                    'refunded_at': item.refunded_at.isoformat() if item.refunded_at else None,
                })
            
            results.append({
                'id': transaction.id,
                'transaction_number': transaction.transaction_number,
                'member_name': transaction.customer_display_name,
                'member_id': transaction.member.id if transaction.member else None,
                'total_amount': str(transaction.total_amount),
                'payment_method': transaction.get_payment_method_display(),
                'created_at': timezone.localtime(transaction.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                'items_count': transaction.items.count(),
                'items': items,
            })
        
        return JsonResponse({'success': True, 'transactions': results})
    except Exception as e:
        return JsonResponse({'success': False, 'error': 'Server error occurred'})


@login_required
@require_http_methods(["GET"])
def api_search_transactions(request):
    """Search transactions with filters for admin management"""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        transaction_number = request.GET.get('transaction_number', '').strip()
        status = request.GET.get('status', '').strip()
        payment_method = request.GET.get('payment_method', '').strip()
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        
        # Build query
        transactions_qs = Transaction.objects.select_related('member').prefetch_related('items').all()
        
        if transaction_number:
            transactions_qs = transactions_qs.filter(transaction_number__icontains=transaction_number)
        if status:
            transactions_qs = transactions_qs.filter(status=status)
        if payment_method:
            transactions_qs = transactions_qs.filter(payment_method=payment_method)
        if date_from:
            try:
                from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
                transactions_qs = transactions_qs.filter(created_at__date__gte=from_date)
            except ValueError:
                pass
        if date_to:
            try:
                to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
                transactions_qs = transactions_qs.filter(created_at__date__lte=to_date)
            except ValueError:
                pass
        
        # Order and limit (pk tie-breaker avoids unstable ordering / duplicate-looking pages)
        transactions_qs = transactions_qs.order_by('-created_at', '-id')[:50]
        
        results = []
        for transaction in transactions_qs:
            local_created_at = timezone.localtime(transaction.created_at)
            voidable_items_count = transaction.items.filter(refunded_at__isnull=True).count()
            results.append({
                'id': transaction.id,
                'transaction_number': transaction.transaction_number,
                'member_name': transaction.customer_display_name,
                'member_rfid': mask_rfid(transaction.member.rfid_card_number) if transaction.member and transaction.member.rfid_card_number else None,
                'date': local_created_at.strftime('%Y-%m-%d'),
                'time': local_created_at.strftime('%H:%M:%S'),
                'amount': str(transaction.total_amount),
                'payment_method': transaction.payment_method,
                'payment_method_display': transaction.get_payment_method_display(),
                'status': transaction.status,
                'status_display': transaction.get_status_display(),
                'amount_paid': str(transaction.amount_paid),
                'amount_from_balance': str(transaction.amount_from_balance),
                'notes': transaction.notes or '',
                'can_void': _transaction_can_void_item(transaction),
                'voidable_items_count': voidable_items_count,
            })
        
        return JsonResponse({'success': True, 'transactions': results})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'})


@login_required
@require_http_methods(["GET"])
def api_get_transaction(request, transaction_id):
    """Get transaction details by ID"""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        transaction = Transaction.objects.select_related('member').prefetch_related('items').get(id=transaction_id)
        
        items = []
        for item in transaction.items.all():
            items.append({
                'id': item.id,
                'product_name': item.product_name,
                'product_barcode': item.product_barcode,
                'quantity': item.quantity,
                'unit_price': str(item.unit_price),
                'total_price': str(item.total_price),
                'refunded_at': item.refunded_at.isoformat() if item.refunded_at else None,
            })
        
        return JsonResponse({
            'success': True,
            'transaction': {
                'id': transaction.id,
                'transaction_number': transaction.transaction_number,
                'member_id': transaction.member.id if transaction.member else None,
                'member_name': transaction.customer_display_name,
                'member_rfid': mask_rfid(transaction.member.rfid_card_number) if transaction.member and transaction.member.rfid_card_number else None,
                'subtotal': str(transaction.subtotal),
                'vatable_sale': str(transaction.vatable_sale),
                'vat_amount': str(transaction.vat_amount),
                'total_amount': str(transaction.total_amount),
                'payment_method': transaction.payment_method,
                'payment_method_display': transaction.get_payment_method_display(),
                'amount_paid': str(transaction.amount_paid),
                'amount_from_balance': str(transaction.amount_from_balance),
                'status': transaction.status,
                'status_display': transaction.get_status_display(),
                'notes': transaction.notes or '',
                'created_at': timezone.localtime(transaction.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                'can_void': _transaction_can_void_item(transaction),
                'items': items,
            }
        })
    except Transaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'})


@login_required
@require_http_methods(["POST"])
def api_void_transaction_item(request):
    """Mark a specific transaction item as void by creating a refund request."""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    transaction_id = data.get('transaction_id')
    item_id = data.get('item_id')
    reason_details = (data.get('reason') or '').strip()

    if not transaction_id:
        return JsonResponse({'success': False, 'error': 'Transaction ID is required'}, status=400)
    if not item_id:
        return JsonResponse({'success': False, 'error': 'Please select an item to void.'}, status=400)

    try:
        transaction_id = int(transaction_id)
        item_id = int(item_id)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid transaction or item.'}, status=400)

    try:
        with db_transaction.atomic():
            txn = Transaction.objects.select_for_update().get(id=transaction_id)
            item = TransactionItem.objects.select_for_update().get(id=item_id, transaction=txn)

            if txn.status not in VOID_ITEM_ALLOWED_STATUSES:
                return JsonResponse({
                    'success': False,
                    'error': f'Cannot void item for transaction in {txn.get_status_display()} status.',
                }, status=400)

            if item.refunded_at:
                return JsonResponse({
                    'success': False,
                    'error': 'This item has already been refunded.',
                }, status=400)

            refund_reason_obj, _ = RefundReason.objects.get_or_create(
                transaction=txn,
                defaults={
                    'reason_type': 'overcharged',
                    'details': '',
                }
            )
            refund_reason_obj.refund_items.set([item])
            detail_note = reason_details or 'Void item requested via Manage Transactions.'
            refund_reason_obj.details = (
                f'Void item request by {request.user.username}: {item.product_name} '
                f'(x{item.quantity}, {item.product_barcode}). {detail_note}'
            )
            refund_reason_obj.save(update_fields=['details', 'updated_at'])

            txn.status = 'refund_requested'
            void_note = (
                f'Void item requested by {request.user.username}: {item.product_name} '
                f'(x{item.quantity}). Awaiting refund approval.'
            )
            txn.notes = f'{txn.notes} | {void_note}' if txn.notes else void_note
            txn.save(update_fields=['status', 'notes', 'updated_at'])

    except Transaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)
    except TransactionItem.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Selected item was not found in this transaction.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'}, status=500)

    return JsonResponse({
        'success': True,
        'message': f'Void request submitted for "{item.product_name}". Proceed to refund approval flow.',
        'transaction': {
            'id': txn.id,
            'transaction_number': txn.transaction_number,
            'status': txn.status,
            'status_display': txn.get_status_display(),
        }
    })


# Statuses where a sale typically still holds deducted stock / unpaid debit liability.
_TXN_DELETE_REVERSAL_STATUSES = frozenset({
    'completed',
    'partially_refunded',
    'refund_requested',
    'return_window',
    'return_expired',
})


@login_required
@require_http_methods(["POST"])
def api_delete_transaction(request):
    """Permanently delete a sale transaction. Admin role only; written to the audit trail.

    For active sale statuses, restores stock on non-refunded lines and credits back
    remaining debit amounts. Refuses delete when the sale is tied to a credit settlement.
    """
    if not is_admin_user(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    transaction_id = data.get('id') or data.get('transaction_id')
    if not transaction_id:
        return JsonResponse({'success': False, 'error': 'Transaction ID is required'}, status=400)

    try:
        transaction_id = int(transaction_id)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid transaction ID'}, status=400)

    try:
        with db_transaction.atomic():
            txn = (
                Transaction.objects.select_for_update()
                .select_related('member')
                .prefetch_related('items__product')
                .get(id=transaction_id)
            )

            if CreditPaymentLine.objects.filter(item__transaction_id=txn.pk).exists():
                return JsonResponse({
                    'success': False,
                    'error': (
                        'This transaction has credit payment settlement lines and cannot be deleted. '
                        'Reverse or adjust the credit payment first.'
                    ),
                }, status=400)

            txn_number = txn.transaction_number or str(txn.pk)
            txn_amount = str(txn.total_amount)
            txn_status = txn.status
            txn_payment = txn.payment_method
            member_name = txn.customer_display_name
            txn_pk = txn.pk

            restocked_lines = 0
            balance_restored = Decimal('0.00')

            if txn.status in _TXN_DELETE_REVERSAL_STATUSES:
                active_items = [
                    item for item in txn.items.all()
                    if item.refunded_at is None
                ]

                for item in active_items:
                    if not item.product_id:
                        continue
                    product = Product.objects.select_for_update().get(pk=item.product_id)
                    before_snap = capture_stock_snapshot(product)
                    product.stock_quantity += item.quantity
                    product.save(update_fields=['stock_quantity', 'updated_at'])
                    record_stock_history(
                        product,
                        ProductStockHistory.CHANGE_REFUND,
                        before_snap,
                        note=f'Delete transaction {txn_number} — restock',
                        user=request.user,
                    )
                    restocked_lines += 1

                if txn.payment_method == 'debit' and txn.member_id and active_items:
                    restore_amount = sum(
                        (Decimal(str(item.total_price or 0)) for item in active_items),
                        Decimal('0.00'),
                    ).quantize(Decimal('0.01'))
                    if restore_amount > 0:
                        member = Member.objects.select_for_update().get(pk=txn.member_id)
                        balance_before = member.balance
                        member.add_balance(restore_amount)
                        member.refresh_from_db(fields=['balance'])
                        BalanceTransaction.objects.create(
                            member=member,
                            transaction_type='deposit',
                            amount=restore_amount,
                            balance_before=balance_before,
                            balance_after=member.balance,
                            notes=f'Delete transaction {txn_number} — restored debit amount',
                        )
                        balance_restored = restore_amount

            txn.delete()
    except Transaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)
    except Exception:
        return JsonResponse({
            'success': False,
            'error': 'Could not delete this transaction. It may still be linked to other records.',
        }, status=400)

    try:
        from admin_panel.audit import mark_audit_recorded, record_audit
        record_audit(
            WebsiteAuditLog.Action.TRANSACTION,
            actor=request.user,
            description=(
                f'Deleted transaction {txn_number} '
                f'({member_name}, ₱{txn_amount}, {txn_status})'
            ),
            request=request,
            object_type='Transaction',
            object_id=txn_pk,
            metadata={
                'transaction_number': txn_number,
                'member': member_name,
                'amount': txn_amount,
                'status': txn_status,
                'payment_method': txn_payment,
                'restocked_lines': restocked_lines,
                'balance_restored': str(balance_restored),
            },
        )
        mark_audit_recorded(request)
    except Exception:
        pass

    return JsonResponse({
        'success': True,
        'message': f'Transaction {txn_number} deleted successfully.',
    })


@login_required
@require_http_methods(["POST"])
def api_update_transaction(request):
    """Update a transaction without using the Django admin UI"""
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)
    
    transaction_id = data.get('transaction_id')
    if not transaction_id:
        return JsonResponse({'success': False, 'error': 'Transaction ID is required'}, status=400)
    
    try:
        transaction = Transaction.objects.get(id=transaction_id)
    except Transaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)
    
    # Update fields
    new_status = data.get('status', '').strip() if 'status' in data else None

    # ── Partial-refund approval ──────────────────────────────────────────
    # When an admin approves a mobile refund request (refund_requested → refunded),
    # credit only the selected items' totals (not necessarily the full transaction amount).
    if new_status == 'refunded' and transaction.status == 'refund_requested':
        from transactions.models import RefundReason, RefundReturnWindow
        from datetime import timedelta
        from admin_panel.models import ReportScheduleConfig

        _rw_config = ReportScheduleConfig.get()
        _return_days = _rw_config.return_window_days

        # ── STEP 1: Set to "return_window" — member must return item within configured days ──
        # Create the return window tracker
        return_deadline = timezone.now() + timedelta(days=_return_days)
        try:
            rw = transaction.return_window  # already exists — update it
            rw.is_returned = False
            rw.return_confirmed_at = None
            rw.return_deadline = return_deadline
            rw.approved_by = request.user
            rw.save()
        except RefundReturnWindow.DoesNotExist:
            rw = RefundReturnWindow.objects.create(
                transaction=transaction,
                return_deadline=return_deadline,
                approved_by=request.user,
            )

        transaction.status = 'return_window'
        transaction.notes = (
            f"Refund approved by {request.user.username}. "
            f"Member must return item(s) by {return_deadline.strftime('%b %d, %Y')}."
        )
        transaction.save()

        return JsonResponse({
            'success': True,
            'message': (
                f'Refund approved. Member has {_return_days} day(s) to return '
                f'the item(s) before {return_deadline.strftime("%b %d, %Y")}. '
                f'Refund money will only be credited after item return is confirmed.'
            ),
            'new_status': 'return_window',
            'transaction': {
                'id': transaction.id,
                'transaction_number': transaction.transaction_number,
                'status': 'return_window',
                'status_display': transaction.get_status_display(),
                'return_deadline': return_deadline.strftime('%b %d, %Y %H:%M'),
            }
        })

    if new_status == 'refunded' and transaction.status == 'return_window':
        # ── Prevent direct status jump — use api_confirm_return instead ──
        return JsonResponse({
            'success': False,
            'error': 'Use the "Confirm Return" action to complete this refund.'
        }, status=400)

    if new_status == 'completed' and transaction.status == 'refund_requested':
        # Reject: restore to completed
        transaction.status = 'completed'
        transaction.notes = (transaction.notes + ' | Refund request rejected.' if transaction.notes
                             else 'Refund request rejected.')
        transaction.save()
        return JsonResponse({
            'success': True,
            'message': f'Refund request for {transaction.transaction_number} rejected.',
            'transaction': {
                'id': transaction.id,
                'transaction_number': transaction.transaction_number,
                'status': 'completed',
                'status_display': transaction.get_status_display(),
            }
        })
    # Legacy direct-refund path (refund_requested → refunded) — kept for backward compat but
    # the dashboard now routes through return_window first.
    if new_status == 'refunded' and transaction.status not in ('refund_requested', 'return_window'):
        pass  # fall through to generic status update below

    # ── ORIGINAL direct-refund block removed; preserved as dead branch for safety ──
    if False and new_status == 'refunded' and transaction.status == 'refund_requested':
        from transactions.models import RefundReason
        member = transaction.member

        # Determine which items to refund and the amount to credit back
        try:
            refund_reason_obj = transaction.refund_reason  # OneToOne
            selected_items = list(refund_reason_obj.refund_items.all())
        except RefundReason.DoesNotExist:
            selected_items = []

        all_items = list(transaction.items.all())


        # Items whose stock will be restored and whose value is credited
        items_to_refund = selected_items if selected_items else all_items

        refund_amount = sum(
            (item.total_price for item in items_to_refund),
            Decimal('0.00')
        )

        # Credit member balance
        balance_after = None
        if member and refund_amount > 0:
            balance_before = member.balance
            member.add_balance(refund_amount)
            balance_after = member.balance

            reason_label = ''
            try:
                reason_label = refund_reason_obj.get_reason_type_display()
            except Exception:
                pass

            item_names = ', '.join(i.product_name for i in items_to_refund)
            BalanceTransaction.objects.create(
                member=member,
                transaction_type='deposit',
                amount=refund_amount,
                balance_before=balance_before,
                balance_after=balance_after,
                notes=(
                    f"Partial refund for transaction {transaction.transaction_number} "
                    f"({len(items_to_refund)} of {len(all_items)} item(s): {item_names})"
                    + (f". Reason: {reason_label}" if reason_label else '')
                ) if selected_items else (
                    f"Refund for transaction {transaction.transaction_number}"
                    + (f". Reason: {reason_label}" if reason_label else '')
                )
            )

        # Restock logic: admin can choose whether to restore stock.
        # Defaults: DON'T restock if reason is 'expired'; DO restock for all other reasons.
        try:
            reason_type_for_default = refund_reason_obj.reason_type
        except Exception:
            reason_type_for_default = 'other'
        default_restock = reason_type_for_default != 'expired'
        restock = data.get('restock', default_restock)

        if restock:
            for item in items_to_refund:
                if item.product:
                    _before_snap = capture_stock_snapshot(item.product)
                    item.product.stock_quantity += item.quantity
                    item.product.save()
                    record_stock_history(
                        item.product,
                        ProductStockHistory.CHANGE_REFUND,
                        _before_snap,
                        note=f'Refund restock — transaction {transaction.transaction_number}',
                        user=request.user,
                    )

        transaction.status = 'refunded'
        restock_note = '' if restock else ' (stock not restored)'
        transaction.notes = (
            f"Partially refunded ({len(items_to_refund)} item(s)){restock_note}." if selected_items
            else f"Refunded{restock_note}."
        )
        transaction.save()

        # Send approval email to member
        if member and getattr(member, 'email', None):
            try:
                reason_display = ''
                try:
                    reason_display = refund_reason_obj.get_reason_type_display()
                except Exception:
                    pass
                send_refund_approval_email(
                    transaction=transaction,
                    member=member,
                    refund_amount=refund_amount,
                    balance_after=balance_after,
                    reason_display=reason_display or None,
                    is_partial=bool(selected_items),
                    items_refunded=items_to_refund,
                )
            except Exception:
                pass  # Email failure must never block the API response

        return JsonResponse({
            'success': True,
            'message': f'Refund of ₱{refund_amount} approved for transaction {transaction.transaction_number}.'
                       + ('' if restock else ' Inventory was NOT restocked.'),
            'transaction': {
                'id': transaction.id,
                'transaction_number': transaction.transaction_number,
                'status': transaction.status,
                'status_display': transaction.get_status_display(),
                'payment_method': transaction.payment_method,
                'payment_method_display': transaction.get_payment_method_display(),
                'total_amount': str(transaction.total_amount),
                'refund_amount': str(refund_amount),
            }
        })
    # ────────────────────────────────────────────────────────────────────

    if new_status and new_status in dict(Transaction.STATUS_CHOICES):
        transaction.status = new_status

    if 'payment_method' in data:
        payment_method = data.get('payment_method', '').strip()
        if payment_method in dict(Transaction.PAYMENT_METHODS):
            transaction.payment_method = payment_method
    
    if 'amount_paid' in data:
        try:
            amount_paid = Decimal(str(data.get('amount_paid', '0')))
            if amount_paid >= 0:
                transaction.amount_paid = amount_paid
        except (InvalidOperation, TypeError, ValueError):
            pass
    
    if 'amount_from_balance' in data:
        try:
            amount_from_balance = Decimal(str(data.get('amount_from_balance', '0')))
            if amount_from_balance >= 0:
                transaction.amount_from_balance = amount_from_balance
        except (InvalidOperation, TypeError, ValueError):
            pass
    
    if 'notes' in data:
        transaction.notes = (data.get('notes') or '').strip()
    
    transaction.save()
    
    return JsonResponse({
        'success': True,
        'message': 'Transaction updated successfully',
        'transaction': {
            'id': transaction.id,
            'transaction_number': transaction.transaction_number,
            'status': transaction.status,
            'status_display': transaction.get_status_display(),
            'payment_method': transaction.payment_method,
            'payment_method_display': transaction.get_payment_method_display(),
            'total_amount': str(transaction.total_amount),
        }
    })


@login_required
@require_http_methods(["POST"])
def api_confirm_return(request):
    """Admin confirms that the member has physically returned the item(s).
    This triggers the actual refund money processing.
    Only valid when transaction is in 'return_window' status.
    """
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    transaction_id = data.get('transaction_id')
    if not transaction_id:
        return JsonResponse({'success': False, 'error': 'Transaction ID is required'}, status=400)

    try:
        txn = Transaction.objects.get(id=transaction_id)
    except Transaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)

    if txn.status != 'return_window':
        return JsonResponse({
            'success': False,
            'error': f'Transaction is not in Return Window status (current: {txn.get_status_display()}).'
        }, status=400)

    from transactions.models import RefundReason, RefundReturnWindow

    # Check if window has expired
    try:
        rw = txn.return_window
        if rw.is_expired():
            txn.status = 'return_expired'
            txn.save()
            return JsonResponse({
                'success': False,
                'error': 'The 3-day return window has expired. The refund cannot be processed.'
            }, status=400)
    except RefundReturnWindow.DoesNotExist:
        pass

    member = txn.member

    # Determine items to refund
    try:
        refund_reason_obj = txn.refund_reason
        selected_items = list(refund_reason_obj.refund_items.all())
    except RefundReason.DoesNotExist:
        refund_reason_obj = None
        selected_items = []

    all_items = list(txn.items.all())
    items_to_refund = selected_items if selected_items else all_items

    refund_amount = sum((item.total_price for item in items_to_refund), Decimal('0.00'))

    # Credit member balance
    balance_after = None
    if member and refund_amount > 0:
        balance_before = member.balance
        member.add_balance(refund_amount)
        balance_after = member.balance

        reason_label = ''
        if refund_reason_obj:
            try:
                reason_label = refund_reason_obj.get_reason_type_display()
            except Exception:
                pass

        item_names = ', '.join(i.product_name for i in items_to_refund)
        BalanceTransaction.objects.create(
            member=member,
            transaction_type='deposit',
            amount=refund_amount,
            balance_before=balance_before,
            balance_after=balance_after,
            notes=(
                f"Refund for transaction {txn.transaction_number} "
                f"({len(items_to_refund)} item(s): {item_names})"
                + (f". Reason: {reason_label}" if reason_label else '')
            )
        )

    # Restock logic
    try:
        reason_type_for_default = refund_reason_obj.reason_type if refund_reason_obj else 'other'
    except Exception:
        reason_type_for_default = 'other'
    default_restock = reason_type_for_default != 'expired'
    restock = data.get('restock', default_restock)

    if restock:
        for item in items_to_refund:
            if item.product:
                _before_snap = capture_stock_snapshot(item.product)
                item.product.stock_quantity += item.quantity
                item.product.save()
                record_stock_history(
                    item.product,
                    ProductStockHistory.CHANGE_REFUND,
                    _before_snap,
                    note=f'Refund restock — transaction {txn.transaction_number}',
                    user=request.user,
                )

    # Mark return window as confirmed
    try:
        rw = txn.return_window
        rw.is_returned = True
        rw.return_confirmed_at = timezone.now()
        rw.save()
    except RefundReturnWindow.DoesNotExist:
        pass

    restock_note = '' if restock else ' (stock not restored)'
    txn.status = 'refunded'
    txn.notes = (
        f"Refunded after item return confirmed by {request.user.username}{restock_note}."
        + (f" ({len(items_to_refund)} of {len(all_items)} item(s))" if selected_items else '')
    )
    txn.save()

    # Send refund confirmation email
    if member and getattr(member, 'email', None):
        try:
            reason_display = reason_label if refund_reason_obj else ''
            send_refund_approval_email(
                transaction=txn,
                member=member,
                refund_amount=refund_amount,
                balance_after=balance_after,
                reason_display=reason_display or None,
                is_partial=bool(selected_items),
                items_refunded=items_to_refund,
            )
        except Exception:
            pass

    return JsonResponse({
        'success': True,
        'message': (
            f'Item return confirmed. Refund of ₱{refund_amount} credited to '
            f'{member.full_name if member else "member"} account.'
            + ('' if restock else ' Inventory was NOT restocked.')
        ),
        'transaction': {
            'id': txn.id,
            'transaction_number': txn.transaction_number,
            'status': 'refunded',
            'status_display': txn.get_status_display(),
            'refund_amount': str(refund_amount),
        }
    })


@login_required
@require_http_methods(["POST"])
def api_expire_return_window(request):
    """Admin manually marks a return window as expired (member did not return item in 3 days).
    The refund is NOT processed.
    """
    if not is_cashier_or_admin(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON payload'}, status=400)

    transaction_id = data.get('transaction_id')
    if not transaction_id:
        return JsonResponse({'success': False, 'error': 'Transaction ID is required'}, status=400)

    try:
        txn = Transaction.objects.get(id=transaction_id)
    except Transaction.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)

    if txn.status != 'return_window':
        return JsonResponse({
            'success': False,
            'error': f'Transaction is not in Return Window status (current: {txn.get_status_display()}).'
        }, status=400)

    txn.status = 'return_expired'
    txn.notes = (
        (txn.notes + ' | ' if txn.notes else '') +
        f'Return window expired — no refund processed. Marked by {request.user.username}.'
    )
    txn.save()

    return JsonResponse({
        'success': True,
        'message': f'Return window closed for {txn.transaction_number}. No refund will be processed.',
        'transaction': {
            'id': txn.id,
            'transaction_number': txn.transaction_number,
            'status': 'return_expired',
            'status_display': txn.get_status_display(),
        }
    })


@login_required
@require_http_methods(["POST"])
def api_process_refund(request):
    """Process a refund for a transaction
    
    Access control:
    - Regular members: can only refund their own transactions (member PIN)
    - Admin, cashier, staff: can refund any transaction (staff PIN)
    """
    try:
        data = json.loads(request.body)
        transaction_id = data.get('transaction_id')
        refund_reason = data.get('reason', '').strip()
        pin = data.get('pin', '').strip()

        if not transaction_id:
            return JsonResponse({'success': False, 'error': 'Transaction ID is required'})

        if not pin:
            return JsonResponse({'success': False, 'error': 'PIN is required to process a refund', 'pin_required': True}, status=400)

        try:
            transaction = Transaction.objects.select_related('member').prefetch_related('items').get(
                id=transaction_id,
                status__in=['completed', 'partially_refunded'],
            )
        except Transaction.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Transaction not found or not eligible for refund (must be completed or partially refunded)',
            })

        has_full_access = is_cashier_or_admin(request.user)
        txn_member = transaction.member

        # --- PIN verification ---
        if has_full_access:
            session_member = _get_active_member_for_refill_user(request.user)
            if not _resolve_refund_pin_authorizer(pin, session_member):
                return JsonResponse({
                    'success': False,
                    'error': (
                        'Incorrect PIN. Enter the 4-digit PIN of an active '
                        'admin, cashier, or staff member with a PIN set.'
                    ),
                    'pin_wrong': True,
                }, status=403)
        elif txn_member is not None:
            if txn_member.is_pin_locked:
                return JsonResponse({'success': False, 'error': 'Your PIN is locked due to too many failed attempts. Please contact an administrator.', 'pin_locked': True}, status=403)

            if not txn_member.pin_hash:
                return JsonResponse({'success': False, 'error': 'No PIN set on your account. Please set a PIN before processing refunds.', 'pin_not_set': True}, status=403)

            if not txn_member.check_pin(pin):
                txn_member.pin_attempts = (txn_member.pin_attempts or 0) + 1
                if txn_member.pin_attempts >= 3:
                    txn_member.is_pin_locked = True
                    txn_member.save()
                    return JsonResponse({'success': False, 'error': 'Incorrect PIN. Your PIN has been locked due to too many failed attempts.', 'pin_locked': True}, status=403)
                txn_member.save()
                remaining = 3 - txn_member.pin_attempts
                return JsonResponse({'success': False, 'error': f'Incorrect PIN. {remaining} attempt(s) remaining before lock-out.', 'pin_wrong': True}, status=403)

            txn_member.pin_attempts = 0
            txn_member.save()
        else:
            return JsonResponse({'success': False, 'error': 'You do not have permission to refund guest transactions.'}, status=403)
        # --- End PIN verification ---
        
        # Enforce the admin-configured refund window
        from admin_panel.models import ReportScheduleConfig
        _refund_config = ReportScheduleConfig.get()
        cutoff_time = timezone.now() - timedelta(days=_refund_config.refund_window_days)
        if transaction.created_at < cutoff_time:
            window_label = f"{_refund_config.refund_window_days} day{'s' if _refund_config.refund_window_days != 1 else ''}"
            return JsonResponse({
                'success': False,
                'error': f'Refund requests are only allowed within {window_label} of purchase. This transaction is no longer eligible for a refund.'
            }, status=400)
        
        # Check access control: regular members can only refund their own transactions
        if not has_full_access:
            # Get member associated with the logged-in user
            try:
                user_member = Member.objects.get(user=request.user, is_active=True)
            except Member.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'You do not have permission to process refunds'}, status=403)
            except Member.MultipleObjectsReturned:
                user_member = Member.objects.filter(user=request.user, is_active=True).first()
                if not user_member:
                    return JsonResponse({'success': False, 'error': 'You do not have permission to process refunds'}, status=403)
            
            # Check if the transaction belongs to the user
            if transaction.member != user_member:
                return JsonResponse({'success': False, 'error': 'You can only refund your own transactions'}, status=403)
        
        member = transaction.member

        # Determine which items to refund (partial or full)
        refund_item_ids_raw = data.get('refund_item_ids')
        # restock_item_ids: list of item IDs that should have stock restored.
        # None means "not provided" → restock all refunded items (backward compat).
        restock_item_ids = data.get('restock_item_ids', None)
        all_items = list(transaction.items.order_by('id'))
        already_refunded = [i for i in all_items if i.refunded_at]
        if already_refunded and refund_item_ids_raw is not None:
            refund_id_set_pre = {int(x) for x in refund_item_ids_raw}
            overlap = [i for i in already_refunded if i.id in refund_id_set_pre]
            if overlap:
                names = ', '.join(i.product_name for i in overlap)
                return JsonResponse(
                    {
                        'success': False,
                        'error': f'Already refunded: {names}',
                    },
                    status=400,
                )

        if refund_item_ids_raw is not None:
            try:
                refund_item_ids = [int(x) for x in refund_item_ids_raw]
            except (TypeError, ValueError):
                return JsonResponse(
                    {'success': False, 'error': 'Invalid item selection for refund'},
                    status=400,
                )
            if not refund_item_ids:
                return JsonResponse(
                    {'success': False, 'error': 'Please select at least one item to refund'},
                    status=400,
                )
            refund_id_set = set(refund_item_ids)
            items_to_refund = [
                i for i in all_items
                if i.id in refund_id_set and not i.refunded_at
            ]
            if not items_to_refund:
                return JsonResponse(
                    {'success': False, 'error': 'Selected items were not found on this transaction'},
                    status=400,
                )
            if len(items_to_refund) != len(refund_item_ids):
                return JsonResponse(
                    {'success': False, 'error': 'One or more selected items are invalid for this transaction'},
                    status=400,
                )
        else:
            # Backward compat: clients that omit refund_item_ids refund all remaining lines.
            items_to_refund = [i for i in all_items if not i.refunded_at]

        if not items_to_refund:
            return JsonResponse(
                {'success': False, 'error': 'No refundable items left on this transaction'},
                status=400,
            )

        refund_amount = sum(
            (item.total_price for item in items_to_refund),
            Decimal('0.00')
        )
        is_partial = (len(already_refunded) + len(items_to_refund)) < len(all_items)

        # Capture balances before refund for receipt
        balance_before = None
        balance_after = None
        refund_now = timezone.now()
        txn = transaction

        with db_transaction.atomic():
            # Process refund - ALL refunds go directly to card balance regardless of payment method
            if member and refund_amount > 0:
                balance_before = member.balance
                member.add_balance(refund_amount)
                balance_after = member.balance

                item_names = ', '.join(i.product_name for i in items_to_refund)
                if is_partial:
                    notes = (
                        f"Partial refund for transaction {txn.transaction_number} "
                        f"({len(items_to_refund)} of {len(all_items)} item(s): {item_names}) "
                        f"(Original: {txn.get_payment_method_display()})"
                    )
                else:
                    notes = (
                        f"Refund for transaction {txn.transaction_number} "
                        f"(Original: {txn.get_payment_method_display()})"
                    )
                if refund_reason:
                    notes += f". {refund_reason}"

                BalanceTransaction.objects.create(
                    member=member,
                    transaction_type='deposit',
                    amount=refund_amount,
                    balance_before=balance_before,
                    balance_after=balance_after,
                    notes=notes,
                )

            for item in items_to_refund:
                item.refunded_at = refund_now
                item.save(update_fields=['refunded_at'])

                if item.product:
                    should_restock = (
                        restock_item_ids is None or item.id in restock_item_ids
                    )
                    if should_restock:
                        _before_snap = capture_stock_snapshot(item.product)
                        item.product.stock_quantity += item.quantity
                        item.product.save()
                        record_stock_history(
                            item.product,
                            ProductStockHistory.CHANGE_REFUND,
                            _before_snap,
                            note=f'Refund restock — transaction {txn.transaction_number}',
                            user=request.user,
                        )

            total_refunded_lines = txn.items.filter(refunded_at__isnull=False).count()
            if is_partial:
                txn.status = 'partially_refunded'
                refunded_lines = list(txn.items.filter(refunded_at__isnull=False).order_by('id'))
                item_names = ', '.join(i.product_name for i in refunded_lines)
                refunded_ids = ','.join(str(i.id) for i in refunded_lines)
                base = (
                    f"Partially refunded ({total_refunded_lines} of {len(all_items)} item(s): "
                    f"{item_names})."
                )
                if refund_reason:
                    txn.notes = f"{base} {refund_reason} [refund_items:{refunded_ids}]"
                else:
                    txn.notes = f"{base} [refund_items:{refunded_ids}]"
            else:
                txn.status = 'refunded'
                txn.notes = f"Refunded. {refund_reason}" if refund_reason else "Refunded"

            txn.recalculate_totals_from_active_items(save=False)
            txn.save(
                update_fields=[
                    'status', 'notes', 'subtotal', 'vat_amount',
                    'vatable_sale', 'total_amount', 'updated_at',
                ]
            )

            refund_reason_obj, _ = RefundReason.objects.get_or_create(
                transaction=txn,
                defaults={'reason_type': 'other', 'details': refund_reason or ''},
            )
            if refund_reason:
                refund_reason_obj.details = refund_reason
                refund_reason_obj.save(update_fields=['details', 'updated_at'])
            refund_reason_obj.refund_items.set(
                txn.items.filter(refunded_at__isnull=False)
            )
        
        # Refresh member to get updated balances
        if member:
            member.refresh_from_db()
        
        receipt_item_ids = [item.id for item in items_to_refund]
        receipt_items_qs = ','.join(str(i) for i in receipt_item_ids)
        receipt_url = f'/refund-receipt/{transaction.id}/?items={receipt_items_qs}'

        # Generate refund receipt data - pass request for proper template rendering
        receipt_data = generate_refund_receipt_data(
            transaction,
            refund_reason,
            member,
            balance_before,
            balance_after,
            request=request,
            receipt_item_ids=receipt_item_ids,
        )

        try:
            from admin_panel.audit import mark_audit_recorded, record_audit
            record_audit(
                'REFUND',
                actor=request.user,
                description=(
                    f"{'Partial ' if is_partial else ''}Refund ₱{refund_amount} "
                    f"for {transaction.transaction_number}"
                ),
                request=request,
                object_type='Transaction',
                object_id=transaction.pk,
                metadata={
                    'transaction_number': transaction.transaction_number,
                    'refund_amount': str(refund_amount),
                    'is_partial': is_partial,
                    'reason': (refund_reason or '')[:200],
                },
            )
            mark_audit_recorded(request)
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully processed refund of ₱{refund_amount} for transaction {transaction.transaction_number}',
            'transaction': {
                'id': transaction.id,
                'transaction_number': transaction.transaction_number,
                'refund_amount': str(refund_amount),
                'is_partial': is_partial,
            },
            'receipt_url': receipt_url,
            'receipt_item_ids': receipt_item_ids,
            'receipt': receipt_data
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Server error: {str(e)}'})


@login_required
@require_http_methods(["GET"])
def generate_daily_report_pdf(request):
    """Generate and download a sales and stock report as PDF or Excel (supports single date or date range)."""
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to generate reports.')
        return redirect('dashboard')

    today = timezone.now().date()

    # Support date range: date_from / date_to, with fallback to single 'date' param
    date_from_str = request.GET.get('date_from', '') or request.GET.get('date', '')
    date_to_str = request.GET.get('date_to', '')

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else today
    except ValueError:
        messages.error(request, 'Invalid start date format. Use YYYY-MM-DD')
        return redirect('dashboard')

    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else date_from
    except ValueError:
        messages.error(request, 'Invalid end date format. Use YYYY-MM-DD')
        return redirect('dashboard')

    if date_to < date_from:
        date_from, date_to = date_to, date_from

    is_range = date_from != date_to
    requested_format = (request.GET.get('format') or 'pdf').strip().lower()
    if requested_format not in ('pdf', 'excel'):
        requested_format = 'pdf'

    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=18)

    # Container for the 'Flowable' objects
    elements = []
    styles = getSampleStyleSheet()

    # Use "PHP" instead of peso sign for better font compatibility in PDF
    currency_symbol = "PHP "
    # Keep report colors aligned with the system theme palette.
    pdf_primary = colors.HexColor('#ED1C24')
    pdf_primary_dark = colors.HexColor('#C4121A')
    pdf_heading = colors.HexColor('#166534')
    pdf_row_alt = colors.HexColor('#FEF7D5')

    # Define custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=pdf_primary_dark,
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=pdf_heading,
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold',
        keepWithNext=True,
    )

    # Title
    report_title = "Sales & Stock Report" if is_range else "Daily Sales & Stock Report"
    title = Paragraph(report_title, title_style)
    elements.append(title)

    if is_range:
        date_label = f"Report Period: {date_from.strftime('%B %d, %Y')} \u2013 {date_to.strftime('%B %d, %Y')}"
    else:
        date_label = f"Report Date: {date_from.strftime('%B %d, %Y')}"
    date_para = Paragraph(date_label, styles['Normal'])
    elements.append(date_para)
    elements.append(Spacer(1, 0.3*inch))

    # ===== SALES SUMMARY =====
    elements.append(Paragraph("Sales Summary", heading_style))

    # Same sale statuses / store-local period bounds as the dashboard
    daily_transactions = get_report_transactions(date_from, date_to)
    total_transactions = daily_transactions.count()

    # Header totals are net of refunded lines (recalculated on partial refund)
    agg = daily_transactions.aggregate(
        total_amount=Sum('total_amount'),
        total_vatable=Sum('vatable_sale'),
        total_vat=Sum('vat_amount'),
    )
    
    # Convert to Decimal, handling None values
    def to_decimal(value):
        if value is None:
            return Decimal('0.00')
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value))
    
    total_revenue = to_decimal(agg['total_amount'])
    total_vatable = to_decimal(agg['total_vatable'])
    total_vat = to_decimal(agg['total_vat'])
    total_vat_exempt = total_revenue - total_vatable - total_vat

    # Resolve the active tax label from the database (e.g. "VAT 12%" or "EVAT 5%").
    tax_enabled = bool(KioskConfig.get().tax_enabled)
    active_tax_label = _active_tax_label() if tax_enabled else ''

    # Payment method breakdown
    payment_breakdown = daily_transactions.values('payment_method').annotate(
        count=Count('id'),
        total=Sum('total_amount')
    ).order_by('-total')
    
    payment_labels = dict(Transaction.PAYMENT_METHODS)
    
    # Sales summary table — tax rows only when tax is enabled system-wide
    sales_data = [
        ['Metric', 'Value'],
        ['Total Transactions', f"{total_transactions:,}"],
        ['Total Revenue', f"{currency_symbol}{float(total_revenue):,.2f}"],
    ]
    if tax_enabled:
        sales_data.extend([
            ['Vatable Sales', f"{currency_symbol}{float(total_vatable):,.2f}"],
            [active_tax_label, f"{currency_symbol}{float(total_vat):,.2f}"],
            ['VAT-Exempt Sales', f"{currency_symbol}{float(total_vat_exempt):,.2f}"],
        ])
    
    sales_table = Table(sales_data, colWidths=[3*inch, 2*inch])
    sales_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), pdf_primary),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(sales_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Payment method breakdown
    if payment_breakdown:
        elements.append(Paragraph("Payment Method Breakdown", heading_style))
        payment_data = [['Payment Method', 'Count', 'Total Amount']]
        for entry in payment_breakdown:
            method_label = payment_labels.get(entry['payment_method'], entry['payment_method'].title())
            total_amount = entry['total'] if entry['total'] is not None else Decimal('0.00')
            payment_data.append([
                method_label,
                f"{entry['count']:,}",
                f"{currency_symbol}{float(total_amount):,.2f}"
            ])
        
        payment_table = Table(payment_data, colWidths=[2.5*inch, 1.25*inch, 1.25*inch])
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), pdf_primary),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(payment_table)
        elements.append(Spacer(1, 0.2*inch))

    top_customers = get_top_customers_for_period(date_from, date_to)
    top_customers_period_label = (
        f"{date_from.strftime('%B %d, %Y')} – {date_to.strftime('%B %d, %Y')}"
        if is_range else date_from.strftime('%B %d, %Y')
    )
    append_top_customers_pdf(
        elements,
        top_customers,
        heading_style=heading_style,
        styles=styles,
        currency_symbol=currency_symbol,
        header_color=pdf_primary,
        header_dark_color=pdf_primary_dark,
        row_alt_color=pdf_row_alt,
        period_label=top_customers_period_label,
    )
    
    # All Products Sold — non-refunded line items; revenue = Sum(total_price)
    all_products_sold = list(
        get_report_sale_items(date_from, date_to)
        .values('product_name', 'product_barcode')
        .annotate(
            quantity_sold=Sum('quantity'),
            total_revenue=Sum('total_price'),
        )
        .order_by('-quantity_sold')
    )

    if all_products_sold:
        elements.append(Paragraph("All Products Sold", heading_style))
        products_data = [['#', 'Product Name', 'Barcode', 'Price', 'Qty Sold', 'Revenue']]
        for idx, product in enumerate(all_products_sold, start=1):
            quantity = product['quantity_sold'] if product['quantity_sold'] is not None else 0
            revenue = product['total_revenue'] if product['total_revenue'] is not None else Decimal('0.00')
            # Effective average unit price so Price × Qty aligns with Revenue
            if quantity:
                unit_price = (Decimal(str(revenue)) / Decimal(quantity)).quantize(Decimal('0.01'))
            else:
                unit_price = Decimal('0.00')
            products_data.append([
                str(idx),
                product['product_name'][:35],
                product['product_barcode'],
                f"{currency_symbol}{float(unit_price):,.2f}",
                f"{quantity:,}",
                f"{currency_symbol}{float(revenue):,.2f}"
            ])
        # Totals footer
        total_qty_sold = sum(
            (p['quantity_sold'] or 0) for p in all_products_sold
        )
        total_rev_sold = sum(
            (float(p['total_revenue']) if p['total_revenue'] is not None else 0.0)
            for p in all_products_sold
        )
        products_data.append([
            '', 'TOTAL', '', '', f"{total_qty_sold:,}", f"{currency_symbol}{total_rev_sold:,.2f}"
        ])
        totals_idx = len(products_data) - 1
        products_table = Table(
            products_data,
            colWidths=[0.3*inch, 2.0*inch, 1.0*inch, 0.85*inch, 0.7*inch, 1.15*inch],
            repeatRows=1,
        )
        prod_style = [
            ('BACKGROUND', (0, 0), (-1, 0), pdf_primary),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (2, -1), 'LEFT'),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, totals_idx - 1), 9),
            ('TOPPADDING', (0, 1), (-1, totals_idx - 1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, totals_idx - 1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, totals_idx - 1), [colors.white, pdf_row_alt]),
            ('BACKGROUND', (0, totals_idx), (-1, totals_idx), pdf_primary_dark),
            ('TEXTCOLOR', (0, totals_idx), (-1, totals_idx), colors.whitesmoke),
            ('FONTNAME', (0, totals_idx), (-1, totals_idx), 'Helvetica-Bold'),
            ('FONTSIZE', (0, totals_idx), (-1, totals_idx), 9),
            ('TOPPADDING', (0, totals_idx), (-1, totals_idx), 7),
            ('BOTTOMPADDING', (0, totals_idx), (-1, totals_idx), 7),
            ('LINEABOVE', (0, totals_idx), (-1, totals_idx), 1.5, pdf_primary_dark),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]
        products_table.setStyle(TableStyle(prod_style))
        elements.append(products_table)
        elements.append(Spacer(1, 0.2*inch))

    # Product Sales Summary: subtotal from DB Sum(total_price); price = effective average
    product_summary_period_label = (
        f"{date_from.strftime('%B %d, %Y')} – {date_to.strftime('%B %d, %Y')}"
        if is_range else date_from.strftime('%B %d, %Y')
    )
    product_sales_summary = get_product_sales_summary_for_period(date_from, date_to)
    append_product_sales_summary_pdf(
        elements,
        product_sales_summary,
        heading_style=heading_style,
        styles=styles,
        currency_symbol=currency_symbol,
        header_color=pdf_primary,
        header_dark_color=pdf_primary_dark,
        row_alt_color=pdf_row_alt,
        period_label=product_summary_period_label,
    )

    wholesale_summary = get_wholesale_sales_for_period(date_from, date_to)
    wholesale_period_label = (
        f"{date_from.strftime('%B %d, %Y')} – {date_to.strftime('%B %d, %Y')}"
        if is_range else date_from.strftime('%B %d, %Y')
    )
    append_wholesale_sales_pdf(
        elements,
        wholesale_summary,
        heading_style=heading_style,
        styles=styles,
        currency_symbol=currency_symbol,
        header_color=pdf_primary,
        header_dark_color=pdf_primary_dark,
        row_alt_color=pdf_row_alt,
        period_label=wholesale_period_label,
    )

    # ===== WALK-IN CUSTOMERS SUMMARY =====
    walk_in_summary = get_walk_in_summary_for_period(date_from, date_to)
    walk_in_period_label = (
        f"{date_from.strftime('%B %d, %Y')} – {date_to.strftime('%B %d, %Y')}"
        if is_range else date_from.strftime('%B %d, %Y')
    )
    append_walk_in_summary_pdf(
        elements,
        walk_in_summary,
        heading_style=heading_style,
        styles=styles,
        currency_symbol=currency_symbol,
        header_color=pdf_primary,
        header_dark_color=pdf_primary_dark,
        row_alt_color=pdf_row_alt,
        period_label=walk_in_period_label,
    )

    # ===== GIVEAWAY PRODUCTS SUMMARY =====
    giveaway_summary = get_giveaway_summary_for_period(date_from, date_to)
    elements.append(Paragraph("Giveaway Products Summary", heading_style))

    giveaway_metrics = [
        ['Metric', 'Value'],
        ['Total Units Given Away', f"{giveaway_summary['total_units']:,}"],
        ['Distinct Products', f"{len(giveaway_summary['products']):,}"],
        ['Estimated Value (List Price)', f"{currency_symbol}{float(giveaway_summary['total_est_value']):,.2f}"],
    ]
    giveaway_metrics_table = Table(giveaway_metrics, colWidths=[3*inch, 2*inch])
    giveaway_metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3730a3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#eef2ff')),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eef2ff')]),
    ]))
    elements.append(giveaway_metrics_table)
    elements.append(Spacer(1, 0.15*inch))

    if giveaway_summary['products']:
        giveaway_data = [['#', 'Product Name', 'Barcode', 'Unit Price', 'Qty Given', 'Est. Value']]
        for idx, product in enumerate(giveaway_summary['products'], start=1):
            giveaway_data.append([
                str(idx),
                product['name'][:35],
                product['barcode'],
                f"{currency_symbol}{float(product['price']):,.2f}",
                f"{product['quantity_given']:,}",
                f"{currency_symbol}{float(product['est_value']):,.2f}",
            ])
        giveaway_data.append([
            '', 'TOTAL', '', '', f"{giveaway_summary['total_units']:,}",
            f"{currency_symbol}{float(giveaway_summary['total_est_value']):,.2f}",
        ])
        giveaway_totals_idx = len(giveaway_data) - 1
        giveaway_table = Table(
            giveaway_data,
            colWidths=[0.3*inch, 1.75*inch, 1*inch, 0.85*inch, 0.75*inch, 1*inch],
            repeatRows=1,
        )
        giveaway_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3730a3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (2, -1), 'LEFT'),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, giveaway_totals_idx - 1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, giveaway_totals_idx - 1), colors.HexColor('#eef2ff')),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, giveaway_totals_idx - 1), [colors.white, colors.HexColor('#eef2ff')]),
            ('BACKGROUND', (0, giveaway_totals_idx), (-1, giveaway_totals_idx), colors.HexColor('#312e81')),
            ('TEXTCOLOR', (0, giveaway_totals_idx), (-1, giveaway_totals_idx), colors.whitesmoke),
            ('FONTNAME', (0, giveaway_totals_idx), (-1, giveaway_totals_idx), 'Helvetica-Bold'),
            ('LINEABOVE', (0, giveaway_totals_idx), (-1, giveaway_totals_idx), 1.5, colors.HexColor('#312e81')),
        ]))
        elements.append(giveaway_table)
    else:
        period_label = 'this period' if is_range else 'this date'
        elements.append(Paragraph(f"No giveaway products recorded for {period_label}.", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    elements.append(PageBreak())
    
    # ===== STOCK SUMMARY =====
    elements.append(Paragraph("Stock Summary", heading_style))
    
    # Total products
    total_products = Product.objects.filter(is_active=True).count()
    low_stock_count = Product.objects.filter(is_active=True, stock_quantity__lte=F('low_stock_threshold')).exclude(stock_quantity=0).count()
    out_of_stock_count = Product.objects.filter(is_active=True, stock_quantity=0).count()
    
    stock_summary_data = [
        ['Metric', 'Value'],
        ['Total Active Products', f"{total_products:,}"],
        ['Low Stock Items', f"{low_stock_count:,}"],
        ['Out of Stock Items', f"{out_of_stock_count:,}"],
    ]
    
    stock_summary_table = Table(stock_summary_data, colWidths=[3*inch, 2*inch])
    stock_summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), pdf_primary),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(stock_summary_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # Low Stock Products
    low_stock_products = Product.objects.filter(
        is_active=True,
        stock_quantity__lte=F('low_stock_threshold')
    ).order_by('stock_quantity', 'name')
    
    if low_stock_products.exists():
        elements.append(Paragraph("Low Stock & Out of Stock Products", heading_style))
        low_stock_data = [['Product Name', 'Barcode', 'Current Stock', 'Status']]
        
        for product in low_stock_products[:50]:  # Limit to 50 for PDF size
            status = "Out of Stock" if product.stock_quantity == 0 else "Low Stock"
            low_stock_data.append([
                product.name[:30],
                product.barcode,
                f"{product.stock_quantity:,}",
                status
            ])
        
        low_stock_table = Table(low_stock_data, colWidths=[2.25*inch, 1.25*inch, 0.9*inch, 1*inch])
        low_stock_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d32f2f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(low_stock_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Category Stock Summary
    category_stock = list(Product.objects.filter(is_active=True).values(
        'category__name'
    ).annotate(
        product_count=Count('id'),
        total_stock=Sum('stock_quantity'),
        low_stock_count=Count('id', filter=Q(stock_quantity__lte=F('low_stock_threshold')))
    ).order_by('category__name'))

    if category_stock:
        elements.append(Paragraph("Stock by Category", heading_style))

        # Dynamically size category column based on longest name
        max_name_len = max((len(cat['category__name'] or 'Uncategorized') for cat in category_stock), default=10)
        # Usable page width: A4 (595.27pt) minus left+right margins (30+30 = 60pt) = 535.27pt
        usable_width = 535.27
        # Category col: between 140pt and 240pt depending on name length
        cat_col = min(240, max(140, max_name_len * 7))
        remaining = usable_width - cat_col
        num_col     = remaining * 0.25
        stock_col   = remaining * 0.375
        low_col     = remaining * 0.375

        # Aggregate totals directly from DB (source of truth, not from annotation)
        db_totals = Product.objects.filter(is_active=True).aggregate(
            total_products=Count('id'),
            total_stock=Sum('stock_quantity'),
            total_low=Count('id', filter=Q(stock_quantity__lte=F('low_stock_threshold'))),
        )
        total_all_products = db_totals['total_products'] or 0
        total_all_stock    = db_totals['total_stock'] or 0
        total_all_low      = db_totals['total_low'] or 0

        category_data = [['Category', 'Products', 'Total Stock', 'Low Stock Items']]
        for cat in category_stock:
            category_name = cat['category__name'] or 'Uncategorized'
            category_data.append([
                category_name,
                f"{cat['product_count']:,}",
                f"{cat['total_stock'] or 0:,}",
                f"{cat['low_stock_count'] or 0:,}",
            ])
        # Totals footer row
        category_data.append([
            'TOTAL',
            f"{total_all_products:,}",
            f"{total_all_stock:,}",
            f"{total_all_low:,}",
        ])

        totals_row_idx = len(category_data) - 1
        category_table = Table(
            category_data,
            colWidths=[cat_col, num_col, stock_col, low_col],
            repeatRows=1,
        )
        cat_style = [
            # Header
            ('BACKGROUND',  (0, 0), (-1, 0),              pdf_primary),
            ('TEXTCOLOR',   (0, 0), (-1, 0),              colors.whitesmoke),
            ('FONTNAME',    (0, 0), (-1, 0),              'Helvetica-Bold'),
            ('FONTSIZE',    (0, 0), (-1, 0),              10),
            ('BOTTOMPADDING', (0, 0), (-1, 0),            10),
            ('TOPPADDING',  (0, 0), (-1, 0),              8),
            # Data rows
            ('FONTSIZE',    (0, 1), (-1, totals_row_idx - 1), 9),
            ('TOPPADDING',  (0, 1), (-1, totals_row_idx - 1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, totals_row_idx - 1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, totals_row_idx - 1), [colors.white, pdf_row_alt]),
            # Totals footer
            ('BACKGROUND',  (0, totals_row_idx), (-1, totals_row_idx), pdf_primary_dark),
            ('TEXTCOLOR',   (0, totals_row_idx), (-1, totals_row_idx), colors.whitesmoke),
            ('FONTNAME',    (0, totals_row_idx), (-1, totals_row_idx), 'Helvetica-Bold'),
            ('FONTSIZE',    (0, totals_row_idx), (-1, totals_row_idx), 9),
            ('TOPPADDING',  (0, totals_row_idx), (-1, totals_row_idx), 7),
            ('BOTTOMPADDING', (0, totals_row_idx), (-1, totals_row_idx), 7),
            # Alignment
            ('ALIGN',  (0, 0), (0, -1), 'LEFT'),
            ('ALIGN',  (1, 0), (-1, -1), 'CENTER'),
            # Grid
            ('GRID',   (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEABOVE', (0, totals_row_idx), (-1, totals_row_idx), 1.5, pdf_primary_dark),
        ]
        # Highlight rows that have any low-stock items
        for i, cat in enumerate(category_stock, start=1):
            if (cat['low_stock_count'] or 0) > 0:
                cat_style.append(('TEXTCOLOR', (3, i), (3, i), colors.HexColor('#c62828')))
                cat_style.append(('FONTNAME',  (3, i), (3, i), 'Helvetica-Bold'))

        category_table.setStyle(TableStyle(cat_style))
        elements.append(category_table)
        elements.append(Spacer(1, 0.2*inch))
    
    elements.append(PageBreak())
    
    # ===== RECENT TRANSACTIONS =====
    elements.append(Paragraph("Recent Transactions (Last 50)", heading_style))
    
    recent_transactions = list(daily_transactions.order_by('-created_at')[:50])
    
    if recent_transactions:
        if tax_enabled:
            transactions_data = [
                ['Transaction #', 'Member', 'Method', 'Time', 'Amount', 'Vatable Sale', active_tax_label]
            ]
            txn_col_widths = [
                1.3 * inch, 1.2 * inch, 0.6 * inch, 0.65 * inch,
                0.95 * inch, 0.9 * inch, 0.85 * inch,
            ]
        else:
            transactions_data = [['Transaction #', 'Member', 'Method', 'Time', 'Amount']]
            txn_col_widths = [1.5 * inch, 1.5 * inch, 0.8 * inch, 0.85 * inch, 1.5 * inch]

        for txn in recent_transactions:
            member_name = txn.customer_display_name
            if len(member_name) > 20:
                member_name = member_name[:17] + '...'
            
            method_short = {
                'cash': 'Cash',
                'debit': 'Debit'
            }.get(txn.payment_method, txn.payment_method.title())
            
            time_str = timezone.localtime(txn.created_at).strftime('%H:%M:%S')
            amount = Decimal(str(txn.total_amount)) if txn.total_amount is not None else Decimal('0.00')
            row = [
                txn.transaction_number[:15],
                member_name,
                method_short,
                time_str,
                f"{currency_symbol}{float(amount):,.2f}",
            ]
            if tax_enabled:
                vatable_amt = (
                    Decimal(str(txn.vatable_sale)) if txn.vatable_sale is not None else Decimal('0.00')
                )
                vat_amt = (
                    Decimal(str(txn.vat_amount)) if txn.vat_amount is not None else Decimal('0.00')
                )
                row.extend([
                    f"{currency_symbol}{float(vatable_amt):,.2f}",
                    f"{currency_symbol}{float(vat_amt):,.2f}",
                ])
            transactions_data.append(row)
        
        txn_table = Table(transactions_data, colWidths=txn_col_widths)
        txn_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), pdf_primary),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (3, -1), 'CENTER'),
            ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(txn_table)
    else:
        elements.append(Paragraph("No transactions for this period.", styles['Normal']))

    # ===== ALL PRODUCTS INVENTORY =====
    elements.append(PageBreak())
    elements.append(Paragraph("All Products & Current Stock Quantities", heading_style))

    all_products = list(Product.objects.filter(is_active=True).select_related('category').order_by(
        'category__name', 'name'
    ))

    # Build a lookup: barcode -> qty sold and revenue for the period
    sold_lookup = {
        p['product_barcode']: {
            'times': p['quantity_sold'] or 0,
            'revenue': float(p['total_revenue']) if p['total_revenue'] is not None else 0.0,
        }
        for p in all_products_sold
    } if all_products_sold else {}

    if all_products:
        inventory_data = [['#', 'Product Name', 'Category', 'Unit Price', 'Stock Qty', 'Qty Sold', 'Total Sales']]
        for idx, prod in enumerate(all_products, start=1):
            info = sold_lookup.get(prod.barcode, {'times': 0, 'revenue': 0.0})
            inventory_data.append([
                str(idx),
                prod.name[:35],
                (prod.category.name if prod.category else 'Uncategorized')[:22],
                f"{currency_symbol}{float(prod.price):,.2f}",
                f"{prod.stock_quantity:,}",
                f"{info['times']:,}",
                f"{currency_symbol}{info['revenue']:,.2f}",
            ])
        # Totals footer
        inv_total_stock = sum(p.stock_quantity for p in all_products)
        inv_total_times = sum(sold_lookup.get(p.barcode, {}).get('times', 0) for p in all_products)
        inv_total_sales = sum(sold_lookup.get(p.barcode, {}).get('revenue', 0.0) for p in all_products)
        inventory_data.append([
            '', 'TOTAL', '', '', f"{inv_total_stock:,}", f"{inv_total_times:,}", f"{currency_symbol}{inv_total_sales:,.2f}"
        ])
        totals_inv_idx = len(inventory_data) - 1

        inventory_table = Table(
            inventory_data,
            colWidths=[0.3*inch, 1.9*inch, 1.15*inch, 0.95*inch, 0.8*inch, 0.8*inch, 0.95*inch],
            repeatRows=1,
        )
        inv_style = [
            ('BACKGROUND', (0, 0), (-1, 0), pdf_primary),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 1), (2, -1), 'LEFT'),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 1), (-1, totals_inv_idx - 1), 8),
            ('TOPPADDING', (0, 1), (-1, totals_inv_idx - 1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, totals_inv_idx - 1), 3),
            ('ROWBACKGROUNDS', (0, 1), (-1, totals_inv_idx - 1), [colors.white, pdf_row_alt]),
            ('BACKGROUND', (0, totals_inv_idx), (-1, totals_inv_idx), pdf_primary_dark),
            ('TEXTCOLOR', (0, totals_inv_idx), (-1, totals_inv_idx), colors.whitesmoke),
            ('FONTNAME', (0, totals_inv_idx), (-1, totals_inv_idx), 'Helvetica-Bold'),
            ('FONTSIZE', (0, totals_inv_idx), (-1, totals_inv_idx), 9),
            ('TOPPADDING', (0, totals_inv_idx), (-1, totals_inv_idx), 7),
            ('BOTTOMPADDING', (0, totals_inv_idx), (-1, totals_inv_idx), 7),
            ('LINEABOVE', (0, totals_inv_idx), (-1, totals_inv_idx), 1.5, pdf_primary_dark),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]

        inventory_table.setStyle(TableStyle(inv_style))
        elements.append(inventory_table)
    else:
        elements.append(Paragraph("No active products found.", styles['Normal']))

    if requested_format == 'excel':
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        header_fill = PatternFill(start_color='1F7A3A', end_color='1F7A3A', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        title_font = Font(size=14, bold=True, color='166534')
        title_fill = PatternFill(start_color='ECFDF3', end_color='ECFDF3', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        def style_header_row(ws, row_number, col_count):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row_number, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        def apply_table_borders(ws, start_row, end_row, col_count):
            for row in range(start_row, end_row + 1):
                for col in range(1, col_count + 1):
                    ws.cell(row=row, column=col).border = thin_border

        def auto_size_columns(ws, max_col):
            for idx in range(1, max_col + 1):
                col_letter = get_column_letter(idx)
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    value = ws.cell(row=row, column=idx).value
                    if value is None:
                        continue
                    max_length = max(max_length, len(str(value)))
                ws.column_dimensions[col_letter].width = min(max(12, max_length + 2), 45)

        def add_sheet(sheet_name, title, headers, rows, currency_cols=None, date_cols=None):
            ws = wb.create_sheet(title=sheet_name)
            ws.freeze_panes = 'A3'
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[1].height = 24

            ws.append(headers)
            style_header_row(ws, 2, len(headers))

            for row_data in rows:
                ws.append(row_data)

            data_end_row = ws.max_row
            if data_end_row >= 3:
                table_ref = f"A2:{get_column_letter(len(headers))}{data_end_row}"
                tbl = XLTable(displayName=f"T_{sheet_name.replace(' ', '_')}", ref=table_ref)
                tbl.tableStyleInfo = TableStyleInfo(
                    name='TableStyleMedium9',
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                ws.add_table(tbl)
                apply_table_borders(ws, 2, data_end_row, len(headers))

            currency_cols = currency_cols or []
            date_cols = date_cols or []
            for row in range(3, ws.max_row + 1):
                for col in currency_cols:
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
                for col in date_cols:
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

            auto_size_columns(ws, len(headers))
            return ws

        if is_range:
            report_period = (
                f"{date_from.strftime('%B %d, %Y')} – {date_to.strftime('%B %d, %Y')}"
            )
            summary_title = f'Sales Summary — {report_period}'
        else:
            report_period = date_from.strftime('%B %d, %Y')
            summary_title = f'Sales Summary — Report Date: {report_period}'

        summary_rows = [
            ['Total Transactions', total_transactions],
            ['Total Revenue', float(total_revenue)],
        ]
        if tax_enabled:
            summary_rows.extend([
                ['Vatable Sales', float(total_vatable)],
                [active_tax_label, float(total_vat)],
                ['VAT-Exempt Sales', float(total_vat_exempt)],
            ])
        add_sheet('Summary', summary_title, ['Metric', 'Value'], summary_rows, currency_cols=[2])

        payment_rows = []
        for entry in payment_breakdown:
            method_label = payment_labels.get(entry['payment_method'], entry['payment_method'].title())
            total_amount = entry['total'] if entry['total'] is not None else Decimal('0.00')
            payment_rows.append([method_label, entry['count'], float(total_amount)])
        add_sheet('Payments', 'Payment Method Breakdown', ['Payment Method', 'Count', 'Total Amount'], payment_rows, currency_cols=[3])

        add_sheet(
            'Top Customers',
            'Top 10 Customers',
            ['#', 'Customer Name', 'Type', 'Orders', 'Revenue', 'Last Visit'],
            top_customers_excel_rows(top_customers),
            currency_cols=[5],
        )

        sold_rows = []
        for idx, product in enumerate(all_products_sold, start=1):
            quantity = product['quantity_sold'] if product['quantity_sold'] is not None else 0
            revenue = product['total_revenue'] if product['total_revenue'] is not None else Decimal('0.00')
            if quantity:
                unit_price = float(
                    (Decimal(str(revenue)) / Decimal(quantity)).quantize(Decimal('0.01'))
                )
            else:
                unit_price = 0.0
            sold_rows.append([
                idx, product['product_name'], product['product_barcode'],
                unit_price, quantity, float(revenue),
            ])
        add_sheet(
            'Products Sold',
            'All Products Sold',
            ['#', 'Product Name', 'Barcode', 'Price', 'Qty Sold', 'Revenue'],
            sold_rows,
            currency_cols=[4, 6],
        )

        add_sheet(
            'Product Summary',
            'Product Sales Summary',
            ['#', 'Product', 'Selling Price', 'Qty Sold', 'Subtotal'],
            product_sales_summary_excel_rows(product_sales_summary),
            currency_cols=[3, 5],
        )

        add_sheet(
            'Wholesale Summary',
            'Wholesale Sales Summary',
            ['Metric', 'Value'],
            wholesale_excel_metric_rows(wholesale_summary),
        )
        add_sheet(
            'Wholesale Products',
            'Wholesale Products Sold',
            [
                '#', 'Product Name', 'Barcode', 'Unit', 'Pcs/Package',
                'Price', 'Packages Sold', 'Pieces Sold', 'Revenue',
            ],
            wholesale_excel_product_rows(wholesale_summary),
            currency_cols=[6, 9],
        )

        add_sheet(
            'Walk-in Summary',
            'Walk-in Customers Summary',
            ['Metric', 'Value'],
            walk_in_excel_metric_rows(walk_in_summary),
        )

        add_sheet(
            'Walk-in Customers',
            'Walk-in Customers (Period)',
            ['#', 'Customer Name', 'Orders', 'Revenue', 'Last Visit'],
            walk_in_excel_customer_rows(walk_in_summary),
            currency_cols=[4],
        )

        giveaway_rows = []
        for idx, product in enumerate(giveaway_summary['products'], start=1):
            giveaway_rows.append([
                idx,
                product['name'],
                product['barcode'],
                float(product['price']),
                product['quantity_given'],
                float(product['est_value']),
            ])
        add_sheet(
            'Giveaways',
            'Giveaway Products Summary',
            ['#', 'Product Name', 'Barcode', 'Unit Price', 'Qty Given', 'Est. Value'],
            giveaway_rows,
            currency_cols=[4, 6],
        )

        category_rows = [
            [
                cat['category__name'] or 'Uncategorized',
                cat['product_count'] or 0,
                cat['total_stock'] or 0,
                cat['low_stock_count'] or 0,
            ]
            for cat in category_stock
        ]
        add_sheet(
            'Stock by Category',
            'Stock by Category',
            ['Category', 'Products', 'Total Stock', 'Low Stock Items'],
            category_rows,
        )

        transaction_rows = []
        for txn in recent_transactions:
            member_name = txn.customer_display_name
            is_walk_in = bool(txn.walk_in_customer_id or (txn.guest_customer_name or '').strip())
            method_short = {'cash': 'Cash', 'debit': 'Debit'}.get(txn.payment_method, txn.payment_method.title())
            amount = Decimal(str(txn.total_amount)) if txn.total_amount is not None else Decimal('0.00')
            row = [
                txn.transaction_number,
                member_name,
                'Walk-in' if is_walk_in else 'Member/Guest',
                method_short,
                timezone.localtime(txn.created_at).strftime('%Y-%m-%d %H:%M:%S'),
                float(amount),
            ]
            if tax_enabled:
                vat_amt = Decimal(str(txn.vat_amount)) if txn.vat_amount is not None else Decimal('0.00')
                vatable_amt = (
                    Decimal(str(txn.vatable_sale)) if txn.vatable_sale is not None else Decimal('0.00')
                )
                row.extend([float(vatable_amt), float(vat_amt)])
            transaction_rows.append(row)
        txn_headers = ['Transaction #', 'Customer', 'Type', 'Method', 'Time', 'Amount']
        txn_currency_cols = [6]
        if tax_enabled:
            txn_headers.extend(['Vatable Sale', active_tax_label])
            txn_currency_cols = [6, 7, 8]
        add_sheet(
            'Transactions',
            'Recent Transactions (Last 50)',
            txn_headers,
            transaction_rows,
            currency_cols=txn_currency_cols,
            date_cols=[5],
        )

        inventory_rows = []
        for idx, prod in enumerate(all_products, start=1):
            info = sold_lookup.get(prod.barcode, {'times': 0, 'revenue': 0.0})
            inventory_rows.append([
                idx,
                prod.name,
                prod.category.name if prod.category else 'Uncategorized',
                float(prod.price),
                prod.stock_quantity,
                info['times'],
                float(info['revenue']),
            ])
        add_sheet(
            'Inventory',
            'All Products and Current Stock Quantities',
            ['#', 'Product Name', 'Category', 'Unit Price', 'Stock Qty', 'Qty Sold', 'Total Sales'],
            inventory_rows,
            currency_cols=[4, 7],
        )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        if is_range:
            filename = f'report_{date_from.strftime("%Y%m%d")}_to_{date_to.strftime("%Y%m%d")}.xlsx'
        else:
            filename = f'daily_report_{date_from.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    # Create HTTP response with PDF
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    if is_range:
        filename = f'report_{date_from.strftime("%Y%m%d")}_to_{date_to.strftime("%Y%m%d")}.pdf'
    else:
        filename = f'daily_report_{date_from.strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def _barcode_entries_for_product(product):
    """Build printable barcode cards for a product (piece + active wholesale units)."""
    category_name = product.category.name if product.category else 'Uncategorized'
    entries = [{
        'display_name': product.name,
        'barcode': product.barcode,
        'price': product.price,
        'category_name': category_name,
        'badge': 'Piece',
        'is_wholesale': False,
    }]
    seen_barcodes = {product.barcode}

    for unit in product.sale_units.all():
        if not unit.is_active or unit.barcode in seen_barcodes:
            continue
        if unit.sale_mode == ProductSaleUnit.SALE_MODE_WHOLESALE:
            seen_barcodes.add(unit.barcode)
            entries.append({
                'display_name': product.name,
                'barcode': unit.barcode,
                'price': unit.price,
                'category_name': category_name,
                'badge': unit.unit_label or 'Wholesale',
                'is_wholesale': True,
            })
        elif unit.sale_mode == ProductSaleUnit.SALE_MODE_RETAIL:
            seen_barcodes.add(unit.barcode)
            entries.append({
                'display_name': product.name,
                'barcode': unit.barcode,
                'price': unit.price,
                'category_name': category_name,
                'badge': unit.unit_label or 'Piece',
                'is_wholesale': False,
            })

    return entries


def _barcode_entries_for_products(products):
    """Expand a product queryset/list into ordered printable barcode cards."""
    entries = []
    for product in products:
        entries.extend(_barcode_entries_for_product(product))
    return entries


def _estimate_barcode_cells_per_page(
    *,
    page_height,
    top_margin,
    bottom_margin,
    items_per_row,
    img_height,
    header_block=1.35 * inch,
):
    """Approximate barcode labels that fit below a header block on one sheet."""
    # Text lines (name, badge, code, price, category) + table cell padding.
    row_height = img_height + 0.78 * inch + 22
    usable = page_height - top_margin - bottom_margin - header_block
    rows = max(1, int(usable // row_height))
    return rows * items_per_row


# 8.5 × 13 in coupon / long bond — common label stock in PH retail.
BARCODE_SHEET_PAGE_SIZE = (8.5 * inch, 13 * inch)
BARCODE_SHEET_HEADER_FIRST = 2.0 * inch
BARCODE_SHEET_HEADER_CONTINUED = 0.5 * inch


def _repeat_entries_to_fill_page(entries, target_count):
    """Cycle barcode entries until the page grid is full (single-product label sheets)."""
    if not entries or target_count <= len(entries):
        return list(entries)
    filled = []
    idx = 0
    while len(filled) < target_count:
        filled.append(entries[idx % len(entries)])
        idx += 1
    return filled


def _filter_barcode_entries_by_type(entries, label_type):
    """Keep piece, wholesale, or both label variants for single-product sheets."""
    if label_type == 'piece':
        filtered = [entry for entry in entries if not entry.get('is_wholesale')]
    elif label_type == 'wholesale':
        filtered = [entry for entry in entries if entry.get('is_wholesale')]
    else:
        filtered = list(entries)
    return filtered


def _parse_barcode_page_count(raw_value, *, default=1, maximum=50):
    try:
        page_count = int(raw_value)
    except (TypeError, ValueError):
        page_count = default
    return max(1, min(maximum, page_count))


def _barcode_label_type_label(label_type):
    return {
        'both': 'Piece + Wholesale',
        'piece': 'Piece only',
        'wholesale': 'Wholesale only',
    }.get(label_type, 'Piece + Wholesale')


@login_required
@require_http_methods(["GET"])
def download_product_barcodes_pdf(request):
    """Download a printable PDF of products with barcodes, filterable by category or specific product selection."""
    if not is_cashier_or_admin(request.user):
        messages.warning(request, 'You do not have permission to access this page.')
        return redirect('inventory_management')

    price_order = request.GET.get('price_order', 'asc')
    product_ids = request.GET.getlist('product_ids')
    bc_mode = request.GET.get('bc_mode', 'category')
    grouped_output = False
    filter_label = ''
    filename_label = 'all'

    if bc_mode == 'by_category':
        # Grouped by category mode — all products, ordered by category then price
        grouped_output = True
        if price_order == 'desc':
            products = Product.objects.filter(is_active=True).select_related('category').prefetch_related(
                'sale_units'
            ).order_by('category__name', '-price', 'name')
        else:
            products = Product.objects.filter(is_active=True).select_related('category').prefetch_related(
                'sale_units'
            ).order_by('category__name', 'price', 'name')
        filter_label = 'All Categories (Grouped)'
    elif product_ids:
        # Specific product selection mode
        safe_ids = [int(pid) for pid in product_ids if str(pid).strip().isdigit()]
        products = Product.objects.filter(id__in=safe_ids, is_active=True).select_related(
            'category'
        ).prefetch_related('sale_units')
        filter_label = f'Selected: {products.count()} product(s)'
        if price_order == 'desc':
            products = products.order_by('-price', 'name')
        else:
            products = products.order_by('price', 'name')
    else:
        # Category filter mode
        category_ids = request.GET.getlist('category_ids')
        safe_category_ids = [int(cid) for cid in category_ids if str(cid).strip().isdigit()]
        products = Product.objects.filter(is_active=True).select_related('category').prefetch_related(
            'sale_units'
        )
        if safe_category_ids:
            products = products.filter(category_id__in=safe_category_ids)
            selected_categories = list(
                Category.objects.filter(id__in=safe_category_ids).values_list('name', flat=True)
            )
            if selected_categories:
                if len(selected_categories) <= 3:
                    filter_label = f'Categories: {", ".join(selected_categories)}'
                else:
                    filter_label = f'Categories: {len(selected_categories)} selected'
                if len(selected_categories) == 1:
                    filename_label = selected_categories[0].replace(' ', '_')
            else:
                filter_label = 'Categories: All'
        else:
            filter_label = 'Categories: All'
        grouped_output = True
        if price_order == 'desc':
            products = products.order_by('category__name', '-price', 'name')
        else:
            products = products.order_by('category__name', 'price', 'name')
        filter_label = f'{filter_label} (Grouped)'

    buffer = io.BytesIO()
    page_width, page_height = BARCODE_SHEET_PAGE_SIZE
    margin = 25
    top_margin = 36

    doc = SimpleDocTemplate(
        buffer,
        pagesize=BARCODE_SHEET_PAGE_SIZE,
        rightMargin=margin,
        leftMargin=margin,
        topMargin=top_margin,
        bottomMargin=margin,
    )

    # Styles
    name_style = ParagraphStyle(
        'PName', fontSize=11, alignment=TA_CENTER, leading=13, fontName='Helvetica-Bold',
        wordWrap='CJK'
    )
    price_style = ParagraphStyle(
        'PPrice', fontSize=10, alignment=TA_CENTER,
        textColor=colors.HexColor('#ED1C24'), fontName='Helvetica-Bold'
    )
    cat_style = ParagraphStyle(
        'PCat', fontSize=6, alignment=TA_CENTER,
        textColor=colors.HexColor('#64748b')
    )
    code_style = ParagraphStyle(
        'PCode', fontSize=6, alignment=TA_CENTER,
        textColor=colors.HexColor('#475569')
    )
    title_style = ParagraphStyle(
        'DocTitle', fontSize=14, alignment=TA_CENTER,
        fontName='Helvetica-Bold', spaceAfter=4
    )
    sub_style = ParagraphStyle(
        'DocSub', fontSize=8, alignment=TA_CENTER,
        textColor=colors.HexColor('#64748b'), spaceAfter=8
    )
    cat_header_style = ParagraphStyle(
        'CatHeader', fontSize=11, alignment=TA_LEFT,
        fontName='Helvetica-Bold', textColor=colors.white,
        leftIndent=6, spaceAfter=0
    )
    cat_count_style = ParagraphStyle(
        'CatCount', fontSize=8, alignment=TA_LEFT,
        textColor=colors.HexColor('#d1fae5'),
        leftIndent=6
    )
    badge_style = ParagraphStyle(
        'PBadge', fontSize=7, alignment=TA_CENTER,
        textColor=colors.HexColor('#E6C200'), fontName='Helvetica-Bold',
    )
    wholesale_badge_style = ParagraphStyle(
        'PWholesaleBadge', fontSize=7, alignment=TA_CENTER,
        textColor=colors.HexColor('#92400e'), fontName='Helvetica-Bold',
    )

    elements = []

    # Document header
    elements.append(Paragraph('Product Barcode Sheet', title_style))
    try:
        cols_param = int(request.GET.get('columns', '2'))
    except (TypeError, ValueError):
        cols_param = 2
    items_per_row = max(1, min(4, cols_param))

    if items_per_row <= 2:
        bc_module_height, img_height = 12.0, 0.9 * inch
    elif items_per_row == 3:
        bc_module_height, img_height = 9.0, 0.62 * inch
    else:
        bc_module_height, img_height = 7.0, 0.48 * inch

    products_list = list(products)
    flat_barcode_entries = None
    single_product_base_entries = None
    single_product_page_count = 1
    single_product_label_type = 'both'
    single_product_cells_first = 0
    single_product_cells_continued = 0
    fill_single_product_page = (
        not grouped_output
        and bc_mode == 'specific'
        and len(products_list) == 1
    )
    if fill_single_product_page:
        label_type = (request.GET.get('label_type') or 'both').strip().lower()
        if label_type not in ('both', 'piece', 'wholesale'):
            label_type = 'both'
        single_product_label_type = label_type
        single_product_page_count = _parse_barcode_page_count(request.GET.get('page_count'))

        single_product_base_entries = _filter_barcode_entries_by_type(
            _barcode_entries_for_products(products_list),
            single_product_label_type,
        )
        if single_product_base_entries:
            estimate_kwargs = dict(
                page_height=page_height,
                top_margin=top_margin,
                bottom_margin=margin,
                items_per_row=items_per_row,
                img_height=img_height,
            )
            single_product_cells_first = _estimate_barcode_cells_per_page(
                **estimate_kwargs,
                header_block=BARCODE_SHEET_HEADER_FIRST,
            )
            single_product_cells_continued = _estimate_barcode_cells_per_page(
                **estimate_kwargs,
                header_block=BARCODE_SHEET_HEADER_CONTINUED,
            )
        else:
            fill_single_product_page = False

    filter_parts = [
        filter_label,
        f'Paper: 8.5×13 in',
        f'Price Order: {"Low to High" if price_order == "asc" else "High to Low"}',
        f'Barcodes per row: {items_per_row}',
        f'Generated: {timezone.now().strftime("%b %d, %Y %I:%M %p")}',
    ]
    if fill_single_product_page:
        filter_parts.append(f'Pages: {single_product_page_count}')
        filter_parts.append(f'Labels: {_barcode_label_type_label(single_product_label_type)}')
    elements.append(Paragraph('  |  '.join(filter_parts), sub_style))
    elements.append(Spacer(1, 6))

    col_width = (page_width - margin * 2) / items_per_row
    total_table_width = col_width * items_per_row
    img_width = col_width * 0.92

    def build_barcode_cell(entry):
        """Build a single printable barcode cell (piece or wholesale)."""
        try:
            bc_class = barcode_lib.get_barcode_class('code128')
            bc_buf = io.BytesIO()
            bc_instance = bc_class(entry['barcode'], writer=ImageWriter())
            bc_instance.write(bc_buf, options={
                'write_text': False,
                'module_height': bc_module_height,
                'quiet_zone': 2.0,
            })
            bc_buf.seek(0)
            bc_img = RLImage(bc_buf, width=img_width, height=img_height)
        except Exception:
            bc_img = Spacer(img_width, img_height)

        price_str = f'PHP {float(entry["price"]):,.2f}'
        product_name = entry['display_name'][:60]
        badge_text = entry.get('badge') or ('Wholesale' if entry.get('is_wholesale') else 'Piece')
        badge_para_style = wholesale_badge_style if entry.get('is_wholesale') else badge_style
        return [
            Paragraph(product_name, name_style),
            Paragraph(badge_text, badge_para_style),
            Spacer(1, 3),
            bc_img,
            Paragraph(entry['barcode'], code_style),
            Spacer(1, 2),
            Paragraph(price_str, price_style),
            Paragraph(entry['category_name'], cat_style),
        ]

    no_style = ParagraphStyle('NoResult', fontSize=11, alignment=TA_CENTER, textColor=colors.grey)

    page_title_style = ParagraphStyle(
        'PageTitle',
        parent=sub_style,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#166534'),
        spaceAfter=6,
    )

    def append_barcode_grid_table(entry_list, *, pad_partial_row=True):
        row_cells = []
        all_rows = []
        for entry in entry_list:
            row_cells.append(build_barcode_cell(entry))
            if len(row_cells) == items_per_row:
                all_rows.append(list(row_cells))
                row_cells = []
        if row_cells:
            if pad_partial_row:
                while len(row_cells) < items_per_row:
                    row_cells.append('')
            all_rows.append(row_cells)
        if not all_rows:
            return False
        table = Table(all_rows, colWidths=[col_width] * items_per_row)
        table.setStyle(TableStyle([
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ]))
        elements.append(table)
        return True

    if grouped_output:
        # Group products by category and render each group with a header banner
        from itertools import groupby
        products_list = list(products)
        if not products_list:
            elements.append(Paragraph('No products found.', no_style))
        else:
            def cat_key(p):
                return p.category.name if p.category else 'Uncategorized'

            for cat_name, group in groupby(products_list, key=cat_key):
                group_products = list(group)
                group_entries = _barcode_entries_for_products(group_products)

                # --- Category banner row (full width) ---
                banner_content = [
                    Paragraph(f'  {cat_name.upper()}', cat_header_style),
                    Paragraph(
                        f'  {len(group_entries)} barcode{"s" if len(group_entries) != 1 else ""}',
                        cat_count_style,
                    ),
                ]
                banner_table = Table(
                    [[banner_content]],
                    colWidths=[total_table_width],
                )
                banner_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#ED1C24')),
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#E6C200')),
                    ('TOPPADDING', (0, 0), (-1, -1), 7),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
                    ('LEFTPADDING', (0, 0), (-1, -1), 10),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(banner_table)

                # --- Product barcode rows for this category ---
                row_cells = []
                all_rows = []
                for entry in group_entries:
                    row_cells.append(build_barcode_cell(entry))
                    if len(row_cells) == items_per_row:
                        all_rows.append(list(row_cells))
                        row_cells = []
                if row_cells:
                    while len(row_cells) < items_per_row:
                        row_cells.append('')
                    all_rows.append(row_cells)

                prod_table = Table(all_rows, colWidths=[col_width] * items_per_row)
                prod_table.setStyle(TableStyle([
                    ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#ED1C24')),
                    ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                ]))
                elements.append(prod_table)
                elements.append(Spacer(1, 10))  # gap between categories
    else:
        if fill_single_product_page and single_product_base_entries:
            for page_idx in range(single_product_page_count):
                if page_idx > 0:
                    elements.append(PageBreak())
                    elements.append(Paragraph(
                        f'Page {page_idx + 1} of {single_product_page_count}',
                        page_title_style,
                    ))
                    elements.append(Spacer(1, 4))
                    target_cells = single_product_cells_continued
                else:
                    target_cells = single_product_cells_first
                page_entries = _repeat_entries_to_fill_page(
                    single_product_base_entries, target_cells
                )
                if not append_barcode_grid_table(page_entries, pad_partial_row=False):
                    elements.append(Paragraph(
                        'No barcodes available for the selected label type.',
                        no_style,
                    ))
                    break
        else:
            barcode_entries = flat_barcode_entries
            if barcode_entries is None:
                barcode_entries = _barcode_entries_for_products(products_list)
            if not append_barcode_grid_table(barcode_entries, pad_partial_row=True):
                elements.append(Paragraph('No products found for the selected filters.', no_style))

    doc.build(elements)
    buffer.seek(0)

    col_suffix = f'_cols{items_per_row}'
    if grouped_output:
        filename = f'product_barcodes_by_category{col_suffix}.pdf'
    elif product_ids:
        filename = f'product_barcodes_selected_{len(product_ids)}_products{col_suffix}.pdf'
    else:
        filename = f'product_barcodes_{filename_label}{col_suffix}.pdf'
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    # Signal the browser that the download is complete so JS can stop the spinner
    download_token = request.GET.get('download_token', '')
    if download_token:
        set_secure_cookie(
            response,
            'bc_download_done',
            download_token,
            max_age=60,
            # JS polling on inventory.html needs to read this cookie.
            httponly=False,
            # Allow local HTTP (127.0.0.1) while still using secure cookies on HTTPS.
            secure=request.is_secure(),
            samesite='Lax',
        )
    return response


# ---------------------------------------------------------------------------
# Staff / Cashier Individual Sales Report
# ---------------------------------------------------------------------------

STAFF_SALE_ROLE_SLUGS = ('cashier', 'staff', 'admin')


def _staff_sales_date_range_from_request(request):
    """Parse date_from / date_to GET params into aware datetime bounds."""
    today = _store_local_today()
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    try:
        date_from = date_type.fromisoformat(date_from_str)
    except ValueError:
        date_from = date_type(today.year, today.month, 1)
    try:
        date_to = date_type.fromisoformat(date_to_str)
    except ValueError:
        date_to = today

    current_tz = timezone.get_current_timezone()
    range_start_aware = timezone.make_aware(
        datetime.combine(date_from, datetime.min.time()), current_tz
    )
    range_end_aware = timezone.make_aware(
        datetime.combine(date_to + timedelta(days=1), datetime.min.time()), current_tz
    )
    return date_from, date_to, range_start_aware, range_end_aware


def _staff_members_qs():
    return (
        Member.objects.filter(
            member_role__slug__in=STAFF_SALE_ROLE_SLUGS,
            is_active=True,
        )
        .select_related('member_role', 'member_type', 'user')
        .order_by('member_role__sort_order', 'first_name', 'last_name')
    )


def _build_staff_overview_data(range_start_aware, range_end_aware):
    staff_data = []
    for member in _staff_members_qs():
        if not member.user_id:
            staff_data.append({
                'member': member,
                'total_revenue': Decimal('0.00'),
                'txn_count': 0,
            })
            continue
        qs = Transaction.objects.filter(
            processed_by=member.user,
            status__in=DASHBOARD_SALE_STATUSES,
            created_at__gte=range_start_aware,
            created_at__lt=range_end_aware,
        )
        agg = qs.aggregate(
            total_revenue=Coalesce(Sum('total_amount'), Decimal('0.00')),
            txn_count=Count('id'),
        )
        staff_data.append({
            'member': member,
            'total_revenue': agg['total_revenue'],
            'txn_count': agg['txn_count'],
        })
    staff_data.sort(key=lambda x: x['total_revenue'], reverse=True)
    return staff_data


def _staff_member_transactions_qs(staff_member, range_start_aware, range_end_aware):
    if not staff_member.user_id:
        return Transaction.objects.none()
    return (
        Transaction.objects.filter(
            processed_by=staff_member.user,
            status__in=DASHBOARD_SALE_STATUSES,
            created_at__gte=range_start_aware,
            created_at__lt=range_end_aware,
        )
        .select_related('member', 'walk_in_customer')
        .order_by('-created_at')
    )


def _staff_sales_export_format(request):
    requested_format = (request.GET.get('format') or 'excel').strip().lower()
    if requested_format not in ('pdf', 'excel'):
        requested_format = 'excel'
    return requested_format


def _staff_sales_date_slug(date_from, date_to):
    return f'{date_from.strftime("%Y%m%d")}_to_{date_to.strftime("%Y%m%d")}'


def _staff_sales_safe_filename(name):
    return re.sub(r'[^\w\-]+', '_', (name or 'staff').strip())[:60]


def _staff_sales_excel_styles():
    header_fill = PatternFill(start_color='1F7A3A', end_color='1F7A3A', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D1D5DB')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_fill, header_font, thin_border


def _staff_sales_write_excel_header(ws, title, date_from, date_to, user_label, extra_lines=None):
    gen_at = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True, color='166534')
    ws['A2'] = f'Period: {date_from.isoformat()} to {date_to.isoformat()}'
    ws['A3'] = f'Generated: {gen_at} — {user_label}'
    row = 4
    if extra_lines:
        for line in extra_lines:
            ws.cell(row=row, column=1, value=line)
            row += 1
    return row


def _staff_sales_excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _staff_sales_pdf_response(buffer, filename):
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _payment_method_label(method):
    return {'cash': 'Cash', 'debit': 'Debit', 'credit': 'Credit'}.get(method, (method or '').title())


STAFF_SALES_FEATURE_ENABLED = False


def _redirect_staff_sales_disabled():
    """Staff Sales is hidden from the console until it is needed again."""
    return redirect('dashboard')


@login_required
def staff_sales_report(request):
    """Overview: list all cashier & staff members with their sales KPIs."""
    if not STAFF_SALES_FEATURE_ENABLED:
        return _redirect_staff_sales_disabled()
    if not is_admin_user(request.user):
        messages.warning(request, 'You do not have permission to access this page.')
        return redirect('kiosk_home')

    kiosk_config = KioskConfig.get()
    date_from, date_to, range_start_aware, range_end_aware = _staff_sales_date_range_from_request(request)
    staff_data = _build_staff_overview_data(range_start_aware, range_end_aware)
    grand_revenue = sum(s['total_revenue'] for s in staff_data)
    grand_txns = sum(s['txn_count'] for s in staff_data)

    context = {
        'staff_data': staff_data,
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
        'grand_revenue': grand_revenue,
        'grand_txns': grand_txns,
        'kiosk_system_name': kiosk_config.system_name if kiosk_config else 'Admin',
        'nav_active': 'staff_sales',
        **admin_role_badge_context(request),
    }
    return render(request, 'admin_panel/staff_sales_report.html', context)


@login_required
@require_http_methods(['GET'])
def export_staff_sales_overview(request):
    """Download all staff/cashier sales summary as Excel or PDF."""
    if not STAFF_SALES_FEATURE_ENABLED:
        return _redirect_staff_sales_disabled()
    if not is_admin_user(request.user):
        messages.warning(request, 'You do not have permission to export this report.')
        return redirect('kiosk_home')

    kiosk_config = KioskConfig.get()
    store_name = kiosk_config.system_name if kiosk_config else 'Admin'
    date_from, date_to, range_start_aware, range_end_aware = _staff_sales_date_range_from_request(request)
    staff_data = _build_staff_overview_data(range_start_aware, range_end_aware)
    grand_revenue = sum(s['total_revenue'] for s in staff_data)
    grand_txns = sum(s['txn_count'] for s in staff_data)
    export_format = _staff_sales_export_format(request)
    date_slug = _staff_sales_date_slug(date_from, date_to)
    user_label = request.user.get_full_name() or request.user.username

    if export_format == 'excel':
        header_fill, header_font, thin_border = _staff_sales_excel_styles()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Staff Sales'
        hdr_row = _staff_sales_write_excel_header(
            ws,
            f'Staff & Cashier Sales Overview — {store_name}',
            date_from,
            date_to,
            user_label,
            extra_lines=[
                f'Total revenue (PHP): {float(grand_revenue):,.2f}',
                f'Total transactions: {grand_txns}',
                f'Active staff: {len(staff_data)}',
            ],
        )
        hdr_row += 1
        headers = ['#', 'Staff Name', 'Username', 'Role', 'Transactions', 'Total Revenue (PHP)', 'Share (%)']
        for col, val in enumerate(headers, start=1):
            c = ws.cell(row=hdr_row, column=col, value=val)
            c.fill = header_fill
            c.font = header_font
            c.border = thin_border
            c.alignment = Alignment(horizontal='center' if col in (1, 5, 7) else 'left', vertical='center')

        for idx, row in enumerate(staff_data, start=1):
            member = row['member']
            share_pct = (
                float(row['total_revenue'] / grand_revenue * 100)
                if grand_revenue > 0 else 0.0
            )
            ws.append([
                idx,
                f'{member.first_name} {member.last_name}'.strip(),
                member.username or '—',
                member.member_role.name if member.member_role_id else '—',
                row['txn_count'],
                float(row['total_revenue']),
                round(share_pct, 1),
            ])

        total_row = hdr_row + len(staff_data) + 1
        ws.cell(row=total_row, column=1, value='')
        ws.cell(row=total_row, column=2, value='Grand Total')
        ws.cell(row=total_row, column=5, value=grand_txns)
        ws.cell(row=total_row, column=6, value=float(grand_revenue))
        for r in range(hdr_row + 1, total_row + 1):
            for col in range(1, 8):
                cell = ws.cell(row=r, column=col)
                cell.border = thin_border
                if col in (5, 6, 7):
                    cell.number_format = '#,##0.00' if col != 5 else '0'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 1:
                    cell.alignment = Alignment(horizontal='center')
        for col, width in zip('ABCDEFG', [6, 28, 18, 14, 14, 20, 12]):
            ws.column_dimensions[col].width = width

        return _staff_sales_excel_response(wb, f'staff_sales_overview_{date_slug}.xlsx')

    pdf_primary_dark = colors.HexColor('#C4121A')
    pdf_heading = colors.HexColor('#166534')
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'StaffOverviewTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=pdf_primary_dark,
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
    )
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    elements = []
    elements.append(Paragraph('Staff &amp; Cashier Sales Overview', title_style))
    elements.append(Paragraph(f'<i>{escape(store_name)}</i>', styles['Normal']))
    elements.append(Spacer(1, 0.12 * inch))
    elements.append(Paragraph(
        f'Period: {date_from.isoformat()} to {date_to.isoformat()}',
        styles['Normal'],
    ))
    gen_at = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
    elements.append(Paragraph(f'Generated: {escape(gen_at)} — {escape(user_label)}', styles['Normal']))
    elements.append(Spacer(1, 0.1 * inch))
    elements.append(Paragraph(
        f'<b>Total revenue:</b> ₱{float(grand_revenue):,.2f} &nbsp;|&nbsp; '
        f'<b>Transactions:</b> {grand_txns} &nbsp;|&nbsp; '
        f'<b>Staff:</b> {len(staff_data)}',
        styles['Normal'],
    ))
    elements.append(Spacer(1, 0.18 * inch))

    table_data = [['#', 'Staff', 'Role', 'Txns', 'Revenue (PHP)', 'Share']]
    for idx, row in enumerate(staff_data, start=1):
        member = row['member']
        share_pct = (
            f'{float(row["total_revenue"] / grand_revenue * 100):.1f}%'
            if grand_revenue > 0 else '—'
        )
        table_data.append([
            str(idx),
            escape(f'{member.first_name} {member.last_name}'.strip()),
            escape(member.member_role.name if member.member_role_id else '—'),
            str(row['txn_count']),
            f'{float(row["total_revenue"]):,.2f}',
            share_pct,
        ])
    table_data.append(['', 'Grand Total', '', str(grand_txns), f'{float(grand_revenue):,.2f}', ''])

    tbl = Table(table_data, colWidths=[28, 130, 70, 42, 90, 50], repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), pdf_heading),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (3, 0), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d1d5db')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(tbl)
    doc.build(elements)
    buffer.seek(0)
    return _staff_sales_pdf_response(buffer, f'staff_sales_overview_{date_slug}.pdf')


@login_required
def staff_sales_detail(request, member_id):
    """Detail: individual transactions for a specific cashier/staff member."""
    if not STAFF_SALES_FEATURE_ENABLED:
        return _redirect_staff_sales_disabled()
    if not is_admin_user(request.user):
        messages.warning(request, 'You do not have permission to access this page.')
        return redirect('kiosk_home')

    kiosk_config = KioskConfig.get()
    staff_member = get_object_or_404(
        Member.objects.select_related('member_role', 'member_type'),
        pk=member_id,
        member_role__slug__in=STAFF_SALE_ROLE_SLUGS,
    )

    date_from, date_to, range_start_aware, range_end_aware = _staff_sales_date_range_from_request(request)
    txn_qs = _staff_member_transactions_qs(staff_member, range_start_aware, range_end_aware)

    agg = txn_qs.aggregate(
        total_revenue=Coalesce(Sum('total_amount'), Decimal('0.00')),
        total_vat=Coalesce(Sum('vat_amount'), Decimal('0.00')),
        txn_count=Count('id'),
        avg_order=Coalesce(Avg('total_amount'), Decimal('0.00')),
    )

    payment_breakdown = list(
        txn_qs.values('payment_method')
        .annotate(count=Count('id'), revenue=Coalesce(Sum('total_amount'), Decimal('0.00')))
        .order_by('-revenue')
    )

    paginator = Paginator(txn_qs, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'staff_member': staff_member,
        'page_obj': page_obj,
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
        'total_revenue': agg['total_revenue'],
        'total_vat': agg['total_vat'],
        'txn_count': agg['txn_count'],
        'avg_order': agg['avg_order'],
        'payment_breakdown': payment_breakdown,
        'kiosk_system_name': kiosk_config.system_name if kiosk_config else 'Admin',
        'nav_active': 'staff_sales',
        **admin_role_badge_context(request),
    }
    return render(request, 'admin_panel/staff_sales_detail.html', context)


@login_required
@require_http_methods(['GET'])
def export_staff_sales_detail(request, member_id):
    """Download individual staff/cashier sales summary and transactions as Excel or PDF."""
    if not STAFF_SALES_FEATURE_ENABLED:
        return _redirect_staff_sales_disabled()
    if not is_admin_user(request.user):
        messages.warning(request, 'You do not have permission to export this report.')
        return redirect('kiosk_home')

    kiosk_config = KioskConfig.get()
    store_name = kiosk_config.system_name if kiosk_config else 'Admin'
    staff_member = get_object_or_404(
        Member.objects.select_related('member_role', 'member_type'),
        pk=member_id,
        member_role__slug__in=STAFF_SALE_ROLE_SLUGS,
    )

    date_from, date_to, range_start_aware, range_end_aware = _staff_sales_date_range_from_request(request)
    txn_qs = _staff_member_transactions_qs(staff_member, range_start_aware, range_end_aware)
    transactions = list(txn_qs)

    agg = txn_qs.aggregate(
        total_revenue=Coalesce(Sum('total_amount'), Decimal('0.00')),
        total_vat=Coalesce(Sum('vat_amount'), Decimal('0.00')),
        txn_count=Count('id'),
        avg_order=Coalesce(Avg('total_amount'), Decimal('0.00')),
    )
    payment_breakdown = list(
        txn_qs.values('payment_method')
        .annotate(count=Count('id'), revenue=Coalesce(Sum('total_amount'), Decimal('0.00')))
        .order_by('-revenue')
    )

    export_format = _staff_sales_export_format(request)
    date_slug = _staff_sales_date_slug(date_from, date_to)
    user_label = request.user.get_full_name() or request.user.username
    staff_name = f'{staff_member.first_name} {staff_member.last_name}'.strip()
    file_slug = _staff_sales_safe_filename(staff_name)

    if export_format == 'excel':
        header_fill, header_font, thin_border = _staff_sales_excel_styles()
        wb = Workbook()

        ws_summary = wb.active
        ws_summary.title = 'Summary'
        hdr_row = _staff_sales_write_excel_header(
            ws_summary,
            f'Staff Sales Summary — {staff_name}',
            date_from,
            date_to,
            user_label,
            extra_lines=[
                f'Store: {store_name}',
                f'Role: {staff_member.member_role.name if staff_member.member_role_id else "—"}',
                f'Username: {staff_member.username or "—"}',
                f'Total revenue (PHP): {float(agg["total_revenue"]):,.2f}',
                f'Transactions: {agg["txn_count"]}',
                f'Avg order (PHP): {float(agg["avg_order"]):,.2f}',
                f'VAT collected (PHP): {float(agg["total_vat"]):,.2f}',
            ],
        )
        if payment_breakdown:
            hdr_row += 1
            ws_summary.cell(row=hdr_row, column=1, value='Payment method breakdown').font = Font(bold=True)
            hdr_row += 1
            for pm in payment_breakdown:
                ws_summary.cell(
                    row=hdr_row,
                    column=1,
                    value=f'{_payment_method_label(pm["payment_method"])}: '
                          f'₱{float(pm["revenue"]):,.2f} ({pm["count"]} txn)',
                )
                hdr_row += 1

        ws_txn = wb.create_sheet('Transactions')
        txn_hdr = 1
        ws_txn.cell(row=txn_hdr, column=1, value=f'Transactions — {staff_name}').font = Font(
            size=12, bold=True, color='166534'
        )
        txn_hdr = 3
        txn_headers = ['#', 'TXN #', 'Date & Time', 'Customer', 'Payment', 'Status', 'Amount (PHP)']
        for col, val in enumerate(txn_headers, start=1):
            c = ws_txn.cell(row=txn_hdr, column=col, value=val)
            c.fill = header_fill
            c.font = header_font
            c.border = thin_border
            c.alignment = Alignment(horizontal='center' if col in (1, 2) else 'left', vertical='center')

        for idx, txn in enumerate(transactions, start=1):
            ws_txn.append([
                idx,
                txn.transaction_number,
                timezone.localtime(txn.created_at).strftime('%Y-%m-%d %H:%M'),
                txn.customer_display_name,
                _payment_method_label(txn.payment_method),
                txn.get_status_display(),
                float(txn.total_amount or 0),
            ])

        for r in range(txn_hdr + 1, ws_txn.max_row + 1):
            for col in range(1, 8):
                cell = ws_txn.cell(row=r, column=col)
                cell.border = thin_border
                if col == 7:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                elif col == 1:
                    cell.alignment = Alignment(horizontal='center')
        for col, width in zip('ABCDEFG', [6, 16, 18, 28, 12, 16, 16]):
            ws_txn.column_dimensions[col].width = width

        return _staff_sales_excel_response(wb, f'staff_sales_{file_slug}_{date_slug}.xlsx')

    pdf_primary_dark = colors.HexColor('#C4121A')
    pdf_heading = colors.HexColor('#166534')
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'StaffDetailTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=pdf_primary_dark,
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
    )
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    elements = []
    elements.append(Paragraph(f'Staff Sales Report — {escape(staff_name)}', title_style))
    elements.append(Paragraph(f'<i>{escape(store_name)}</i>', styles['Normal']))
    elements.append(Spacer(1, 0.1 * inch))
    elements.append(Paragraph(
        f'Role: {escape(staff_member.member_role.name if staff_member.member_role_id else "—")} &nbsp;|&nbsp; '
        f'Username: {escape(staff_member.username or "—")}',
        styles['Normal'],
    ))
    elements.append(Paragraph(
        f'Period: {date_from.isoformat()} to {date_to.isoformat()}',
        styles['Normal'],
    ))
    gen_at = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M')
    elements.append(Paragraph(f'Generated: {escape(gen_at)} — {escape(user_label)}', styles['Normal']))
    elements.append(Spacer(1, 0.12 * inch))
    elements.append(Paragraph(
        f'<b>Revenue:</b> ₱{float(agg["total_revenue"]):,.2f} &nbsp;|&nbsp; '
        f'<b>Transactions:</b> {agg["txn_count"]} &nbsp;|&nbsp; '
        f'<b>Avg order:</b> ₱{float(agg["avg_order"]):,.2f} &nbsp;|&nbsp; '
        f'<b>VAT:</b> ₱{float(agg["total_vat"]):,.2f}',
        styles['Normal'],
    ))
    if payment_breakdown:
        pm_parts = [
            f'{escape(_payment_method_label(pm["payment_method"]))}: ₱{float(pm["revenue"]):,.2f} ({pm["count"]})'
            for pm in payment_breakdown
        ]
        elements.append(Paragraph(f'<b>Payment:</b> {" · ".join(pm_parts)}', styles['Normal']))
    elements.append(Spacer(1, 0.18 * inch))

    table_data = [['#', 'TXN #', 'Date', 'Customer', 'Pay', 'Status', 'Amount']]
    for idx, txn in enumerate(transactions, start=1):
        table_data.append([
            str(idx),
            escape(txn.transaction_number or ''),
            timezone.localtime(txn.created_at).strftime('%Y-%m-%d %H:%M'),
            escape((txn.customer_display_name or '')[:28]),
            escape(_payment_method_label(txn.payment_method)[:6]),
            escape(txn.get_status_display()[:14]),
            f'{float(txn.total_amount or 0):,.2f}',
        ])

    tbl = Table(
        table_data,
        colWidths=[22, 62, 72, 90, 38, 58, 58],
        repeatRows=1,
    )
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), pdf_heading),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (6, 0), (6, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d1d5db')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    elements.append(tbl)
    doc.build(elements)
    buffer.seek(0)
    return _staff_sales_pdf_response(buffer, f'staff_sales_{file_slug}_{date_slug}.pdf')

@login_required
def website_audit_trail(request):
    """Admin-only monitor of site-wide activity (login, logout, staff actions)."""
    if not is_admin_user(request.user):
        messages.warning(request, 'You do not have permission to access this page.')
        return redirect('kiosk_home')

    kiosk_config = KioskConfig.get()
    qs = WebsiteAuditLog.objects.select_related('actor').all()

    action_filter = (request.GET.get('action') or '').strip()
    actor_q = (request.GET.get('q') or '').strip()
    date_from_raw = (request.GET.get('date_from') or '').strip()
    date_to_raw = (request.GET.get('date_to') or '').strip()

    if action_filter and action_filter in dict(WebsiteAuditLog.Action.choices):
        qs = qs.filter(action=action_filter)

    if actor_q:
        qs = qs.filter(
            Q(actor_label__icontains=actor_q)
            | Q(description__icontains=actor_q)
            | Q(request_path__icontains=actor_q)
            | Q(actor__username__icontains=actor_q)
        )

    today = timezone.localdate()
    date_from = today - timedelta(days=7)
    date_to = today
    try:
        if date_from_raw:
            date_from = date_type.fromisoformat(date_from_raw)
    except ValueError:
        pass
    try:
        if date_to_raw:
            date_to = date_type.fromisoformat(date_to_raw)
    except ValueError:
        pass
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    range_start = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
    range_end = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))
    qs = qs.filter(created_at__gte=range_start, created_at__lte=range_end)

    total_events = qs.count()
    login_count = qs.filter(action=WebsiteAuditLog.Action.LOGIN).count()
    logout_count = qs.filter(action=WebsiteAuditLog.Action.LOGOUT).count()
    failed_count = qs.filter(action=WebsiteAuditLog.Action.LOGIN_FAILED).count()

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'audit_logs': page_obj.object_list,
        'action_choices': WebsiteAuditLog.Action.choices,
        'action_filter': action_filter,
        'actor_q': actor_q,
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
        'total_events': total_events,
        'login_count': login_count,
        'logout_count': logout_count,
        'failed_count': failed_count,
        'kiosk_system_name': kiosk_config.system_name if kiosk_config else 'Admin',
        'nav_active': 'audit',
        **admin_role_badge_context(request),
    }
    return render(request, 'admin_panel/audit_trail.html', context)

